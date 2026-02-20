"""
Build: Sistema de Ventas con Código Único (Bs. 89)
No Somos Ignorantes v1.0
Sales system with unique product codes, inventory tracking, VLOOKUP auto-fill,
executive dashboard with KPIs and charts.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule, DataBarRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

wb = Workbook()

# ============================================================
# COLORS — Executive Dashboard Palette
# ============================================================
# Brand
GOLD = "D4AF37"
DARK_BG = "1A1A2E"
CARD_BG = "16213E"
ACCENT_BLUE = "0F3460"

# Base
WHITE = "FFFFFF"
BLACK = "000000"

# Input/Output
LIGHT_YELLOW = "FFF9C4"
LIGHT_GREEN = "E8F5E9"

# Executive Dashboard palette
RED_CORAL = "E74C3C"
TURQUOISE = "16A085"
ORANGE = "E67E22"
PURPLE = "8E44AD"
LIGHT_GRAY_CHART = "BDC3C7"

# Traffic light
GREEN_OK = "27AE60"
YELLOW_WARN = "F1C40F"
RED_ALERT = "E74C3C"

# Grays
GRAY_50 = "F9FAFB"
GRAY_100 = "F3F4F6"
GRAY_200 = "E5E7EB"
GRAY_300 = "D1D5DB"
GRAY_400 = "9CA3AF"
GRAY_500 = "6B7280"
GRAY_600 = "4B5563"
GRAY_700 = "374151"
GRAY_800 = "1F2937"
GRAY_900 = "111827"

# Category colors
CAT_COLORS = [TURQUOISE, RED_CORAL, ORANGE, PURPLE, "3498DB", "1ABC9C", "F39C12", "2ECC71"]

# ============================================================
# STYLE HELPERS
# ============================================================
thin_border = Border(
    left=Side(style='thin', color=GRAY_300),
    right=Side(style='thin', color=GRAY_300),
    top=Side(style='thin', color=GRAY_300),
    bottom=Side(style='thin', color=GRAY_300)
)
gold_border = Border(
    left=Side(style='thin', color=GOLD),
    right=Side(style='thin', color=GOLD),
    top=Side(style='thin', color=GOLD),
    bottom=Side(style='thin', color=GOLD)
)
subtle_border = Border(
    left=Side(style='thin', color="E0E0E0"),
    right=Side(style='thin', color="E0E0E0"),
    top=Side(style='thin', color="E0E0E0"),
    bottom=Side(style='thin', color="E0E0E0")
)
bottom_border = Border(
    bottom=Side(style='thin', color=GRAY_300)
)

def style_title(cell, size=18):
    """Main dashboard title — sans-serif, bold, black on white"""
    cell.font = Font(name='Calibri', bold=True, color=BLACK, size=size)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def style_kpi_subtitle(cell):
    """KPI subtitle — 10pt, dark gray"""
    cell.font = Font(name='Calibri', size=10, color="4A4A4A")
    cell.alignment = Alignment(horizontal='left', vertical='center')

def style_kpi_number(cell, size=24):
    """Large KPI number — bold, black"""
    cell.font = Font(name='Calibri', bold=True, color=BLACK, size=size)
    cell.alignment = Alignment(horizontal='right', vertical='center')

def style_kpi_label(cell):
    """Small label like Actual/Target — gray"""
    cell.font = Font(name='Calibri', size=9, color="808080")
    cell.alignment = Alignment(horizontal='left', vertical='center')

def style_header(cell, bg=DARK_BG, fg=WHITE, size=11):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

def style_header_gold(cell, size=12):
    cell.font = Font(name='Calibri', bold=True, color=WHITE, size=size)
    cell.fill = PatternFill('solid', fgColor=GOLD)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

def style_input(cell):
    cell.font = Font(name='Calibri', size=11, color="0000FF")
    cell.fill = PatternFill('solid', fgColor=LIGHT_YELLOW)
    cell.border = gold_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

def style_output(cell, bold=False):
    cell.font = Font(name='Calibri', size=11, color=GRAY_900, bold=bold)
    cell.fill = PatternFill('solid', fgColor=LIGHT_GREEN)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='right', vertical='center')

def style_label(cell, bold=False, size=11):
    cell.font = Font(name='Calibri', size=size, color=GRAY_700, bold=bold)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def style_cell(cell, bold=False, color=BLACK, size=11, align='left'):
    cell.font = Font(name='Calibri', size=size, color=color, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = thin_border

def style_banner(cell, bg_color, fg=WHITE, size=11):
    """Colored banner/ribbon for totals"""
    cell.font = Font(name='Calibri', bold=True, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg_color)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

def fill_range_white(ws, min_row, max_row, min_col, max_col):
    """Fill range with white background"""
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=WHITE)

def unlock_cell(cell):
    """Mark cell as unlocked for user input"""
    cell.protection = cell.protection.copy(locked=False)

def lock_cell(cell):
    """Mark cell as locked"""
    cell.protection = cell.protection.copy(locked=True)

# ============================================================
# SHEET 1: DASHBOARD
# ============================================================
ws_dash = wb.active
ws_dash.title = "Dashboard"
ws_dash.sheet_properties.tabColor = RED_CORAL

# Page setup
ws_dash.sheet_view.showGridLines = False

# Column widths
dash_widths = {1: 3, 2: 22, 3: 18, 4: 18, 5: 3, 6: 22, 7: 18, 8: 18, 9: 3}
for c, w in dash_widths.items():
    ws_dash.column_dimensions[get_column_letter(c)].width = w

# White background
fill_range_white(ws_dash, 1, 55, 1, 9)

# Row 2: Main title
ws_dash.merge_cells('B2:H2')
c = ws_dash['B2']
style_title(c, size=20)
c.value = "SISTEMA DE VENTAS — PANEL EJECUTIVO"

# Row 3: Subtitle
ws_dash.merge_cells('B3:H3')
c = ws_dash['B3']
c.font = Font(name='Calibri', size=10, color=GRAY_500)
c.value = "No Somos Ignorantes  |  v1.0  |  Resumen en tiempo real"

# Row 4: separator
ws_dash.row_dimensions[4].height = 6

# ── KPI ROW 1 (Row 5-10): Ventas del Mes + Unidades Vendidas ──
# KPI 1: Ventas del Mes (B5:D10)
ws_dash.merge_cells('B5:D5')
c = ws_dash['B5']
style_banner(c, RED_CORAL, WHITE, 10)
c.value = "VENTAS DEL MES (Bs.)"

ws_dash.merge_cells('B6:D6')
c = ws_dash['B6']
style_kpi_number(c, 28)
c.value = '=IFERROR(SUMPRODUCT((MONTH(Ventas!A4:A203)=MONTH(TODAY()))*(YEAR(Ventas!A4:A203)=YEAR(TODAY()))*Ventas!H4:H203),0)'
c.number_format = '#,##0.00'

ws_dash['B7'].value = "Meta mensual:"
style_kpi_label(ws_dash['B7'])
ws_dash['C7'].value = 50000
ws_dash['C7'].number_format = '#,##0'
style_input(ws_dash['C7'])
unlock_cell(ws_dash['C7'])

ws_dash['B8'].value = "Cumplimiento:"
style_kpi_label(ws_dash['B8'])
ws_dash.merge_cells('C8:D8')
c = ws_dash['C8']
c.value = '=IFERROR(B6/C7,0)'
c.number_format = '0%'
c.font = Font(name='Calibri', bold=True, size=14, color=BLACK)
c.alignment = Alignment(horizontal='left', vertical='center')

# KPI 2: Unidades Vendidas (F5:H10)
ws_dash.merge_cells('F5:H5')
c = ws_dash['F5']
style_banner(c, TURQUOISE, WHITE, 10)
c.value = "UNIDADES VENDIDAS"

ws_dash.merge_cells('F6:H6')
c = ws_dash['F6']
style_kpi_number(c, 28)
c.value = '=IFERROR(SUMPRODUCT((MONTH(Ventas!A4:A203)=MONTH(TODAY()))*(YEAR(Ventas!A4:A203)=YEAR(TODAY()))*Ventas!E4:E203),0)'
c.number_format = '#,##0'

ws_dash['F7'].value = "Transacciones:"
style_kpi_label(ws_dash['F7'])
ws_dash.merge_cells('G7:H7')
c = ws_dash['G7']
c.value = '=IFERROR(SUMPRODUCT((MONTH(Ventas!A4:A203)=MONTH(TODAY()))*(YEAR(Ventas!A4:A203)=YEAR(TODAY()))*(Ventas!H4:H203>0)*1),0)'
c.number_format = '#,##0'
c.font = Font(name='Calibri', size=14, bold=True, color=BLACK)
c.alignment = Alignment(horizontal='left')

ws_dash['F8'].value = "Ticket promedio:"
style_kpi_label(ws_dash['F8'])
ws_dash.merge_cells('G8:H8')
c = ws_dash['G8']
c.value = '=IFERROR(B6/G7,"—")'
c.number_format = '#,##0.00'
c.font = Font(name='Calibri', size=14, bold=True, color=BLACK)
c.alignment = Alignment(horizontal='left')

# Row 9-10: spacing
ws_dash.row_dimensions[9].height = 4
ws_dash.row_dimensions[10].height = 4

# ── KPI ROW 2 (Row 11-16): Margen Bruto + Productos Bajo Stock ──
# KPI 3: Margen Bruto
ws_dash.merge_cells('B11:D11')
c = ws_dash['B11']
style_banner(c, ORANGE, WHITE, 10)
c.value = "GANANCIA BRUTA (Bs.)"

ws_dash.merge_cells('B12:D12')
c = ws_dash['B12']
style_kpi_number(c, 28)
# Ganancia = Ventas - Costo
c.value = ('=IFERROR(SUMPRODUCT((MONTH(Ventas!A4:A203)=MONTH(TODAY()))'
           '*(YEAR(Ventas!A4:A203)=YEAR(TODAY()))*Ventas!H4:H203)'
           '-SUMPRODUCT((MONTH(Ventas!A4:A203)=MONTH(TODAY()))'
           '*(YEAR(Ventas!A4:A203)=YEAR(TODAY()))*Ventas!E4:E203'
           '*IFERROR(VLOOKUP(Ventas!C4:C203,Productos!A4:H103,4,FALSE),0)),0)')
c.number_format = '#,##0.00'

ws_dash['B13'].value = "Margen (%):"
style_kpi_label(ws_dash['B13'])
ws_dash.merge_cells('C13:D13')
c = ws_dash['C13']
c.value = '=IFERROR(B12/B6,0)'
c.number_format = '0.0%'
c.font = Font(name='Calibri', bold=True, size=14, color=BLACK)
c.alignment = Alignment(horizontal='left')

# KPI 4: Productos con Stock Bajo
ws_dash.merge_cells('F11:H11')
c = ws_dash['F11']
style_banner(c, PURPLE, WHITE, 10)
c.value = "ALERTAS DE INVENTARIO"

ws_dash.merge_cells('F12:H12')
c = ws_dash['F12']
style_kpi_number(c, 28)
c.value = '=COUNTIF(Productos!I4:I103,"BAJO")+COUNTIF(Productos!I4:I103,"AGOTADO")'
c.number_format = '#,##0'

ws_dash['F13'].value = "Agotados:"
style_kpi_label(ws_dash['F13'])
ws_dash['G13'].value = '=COUNTIF(Productos!I4:I103,"AGOTADO")'
ws_dash['G13'].number_format = '#,##0'
ws_dash['G13'].font = Font(name='Calibri', bold=True, size=14, color=RED_ALERT)
ws_dash['G13'].alignment = Alignment(horizontal='left')

ws_dash['F14'].value = "Stock bajo:"
style_kpi_label(ws_dash['F14'])
ws_dash['G14'].value = '=COUNTIF(Productos!I4:I103,"BAJO")'
ws_dash['G14'].number_format = '#,##0'
ws_dash['G14'].font = Font(name='Calibri', bold=True, size=14, color=YELLOW_WARN)
ws_dash['G14'].alignment = Alignment(horizontal='left')

# Row 15-16: spacing
ws_dash.row_dimensions[15].height = 4
ws_dash.row_dimensions[16].height = 4

# ── TOTALS BANNER (Row 17-19) ──
ws_dash.merge_cells('B17:D17')
c = ws_dash['B17']
style_banner(c, RED_CORAL, WHITE, 9)
c.value = "TOTAL PRODUCTOS"

ws_dash.merge_cells('B18:D18')
c = ws_dash['B18']
c.value = '=COUNTA(Productos!B4:B103)'
c.number_format = '#,##0'
c.font = Font(name='Calibri', bold=True, size=18, color=BLACK)
c.alignment = Alignment(horizontal='center', vertical='center')

ws_dash.merge_cells('F17:H17')
c = ws_dash['F17']
style_banner(c, TURQUOISE, WHITE, 9)
c.value = "VALOR INVENTARIO (Bs.)"

ws_dash.merge_cells('F18:H18')
c = ws_dash['F18']
c.value = '=IFERROR(SUMPRODUCT(Productos!E4:E103*Productos!G4:G103),0)'
c.number_format = '#,##0.00'
c.font = Font(name='Calibri', bold=True, size=18, color=BLACK)
c.alignment = Alignment(horizontal='center', vertical='center')

# Row 19-20: spacing
ws_dash.row_dimensions[19].height = 6
ws_dash.row_dimensions[20].height = 6

# ── CHART DATA (hidden, Row 22-34): Top 5 Products by Sales ──
ws_dash['B21'].value = "TOP 5 PRODUCTOS POR VENTAS"
ws_dash['B21'].font = Font(name='Calibri', bold=True, size=11, color=GRAY_700)

# We'll create a helper data table for the chart in rows 22-27
ws_dash['B22'].value = "Producto"
ws_dash['C22'].value = "Ventas (Bs.)"
style_header(ws_dash['B22'], bg=GRAY_700)
style_header(ws_dash['C22'], bg=GRAY_700)

# Top products using LARGE/INDEX/MATCH (rows 23-27, top 5)
for i in range(5):
    r = 23 + i
    ws_dash.cell(row=r, column=2).value = f'=IFERROR(INDEX(Productos!B4:B103,MATCH(LARGE(Productos!J4:J103,{i+1}),Productos!J4:J103,0)),"—")'
    ws_dash.cell(row=r, column=2).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws_dash.cell(row=r, column=2).border = thin_border
    ws_dash.cell(row=r, column=3).value = f'=IFERROR(LARGE(Productos!J4:J103,{i+1}),0)'
    ws_dash.cell(row=r, column=3).number_format = '#,##0.00'
    ws_dash.cell(row=r, column=3).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws_dash.cell(row=r, column=3).border = thin_border

# Bar chart: Top 5 Products
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Top 5 Productos por Ventas"
chart1.y_axis.title = "Ventas (Bs.)"
chart1.x_axis.title = None
chart1.legend = None
data1 = Reference(ws_dash, min_col=3, min_row=22, max_row=27)
cats1 = Reference(ws_dash, min_col=2, min_row=23, max_row=27)
chart1.add_data(data1, titles_from_data=True)
chart1.set_categories(cats1)
chart1.shape = 4
chart1.width = 18
chart1.height = 12
# Style bars
s1 = chart1.series[0]
s1.graphicalProperties.solidFill = RED_CORAL
s1.graphicalProperties.line.noFill = True
ws_dash.add_chart(chart1, "B29")

# ── Ventas por Método de Pago (Pie chart) ──
ws_dash['F22'].value = "Método de Pago"
ws_dash['G22'].value = "Total (Bs.)"
style_header(ws_dash['F22'], bg=GRAY_700)
style_header(ws_dash['G22'], bg=GRAY_700)

methods = ["Efectivo", "Transferencia", "QR", "Tarjeta"]
for i, m in enumerate(methods):
    r = 23 + i
    ws_dash.cell(row=r, column=6).value = m
    ws_dash.cell(row=r, column=6).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws_dash.cell(row=r, column=6).border = thin_border
    ws_dash.cell(row=r, column=7).value = f'=IFERROR(SUMPRODUCT((Ventas!I4:I203="{m}")*Ventas!H4:H203),0)'
    ws_dash.cell(row=r, column=7).number_format = '#,##0.00'
    ws_dash.cell(row=r, column=7).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws_dash.cell(row=r, column=7).border = thin_border

chart2 = PieChart()
chart2.title = "Ventas por Método de Pago"
chart2.style = 10
data2 = Reference(ws_dash, min_col=7, min_row=22, max_row=26)
cats2 = Reference(ws_dash, min_col=6, min_row=23, max_row=26)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.width = 16
chart2.height = 12
# Color slices
pie_colors = [TURQUOISE, ORANGE, PURPLE, "3498DB"]
for idx, color in enumerate(pie_colors):
    pt = DataPoint(idx=idx)
    pt.graphicalProperties.solidFill = color
    chart2.series[0].data_points.append(pt)
chart2.dataLabels = DataLabelList()
chart2.dataLabels.showPercent = True
chart2.dataLabels.showVal = False
ws_dash.add_chart(chart2, "F29")

# ============================================================
# SHEET 2: PRODUCTOS (Product Database)
# ============================================================
ws_prod = wb.create_sheet("Productos")
ws_prod.sheet_properties.tabColor = TURQUOISE
ws_prod.sheet_view.showGridLines = False

# Column widths
prod_widths = {1: 16, 2: 30, 3: 18, 4: 16, 5: 16, 6: 12, 7: 14, 8: 14, 9: 14, 10: 16}
for c, w in prod_widths.items():
    ws_prod.column_dimensions[get_column_letter(c)].width = w

# Row 1: Title
ws_prod.merge_cells('A1:J1')
c = ws_prod['A1']
style_title(c, 16)
c.value = "CATÁLOGO DE PRODUCTOS"
c.fill = PatternFill('solid', fgColor=WHITE)
ws_prod.row_dimensions[1].height = 35

# Row 2: Subtitle
ws_prod.merge_cells('A2:J2')
c = ws_prod['A2']
c.value = "Registra tus productos con código único. Las celdas amarillas son editables."
c.font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)
c.fill = PatternFill('solid', fgColor=WHITE)

# Row 3: Headers
headers_prod = [
    ("A", "CÓDIGO"),
    ("B", "PRODUCTO"),
    ("C", "CATEGORÍA"),
    ("D", "COSTO (Bs.)"),
    ("E", "PRECIO VENTA (Bs.)"),
    ("F", "MARGEN (%)"),
    ("G", "STOCK ACTUAL"),
    ("H", "STOCK MÍNIMO"),
    ("I", "ESTADO"),
    ("J", "VENTAS ACUM. (Bs.)")
]
ws_prod.row_dimensions[3].height = 30
for col_letter, label in headers_prod:
    cell = ws_prod[f'{col_letter}3']
    style_header(cell, bg=DARK_BG, fg=GOLD, size=10)
    cell.value = label

# Rows 4-103: Product data rows (100 products max)
for row in range(4, 104):
    # A: Code (auto-generated)
    c = ws_prod.cell(row=row, column=1)
    c.value = f'=IF(B{row}="","","PROD-"&TEXT(ROW()-3,"0000"))'
    c.font = Font(name='Calibri', size=10, color=ACCENT_BLUE, bold=True)
    c.border = thin_border
    c.alignment = Alignment(horizontal='center', vertical='center')

    # B: Product name (INPUT)
    c = ws_prod.cell(row=row, column=2)
    style_input(c)
    c.alignment = Alignment(horizontal='left', vertical='center')
    unlock_cell(c)

    # C: Category (INPUT - dropdown)
    c = ws_prod.cell(row=row, column=3)
    style_input(c)
    unlock_cell(c)

    # D: Cost price (INPUT)
    c = ws_prod.cell(row=row, column=4)
    style_input(c)
    c.number_format = '#,##0.00'
    unlock_cell(c)

    # E: Sale price (INPUT)
    c = ws_prod.cell(row=row, column=5)
    style_input(c)
    c.number_format = '#,##0.00'
    unlock_cell(c)

    # F: Margin (%) (CALCULATED)
    c = ws_prod.cell(row=row, column=6)
    c.value = f'=IFERROR((E{row}-D{row})/D{row},0)'
    c.number_format = '0.0%'
    style_output(c)

    # G: Current stock (INPUT)
    c = ws_prod.cell(row=row, column=7)
    style_input(c)
    c.number_format = '#,##0'
    unlock_cell(c)

    # H: Minimum stock (INPUT)
    c = ws_prod.cell(row=row, column=8)
    style_input(c)
    c.number_format = '#,##0'
    unlock_cell(c)

    # I: Status (CALCULATED)
    c = ws_prod.cell(row=row, column=9)
    c.value = f'=IF(B{row}="","",IF(G{row}=0,"AGOTADO",IF(G{row}<=H{row},"BAJO","NORMAL")))'
    c.font = Font(name='Calibri', size=10, bold=True)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border

    # J: Accumulated sales (CALCULATED from Ventas sheet)
    c = ws_prod.cell(row=row, column=10)
    c.value = f'=IFERROR(SUMPRODUCT((Ventas!C4:C203=A{row})*Ventas!H4:H203),0)'
    c.number_format = '#,##0.00'
    style_output(c)

# Conditional formatting for Status column (I)
ws_prod.conditional_formatting.add('I4:I103',
    CellIsRule(operator='equal', formula=['"AGOTADO"'],
              fill=PatternFill('solid', fgColor="FEE2E2"),
              font=Font(color=RED_ALERT, bold=True)))
ws_prod.conditional_formatting.add('I4:I103',
    CellIsRule(operator='equal', formula=['"BAJO"'],
              fill=PatternFill('solid', fgColor="FEF3C7"),
              font=Font(color="D97706", bold=True)))
ws_prod.conditional_formatting.add('I4:I103',
    CellIsRule(operator='equal', formula=['"NORMAL"'],
              fill=PatternFill('solid', fgColor="D1FAE5"),
              font=Font(color=GREEN_OK, bold=True)))

# Conditional formatting: Margin data bars
ws_prod.conditional_formatting.add('F4:F103',
    DataBarRule(start_type='min', end_type='max',
               color=GREEN_OK))

# Data validation: Category dropdown
categories = '"Alimentos,Bebidas,Limpieza,Electrónica,Ropa,Otros"'
dv_cat = DataValidation(type="list", formula1=categories, allow_blank=True)
dv_cat.prompt = "Selecciona la categoría del producto"
dv_cat.promptTitle = "Categoría"
dv_cat.error = "Selecciona una categoría válida de la lista"
dv_cat.errorTitle = "Categoría inválida"
ws_prod.add_data_validation(dv_cat)
dv_cat.add('C4:C103')

# Data validation: No negative costs/prices/stock
dv_pos = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
dv_pos.prompt = "Ingresa un valor mayor o igual a 0"
dv_pos.promptTitle = "Dato requerido"
dv_pos.error = "El valor no puede ser negativo"
dv_pos.errorTitle = "Valor inválido"
ws_prod.add_data_validation(dv_pos)
dv_pos.add('D4:D103')
dv_pos.add('E4:E103')
dv_pos.add('G4:G103')
dv_pos.add('H4:H103')

# Sample data (5 products)
sample_products = [
    ("Café Molido 500g", "Alimentos", 35, 55, 50, 10),
    ("Jugo Natural 1L", "Bebidas", 8, 15, 120, 20),
    ("Detergente Premium", "Limpieza", 22, 38, 80, 15),
    ("Cable USB-C", "Electrónica", 15, 35, 45, 10),
    ("Camiseta Algodón", "Ropa", 30, 65, 60, 12),
]
for i, (name, cat, cost, price, stock, min_stock) in enumerate(sample_products):
    r = 4 + i
    ws_prod.cell(row=r, column=2).value = name
    ws_prod.cell(row=r, column=3).value = cat
    ws_prod.cell(row=r, column=4).value = cost
    ws_prod.cell(row=r, column=5).value = price
    ws_prod.cell(row=r, column=7).value = stock
    ws_prod.cell(row=r, column=8).value = min_stock

# ============================================================
# SHEET 3: VENTAS (Sales Entry)
# ============================================================
ws_ventas = wb.create_sheet("Ventas")
ws_ventas.sheet_properties.tabColor = ORANGE
ws_ventas.sheet_view.showGridLines = False

# Column widths
ventas_widths = {1: 14, 2: 10, 3: 16, 4: 28, 5: 10, 6: 16, 7: 10, 8: 16, 9: 16, 10: 18}
for c, w in ventas_widths.items():
    ws_ventas.column_dimensions[get_column_letter(c)].width = w

# Row 1: Title
ws_ventas.merge_cells('A1:J1')
c = ws_ventas['A1']
style_title(c, 16)
c.value = "REGISTRO DE VENTAS"
c.fill = PatternFill('solid', fgColor=WHITE)
ws_ventas.row_dimensions[1].height = 35

# Row 2: Instructions
ws_ventas.merge_cells('A2:J2')
c = ws_ventas['A2']
c.value = "Ingresa el código del producto en la columna C. El nombre y precio se llenan automáticamente (VLOOKUP)."
c.font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)
c.fill = PatternFill('solid', fgColor=WHITE)

# Row 3: Headers
headers_ventas = [
    ("A", "FECHA"),
    ("B", "# VENTA"),
    ("C", "CÓDIGO PROD."),
    ("D", "PRODUCTO (auto)"),
    ("E", "CANT."),
    ("F", "PRECIO UNIT. (auto)"),
    ("G", "DESC. (%)"),
    ("H", "TOTAL (Bs.)"),
    ("I", "MÉTODO PAGO"),
    ("J", "VENDEDOR")
]
ws_ventas.row_dimensions[3].height = 30
for col_letter, label in headers_ventas:
    cell = ws_ventas[f'{col_letter}3']
    style_header(cell, bg=DARK_BG, fg=GOLD, size=10)
    cell.value = label

# Rows 4-203: Sales rows (200 max)
for row in range(4, 204):
    # A: Date (INPUT)
    c = ws_ventas.cell(row=row, column=1)
    style_input(c)
    c.number_format = 'DD/MM/YYYY'
    unlock_cell(c)

    # B: Sale number (AUTO)
    c = ws_ventas.cell(row=row, column=2)
    c.value = f'=IF(A{row}="","",ROW()-3)'
    c.font = Font(name='Calibri', size=10, color=GRAY_500)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border

    # C: Product code (INPUT)
    c = ws_ventas.cell(row=row, column=3)
    style_input(c)
    unlock_cell(c)

    # D: Product name (VLOOKUP auto-fill)
    c = ws_ventas.cell(row=row, column=4)
    c.value = f'=IFERROR(VLOOKUP(C{row},Productos!A4:B103,2,FALSE),"")'
    c.font = Font(name='Calibri', size=10, color=GRAY_700)
    c.border = thin_border
    c.alignment = Alignment(horizontal='left', vertical='center')

    # E: Quantity (INPUT)
    c = ws_ventas.cell(row=row, column=5)
    style_input(c)
    c.number_format = '#,##0'
    unlock_cell(c)

    # F: Unit price (VLOOKUP auto-fill)
    c = ws_ventas.cell(row=row, column=6)
    c.value = f'=IFERROR(VLOOKUP(C{row},Productos!A4:E103,5,FALSE),0)'
    c.number_format = '#,##0.00'
    style_output(c)

    # G: Discount % (INPUT)
    c = ws_ventas.cell(row=row, column=7)
    style_input(c)
    c.number_format = '0%'
    unlock_cell(c)

    # H: Total (CALCULATED)
    c = ws_ventas.cell(row=row, column=8)
    c.value = f'=IFERROR(E{row}*F{row}*(1-G{row}),0)'
    c.number_format = '#,##0.00'
    style_output(c, bold=True)

    # I: Payment method (INPUT - dropdown)
    c = ws_ventas.cell(row=row, column=9)
    style_input(c)
    unlock_cell(c)

    # J: Seller (INPUT)
    c = ws_ventas.cell(row=row, column=10)
    style_input(c)
    c.alignment = Alignment(horizontal='left', vertical='center')
    unlock_cell(c)

# Data validation: Payment method dropdown
dv_pay = DataValidation(type="list", formula1='"Efectivo,Transferencia,QR,Tarjeta"', allow_blank=True)
dv_pay.prompt = "Selecciona el método de pago"
dv_pay.promptTitle = "Método de Pago"
ws_ventas.add_data_validation(dv_pay)
dv_pay.add('I4:I203')

# Data validation: Quantity >= 1
dv_qty = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="1", allow_blank=True)
dv_qty.prompt = "Ingresa la cantidad (mínimo 1)"
dv_qty.promptTitle = "Cantidad"
dv_qty.error = "La cantidad debe ser al menos 1"
dv_qty.errorTitle = "Cantidad inválida"
ws_ventas.add_data_validation(dv_qty)
dv_qty.add('E4:E203')

# Data validation: Discount 0-100%
dv_disc = DataValidation(type="decimal", operator="between", formula1="0", formula2="1", allow_blank=True)
dv_disc.prompt = "Ingresa el descuento (0% a 100%)"
dv_disc.promptTitle = "Descuento"
dv_disc.error = "El descuento debe estar entre 0% y 100%"
dv_disc.errorTitle = "Descuento inválido"
ws_ventas.add_data_validation(dv_disc)
dv_disc.add('G4:G203')

# Sample sales data
from datetime import date, timedelta
import random
sample_sales = [
    (date(2025, 1, 15), "PROD-0001", 3, 0, "Efectivo", "Ana"),
    (date(2025, 1, 15), "PROD-0002", 10, 0.05, "QR", "Carlos"),
    (date(2025, 1, 16), "PROD-0003", 2, 0, "Transferencia", "Ana"),
    (date(2025, 1, 16), "PROD-0005", 4, 0.10, "Tarjeta", "María"),
    (date(2025, 1, 17), "PROD-0001", 5, 0, "Efectivo", "Carlos"),
    (date(2025, 1, 17), "PROD-0004", 2, 0, "QR", "Ana"),
    (date(2025, 1, 18), "PROD-0002", 8, 0, "Transferencia", "María"),
    (date(2025, 1, 18), "PROD-0005", 3, 0.05, "Efectivo", "Carlos"),
]
for i, (dt, code, qty, disc, pay, seller) in enumerate(sample_sales):
    r = 4 + i
    ws_ventas.cell(row=r, column=1).value = dt
    ws_ventas.cell(row=r, column=3).value = code
    ws_ventas.cell(row=r, column=5).value = qty
    ws_ventas.cell(row=r, column=7).value = disc
    ws_ventas.cell(row=r, column=9).value = pay
    ws_ventas.cell(row=r, column=10).value = seller

# ============================================================
# SHEET 4: INVENTARIO (Stock Tracking)
# ============================================================
ws_inv = wb.create_sheet("Inventario")
ws_inv.sheet_properties.tabColor = PURPLE
ws_inv.sheet_view.showGridLines = False

inv_widths = {1: 16, 2: 30, 3: 14, 4: 14, 5: 14, 6: 14, 7: 16, 8: 14}
for c, w in inv_widths.items():
    ws_inv.column_dimensions[get_column_letter(c)].width = w

# Row 1: Title
ws_inv.merge_cells('A1:H1')
c = ws_inv['A1']
style_title(c, 16)
c.value = "MOVIMIENTOS DE INVENTARIO"
c.fill = PatternFill('solid', fgColor=WHITE)
ws_inv.row_dimensions[1].height = 35

# Row 2: Instructions
ws_inv.merge_cells('A2:H2')
c = ws_inv['A2']
c.value = "Registra entradas y salidas de inventario. Tipo: ENTRADA = compras/reposición, SALIDA = ajustes/pérdidas."
c.font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)
c.fill = PatternFill('solid', fgColor=WHITE)

# Row 3: Headers
headers_inv = [
    ("A", "FECHA"),
    ("B", "CÓDIGO PROD."),
    ("C", "PRODUCTO (auto)"),
    ("D", "TIPO"),
    ("E", "CANTIDAD"),
    ("F", "COSTO UNIT."),
    ("G", "TOTAL MOVIMIENTO"),
    ("H", "OBSERVACIÓN")
]
ws_inv.row_dimensions[3].height = 30
for i, (col_letter, label) in enumerate(headers_inv):
    cell = ws_inv[f'{col_letter}3']
    style_header(cell, bg=DARK_BG, fg=GOLD, size=10)
    cell.value = label

# Make column C wider for product name
ws_inv.column_dimensions['C'].width = 28

# Rows 4-203: Inventory movement rows
for row in range(4, 204):
    # A: Date (INPUT)
    c = ws_inv.cell(row=row, column=1)
    style_input(c)
    c.number_format = 'DD/MM/YYYY'
    unlock_cell(c)

    # B: Product code (INPUT)
    c = ws_inv.cell(row=row, column=2)
    style_input(c)
    unlock_cell(c)

    # C: Product name (VLOOKUP)
    c = ws_inv.cell(row=row, column=3)
    c.value = f'=IFERROR(VLOOKUP(B{row},Productos!A4:B103,2,FALSE),"")'
    c.font = Font(name='Calibri', size=10, color=GRAY_700)
    c.border = thin_border

    # D: Type (INPUT - dropdown)
    c = ws_inv.cell(row=row, column=4)
    style_input(c)
    unlock_cell(c)

    # E: Quantity (INPUT)
    c = ws_inv.cell(row=row, column=5)
    style_input(c)
    c.number_format = '#,##0'
    unlock_cell(c)

    # F: Unit cost (INPUT)
    c = ws_inv.cell(row=row, column=6)
    style_input(c)
    c.number_format = '#,##0.00'
    unlock_cell(c)

    # G: Total (CALCULATED)
    c = ws_inv.cell(row=row, column=7)
    c.value = f'=IFERROR(E{row}*F{row},0)'
    c.number_format = '#,##0.00'
    style_output(c)

    # H: Observation (INPUT)
    c = ws_inv.cell(row=row, column=8)
    style_input(c)
    c.alignment = Alignment(horizontal='left', vertical='center')
    unlock_cell(c)

# Data validation: Type dropdown
dv_type = DataValidation(type="list", formula1='"ENTRADA,SALIDA"', allow_blank=True)
dv_type.prompt = "ENTRADA = compras/reposición, SALIDA = ajustes/pérdidas"
dv_type.promptTitle = "Tipo de Movimiento"
ws_inv.add_data_validation(dv_type)
dv_type.add('D4:D203')

# ============================================================
# SHEET 5: REPORTES (Sales Reports)
# ============================================================
ws_rep = wb.create_sheet("Reportes")
ws_rep.sheet_properties.tabColor = "3498DB"
ws_rep.sheet_view.showGridLines = False

rep_widths = {1: 3, 2: 22, 3: 18, 4: 18, 5: 18, 6: 3, 7: 22, 8: 18, 9: 18, 10: 3}
for c, w in rep_widths.items():
    ws_rep.column_dimensions[get_column_letter(c)].width = w

fill_range_white(ws_rep, 1, 50, 1, 10)

# Title
ws_rep.merge_cells('B2:I2')
c = ws_rep['B2']
style_title(c, 16)
c.value = "REPORTES DE VENTAS"

# ── Report 1: Sales by Category ──
ws_rep['B4'].value = "VENTAS POR CATEGORÍA"
ws_rep['B4'].font = Font(name='Calibri', bold=True, size=12, color=GRAY_800)

ws_rep['B5'].value = "Categoría"
ws_rep['C5'].value = "Ventas (Bs.)"
ws_rep['D5'].value = "% del Total"
style_header(ws_rep['B5'], bg=DARK_BG, fg=GOLD, size=10)
style_header(ws_rep['C5'], bg=DARK_BG, fg=GOLD, size=10)
style_header(ws_rep['D5'], bg=DARK_BG, fg=GOLD, size=10)

report_cats = ["Alimentos", "Bebidas", "Limpieza", "Electrónica", "Ropa", "Otros"]
for i, cat in enumerate(report_cats):
    r = 6 + i
    ws_rep.cell(row=r, column=2).value = cat
    ws_rep.cell(row=r, column=2).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws_rep.cell(row=r, column=2).border = thin_border

    # Sum sales where product category matches
    ws_rep.cell(row=r, column=3).value = (
        f'=IFERROR(SUMPRODUCT((VLOOKUP(Ventas!C4:C203,Productos!A4:C103,3,FALSE)="{cat}")'
        f'*Ventas!H4:H203),0)'
    )
    ws_rep.cell(row=r, column=3).number_format = '#,##0.00'
    ws_rep.cell(row=r, column=3).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws_rep.cell(row=r, column=3).border = thin_border

    ws_rep.cell(row=r, column=4).value = f'=IFERROR(C{r}/SUM(C6:C11),0)'
    ws_rep.cell(row=r, column=4).number_format = '0.0%'
    ws_rep.cell(row=r, column=4).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws_rep.cell(row=r, column=4).border = thin_border

# Total row
r_tot = 12
ws_rep.cell(row=r_tot, column=2).value = "TOTAL"
ws_rep.cell(row=r_tot, column=2).font = Font(name='Calibri', size=10, bold=True, color=BLACK)
ws_rep.cell(row=r_tot, column=2).border = thin_border
ws_rep.cell(row=r_tot, column=3).value = '=SUM(C6:C11)'
ws_rep.cell(row=r_tot, column=3).number_format = '#,##0.00'
ws_rep.cell(row=r_tot, column=3).font = Font(name='Calibri', size=10, bold=True, color=BLACK)
ws_rep.cell(row=r_tot, column=3).border = thin_border

# Chart: Sales by Category
chart3 = BarChart()
chart3.type = "col"
chart3.style = 10
chart3.title = "Ventas por Categoría"
chart3.y_axis.title = "Bs."
chart3.legend = None
data3 = Reference(ws_rep, min_col=3, min_row=5, max_row=11)
cats3 = Reference(ws_rep, min_col=2, min_row=6, max_row=11)
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
chart3.width = 18
chart3.height = 12
s3 = chart3.series[0]
s3.graphicalProperties.solidFill = TURQUOISE
s3.graphicalProperties.line.noFill = True
ws_rep.add_chart(chart3, "B14")

# ── Report 2: Sales by Seller ──
ws_rep['G4'].value = "VENTAS POR VENDEDOR"
ws_rep['G4'].font = Font(name='Calibri', bold=True, size=12, color=GRAY_800)

ws_rep['G5'].value = "Vendedor"
ws_rep['H5'].value = "Ventas (Bs.)"
ws_rep['I5'].value = "# Transacciones"
style_header(ws_rep['G5'], bg=DARK_BG, fg=GOLD, size=10)
style_header(ws_rep['H5'], bg=DARK_BG, fg=GOLD, size=10)
style_header(ws_rep['I5'], bg=DARK_BG, fg=GOLD, size=10)

# Up to 10 sellers from Config
sellers = ["Ana", "Carlos", "María", "Pedro", "Lucía"]
for i, seller in enumerate(sellers):
    r = 6 + i
    ws_rep.cell(row=r, column=7).value = seller
    ws_rep.cell(row=r, column=7).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws_rep.cell(row=r, column=7).border = thin_border

    ws_rep.cell(row=r, column=8).value = f'=IFERROR(SUMPRODUCT((Ventas!J4:J203="{seller}")*Ventas!H4:H203),0)'
    ws_rep.cell(row=r, column=8).number_format = '#,##0.00'
    ws_rep.cell(row=r, column=8).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws_rep.cell(row=r, column=8).border = thin_border

    ws_rep.cell(row=r, column=9).value = f'=COUNTIF(Ventas!J4:J203,"{seller}")'
    ws_rep.cell(row=r, column=9).number_format = '#,##0'
    ws_rep.cell(row=r, column=9).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws_rep.cell(row=r, column=9).border = thin_border

# ============================================================
# SHEET 6: CONFIG
# ============================================================
ws_conf = wb.create_sheet("Config")
ws_conf.sheet_properties.tabColor = GRAY_500
ws_conf.sheet_view.showGridLines = False

conf_widths = {1: 25, 2: 40, 3: 5}
for c, w in conf_widths.items():
    ws_conf.column_dimensions[get_column_letter(c)].width = w

fill_range_white(ws_conf, 1, 30, 1, 3)

# Version
ws_conf['A1'].value = "v1.0.0"
ws_conf['A1'].font = Font(name='Calibri', size=9, color=GRAY_400)

# Title
ws_conf.merge_cells('A3:B3')
c = ws_conf['A3']
style_title(c, 14)
c.value = "CONFIGURACIÓN"

# Settings
settings = [
    ("Producto", "Sistema de Ventas con Código Único"),
    ("Versión", "v1.0.0"),
    ("Marca", "No Somos Ignorantes"),
    ("Precio", "Bs. 89"),
    ("Protección", "nsi2024"),
    ("", ""),
    ("INSTRUCCIONES", ""),
    ("1.", "Las celdas AMARILLAS son editables — ingresa tus datos ahí."),
    ("2.", "Las celdas VERDES se calculan automáticamente — no las modifiques."),
    ("3.", "En la hoja 'Productos', registra tus productos con nombre, costo y precio."),
    ("4.", "El CÓDIGO se genera automáticamente (PROD-0001, PROD-0002, etc.)."),
    ("5.", "En la hoja 'Ventas', ingresa el código del producto y la cantidad."),
    ("6.", "El nombre y precio se llenan solos con VLOOKUP."),
    ("7.", "El Dashboard muestra KPIs, gráficos y alertas en tiempo real."),
    ("8.", "Para desbloquear una hoja: Revisar → Desproteger hoja → nsi2024"),
]
for i, (label, value) in enumerate(settings):
    r = 5 + i
    ws_conf.cell(row=r, column=1).value = label
    ws_conf.cell(row=r, column=1).font = Font(name='Calibri', size=10, bold=True, color=GRAY_700)
    ws_conf.cell(row=r, column=2).value = value
    ws_conf.cell(row=r, column=2).font = Font(name='Calibri', size=10, color=GRAY_600)

# ============================================================
# SHEET PROTECTION (all sheets)
# ============================================================
for ws in wb.worksheets:
    ws.protection.sheet = True
    ws.protection.password = "nsi2024"
    ws.protection.enable()

# Explicitly unlock input cells on Productos
for row in range(4, 104):
    for col in [2, 3, 4, 5, 7, 8]:  # B, C, D, E, G, H
        unlock_cell(ws_prod.cell(row=row, column=col))

# Explicitly unlock input cells on Ventas
for row in range(4, 204):
    for col in [1, 3, 5, 7, 9, 10]:  # A, C, E, G, I, J
        unlock_cell(ws_ventas.cell(row=row, column=col))

# Explicitly unlock input cells on Inventario
for row in range(4, 204):
    for col in [1, 2, 4, 5, 6, 8]:  # A, B, D, E, F, H
        unlock_cell(ws_inv.cell(row=row, column=col))

# Unlock Dashboard meta input (C7)
unlock_cell(ws_dash['C7'])

# ============================================================
# SAVE
# ============================================================
output_path = r"D:\Landing-Page_marketplace\excel_products\Sistema_Ventas_Codigo_Unico_NSI.xlsx"
wb.save(output_path)
print(f"[OK] Saved: {output_path}")

# ============================================================
# VERIFICATION
# ============================================================
from openpyxl import load_workbook
wb2 = load_workbook(output_path)

formula_count = 0
for ws in wb2.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formula_count += 1

print(f"[OK] Sheets: {wb2.sheetnames}")
print(f"[OK] Total formulas: {formula_count}")
print(f"[OK] Charts: Dashboard={len(ws_dash._charts)}, Reportes={len(ws_rep._charts)}")
print(f"[OK] Protection: all sheets protected with password nsi2024")
print(f"[OK] Sample data: 5 products, 8 sales transactions")
