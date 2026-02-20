"""
Build: Calculadora de ARPU (Bs. 49)
No Somos Ignorantes v1.0
ARPU calculator: Dashboard with subscription metrics, customer database,
revenue tracking, segment analysis, LTV estimation.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint, SeriesLabel
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

GOLD = "D4AF37"; DARK_BG = "1A1A2E"; ACCENT_BLUE = "0F3460"
WHITE = "FFFFFF"; BLACK = "000000"
LIGHT_YELLOW = "FFF9C4"; LIGHT_GREEN = "E8F5E9"
RED_CORAL = "E74C3C"; TURQUOISE = "16A085"; ORANGE = "E67E22"; PURPLE = "8E44AD"
LIGHT_GRAY = "BDC3C7"; GREEN_OK = "27AE60"; RED_ALERT = "E74C3C"
GRAY_300 = "D1D5DB"; GRAY_400 = "9CA3AF"; GRAY_500 = "6B7280"
GRAY_600 = "4B5563"; GRAY_700 = "374151"; GRAY_900 = "111827"

thin_border = Border(
    left=Side(style='thin', color=GRAY_300), right=Side(style='thin', color=GRAY_300),
    top=Side(style='thin', color=GRAY_300), bottom=Side(style='thin', color=GRAY_300))
gold_border = Border(
    left=Side(style='thin', color=GOLD), right=Side(style='thin', color=GOLD),
    top=Side(style='thin', color=GOLD), bottom=Side(style='thin', color=GOLD))

def style_title(cell, size=18):
    cell.font = Font(name='Calibri', bold=True, color=BLACK, size=size)
    cell.alignment = Alignment(horizontal='left', vertical='center')
def style_header(cell, bg=DARK_BG, fg=WHITE, size=10):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
def style_input(cell):
    cell.font = Font(name='Calibri', size=11, color="0000FF")
    cell.fill = PatternFill('solid', fgColor=LIGHT_YELLOW)
    cell.border = gold_border
    cell.alignment = Alignment(horizontal='center', vertical='center')
def style_output(cell, bold=False):
    cell.font = Font(name='Calibri', size=11, color=GRAY_900, bold=bold)
    cell.fill = PatternFill('solid', fgColor=LIGHT_GREEN)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='right', vertical='center')
def style_banner(cell, bg_color, fg=WHITE, size=10):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg_color)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
def fill_white(ws, r1, r2, c1, c2):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=WHITE)
def unlock_cell(cell):
    cell.protection = cell.protection.copy(locked=False)

MONTHS_ES = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
MAX_CUSTOMERS = 200
MAX_REVENUE = 300

# ============================================================
# SHEET 1: DASHBOARD
# ============================================================
ws = wb.active
ws.title = "Dashboard"
ws.sheet_properties.tabColor = RED_CORAL
ws.sheet_view.showGridLines = False

dw = {1:3, 2:22, 3:16, 4:16, 5:3, 6:22, 7:16, 8:16, 9:3}
for c, w in dw.items():
    ws.column_dimensions[get_column_letter(c)].width = w
fill_white(ws, 1, 55, 1, 9)

ws.merge_cells('B2:H2')
style_title(ws['B2'], 20)
ws['B2'].value = "CALCULADORA DE ARPU"
ws.merge_cells('B3:H3')
ws['B3'].font = Font(name='Calibri', size=10, color=GRAY_500)
ws['B3'].value = "No Somos Ignorantes  |  v1.0  |  Ingreso Promedio por Usuario"

ws.row_dimensions[4].height = 6

# KPI 1: ARPU Mensual
ws.merge_cells('B5:D5')
style_banner(ws['B5'], RED_CORAL, WHITE, 10)
ws['B5'].value = "ARPU MENSUAL (Bs.)"
ws.merge_cells('B6:D6')
c = ws['B6']
c.font = Font(name='Calibri', bold=True, color=BLACK, size=26)
c.alignment = Alignment(horizontal='right', vertical='center')
c.value = ('=IFERROR(SUMPRODUCT((MONTH(Ingresos!A4:A303)=MONTH(TODAY()))'
           '*(YEAR(Ingresos!A4:A303)=YEAR(TODAY()))*Ingresos!E4:E303)'
           '/COUNTIF(Clientes!F4:F203,"Activo"),0)')
c.number_format = '#,##0.00'

# KPI 2: Usuarios Activos
ws.merge_cells('F5:H5')
style_banner(ws['F5'], TURQUOISE, WHITE, 10)
ws['F5'].value = "USUARIOS ACTIVOS"
ws.merge_cells('F6:H6')
c = ws['F6']
c.font = Font(name='Calibri', bold=True, color=BLACK, size=26)
c.alignment = Alignment(horizontal='right', vertical='center')
c.value = '=COUNTIF(Clientes!F4:F203,"Activo")'
c.number_format = '#,##0'

ws.row_dimensions[7].height = 4

ws['B7'].value = "ARPU anterior:"
ws['B7'].font = Font(name='Calibri', size=9, color="808080")
ws['C7'].value = 0
style_input(ws['C7'])
ws['C7'].number_format = '#,##0.00'
unlock_cell(ws['C7'])

ws['B8'].value = "Crecimiento ARPU:"
ws['B8'].font = Font(name='Calibri', size=9, color="808080")
ws.merge_cells('C8:D8')
c = ws['C8']
c.value = '=IFERROR((B6-C7)/C7,0)'
c.number_format = '+0.0%;-0.0%'
c.font = Font(name='Calibri', bold=True, size=14, color=BLACK)

ws['F7'].value = "Total clientes:"
ws['F7'].font = Font(name='Calibri', size=9, color="808080")
ws['G7'].value = '=COUNTA(Clientes!B4:B203)'
ws['G7'].number_format = '#,##0'
ws['G7'].font = Font(name='Calibri', size=11, color=GRAY_700)

ws['F8'].value = "Tasa retencion:"
ws['F8'].font = Font(name='Calibri', size=9, color="808080")
ws.merge_cells('G8:H8')
c = ws['G8']
c.value = '=IFERROR(F6/G7,0)'
c.number_format = '0.0%'
c.font = Font(name='Calibri', bold=True, size=14, color=BLACK)

ws.row_dimensions[9].height = 6

# KPI 3: MRR (Monthly Recurring Revenue)
ws.merge_cells('B10:D10')
style_banner(ws['B10'], ORANGE, WHITE, 10)
ws['B10'].value = "MRR - INGRESO RECURRENTE (Bs.)"
ws.merge_cells('B11:D11')
c = ws['B11']
c.font = Font(name='Calibri', bold=True, color=BLACK, size=26)
c.alignment = Alignment(horizontal='right', vertical='center')
c.value = ('=IFERROR(SUMPRODUCT((MONTH(Ingresos!A4:A303)=MONTH(TODAY()))'
           '*(YEAR(Ingresos!A4:A303)=YEAR(TODAY()))*Ingresos!E4:E303),0)')
c.number_format = '#,##0.00'

# KPI 4: LTV Estimado
ws.merge_cells('F10:H10')
style_banner(ws['F10'], PURPLE, WHITE, 10)
ws['F10'].value = "LTV ESTIMADO (Bs.)"
ws.merge_cells('F11:H11')
c = ws['F11']
c.font = Font(name='Calibri', bold=True, color=BLACK, size=26)
c.alignment = Alignment(horizontal='right', vertical='center')
# LTV = ARPU * Avg lifetime months (input)
ws['F12'].value = "Vida promedio (meses):"
ws['F12'].font = Font(name='Calibri', size=9, color="808080")
ws['H12'].value = 12
style_input(ws['H12'])
ws['H12'].number_format = '#,##0'
unlock_cell(ws['H12'])

c = ws['F11']
c.value = '=IFERROR(B6*H12,0)'
c.number_format = '#,##0.00'

ws.row_dimensions[13].height = 8

# Monthly ARPU trend
ws['B14'].value = "ARPU MENSUAL"
ws['B14'].font = Font(name='Calibri', bold=True, size=11, color=GRAY_700)

ws['B15'].value = "Mes"
ws['C15'].value = "Ingresos"
ws['D15'].value = "Usuarios"
style_header(ws['B15'], bg=GRAY_700)
style_header(ws['C15'], bg=GRAY_700)
style_header(ws['D15'], bg=GRAY_700)

for i, m in enumerate(MONTHS_ES):
    r = 16 + i
    ws.cell(row=r, column=2).value = m
    ws.cell(row=r, column=2).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws.cell(row=r, column=2).border = thin_border

    ws.cell(row=r, column=3).value = (
        f'=IFERROR(SUMPRODUCT((MONTH(Ingresos!A4:A303)={i+1})'
        f'*(YEAR(Ingresos!A4:A303)=YEAR(TODAY()))*Ingresos!E4:E303),0)')
    ws.cell(row=r, column=3).number_format = '#,##0'
    ws.cell(row=r, column=3).border = thin_border

    # Count distinct active users that month (simplified: count revenue entries)
    ws.cell(row=r, column=4).value = (
        f'=IFERROR(SUMPRODUCT((MONTH(Ingresos!A4:A303)={i+1})'
        f'*(YEAR(Ingresos!A4:A303)=YEAR(TODAY()))*(Ingresos!E4:E303>0)*1),0)')
    ws.cell(row=r, column=4).number_format = '#,##0'
    ws.cell(row=r, column=4).border = thin_border

# Line chart: ARPU trend
chart1 = LineChart()
chart1.title = "Evolucion de Ingresos Mensuales"
chart1.style = 10
chart1.y_axis.title = "Bs."
chart1.legend.position = 'b'
chart1.width = 22
chart1.height = 13

d1 = Reference(ws, min_col=3, min_row=15, max_row=27)
cats1 = Reference(ws, min_col=2, min_row=16, max_row=27)
chart1.add_data(d1, titles_from_data=True)
chart1.set_categories(cats1)
chart1.series[0].graphicalProperties.line.solidFill = RED_CORAL
chart1.series[0].graphicalProperties.line.width = 28000
ws.add_chart(chart1, "B29")

# Segment analysis
ws['F14'].value = "ARPU POR SEGMENTO"
ws['F14'].font = Font(name='Calibri', bold=True, size=11, color=GRAY_700)

ws['F15'].value = "Segmento"
ws['G15'].value = "Usuarios"
ws['H15'].value = "ARPU"
style_header(ws['F15'], bg=GRAY_700)
style_header(ws['G15'], bg=GRAY_700)
style_header(ws['H15'], bg=GRAY_700)

segments = ["Basico", "Estandar", "Premium", "Enterprise"]
for i, seg in enumerate(segments):
    r = 16 + i
    ws.cell(row=r, column=6).value = seg
    ws.cell(row=r, column=6).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws.cell(row=r, column=6).border = thin_border

    ws.cell(row=r, column=7).value = f'=COUNTIF(Clientes!E4:E203,"{seg}")'
    ws.cell(row=r, column=7).number_format = '#,##0'
    ws.cell(row=r, column=7).border = thin_border

    ws.cell(row=r, column=8).value = (
        f'=IFERROR(SUMPRODUCT((Clientes!E4:E203="{seg}")*Clientes!G4:G203)'
        f'/COUNTIF(Clientes!E4:E203,"{seg}"),0)')
    ws.cell(row=r, column=8).number_format = '#,##0.00'
    ws.cell(row=r, column=8).border = thin_border

# Bar chart: ARPU by segment
chart2 = BarChart()
chart2.type = "col"
chart2.style = 10
chart2.title = "ARPU por Segmento"
chart2.y_axis.title = "Bs."
chart2.legend = None
chart2.width = 16
chart2.height = 13
d2 = Reference(ws, min_col=8, min_row=15, max_row=19)
cats2 = Reference(ws, min_col=6, min_row=16, max_row=19)
chart2.add_data(d2, titles_from_data=True)
chart2.set_categories(cats2)
s = chart2.series[0]
s.graphicalProperties.solidFill = TURQUOISE
s.graphicalProperties.line.noFill = True
ws.add_chart(chart2, "F29")

# ============================================================
# SHEET 2: CLIENTES
# ============================================================
ws_cl = wb.create_sheet("Clientes")
ws_cl.sheet_properties.tabColor = TURQUOISE
ws_cl.sheet_view.showGridLines = False

cl_w = {1:6, 2:25, 3:25, 4:16, 5:14, 6:12, 7:16}
for c, w in cl_w.items():
    ws_cl.column_dimensions[get_column_letter(c)].width = w

ws_cl.merge_cells('A1:G1')
style_title(ws_cl['A1'], 16)
ws_cl['A1'].value = "BASE DE CLIENTES"
ws_cl['A1'].fill = PatternFill('solid', fgColor=WHITE)
ws_cl.row_dimensions[1].height = 35

ws_cl.merge_cells('A2:G2')
ws_cl['A2'].value = "Registra tus clientes/suscriptores. El segmento y estado se usan para calcular ARPU."
ws_cl['A2'].font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)
ws_cl['A2'].fill = PatternFill('solid', fgColor=WHITE)

headers_cl = [("A","#"), ("B","NOMBRE/EMPRESA"), ("C","EMAIL/CONTACTO"),
              ("D","FECHA REGISTRO"), ("E","SEGMENTO"), ("F","ESTADO"),
              ("G","INGRESO MENSUAL")]
ws_cl.row_dimensions[3].height = 28
for col_letter, label in headers_cl:
    style_header(ws_cl[f'{col_letter}3'], bg=DARK_BG, fg=GOLD, size=10)
    ws_cl[f'{col_letter}3'].value = label

for row in range(4, 4 + MAX_CUSTOMERS):
    ws_cl.cell(row=row, column=1).value = f'=IF(B{row}="","",ROW()-3)'
    ws_cl.cell(row=row, column=1).font = Font(name='Calibri', size=10, color=GRAY_500)
    ws_cl.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    ws_cl.cell(row=row, column=1).border = thin_border

    for col in [2, 3, 4, 5, 6, 7]:
        c = ws_cl.cell(row=row, column=col)
        style_input(c)
        unlock_cell(c)
        if col in [2, 3]:
            c.alignment = Alignment(horizontal='left', vertical='center')
        elif col == 4:
            c.number_format = 'DD/MM/YYYY'
        elif col == 7:
            c.number_format = '#,##0.00'

# Dropdowns
dv_seg = DataValidation(type="list", formula1='"Basico,Estandar,Premium,Enterprise"', allow_blank=True)
dv_seg.prompt = "Selecciona el segmento del cliente"
dv_seg.promptTitle = "Segmento"
ws_cl.add_data_validation(dv_seg)
dv_seg.add('E4:E203')

dv_st = DataValidation(type="list", formula1='"Activo,Inactivo,Cancelado"', allow_blank=True)
dv_st.prompt = "Estado del cliente"
dv_st.promptTitle = "Estado"
ws_cl.add_data_validation(dv_st)
dv_st.add('F4:F203')

# Conditional formatting: Status
ws_cl.conditional_formatting.add('F4:F203',
    CellIsRule(operator='equal', formula=['"Activo"'],
              fill=PatternFill('solid', fgColor="D1FAE5"),
              font=Font(color=GREEN_OK, bold=True)))
ws_cl.conditional_formatting.add('F4:F203',
    CellIsRule(operator='equal', formula=['"Cancelado"'],
              fill=PatternFill('solid', fgColor="FEE2E2"),
              font=Font(color=RED_ALERT, bold=True)))

# Sample customers
from datetime import date
sample_cl = [
    ("TechCorp SRL", "tech@corp.com", date(2024,3,1), "Premium", "Activo", 500),
    ("Cafe Delicia", "info@cafe.com", date(2024,5,15), "Estandar", "Activo", 200),
    ("Libreria Sol", "sol@lib.com", date(2024,6,1), "Basico", "Activo", 99),
    ("MegaStore", "mega@store.com", date(2024,1,10), "Enterprise", "Activo", 1500),
    ("Juan Perez", "juan@mail.com", date(2024,7,20), "Basico", "Activo", 99),
    ("Fashion Plus", "info@fashion.com", date(2024,4,5), "Estandar", "Activo", 200),
    ("Tech Solutions", "hello@ts.com", date(2024,2,15), "Premium", "Inactivo", 0),
    ("Restaurante Gourmet", "gourmet@rest.com", date(2024,8,1), "Estandar", "Activo", 200),
    ("Consultora ABC", "abc@consult.com", date(2024,9,10), "Premium", "Activo", 500),
    ("Minimarket Don Pedro", "pedro@mini.com", date(2024,3,25), "Basico", "Cancelado", 0),
]
for i, (name, email, dt, seg, status, rev) in enumerate(sample_cl):
    r = 4 + i
    ws_cl.cell(row=r, column=2).value = name
    ws_cl.cell(row=r, column=3).value = email
    ws_cl.cell(row=r, column=4).value = dt
    ws_cl.cell(row=r, column=5).value = seg
    ws_cl.cell(row=r, column=6).value = status
    ws_cl.cell(row=r, column=7).value = rev

# ============================================================
# SHEET 3: INGRESOS
# ============================================================
ws_ing = wb.create_sheet("Ingresos")
ws_ing.sheet_properties.tabColor = ORANGE
ws_ing.sheet_view.showGridLines = False

ing_w = {1:14, 2:25, 3:14, 4:18, 5:18, 6:20}
for c, w in ing_w.items():
    ws_ing.column_dimensions[get_column_letter(c)].width = w

ws_ing.merge_cells('A1:F1')
style_title(ws_ing['A1'], 16)
ws_ing['A1'].value = "REGISTRO DE INGRESOS"
ws_ing['A1'].fill = PatternFill('solid', fgColor=WHITE)
ws_ing.row_dimensions[1].height = 35

ws_ing.merge_cells('A2:F2')
ws_ing['A2'].value = "Registra cada pago/ingreso recibido de tus clientes."
ws_ing['A2'].font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)
ws_ing['A2'].fill = PatternFill('solid', fgColor=WHITE)

headers_ing = [("A","FECHA"), ("B","CLIENTE"), ("C","SEGMENTO"),
               ("D","CONCEPTO"), ("E","MONTO (Bs.)"), ("F","METODO PAGO")]
ws_ing.row_dimensions[3].height = 28
for col_letter, label in headers_ing:
    style_header(ws_ing[f'{col_letter}3'], bg=DARK_BG, fg=GOLD, size=10)
    ws_ing[f'{col_letter}3'].value = label

for row in range(4, 4 + MAX_REVENUE):
    for col in range(1, 7):
        c = ws_ing.cell(row=row, column=col)
        style_input(c)
        unlock_cell(c)
        if col == 1: c.number_format = 'DD/MM/YYYY'
        elif col in [2, 4]: c.alignment = Alignment(horizontal='left', vertical='center')
        elif col == 5: c.number_format = '#,##0.00'

dv_pay = DataValidation(type="list", formula1='"Transferencia,QR,Efectivo,Tarjeta"', allow_blank=True)
ws_ing.add_data_validation(dv_pay)
dv_pay.add('F4:F303')

dv_seg2 = DataValidation(type="list", formula1='"Basico,Estandar,Premium,Enterprise"', allow_blank=True)
ws_ing.add_data_validation(dv_seg2)
dv_seg2.add('C4:C303')

# Sample revenue data
sample_ing = [
    (date(2025,1,1), "TechCorp SRL", "Premium", "Suscripcion mensual", 500, "Transferencia"),
    (date(2025,1,1), "Cafe Delicia", "Estandar", "Suscripcion mensual", 200, "QR"),
    (date(2025,1,1), "Libreria Sol", "Basico", "Suscripcion mensual", 99, "Transferencia"),
    (date(2025,1,1), "MegaStore", "Enterprise", "Suscripcion mensual", 1500, "Transferencia"),
    (date(2025,1,1), "Juan Perez", "Basico", "Suscripcion mensual", 99, "QR"),
    (date(2025,1,1), "Fashion Plus", "Estandar", "Suscripcion mensual", 200, "Transferencia"),
    (date(2025,1,1), "Restaurante Gourmet", "Estandar", "Suscripcion mensual", 200, "QR"),
    (date(2025,1,1), "Consultora ABC", "Premium", "Suscripcion mensual", 500, "Transferencia"),
    (date(2025,2,1), "TechCorp SRL", "Premium", "Suscripcion mensual", 500, "Transferencia"),
    (date(2025,2,1), "Cafe Delicia", "Estandar", "Suscripcion mensual", 200, "QR"),
    (date(2025,2,1), "MegaStore", "Enterprise", "Suscripcion mensual", 1500, "Transferencia"),
    (date(2025,2,1), "Juan Perez", "Basico", "Suscripcion mensual", 99, "QR"),
]
for i, (dt, cli, seg, concept, amount, pay) in enumerate(sample_ing):
    r = 4 + i
    ws_ing.cell(row=r, column=1).value = dt
    ws_ing.cell(row=r, column=2).value = cli
    ws_ing.cell(row=r, column=3).value = seg
    ws_ing.cell(row=r, column=4).value = concept
    ws_ing.cell(row=r, column=5).value = amount
    ws_ing.cell(row=r, column=6).value = pay

# ============================================================
# SHEET 4: SEGMENTOS
# ============================================================
ws_seg = wb.create_sheet("Segmentos")
ws_seg.sheet_properties.tabColor = PURPLE
ws_seg.sheet_view.showGridLines = False

seg_w = {1:4, 2:18, 3:16, 4:16, 5:16, 6:16, 7:16}
for c, w in seg_w.items():
    ws_seg.column_dimensions[get_column_letter(c)].width = w
fill_white(ws_seg, 1, 30, 1, 7)

ws_seg.merge_cells('B1:G1')
style_title(ws_seg['B1'], 16)
ws_seg['B1'].value = "ANALISIS POR SEGMENTO"
ws_seg.row_dimensions[1].height = 35

headers_seg = [("B","SEGMENTO"), ("C","PRECIO PLAN"), ("D","CLIENTES ACTIVOS"),
               ("E","INGRESOS MES"), ("F","ARPU"), ("G","% INGRESOS")]
ws_seg.row_dimensions[3].height = 28
for col_letter, label in headers_seg:
    style_header(ws_seg[f'{col_letter}3'], bg=DARK_BG, fg=GOLD, size=10)
    ws_seg[f'{col_letter}3'].value = label

segs = ["Basico", "Estandar", "Premium", "Enterprise"]
prices = [99, 200, 500, 1500]
for i, (seg, price) in enumerate(zip(segs, prices)):
    r = 4 + i
    ws_seg.cell(row=r, column=2).value = seg
    ws_seg.cell(row=r, column=2).font = Font(name='Calibri', size=11, bold=True, color=GRAY_700)
    ws_seg.cell(row=r, column=2).border = thin_border

    c = ws_seg.cell(row=r, column=3)
    c.value = price
    c.number_format = '#,##0.00'
    style_input(c)
    unlock_cell(c)

    c = ws_seg.cell(row=r, column=4)
    c.value = f'=COUNTIFS(Clientes!E4:E203,"{seg}",Clientes!F4:F203,"Activo")'
    c.number_format = '#,##0'
    style_output(c)

    c = ws_seg.cell(row=r, column=5)
    c.value = (f'=IFERROR(SUMPRODUCT((Ingresos!C4:C303="{seg}")'
               f'*(MONTH(Ingresos!A4:A303)=MONTH(TODAY()))'
               f'*(YEAR(Ingresos!A4:A303)=YEAR(TODAY()))*Ingresos!E4:E303),0)')
    c.number_format = '#,##0.00'
    style_output(c)

    c = ws_seg.cell(row=r, column=6)
    c.value = f'=IFERROR(E{r}/D{r},0)'
    c.number_format = '#,##0.00'
    style_output(c, bold=True)

    c = ws_seg.cell(row=r, column=7)
    c.value = f'=IFERROR(E{r}/SUM(E$4:E$7),0)'
    c.number_format = '0.0%'
    style_output(c)

# Total row
r_tot = 8
ws_seg.cell(row=r_tot, column=2).value = "TOTAL"
ws_seg.cell(row=r_tot, column=2).font = Font(name='Calibri', bold=True, size=11, color=BLACK)
ws_seg.cell(row=r_tot, column=2).border = thin_border
for col in [4, 5]:
    c = ws_seg.cell(row=r_tot, column=col)
    c.value = f'=SUM({get_column_letter(col)}4:{get_column_letter(col)}7)'
    c.number_format = '#,##0.00' if col == 5 else '#,##0'
    style_output(c, bold=True)

# Data bars
ws_seg.conditional_formatting.add('G4:G7',
    DataBarRule(start_type='min', end_type='max', color=PURPLE))

# ============================================================
# SHEET 5: CONFIG
# ============================================================
ws_conf = wb.create_sheet("Config")
ws_conf.sheet_properties.tabColor = GRAY_500
ws_conf.sheet_view.showGridLines = False
cw = {1:25, 2:45}
for c, w in cw.items():
    ws_conf.column_dimensions[get_column_letter(c)].width = w
fill_white(ws_conf, 1, 25, 1, 3)
ws_conf['A1'].value = "v1.0.0"
ws_conf['A1'].font = Font(name='Calibri', size=9, color=GRAY_400)
ws_conf.merge_cells('A3:B3')
style_title(ws_conf['A3'], 14)
ws_conf['A3'].value = "CONFIGURACION"

settings = [
    ("Producto", "Calculadora de ARPU"),
    ("Version", "v1.0.0"),
    ("Marca", "No Somos Ignorantes"),
    ("Precio", "Bs. 49"),
    ("Proteccion", "nsi2024"),
    ("", ""),
    ("INSTRUCCIONES", ""),
    ("1.", "En 'Clientes', registra tus clientes con segmento y estado."),
    ("2.", "En 'Ingresos', registra cada pago recibido."),
    ("3.", "En 'Segmentos', define los precios de cada plan."),
    ("4.", "ARPU = Ingreso Total / Usuarios Activos"),
    ("5.", "LTV = ARPU x Vida Promedio del Cliente (en meses)"),
    ("6.", "Para desbloquear: Revisar -> Desproteger hoja -> nsi2024"),
]
for i, (label, value) in enumerate(settings):
    r = 5 + i
    ws_conf.cell(row=r, column=1).value = label
    ws_conf.cell(row=r, column=1).font = Font(name='Calibri', size=10, bold=True, color=GRAY_700)
    ws_conf.cell(row=r, column=2).value = value
    ws_conf.cell(row=r, column=2).font = Font(name='Calibri', size=10, color=GRAY_600)

# ============================================================
# PROTECTION
# ============================================================
for sheet in wb.worksheets:
    sheet.protection.sheet = True
    sheet.protection.password = "nsi2024"
    sheet.protection.enable()

# Unlock inputs
unlock_cell(ws['C7'])
unlock_cell(ws['H12'])

for row in range(4, 4+MAX_CUSTOMERS):
    for col in [2, 3, 4, 5, 6, 7]:
        unlock_cell(ws_cl.cell(row=row, column=col))

for row in range(4, 4+MAX_REVENUE):
    for col in range(1, 7):
        unlock_cell(ws_ing.cell(row=row, column=col))

for i in range(4):
    unlock_cell(ws_seg.cell(row=4+i, column=3))

# ============================================================
# SAVE & VERIFY
# ============================================================
out = r"D:\Landing-Page_marketplace\excel_products\Calculadora_ARPU_NSI.xlsx"
wb.save(out)
print(f"[OK] Saved: {out}")

from openpyxl import load_workbook
wb2 = load_workbook(out)
fc = sum(1 for s in wb2.worksheets for row in s.iter_rows() for cell in row
         if isinstance(cell.value, str) and cell.value.startswith('='))
print(f"[OK] Sheets: {wb2.sheetnames}")
print(f"[OK] Total formulas: {fc}")
print(f"[OK] Charts: Dashboard={len(ws._charts)}")
