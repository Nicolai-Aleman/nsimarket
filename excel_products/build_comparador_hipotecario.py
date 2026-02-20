"""
Build: Comparador de Creditos Hipotecarios (Bs. 49)
No Somos Ignorantes v1.0
Compare up to 5 mortgage offers, full amortization, sensitivity analysis, winner finder.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule, DataBarRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ============================================================
# COLOR PALETTE
# ============================================================
GOLD = "D4AF37"
DARK_BG = "1A1A2E"
CARD_BG = "16213E"
ACCENT_BLUE = "0F3460"
WHITE = "FFFFFF"
BLACK = "000000"
LIGHT_YELLOW = "FFF9C4"
LIGHT_GREEN = "E8F5E9"
SUCCESS = "10B981"
WARNING = "F59E0B"
ERROR_RED = "EF4444"
INFO_BLUE = "3B82F6"
GRAY_100 = "F3F4F6"
GRAY_200 = "E5E7EB"
GRAY_300 = "D1D5DB"
GRAY_500 = "6B7280"
GRAY_700 = "374151"
GRAY_900 = "111827"

# Bank-specific colors (5 distinct)
BANK_COLORS = ["2563EB", "DC2626", "059669", "D97706", "7C3AED"]
BANK_COLORS_LIGHT = ["DBEAFE", "FEE2E2", "D1FAE5", "FEF3C7", "EDE9FE"]
BANK_NAMES_DEFAULT = ["Banco Nacional de Bolivia", "Banco Mercantil Santa Cruz",
                      "Banco Union", "BancoSol", "Banco BISA"]
WINNER_GREEN = "047857"
WINNER_GREEN_LIGHT = "A7F3D0"

# ============================================================
# STYLE HELPERS
# ============================================================
thin_border = Border(
    left=Side(style='thin', color=GRAY_300),
    right=Side(style='thin', color=GRAY_300),
    top=Side(style='thin', color=GRAY_300),
    bottom=Side(style='thin', color=GRAY_300)
)
gold_border = Border(
    left=Side(style='thin', color=GOLD),
    right=Side(style='thin', color=GOLD),
    top=Side(style='thin', color=GOLD),
    bottom=Side(style='thin', color=GOLD)
)
thick_bottom = Border(bottom=Side(style='medium', color=GRAY_700))

def style_header(cell, bg=DARK_BG, fg=WHITE, size=12):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')

def style_subheader(cell, bg=ACCENT_BLUE, fg=WHITE):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=11)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def style_input(cell):
    cell.font = Font(name='Calibri', size=11, color="0000FF")
    cell.fill = PatternFill('solid', fgColor=LIGHT_YELLOW)
    cell.border = gold_border
    cell.alignment = Alignment(horizontal='right', vertical='center')

def style_output(cell, bold=False):
    cell.font = Font(name='Calibri', size=11, color=GRAY_900, bold=bold)
    cell.fill = PatternFill('solid', fgColor=LIGHT_GREEN)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='right', vertical='center')

def style_label(cell, bold=False, indent=0):
    cell.font = Font(name='Calibri', size=11, color=GRAY_700, bold=bold)
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=indent)

def style_section_header(cell, bg, fg=WHITE):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=11)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def fmt_currency(cell):
    cell.number_format = '#,##0.00'

def fmt_pct(cell):
    cell.number_format = '0.00%'

def fmt_int(cell):
    cell.number_format = '#,##0'

def rh(ws, row, height):
    ws.row_dimensions[row].height = height

def fill_row(ws, row, cols, color):
    for c in cols:
        ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=color)

NUM_BANKS = 5
AMORT_MONTHS = 360  # 30 years max

# Bank sheet names
BANK_SHEETS = [f"Banco{i+1}" for i in range(NUM_BANKS)]

# ============================================================
# SHEET 1: DASHBOARD
# ============================================================
ws = wb.active
ws.title = "Dashboard"
ws.sheet_properties.tabColor = GOLD

ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 18
ws.column_dimensions['G'].width = 18
ws.column_dimensions['H'].width = 18
ws.column_dimensions['I'].width = 2

ws.freeze_panes = 'B6'

# ── HEADER BANNER ──
rh(ws, 1, 8)
ws.merge_cells('B1:H1')
fill_row(ws, 1, range(2, 9), DARK_BG)

ws.merge_cells('B2:H2')
h = ws['B2']
h.value = "COMPARADOR DE CREDITOS HIPOTECARIOS"
h.font = Font(name='Calibri', bold=True, color=GOLD, size=20)
h.fill = PatternFill('solid', fgColor=DARK_BG)
h.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws, 2, range(2, 9), DARK_BG)
rh(ws, 2, 42)

ws.merge_cells('B3:H3')
sub = ws['B3']
sub.value = "No Somos Ignorantes  |  v1.0  |  Encuentra el mejor credito para tu hogar"
sub.font = Font(name='Calibri', color=GRAY_300, size=10, italic=True)
sub.fill = PatternFill('solid', fgColor=DARK_BG)
sub.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws, 3, range(2, 9), DARK_BG)
rh(ws, 3, 24)

rh(ws, 4, 8)

# ── GLOBAL INPUTS ──
r = 5
rh(ws, r, 30)
ws.merge_cells(f'B{r}:H{r}')
sec = ws[f'B{r}']
sec.value = "DATOS DEL PRESTAMO"
style_section_header(sec, ACCENT_BLUE)
fill_row(ws, r, range(3, 9), ACCENT_BLUE)

# Monto del prestamo
rh(ws, 6, 32)
lbl = ws.cell(row=6, column=2, value="MONTO DEL PRESTAMO (Bs.)")
lbl.font = Font(name='Calibri', bold=True, color=GRAY_900, size=12)
lbl.alignment = Alignment(horizontal='left', vertical='center')
ws.merge_cells('C6:D6')
amt = ws['C6']
amt.value = 350000
style_input(amt)
amt.font = Font(name='Calibri', size=14, bold=True, color="0000FF")
fmt_currency(amt)

# Plazo
rh(ws, 7, 32)
lbl2 = ws.cell(row=7, column=2, value="PLAZO DESEADO (meses)")
lbl2.font = Font(name='Calibri', bold=True, color=GRAY_900, size=12)
lbl2.alignment = Alignment(horizontal='left', vertical='center')
ws.merge_cells('C7:D7')
term = ws['C7']
term.value = 240
style_input(term)
term.font = Font(name='Calibri', size=14, bold=True, color="0000FF")
fmt_int(term)

# Plazo in years display
lbl3 = ws.cell(row=7, column=5, value='=IFERROR(C7/12&" anios","")')
lbl3.font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)

# Data validations for global inputs
dv_amount = DataValidation(type="whole", operator="between", formula1=10000, formula2=5000000)
dv_amount.error = "El monto debe estar entre Bs. 10,000 y Bs. 5,000,000"
dv_amount.errorTitle = "Monto invalido"
dv_amount.prompt = "Ingresa el monto del prestamo hipotecario"
dv_amount.promptTitle = "Monto del Prestamo"
ws.add_data_validation(dv_amount)
dv_amount.add('C6')

dv_term = DataValidation(type="whole", operator="between", formula1=12, formula2=360)
dv_term.error = "El plazo debe estar entre 12 y 360 meses (1-30 anios)"
dv_term.errorTitle = "Plazo invalido"
dv_term.prompt = "Ingresa el plazo en meses (ej: 240 = 20 anios)"
dv_term.promptTitle = "Plazo en Meses"
ws.add_data_validation(dv_term)
dv_term.add('C7')

rh(ws, 8, 8)

# ── COMPARISON TABLE ──
r = 9
rh(ws, r, 30)
ws.merge_cells(f'B{r}:H{r}')
sec = ws[f'B{r}']
sec.value = "COMPARACION RAPIDA"
style_section_header(sec, CARD_BG, fg=GOLD)
fill_row(ws, r, range(3, 9), CARD_BG)

# Table headers
r = 10
rh(ws, r, 28)
comp_headers = ['', 'BANCO', 'TASA ANUAL', 'CUOTA MENSUAL', 'TOTAL A PAGAR',
                'TOTAL INTERESES', 'CAE', 'DIFERENCIA']
for i, hdr in enumerate(comp_headers):
    c = ws.cell(row=r, column=i + 1, value=hdr)
    if i >= 1:
        style_header(c, bg=CARD_BG, fg=GOLD, size=9)
        c.border = Border(bottom=Side(style='medium', color=GOLD))

# Bank comparison rows (pulling from individual bank sheets)
for b in range(NUM_BANKS):
    row = 11 + b
    rh(ws, row, 30)
    bg = BANK_COLORS_LIGHT[b]

    # Bank name
    name_c = ws.cell(row=row, column=2)
    name_c.value = f'={BANK_SHEETS[b]}!C3'
    name_c.font = Font(name='Calibri', bold=True, color=BANK_COLORS[b], size=11)
    name_c.fill = PatternFill('solid', fgColor=bg)
    name_c.border = thin_border
    name_c.alignment = Alignment(horizontal='left', vertical='center')

    # Tasa anual
    rate_c = ws.cell(row=row, column=3)
    rate_c.value = f'={BANK_SHEETS[b]}!C5'
    style_output(rate_c)
    rate_c.fill = PatternFill('solid', fgColor=bg)
    fmt_pct(rate_c)

    # Cuota mensual (PMT)
    cuota_c = ws.cell(row=row, column=4)
    cuota_c.value = f'={BANK_SHEETS[b]}!C11'
    style_output(cuota_c, bold=True)
    cuota_c.fill = PatternFill('solid', fgColor=bg)
    fmt_currency(cuota_c)

    # Total a pagar
    total_c = ws.cell(row=row, column=5)
    total_c.value = f'={BANK_SHEETS[b]}!C12'
    style_output(total_c)
    total_c.fill = PatternFill('solid', fgColor=bg)
    fmt_currency(total_c)

    # Total intereses
    int_c = ws.cell(row=row, column=6)
    int_c.value = f'={BANK_SHEETS[b]}!C13'
    style_output(int_c)
    int_c.fill = PatternFill('solid', fgColor=bg)
    fmt_currency(int_c)

    # CAE
    cae_c = ws.cell(row=row, column=7)
    cae_c.value = f'={BANK_SHEETS[b]}!C14'
    style_output(cae_c)
    cae_c.fill = PatternFill('solid', fgColor=bg)
    fmt_pct(cae_c)

    # Diferencia vs best (first valid bank as reference)
    diff_c = ws.cell(row=row, column=8)
    diff_c.value = f'=IFERROR(E{row}-MIN($E$11:$E$15),"")'
    style_output(diff_c)
    diff_c.fill = PatternFill('solid', fgColor=bg)
    fmt_currency(diff_c)

# Conditional formatting: highlight winner row (min total)
for b in range(NUM_BANKS):
    row = 11 + b
    rng = f'B{row}:H{row}'
    ws.conditional_formatting.add(rng,
        FormulaRule(formula=[f'$E{row}=MIN($E$11:$E$15)'],
                    fill=PatternFill('solid', fgColor=WINNER_GREEN_LIGHT),
                    font=Font(bold=True, color=WINNER_GREEN)))

# Data bars on total column
ws.conditional_formatting.add('E11:E15',
    DataBarRule(start_type='min', end_type='max', color=ACCENT_BLUE))

# Data bars on cuota
ws.conditional_formatting.add('D11:D15',
    DataBarRule(start_type='min', end_type='max', color=SUCCESS))

# Conditional: difference = 0 is green, > 0 is red
ws.conditional_formatting.add('H11:H15',
    CellIsRule(operator='equal', formula=['0'],
              fill=PatternFill('solid', fgColor=WINNER_GREEN_LIGHT),
              font=Font(bold=True, color=WINNER_GREEN)))
ws.conditional_formatting.add('H11:H15',
    CellIsRule(operator='greaterThan', formula=['0'],
              fill=PatternFill('solid', fgColor="FEE2E2"),
              font=Font(bold=True, color=ERROR_RED)))

rh(ws, 16, 8)

# ── BEST OPTION FINDER ──
r = 17
rh(ws, r, 36)
ws.merge_cells(f'B{r}:H{r}')
best = ws[f'B{r}']
best.value = '=IFERROR("MEJOR OPCION: "&INDEX(B11:B15,MATCH(MIN(E11:E15),E11:E15,0))&" - Ahorras Bs. "&TEXT(MAX(E11:E15)-MIN(E11:E15),"#,##0")&" vs la peor opcion","Ingresa datos en las hojas de bancos")'
best.font = Font(name='Calibri', bold=True, color=WINNER_GREEN, size=14)
best.fill = PatternFill('solid', fgColor=WINNER_GREEN_LIGHT)
best.alignment = Alignment(horizontal='center', vertical='center')
best.border = Border(
    left=Side(style='medium', color=WINNER_GREEN),
    right=Side(style='medium', color=WINNER_GREEN),
    top=Side(style='medium', color=WINNER_GREEN),
    bottom=Side(style='medium', color=WINNER_GREEN))
fill_row(ws, r, range(3, 9), WINNER_GREEN_LIGHT)

rh(ws, 18, 8)

# ── KEY METRICS SUMMARY ──
r = 19
rh(ws, r, 30)
ws.merge_cells(f'B{r}:H{r}')
sec = ws[f'B{r}']
sec.value = "METRICAS CLAVE"
style_section_header(sec, INFO_BLUE)
fill_row(ws, r, range(3, 9), INFO_BLUE)

metrics = [
    ("Cuota mas baja (Bs.)", '=IFERROR(MIN(D11:D15),0)', True),
    ("Cuota mas alta (Bs.)", '=IFERROR(MAX(D11:D15),0)', True),
    ("Diferencia mensual (Bs.)", '=IFERROR(MAX(D11:D15)-MIN(D11:D15),0)', True),
    ("Menor costo total (Bs.)", '=IFERROR(MIN(E11:E15),0)', True),
    ("Mayor costo total (Bs.)", '=IFERROR(MAX(E11:E15),0)', True),
    ("Ahorro maximo posible (Bs.)", '=IFERROR(MAX(E11:E15)-MIN(E11:E15),0)', True),
    ("Tasa promedio (%)", '=IFERROR(AVERAGE(C11:C15),0)', False),
]

for i, (label, formula, is_cur) in enumerate(metrics):
    row = 20 + i
    rh(ws, row, 26)
    lbl = ws.cell(row=row, column=2, value=label)
    style_label(lbl, bold=True)
    lbl.fill = PatternFill('solid', fgColor=GRAY_100 if i % 2 == 0 else WHITE)
    lbl.border = thin_border
    ws.merge_cells(f'C{row}:D{row}')
    val = ws.cell(row=row, column=3)
    val.value = formula
    style_output(val, bold=True)
    if is_cur:
        fmt_currency(val)
    else:
        fmt_pct(val)

# Highlight savings row
ws.cell(row=25, column=2).font = Font(name='Calibri', bold=True, color=SUCCESS, size=12)
ws.cell(row=25, column=3).font = Font(name='Calibri', bold=True, color=SUCCESS, size=14)

rh(ws, 27, 8)

# ── CHARTS ──
# Bar chart: Total a pagar by bank
bar1 = BarChart()
bar1.type = "col"
bar1.title = "Total a Pagar por Banco (Bs.)"
bar1.y_axis.title = "Total (Bs.)"
bar1.style = 10
bar1.width = 22
bar1.height = 14

data1 = Reference(ws, min_col=5, min_row=10, max_row=15)
cats1 = Reference(ws, min_col=2, min_row=11, max_row=15)
bar1.add_data(data1, titles_from_data=True)
bar1.set_categories(cats1)

for i, color in enumerate(BANK_COLORS):
    if i < len(bar1.series[0].data_points if hasattr(bar1.series[0], 'data_points') else []):
        pass
# Color individual bars
for i in range(NUM_BANKS):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = BANK_COLORS[i]
    bar1.series[0].data_points.append(pt)

bar1.dataLabels = DataLabelList()
bar1.dataLabels.showVal = True
bar1.dataLabels.numFmt = '#,##0'
ws.add_chart(bar1, "B28")

# Bar chart: Cuota mensual by bank
bar2 = BarChart()
bar2.type = "col"
bar2.title = "Cuota Mensual por Banco (Bs.)"
bar2.y_axis.title = "Cuota (Bs.)"
bar2.style = 10
bar2.width = 22
bar2.height = 14

data2 = Reference(ws, min_col=4, min_row=10, max_row=15)
cats2 = Reference(ws, min_col=2, min_row=11, max_row=15)
bar2.add_data(data2, titles_from_data=True)
bar2.set_categories(cats2)

for i in range(NUM_BANKS):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = BANK_COLORS[i]
    bar2.series[0].data_points.append(pt)

bar2.dataLabels = DataLabelList()
bar2.dataLabels.showVal = True
bar2.dataLabels.numFmt = '#,##0'
ws.add_chart(bar2, "B44")

# Stacked bar: Capital vs Interest breakdown
bar3 = BarChart()
bar3.type = "col"
bar3.grouping = "stacked"
bar3.title = "Desglose: Capital vs Intereses por Banco"
bar3.y_axis.title = "Monto (Bs.)"
bar3.style = 10
bar3.width = 22
bar3.height = 14

# We need capital (=amount) and interest columns side by side
# Put helper data in a hidden area
helper_start = 60
ws.cell(row=helper_start, column=2, value="Banco")
ws.cell(row=helper_start, column=3, value="Capital")
ws.cell(row=helper_start, column=4, value="Intereses")
for b in range(NUM_BANKS):
    r_h = helper_start + 1 + b
    ws.cell(row=r_h, column=2).value = f'=B{11+b}'
    ws.cell(row=r_h, column=3).value = f'=C6'  # loan amount (same for all)
    ws.cell(row=r_h, column=4).value = f'=F{11+b}'  # total interest
    fmt_currency(ws.cell(row=r_h, column=3))
    fmt_currency(ws.cell(row=r_h, column=4))

cap_data = Reference(ws, min_col=3, min_row=helper_start, max_row=helper_start + NUM_BANKS)
int_data = Reference(ws, min_col=4, min_row=helper_start, max_row=helper_start + NUM_BANKS)
cats3 = Reference(ws, min_col=2, min_row=helper_start + 1, max_row=helper_start + NUM_BANKS)
bar3.add_data(cap_data, titles_from_data=True)
bar3.add_data(int_data, titles_from_data=True)
bar3.set_categories(cats3)
bar3.series[0].graphicalProperties.solidFill = SUCCESS
bar3.series[1].graphicalProperties.solidFill = ERROR_RED
ws.add_chart(bar3, "B60")


# ============================================================
# SHEETS 2-6: INDIVIDUAL BANK SHEETS
# ============================================================
for b in range(NUM_BANKS):
    ws_b = wb.create_sheet(BANK_SHEETS[b])
    ws_b.sheet_properties.tabColor = BANK_COLORS[b]

    ws_b.column_dimensions['A'].width = 2
    ws_b.column_dimensions['B'].width = 35
    ws_b.column_dimensions['C'].width = 22
    ws_b.column_dimensions['D'].width = 22
    ws_b.column_dimensions['E'].width = 2

    ws_b.freeze_panes = 'B5'

    # Header
    ws_b.merge_cells('B1:D1')
    fill_row(ws_b, 1, range(2, 5), DARK_BG)
    rh(ws_b, 1, 8)

    ws_b.merge_cells('B2:D2')
    h = ws_b['B2']
    h.value = f"BANCO {b+1} - DATOS DEL CREDITO"
    h.font = Font(name='Calibri', bold=True, color=BANK_COLORS[b], size=16)
    h.fill = PatternFill('solid', fgColor=DARK_BG)
    h.alignment = Alignment(horizontal='center', vertical='center')
    fill_row(ws_b, 2, range(2, 5), DARK_BG)
    rh(ws_b, 2, 38)

    # Bank name input
    rh(ws_b, 3, 32)
    lbl = ws_b.cell(row=3, column=2, value="Nombre del Banco")
    style_label(lbl, bold=True)
    name_c = ws_b.cell(row=3, column=3, value=BANK_NAMES_DEFAULT[b])
    style_input(name_c)
    name_c.alignment = Alignment(horizontal='left', vertical='center')
    name_c.font = Font(name='Calibri', size=12, bold=True, color="0000FF")

    rh(ws_b, 4, 8)

    # ── INPUT SECTION ──
    # Row 5: Tasa anual
    inputs_config = [
        (5, "Tasa de interes anual (%)", 0.065 + b * 0.005, True),
        (6, "Comision de apertura (%)", 0.01 + b * 0.002, True),
        (7, "Seguro de desgravamen anual (%)", 0.0005 + b * 0.0001, True),
        (8, "Otros costos fijos (Bs.)", 500 + b * 200, False),
    ]

    # Data validations
    dv_rate = DataValidation(type="decimal", operator="between", formula1=0, formula2=1)
    dv_rate.error = "La tasa debe estar entre 0% y 100%"
    dv_rate.errorTitle = "Tasa invalida"
    dv_rate.prompt = "Ingresa la tasa anual como decimal (ej: 0.065 = 6.5%)"
    dv_rate.promptTitle = "Tasa de Interes Anual"
    ws_b.add_data_validation(dv_rate)

    dv_pct = DataValidation(type="decimal", operator="between", formula1=0, formula2=0.5)
    dv_pct.error = "El porcentaje debe estar entre 0% y 50%"
    dv_pct.errorTitle = "Porcentaje invalido"
    ws_b.add_data_validation(dv_pct)

    dv_cost = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1=0)
    dv_cost.error = "El costo no puede ser negativo"
    dv_cost.errorTitle = "Costo invalido"
    ws_b.add_data_validation(dv_cost)

    for (row, label, default, is_pct) in inputs_config:
        rh(ws_b, row, 30)
        lbl = ws_b.cell(row=row, column=2, value=label)
        style_label(lbl, bold=True)
        lbl.fill = PatternFill('solid', fgColor=BANK_COLORS_LIGHT[b])
        lbl.border = thin_border

        val = ws_b.cell(row=row, column=3, value=default)
        style_input(val)
        if is_pct:
            fmt_pct(val)
            if row == 5:
                dv_rate.add(f'C{row}')
            else:
                dv_pct.add(f'C{row}')
        else:
            fmt_currency(val)
            dv_cost.add(f'C{row}')

    rh(ws_b, 9, 8)

    # ── OUTPUT SECTION ──
    r = 10
    rh(ws_b, r, 30)
    ws_b.merge_cells(f'B{r}:D{r}')
    sec = ws_b[f'B{r}']
    sec.value = "RESULTADOS DEL CALCULO"
    style_section_header(sec, BANK_COLORS[b])
    fill_row(ws_b, r, range(3, 5), BANK_COLORS[b])

    # Cuota mensual (PMT formula)
    # PMT(rate/12, term, -amount) + insurance + commission amortized
    outputs = [
        (11, "Cuota mensual (Bs.)",
         f'=IFERROR(-PMT(C5/12,Dashboard!$C$7,Dashboard!$C$6)+(Dashboard!$C$6*C7/12),0)'),
        (12, "Total a pagar (Bs.)",
         f'=IFERROR(C11*Dashboard!$C$7+C8+(Dashboard!$C$6*C6),0)'),
        (13, "Total intereses (Bs.)",
         f'=IFERROR(C12-Dashboard!$C$6,0)'),
        (14, "CAE - Costo Anual Equivalente",
         f'=IFERROR(((1+(C5/12+C7/12))^12-1)+C6/IFERROR(Dashboard!$C$7/12,1),0)'),
        (15, "Cuota/Ingreso recomendado (%)",
         f'=IFERROR(IF(D15>0,C11/D15,0),0)'),
    ]

    for (row, label, formula) in outputs:
        rh(ws_b, row, 32)
        lbl = ws_b.cell(row=row, column=2, value=label)
        style_label(lbl, bold=True)
        lbl.fill = PatternFill('solid', fgColor=LIGHT_GREEN)
        lbl.border = thin_border

        val = ws_b.cell(row=row, column=3)
        val.value = formula
        style_output(val, bold=True)
        if row == 14 or row == 15:
            fmt_pct(val)
        else:
            fmt_currency(val)

    # Ingreso mensual input for affordability check
    rh(ws_b, 15, 32)
    inc_lbl = ws_b.cell(row=15, column=2, value="Cuota/Ingreso recomendado (%)")
    style_label(inc_lbl, bold=True)
    inc_lbl.fill = PatternFill('solid', fgColor=LIGHT_GREEN)
    inc_lbl.border = thin_border

    inc_input = ws_b.cell(row=15, column=4, value=8000 + b * 500)
    style_input(inc_input)
    fmt_currency(inc_input)
    inc_input_lbl = ws_b.cell(row=15, column=4)

    # Affordability warning
    rh(ws_b, 16, 30)
    ws_b.merge_cells(f'B16:D16')
    afford = ws_b['B16']
    afford.value = f'=IF(C15>0.35,"ALERTA: La cuota supera el 35% de tu ingreso. Riesgo alto.","OK: La cuota esta dentro del rango recomendado (<35%)")'
    afford.font = Font(name='Calibri', bold=True, size=11)
    afford.alignment = Alignment(horizontal='center', vertical='center')
    afford.border = thin_border

    ws_b.conditional_formatting.add('B16:D16',
        FormulaRule(formula=['C15>0.35'],
                    fill=PatternFill('solid', fgColor="FEE2E2"),
                    font=Font(bold=True, color=ERROR_RED)))
    ws_b.conditional_formatting.add('B16:D16',
        FormulaRule(formula=['C15<=0.35'],
                    fill=PatternFill('solid', fgColor=WINNER_GREEN_LIGHT),
                    font=Font(bold=True, color=WINNER_GREEN)))

    rh(ws_b, 17, 8)

    # ── MINI AMORTIZATION (first 12 months + last 12 months) ──
    r = 18
    rh(ws_b, r, 30)
    ws_b.merge_cells(f'B{r}:D{r}')
    sec = ws_b[f'B{r}']
    sec.value = "TABLA DE AMORTIZACION (primeros 12 meses)"
    style_section_header(sec, GRAY_700)
    fill_row(ws_b, r, range(3, 5), GRAY_700)

    r = 19
    rh(ws_b, r, 24)
    amort_headers = ['', 'MES', 'CUOTA', 'CAPITAL', 'INTERES', 'SALDO']
    # Need more columns for mini amort
    ws_b.column_dimensions['D'].width = 18
    ws_b.column_dimensions['E'].width = 18
    ws_b.column_dimensions['F'].width = 18

    mini_headers = ['MES', 'CUOTA (Bs.)', 'A CAPITAL (Bs.)', 'INTERES (Bs.)', 'SALDO (Bs.)']
    for i, hdr in enumerate(mini_headers):
        c = ws_b.cell(row=r, column=2 + i, value=hdr)
        style_header(c, bg=CARD_BG, fg=GOLD, size=9)

    for m in range(12):
        row = 20 + m
        rh(ws_b, row, 22)

        # Month
        ws_b.cell(row=row, column=2, value=m + 1).alignment = Alignment(horizontal='center')
        ws_b.cell(row=row, column=2).border = thin_border
        ws_b.cell(row=row, column=2).font = Font(name='Calibri', size=10)

        # Cuota
        c_cuota = ws_b.cell(row=row, column=3)
        c_cuota.value = f'=IFERROR(C11,0)'
        style_output(c_cuota)
        fmt_currency(c_cuota)

        # Interest
        c_int = ws_b.cell(row=row, column=5)
        if m == 0:
            c_int.value = f'=IFERROR(ROUND(Dashboard!$C$6*C5/12,2),0)'
        else:
            c_int.value = f'=IFERROR(ROUND(F{row-1}*C5/12,2),0)'
        style_output(c_int)
        fmt_currency(c_int)
        c_int.fill = PatternFill('solid', fgColor="FEE2E2")

        # Capital
        c_cap = ws_b.cell(row=row, column=4)
        c_cap.value = f'=IFERROR(C{row}-E{row},0)'
        style_output(c_cap)
        fmt_currency(c_cap)
        c_cap.fill = PatternFill('solid', fgColor=LIGHT_GREEN)

        # Saldo
        c_sal = ws_b.cell(row=row, column=6)
        if m == 0:
            c_sal.value = f'=IFERROR(Dashboard!$C$6-D{row},0)'
        else:
            c_sal.value = f'=IFERROR(F{row-1}-D{row},0)'
        style_output(c_sal)
        fmt_currency(c_sal)

    # Totals
    tot_row = 32
    rh(ws_b, tot_row, 26)
    ws_b.cell(row=tot_row, column=2, value="TOTAL 12 MESES").font = Font(name='Calibri', bold=True, size=10, color=WHITE)
    ws_b.cell(row=tot_row, column=2).fill = PatternFill('solid', fgColor=GRAY_700)
    ws_b.cell(row=tot_row, column=2).alignment = Alignment(horizontal='center')
    for col in [3, 4, 5, 6]:
        c = ws_b.cell(row=tot_row, column=col)
        if col <= 5:
            c.value = f'=SUM({get_column_letter(col)}20:{get_column_letter(col)}31)'
        else:
            c.value = f'=F31'
        c.font = Font(name='Calibri', bold=True, size=10, color=WHITE)
        c.fill = PatternFill('solid', fgColor=GRAY_700)
        fmt_currency(c)


# ============================================================
# SHEET 7: AMORTIZACION (Full amortization for best bank)
# ============================================================
ws_a = wb.create_sheet("Amortizacion")
ws_a.sheet_properties.tabColor = INFO_BLUE

ws_a.column_dimensions['A'].width = 10
ws_a.column_dimensions['B'].width = 20
ws_a.column_dimensions['C'].width = 20
ws_a.column_dimensions['D'].width = 20
ws_a.column_dimensions['E'].width = 20
ws_a.column_dimensions['F'].width = 20
ws_a.column_dimensions['G'].width = 20

ws_a.freeze_panes = 'A6'

# Header
ws_a.merge_cells('A1:G1')
h = ws_a['A1']
h.value = "TABLA DE AMORTIZACION COMPLETA"
h.font = Font(name='Calibri', bold=True, color=GOLD, size=18)
h.fill = PatternFill('solid', fgColor=DARK_BG)
h.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws_a, 1, range(1, 8), DARK_BG)
rh(ws_a, 1, 40)

ws_a.merge_cells('A2:G2')
sub = ws_a['A2']
sub.value = "Amortizacion Frances (cuota fija) - Selecciona el banco a analizar"
sub.font = Font(name='Calibri', color=GRAY_300, size=10, italic=True)
sub.fill = PatternFill('solid', fgColor=DARK_BG)
sub.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws_a, 2, range(1, 8), DARK_BG)
rh(ws_a, 2, 24)

# Bank selector
rh(ws_a, 3, 32)
ws_a.cell(row=3, column=1, value="Selecciona Banco:").font = Font(name='Calibri', bold=True, size=12)
sel = ws_a.cell(row=3, column=2, value=1)
style_input(sel)
sel.font = Font(name='Calibri', size=14, bold=True, color="0000FF")
fmt_int(sel)

dv_bank = DataValidation(type="whole", operator="between", formula1=1, formula2=5)
dv_bank.error = "Selecciona un banco del 1 al 5"
dv_bank.errorTitle = "Banco invalido"
dv_bank.prompt = "Ingresa el numero de banco (1-5) para ver su amortizacion"
dv_bank.promptTitle = "Seleccionar Banco"
ws_a.add_data_validation(dv_bank)
dv_bank.add('B3')

# Display selected bank info
ws_a.cell(row=3, column=3, value="Banco:").font = Font(name='Calibri', bold=True, size=10)
ws_a.cell(row=3, column=4).value = '=IFERROR(CHOOSE(B3,Banco1!C3,Banco2!C3,Banco3!C3,Banco4!C3,Banco5!C3),"")'
ws_a.cell(row=3, column=4).font = Font(name='Calibri', bold=True, size=12, color=ACCENT_BLUE)

ws_a.cell(row=3, column=5, value="Tasa:").font = Font(name='Calibri', bold=True, size=10)
ws_a.cell(row=3, column=6).value = '=IFERROR(CHOOSE(B3,Banco1!C5,Banco2!C5,Banco3!C5,Banco4!C5,Banco5!C5),"")'
ws_a.cell(row=3, column=6).font = Font(name='Calibri', bold=True, size=12, color=ACCENT_BLUE)
fmt_pct(ws_a.cell(row=3, column=6))

rh(ws_a, 4, 8)

# Column headers
r = 5
rh(ws_a, r, 28)
amort_headers = ['MES', 'SALDO INICIAL', 'CUOTA', 'A CAPITAL', 'A INTERESES', 'SALDO FINAL', '% PAGADO']
for i, hdr in enumerate(amort_headers):
    c = ws_a.cell(row=r, column=i + 1, value=hdr)
    style_header(c, bg=CARD_BG, fg=GOLD, size=9)

# Amortization rows (up to 360 months)
# Using CHOOSE to select rate from the chosen bank
AMORT_DISPLAY = 360

for m in range(AMORT_DISPLAY):
    row = 6 + m
    rh(ws_a, row, 20)

    # Month
    ws_a.cell(row=row, column=1, value=m + 1).alignment = Alignment(horizontal='center')
    ws_a.cell(row=row, column=1).font = Font(name='Calibri', size=9, color=GRAY_700)
    ws_a.cell(row=row, column=1).border = thin_border

    # Starting balance
    sb = ws_a.cell(row=row, column=2)
    if m == 0:
        sb.value = '=Dashboard!C6'
    else:
        sb.value = f'=IF(F{row-1}<=0.01,0,F{row-1})'
    style_output(sb)
    fmt_currency(sb)
    sb.font = Font(name='Calibri', size=9, color=GRAY_900)

    # Monthly rate reference
    rate_ref = 'CHOOSE($B$3,Banco1!$C$5,Banco2!$C$5,Banco3!$C$5,Banco4!$C$5,Banco5!$C$5)/12'

    # Cuota (fixed PMT)
    cuota = ws_a.cell(row=row, column=3)
    cuota.value = f'=IF(B{row}<=0.01,0,IFERROR(-PMT({rate_ref},Dashboard!$C$7,Dashboard!$C$6),0))'
    style_output(cuota)
    fmt_currency(cuota)
    cuota.font = Font(name='Calibri', size=9, color=GRAY_900)

    # Interest
    interest = ws_a.cell(row=row, column=5)
    interest.value = f'=IF(B{row}<=0.01,0,ROUND(B{row}*{rate_ref},2))'
    style_output(interest)
    fmt_currency(interest)
    interest.font = Font(name='Calibri', size=9, color=GRAY_900)
    interest.fill = PatternFill('solid', fgColor="FEE2E2")

    # Capital
    capital = ws_a.cell(row=row, column=4)
    capital.value = f'=IF(B{row}<=0.01,0,MIN(C{row}-E{row},B{row}))'
    style_output(capital)
    fmt_currency(capital)
    capital.font = Font(name='Calibri', size=9, color=GRAY_900)
    capital.fill = PatternFill('solid', fgColor=LIGHT_GREEN)

    # Ending balance
    ending = ws_a.cell(row=row, column=6)
    ending.value = f'=IF(B{row}<=0.01,0,MAX(B{row}-D{row},0))'
    style_output(ending)
    fmt_currency(ending)
    ending.font = Font(name='Calibri', size=9, color=GRAY_900)

    # % paid
    pct_paid = ws_a.cell(row=row, column=7)
    pct_paid.value = f'=IFERROR(IF(Dashboard!C6>0,1-(F{row}/Dashboard!C6),0),0)'
    style_output(pct_paid)
    fmt_pct(pct_paid)
    pct_paid.font = Font(name='Calibri', size=9, color=GRAY_900)

# Highlight when balance reaches zero
ws_a.conditional_formatting.add(f'F6:F{6+AMORT_DISPLAY-1}',
    CellIsRule(operator='lessThanOrEqual', formula=['0.01'],
              fill=PatternFill('solid', fgColor=WINNER_GREEN_LIGHT),
              font=Font(bold=True, color=WINNER_GREEN)))

# Progress data bar
ws_a.conditional_formatting.add(f'G6:G{6+AMORT_DISPLAY-1}',
    DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1,
                color=SUCCESS))

# Summary totals
tot_row = 6 + AMORT_DISPLAY
rh(ws_a, tot_row, 28)
ws_a.cell(row=tot_row, column=1, value="TOTALES").font = Font(name='Calibri', bold=True, color=WHITE, size=10)
ws_a.cell(row=tot_row, column=1).fill = PatternFill('solid', fgColor=DARK_BG)
ws_a.cell(row=tot_row, column=1).alignment = Alignment(horizontal='center')

for col in [3, 4, 5]:
    c = ws_a.cell(row=tot_row, column=col)
    c.value = f'=SUM({get_column_letter(col)}6:{get_column_letter(col)}{6+AMORT_DISPLAY-1})'
    c.font = Font(name='Calibri', bold=True, color=WHITE, size=10)
    c.fill = PatternFill('solid', fgColor=DARK_BG)
    fmt_currency(c)

for col in [2, 6, 7]:
    ws_a.cell(row=tot_row, column=col).fill = PatternFill('solid', fgColor=DARK_BG)

# Charts at the bottom of amort sheet
chart_row = tot_row + 3

# Line chart: balance over time (sample every 12 months)
line1 = LineChart()
line1.title = "Evolucion del Saldo del Prestamo"
line1.y_axis.title = "Saldo (Bs.)"
line1.x_axis.title = "Mes"
line1.style = 10
line1.width = 28
line1.height = 14

bal_data = Reference(ws_a, min_col=6, min_row=5, max_row=6 + AMORT_DISPLAY - 1)
months_cats = Reference(ws_a, min_col=1, min_row=6, max_row=6 + AMORT_DISPLAY - 1)
line1.add_data(bal_data, titles_from_data=True)
line1.set_categories(months_cats)
s = line1.series[0]
s.graphicalProperties.line.solidFill = ACCENT_BLUE
s.graphicalProperties.line.width = 20000
ws_a.add_chart(line1, f"A{chart_row}")

# Area chart: capital vs interest over time
area1 = BarChart()
area1.type = "col"
area1.grouping = "stacked"
area1.title = "Capital vs Intereses Acumulados"
area1.y_axis.title = "Monto (Bs.)"
area1.style = 10
area1.width = 28
area1.height = 14

# Use every 12th row for cleaner chart
helper_row = tot_row + 20
ws_a.cell(row=helper_row, column=1, value="Anio")
ws_a.cell(row=helper_row, column=2, value="Capital Acum.")
ws_a.cell(row=helper_row, column=3, value="Interes Acum.")

for y in range(30):
    r_y = helper_row + 1 + y
    month_row = 6 + (y + 1) * 12 - 1
    if month_row > 6 + AMORT_DISPLAY - 1:
        break
    ws_a.cell(row=r_y, column=1, value=y + 1)
    ws_a.cell(row=r_y, column=2).value = f'=IFERROR(SUM(D6:D{month_row}),0)'
    ws_a.cell(row=r_y, column=3).value = f'=IFERROR(SUM(E6:E{month_row}),0)'
    fmt_currency(ws_a.cell(row=r_y, column=2))
    fmt_currency(ws_a.cell(row=r_y, column=3))

last_helper = helper_row + min(30, AMORT_DISPLAY // 12)
cap_ref = Reference(ws_a, min_col=2, min_row=helper_row, max_row=last_helper)
int_ref = Reference(ws_a, min_col=3, min_row=helper_row, max_row=last_helper)
yr_cats = Reference(ws_a, min_col=1, min_row=helper_row + 1, max_row=last_helper)
area1.add_data(cap_ref, titles_from_data=True)
area1.add_data(int_ref, titles_from_data=True)
area1.set_categories(yr_cats)
area1.series[0].graphicalProperties.solidFill = SUCCESS
area1.series[1].graphicalProperties.solidFill = ERROR_RED
ws_a.add_chart(area1, f"A{chart_row + 16}")


# ============================================================
# SHEET 8: SENSIBILIDAD (What-If Analysis)
# ============================================================
ws_s = wb.create_sheet("Sensibilidad")
ws_s.sheet_properties.tabColor = WARNING

ws_s.column_dimensions['A'].width = 2
ws_s.column_dimensions['B'].width = 18
ws_s.column_dimensions['C'].width = 16
ws_s.column_dimensions['D'].width = 16
ws_s.column_dimensions['E'].width = 16
ws_s.column_dimensions['F'].width = 16
ws_s.column_dimensions['G'].width = 16
ws_s.column_dimensions['H'].width = 16
ws_s.column_dimensions['I'].width = 2

ws_s.freeze_panes = 'C6'

# Header
ws_s.merge_cells('B1:H1')
h = ws_s['B1']
h.value = "ANALISIS DE SENSIBILIDAD"
h.font = Font(name='Calibri', bold=True, color=GOLD, size=18)
h.fill = PatternFill('solid', fgColor=DARK_BG)
h.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws_s, 1, range(2, 9), DARK_BG)
rh(ws_s, 1, 40)

ws_s.merge_cells('B2:H2')
sub = ws_s['B2']
sub.value = "Que pasa si cambia la tasa o el plazo? Analiza diferentes escenarios"
sub.font = Font(name='Calibri', color=GRAY_300, size=10, italic=True)
sub.fill = PatternFill('solid', fgColor=DARK_BG)
sub.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws_s, 2, range(2, 9), DARK_BG)
rh(ws_s, 2, 24)

rh(ws_s, 3, 8)

# ── TABLE 1: Cuota mensual varies by Rate x Term ──
r = 4
ws_s.merge_cells(f'B{r}:H{r}')
sec = ws_s[f'B{r}']
sec.value = "CUOTA MENSUAL (Bs.) segun Tasa y Plazo"
style_section_header(sec, ACCENT_BLUE)
fill_row(ws_s, r, range(3, 9), ACCENT_BLUE)
rh(ws_s, r, 28)

# Term options (columns): 10, 15, 20, 25, 30 years
terms_years = [10, 15, 20, 25, 30]
# Rate options (rows): from 5% to 12%
rates = [0.05, 0.06, 0.065, 0.07, 0.08, 0.09, 0.10, 0.12]

r = 5
rh(ws_s, r, 24)
ws_s.cell(row=r, column=2, value="TASA \\ PLAZO").font = Font(name='Calibri', bold=True, size=9, color=GOLD)
ws_s.cell(row=r, column=2).fill = PatternFill('solid', fgColor=CARD_BG)
ws_s.cell(row=r, column=2).alignment = Alignment(horizontal='center')
ws_s.cell(row=r, column=2).border = thin_border

for j, term_y in enumerate(terms_years):
    c = ws_s.cell(row=r, column=3 + j, value=f"{term_y} anios")
    style_header(c, bg=CARD_BG, fg=GOLD, size=9)

# Store term months in row for formula reference
# Hidden reference row
ref_row = 50
for j, term_y in enumerate(terms_years):
    ws_s.cell(row=ref_row, column=3 + j, value=term_y * 12)

for i, rate in enumerate(rates):
    row = 6 + i
    rh(ws_s, row, 24)

    rate_c = ws_s.cell(row=row, column=2, value=rate)
    rate_c.font = Font(name='Calibri', bold=True, size=10, color=ACCENT_BLUE)
    rate_c.fill = PatternFill('solid', fgColor=GRAY_100 if i % 2 == 0 else WHITE)
    rate_c.border = thin_border
    fmt_pct(rate_c)

    for j, term_y in enumerate(terms_years):
        c = ws_s.cell(row=row, column=3 + j)
        term_months = term_y * 12
        c.value = f'=IFERROR(-PMT($B{row}/12,{term_months},Dashboard!$C$6),0)'
        style_output(c)
        fmt_currency(c)
        c.font = Font(name='Calibri', size=9, color=GRAY_900)
        c.fill = PatternFill('solid', fgColor=GRAY_100 if i % 2 == 0 else WHITE)

# Highlight the cell matching current scenario
for i, rate in enumerate(rates):
    row = 6 + i
    for j, term_y in enumerate(terms_years):
        col = 3 + j
        cell_ref = f'{get_column_letter(col)}{row}'
        ws_s.conditional_formatting.add(cell_ref,
            FormulaRule(
                formula=[f'AND(ABS($B{row}-CHOOSE(Dashboard!$C$8,Banco1!$C$5,Banco2!$C$5,Banco3!$C$5,Banco4!$C$5,Banco5!$C$5))<0.001,{term_y*12}=Dashboard!$C$7)'],
                fill=PatternFill('solid', fgColor=WINNER_GREEN_LIGHT),
                font=Font(bold=True, color=WINNER_GREEN)
            ))

# Data bars on all cuota cells
ws_s.conditional_formatting.add(f'C6:G{6+len(rates)-1}',
    DataBarRule(start_type='min', end_type='max', color=WARNING))

rh(ws_s, 6 + len(rates), 12)

# ── TABLE 2: Total intereses varies by Rate x Term ──
r2 = 6 + len(rates) + 1
ws_s.merge_cells(f'B{r2}:H{r2}')
sec2 = ws_s[f'B{r2}']
sec2.value = "TOTAL INTERESES (Bs.) segun Tasa y Plazo"
style_section_header(sec2, ERROR_RED)
fill_row(ws_s, r2, range(3, 9), ERROR_RED)
rh(ws_s, r2, 28)

r2 += 1
rh(ws_s, r2, 24)
ws_s.cell(row=r2, column=2, value="TASA \\ PLAZO").font = Font(name='Calibri', bold=True, size=9, color=GOLD)
ws_s.cell(row=r2, column=2).fill = PatternFill('solid', fgColor=CARD_BG)
ws_s.cell(row=r2, column=2).alignment = Alignment(horizontal='center')
ws_s.cell(row=r2, column=2).border = thin_border

for j, term_y in enumerate(terms_years):
    c = ws_s.cell(row=r2, column=3 + j, value=f"{term_y} anios")
    style_header(c, bg=CARD_BG, fg=GOLD, size=9)

for i, rate in enumerate(rates):
    row = r2 + 1 + i
    rh(ws_s, row, 24)

    rate_c = ws_s.cell(row=row, column=2, value=rate)
    rate_c.font = Font(name='Calibri', bold=True, size=10, color=ERROR_RED)
    rate_c.fill = PatternFill('solid', fgColor=GRAY_100 if i % 2 == 0 else WHITE)
    rate_c.border = thin_border
    fmt_pct(rate_c)

    for j, term_y in enumerate(terms_years):
        c = ws_s.cell(row=row, column=3 + j)
        term_months = term_y * 12
        c.value = f'=IFERROR(-PMT($B{row}/12,{term_months},Dashboard!$C$6)*{term_months}-Dashboard!$C$6,0)'
        style_output(c)
        fmt_currency(c)
        c.font = Font(name='Calibri', size=9, color=GRAY_900)
        c.fill = PatternFill('solid', fgColor=GRAY_100 if i % 2 == 0 else WHITE)

# Data bars (red for interest)
int_start_row = r2 + 1
int_end_row = r2 + len(rates)
ws_s.conditional_formatting.add(f'C{int_start_row}:G{int_end_row}',
    DataBarRule(start_type='min', end_type='max', color=ERROR_RED))

rh(ws_s, int_end_row + 1, 12)

# ── PREPAYMENT CALCULATOR ──
prep_r = int_end_row + 2
ws_s.merge_cells(f'B{prep_r}:H{prep_r}')
sec3 = ws_s[f'B{prep_r}']
sec3.value = "CALCULADORA DE PREPAGO"
style_section_header(sec3, SUCCESS)
fill_row(ws_s, prep_r, range(3, 9), SUCCESS)
rh(ws_s, prep_r, 28)

prep_items = [
    (prep_r + 1, "Pago extra mensual (Bs.)", None, True, True),
    (prep_r + 2, "Meses que ahorras", None, False, False),
    (prep_r + 3, "Interes que ahorras (Bs.)", None, True, False),
    (prep_r + 4, "Nuevo plazo total (meses)", None, False, False),
]

for (row, label, _, is_cur, is_input_cell) in prep_items:
    rh(ws_s, row, 28)
    lbl = ws_s.cell(row=row, column=2, value=label)
    style_label(lbl, bold=True)
    lbl.fill = PatternFill('solid', fgColor=LIGHT_GREEN if not is_input_cell else LIGHT_YELLOW)
    lbl.border = thin_border

    ws_s.merge_cells(f'C{row}:D{row}')
    val = ws_s.cell(row=row, column=3)
    if is_input_cell:
        val.value = 1000
        style_input(val)
        val.font = Font(name='Calibri', size=14, bold=True, color="0000FF")
        fmt_currency(val)
    else:
        style_output(val, bold=True)

# Formulas for prepayment calculator
pr1 = prep_r + 1  # extra payment input
pr2 = prep_r + 2  # months saved
pr3 = prep_r + 3  # interest saved
pr4 = prep_r + 4  # new term

# New term: use NPER with increased payment
ws_s.cell(row=pr4, column=3).value = f'=IFERROR(ROUND(NPER(CHOOSE(Amortizacion!$B$3,Banco1!$C$5,Banco2!$C$5,Banco3!$C$5,Banco4!$C$5,Banco5!$C$5)/12,-(CHOOSE(Amortizacion!$B$3,Banco1!$C$11,Banco2!$C$11,Banco3!$C$11,Banco4!$C$11,Banco5!$C$11)+C{pr1}),Dashboard!$C$6),0),0)'
fmt_int(ws_s.cell(row=pr4, column=3))

# Months saved
ws_s.cell(row=pr2, column=3).value = f'=IFERROR(Dashboard!C7-C{pr4},0)'
fmt_int(ws_s.cell(row=pr2, column=3))
ws_s.cell(row=pr2, column=3).font = Font(name='Calibri', bold=True, size=14, color=SUCCESS)

# Interest saved
ws_s.cell(row=pr3, column=3).value = f'=IFERROR((CHOOSE(Amortizacion!$B$3,Banco1!$C$11,Banco2!$C$11,Banco3!$C$11,Banco4!$C$11,Banco5!$C$11)*Dashboard!C7-Dashboard!C6)-((CHOOSE(Amortizacion!$B$3,Banco1!$C$11,Banco2!$C$11,Banco3!$C$11,Banco4!$C$11,Banco5!$C$11)+C{pr1})*C{pr4}-Dashboard!C6),0)'
fmt_currency(ws_s.cell(row=pr3, column=3))
ws_s.cell(row=pr3, column=3).font = Font(name='Calibri', bold=True, size=14, color=SUCCESS)


# ============================================================
# SHEET 9: CONFIG
# ============================================================
ws_c = wb.create_sheet("Config")
ws_c.sheet_properties.tabColor = GRAY_500

ws_c.column_dimensions['A'].width = 25
ws_c.column_dimensions['B'].width = 55

ws_c.merge_cells('A1:B1')
h = ws_c['A1']
h.value = "CONFIGURACION"
h.font = Font(name='Calibri', bold=True, color=GOLD, size=16)
h.fill = PatternFill('solid', fgColor=DARK_BG)
h.alignment = Alignment(horizontal='center', vertical='center')
ws_c.cell(row=1, column=2).fill = PatternFill('solid', fgColor=DARK_BG)
rh(ws_c, 1, 36)

config_data = [
    ("Producto", "Comparador de Creditos Hipotecarios"),
    ("Version", "v1.0.0"),
    ("Marca", "No Somos Ignorantes"),
    ("Precio", "Bs. 49"),
    ("Max bancos", "5"),
    ("Max plazo", "360 meses (30 anios)"),
    ("Metodo amortizacion", "Frances (cuota fija)"),
    ("Proteccion", "nsi2024"),
    ("Contacto", "nosomosignorantes@gmail.com"),
]

for i, (key, val) in enumerate(config_data):
    row = 3 + i
    k = ws_c.cell(row=row, column=1, value=key)
    k.font = Font(name='Calibri', bold=True, size=11, color=GRAY_700)
    k.border = thin_border
    v = ws_c.cell(row=row, column=2, value=val)
    v.font = Font(name='Calibri', size=11, color=GRAY_900)
    v.border = thin_border

r = 14
ws_c.merge_cells(f'A{r}:B{r}')
inst = ws_c[f'A{r}']
inst.value = "INSTRUCCIONES DE USO"
inst.font = Font(name='Calibri', bold=True, color=GOLD, size=14)
inst.fill = PatternFill('solid', fgColor=DARK_BG)
inst.alignment = Alignment(horizontal='center')
ws_c.cell(row=r, column=2).fill = PatternFill('solid', fgColor=DARK_BG)
rh(ws_c, r, 32)

instructions = [
    "1. En el Dashboard ingresa el monto del prestamo y el plazo deseado en meses.",
    "2. Ve a cada hoja de banco (Banco1-Banco5) e ingresa los datos de la oferta.",
    "3. Para cada banco ingresa: nombre, tasa anual, comision, seguro y otros costos.",
    "4. Regresa al Dashboard para ver la comparacion automatica.",
    "5. El sistema destaca el banco con menor costo total en verde.",
    "6. Ve a 'Amortizacion' para ver la tabla mes a mes del banco que elijas.",
    "7. Usa 'Sensibilidad' para analizar que pasa si cambian tasas o plazos.",
    "8. La calculadora de prepago muestra cuanto ahorras pagando extra.",
    "",
    "TIPS:",
    "- Pide a cada banco su tasa efectiva anual, no solo la nominal.",
    "- Incluye TODOS los costos: comision, seguros, gastos notariales.",
    "- Una diferencia de 0.5% en tasa puede significar miles de Bs en 20 anios.",
    "- La regla de oro: la cuota no debe superar el 35% de tu ingreso.",
]

for i, line in enumerate(instructions):
    row = 15 + i
    ws_c.merge_cells(f'A{row}:B{row}')
    c = ws_c[f'A{row}']
    c.value = line
    c.font = Font(name='Calibri', size=10, color=GRAY_700)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    rh(ws_c, row, 22)


# ============================================================
# SHEET PROTECTION
# ============================================================
for ws_name in ["Dashboard", "Amortizacion", "Sensibilidad", "Config"]:
    sheet = wb[ws_name]
    sheet.protection.sheet = True
    sheet.protection.password = "nsi2024"
    sheet.protection.enable()

# Dashboard: unlock global inputs
ws['C6'].protection = ws['C6'].protection.copy(locked=False)
ws['C7'].protection = ws['C7'].protection.copy(locked=False)

# Amortizacion: unlock bank selector
ws_a['B3'].protection = ws_a['B3'].protection.copy(locked=False)

# Sensibilidad: unlock prepayment input
ws_s.cell(row=prep_r + 1, column=3).protection = ws_s.cell(row=prep_r + 1, column=3).protection.copy(locked=False)

# Bank sheets: protect outputs, allow inputs
for b in range(NUM_BANKS):
    sheet = wb[BANK_SHEETS[b]]
    sheet.protection.sheet = True
    sheet.protection.password = "nsi2024"
    sheet.protection.enable()
    # Unlock input cells
    for row in [3, 5, 6, 7, 8, 15]:
        if row == 15:
            sheet.cell(row=row, column=4).protection = sheet.cell(row=row, column=4).protection.copy(locked=False)
        else:
            sheet.cell(row=row, column=3).protection = sheet.cell(row=row, column=3).protection.copy(locked=False)


# ============================================================
# SAVE
# ============================================================
OUTPUT = "D:/Landing-Page_marketplace/excel_products/Comparador_Hipotecario_NSI.xlsx"
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print("Sheets:", wb.sheetnames)
print("Done!")
