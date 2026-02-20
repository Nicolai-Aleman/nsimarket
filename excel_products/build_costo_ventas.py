"""
Build: Gestor de Costo de Ventas (Bs. 59)
No Somos Ignorantes v1.0
COGS manager: Dashboard with gross profit KPIs, inventory tracking,
purchase log, sales with COGS, gross margin analysis.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint, SeriesLabel
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# Colors
GOLD = "D4AF37"; DARK_BG = "1A1A2E"; ACCENT_BLUE = "0F3460"
WHITE = "FFFFFF"; BLACK = "000000"
LIGHT_YELLOW = "FFF9C4"; LIGHT_GREEN = "E8F5E9"
RED_CORAL = "E74C3C"; TURQUOISE = "16A085"; ORANGE = "E67E22"; PURPLE = "8E44AD"
LIGHT_GRAY = "BDC3C7"; GREEN_OK = "27AE60"; RED_ALERT = "E74C3C"
GRAY_300 = "D1D5DB"; GRAY_400 = "9CA3AF"; GRAY_500 = "6B7280"
GRAY_600 = "4B5563"; GRAY_700 = "374151"; GRAY_900 = "111827"

thin_border = Border(
    left=Side(style='thin', color=GRAY_300), right=Side(style='thin', color=GRAY_300),
    top=Side(style='thin', color=GRAY_300), bottom=Side(style='thin', color=GRAY_300))
gold_border = Border(
    left=Side(style='thin', color=GOLD), right=Side(style='thin', color=GOLD),
    top=Side(style='thin', color=GOLD), bottom=Side(style='thin', color=GOLD))

def style_title(cell, size=18):
    cell.font = Font(name='Calibri', bold=True, color=BLACK, size=size)
    cell.alignment = Alignment(horizontal='left', vertical='center')
def style_header(cell, bg=DARK_BG, fg=WHITE, size=10):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
def style_input(cell):
    cell.font = Font(name='Calibri', size=11, color="0000FF")
    cell.fill = PatternFill('solid', fgColor=LIGHT_YELLOW)
    cell.border = gold_border
    cell.alignment = Alignment(horizontal='center', vertical='center')
def style_output(cell, bold=False):
    cell.font = Font(name='Calibri', size=11, color=GRAY_900, bold=bold)
    cell.fill = PatternFill('solid', fgColor=LIGHT_GREEN)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='right', vertical='center')
def style_banner(cell, bg_color, fg=WHITE, size=10):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg_color)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
def fill_white(ws, r1, r2, c1, c2):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=WHITE)
def unlock_cell(cell):
    cell.protection = cell.protection.copy(locked=False)

MONTHS_ES = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
MAX_PRODUCTS = 30
MAX_ENTRIES = 200

# ============================================================
# SHEET 1: DASHBOARD
# ============================================================
ws_dash = wb.active
ws_dash.title = "Dashboard"
ws_dash.sheet_properties.tabColor = RED_CORAL
ws_dash.sheet_view.showGridLines = False

dash_w = {1:3, 2:22, 3:16, 4:16, 5:3, 6:22, 7:16, 8:16, 9:3}
for c, w in dash_w.items():
    ws_dash.column_dimensions[get_column_letter(c)].width = w
fill_white(ws_dash, 1, 55, 1, 9)

ws_dash.merge_cells('B2:H2')
c = ws_dash['B2']
style_title(c, 20)
c.value = "GESTOR DE COSTO DE VENTAS"
ws_dash.merge_cells('B3:H3')
c = ws_dash['B3']
c.font = Font(name='Calibri', size=10, color=GRAY_500)
c.value = "No Somos Ignorantes  |  v1.0  |  Analisis de margen bruto"

ws_dash.row_dimensions[4].height = 6

# KPI 1: Ventas Totales
ws_dash.merge_cells('B5:D5')
style_banner(ws_dash['B5'], RED_CORAL, WHITE, 10)
ws_dash['B5'].value = "VENTAS TOTALES (Bs.)"
ws_dash.merge_cells('B6:D6')
c = ws_dash['B6']
c.font = Font(name='Calibri', bold=True, color=BLACK, size=26)
c.alignment = Alignment(horizontal='right', vertical='center')
c.value = '=IFERROR(SUM(Ventas!G4:G203),0)'
c.number_format = '#,##0.00'

# KPI 2: Costo de Ventas (COGS)
ws_dash.merge_cells('F5:H5')
style_banner(ws_dash['F5'], TURQUOISE, WHITE, 10)
ws_dash['F5'].value = "COSTO DE VENTAS (Bs.)"
ws_dash.merge_cells('F6:H6')
c = ws_dash['F6']
c.font = Font(name='Calibri', bold=True, color=BLACK, size=26)
c.alignment = Alignment(horizontal='right', vertical='center')
c.value = '=IFERROR(SUM(Ventas!H4:H203),0)'
c.number_format = '#,##0.00'

ws_dash.row_dimensions[7].height = 4

# KPI 3: Utilidad Bruta
ws_dash.merge_cells('B8:D8')
style_banner(ws_dash['B8'], ORANGE, WHITE, 10)
ws_dash['B8'].value = "UTILIDAD BRUTA (Bs.)"
ws_dash.merge_cells('B9:D9')
c = ws_dash['B9']
c.font = Font(name='Calibri', bold=True, color=BLACK, size=26)
c.alignment = Alignment(horizontal='right', vertical='center')
c.value = '=B6-F6'
c.number_format = '#,##0.00'

ws_dash['B10'].value = "Margen bruto:"
ws_dash['B10'].font = Font(name='Calibri', size=9, color="808080")
ws_dash.merge_cells('C10:D10')
c = ws_dash['C10']
c.value = '=IFERROR(B9/B6,0)'
c.number_format = '0.0%'
c.font = Font(name='Calibri', bold=True, size=14, color=BLACK)

# KPI 4: Rotacion de Inventario
ws_dash.merge_cells('F8:H8')
style_banner(ws_dash['F8'], PURPLE, WHITE, 10)
ws_dash['F8'].value = "ROTACION DE INVENTARIO"
ws_dash.merge_cells('F9:H9')
c = ws_dash['F9']
c.font = Font(name='Calibri', bold=True, color=BLACK, size=26)
c.alignment = Alignment(horizontal='right', vertical='center')
# COGS / Average Inventory
c.value = '=IFERROR(F6/((Inventario!C4+Inventario!D4)/2),0)'
c.number_format = '#,##0.0" veces"'

ws_dash['F10'].value = "Dias de inventario:"
ws_dash['F10'].font = Font(name='Calibri', size=9, color="808080")
ws_dash.merge_cells('G10:H10')
c = ws_dash['G10']
c.value = '=IFERROR(365/F9,0)'
c.number_format = '#,##0" dias"'
c.font = Font(name='Calibri', bold=True, size=14, color=BLACK)

ws_dash.row_dimensions[11].height = 6

# ── Formula COGS: Inv. Inicial + Compras - Inv. Final ──
ws_dash.merge_cells('B12:H12')
c = ws_dash['B12']
style_banner(c, DARK_BG, GOLD, 11)
c.value = "FORMULA: Inv. Inicial + Compras - Inv. Final = Costo de Ventas"

ws_dash['B13'].value = "Inventario Inicial:"
ws_dash['B13'].font = Font(name='Calibri', size=10, bold=True, color=GRAY_700)
ws_dash['D13'].value = '=Inventario!C4'
ws_dash['D13'].number_format = '#,##0.00'
ws_dash['D13'].font = Font(name='Calibri', size=11, bold=True, color=ACCENT_BLUE)

ws_dash['B14'].value = "(+) Compras del Periodo:"
ws_dash['B14'].font = Font(name='Calibri', size=10, bold=True, color=GRAY_700)
ws_dash['D14'].value = '=SUM(Compras!E4:E203)'
ws_dash['D14'].number_format = '#,##0.00'
ws_dash['D14'].font = Font(name='Calibri', size=11, bold=True, color=ACCENT_BLUE)

ws_dash['B15'].value = "(-) Inventario Final:"
ws_dash['B15'].font = Font(name='Calibri', size=10, bold=True, color=GRAY_700)
ws_dash['D15'].value = '=Inventario!D4'
ws_dash['D15'].number_format = '#,##0.00'
ws_dash['D15'].font = Font(name='Calibri', size=11, bold=True, color=RED_CORAL)

ws_dash['B16'].value = "(=) Costo de Ventas:"
ws_dash['B16'].font = Font(name='Calibri', size=10, bold=True, color=BLACK)
ws_dash['D16'].value = '=D13+D14-D15'
ws_dash['D16'].number_format = '#,##0.00'
style_output(ws_dash['D16'], bold=True)

ws_dash.row_dimensions[17].height = 8

# Monthly trend table
ws_dash['B18'].value = "MARGEN BRUTO MENSUAL"
ws_dash['B18'].font = Font(name='Calibri', bold=True, size=11, color=GRAY_700)

ws_dash['B19'].value = "Mes"
ws_dash['C19'].value = "Ventas"
ws_dash['D19'].value = "COGS"
style_header(ws_dash['B19'], bg=GRAY_700)
style_header(ws_dash['C19'], bg=GRAY_700)
style_header(ws_dash['D19'], bg=GRAY_700)

for i, m in enumerate(MONTHS_ES):
    r = 20 + i
    ws_dash.cell(row=r, column=2).value = m
    ws_dash.cell(row=r, column=2).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws_dash.cell(row=r, column=2).border = thin_border

    ws_dash.cell(row=r, column=3).value = (
        f'=IFERROR(SUMPRODUCT((MONTH(Ventas!A4:A203)={i+1})*Ventas!G4:G203),0)')
    ws_dash.cell(row=r, column=3).number_format = '#,##0'
    ws_dash.cell(row=r, column=3).border = thin_border

    ws_dash.cell(row=r, column=4).value = (
        f'=IFERROR(SUMPRODUCT((MONTH(Ventas!A4:A203)={i+1})*Ventas!H4:H203),0)')
    ws_dash.cell(row=r, column=4).number_format = '#,##0'
    ws_dash.cell(row=r, column=4).border = thin_border

# Clustered bar chart: Ventas vs COGS
chart1 = BarChart()
chart1.type = "col"
chart1.grouping = "clustered"
chart1.style = 10
chart1.title = "Ventas vs Costo de Ventas Mensual"
chart1.y_axis.title = "Bs."
chart1.legend.position = 'b'
d1 = Reference(ws_dash, min_col=3, min_row=19, max_row=31)
d2 = Reference(ws_dash, min_col=4, min_row=19, max_row=31)
cats = Reference(ws_dash, min_col=2, min_row=20, max_row=31)
chart1.add_data(d1, titles_from_data=True)
chart1.add_data(d2, titles_from_data=True)
chart1.set_categories(cats)
chart1.series[0].graphicalProperties.solidFill = TURQUOISE
chart1.series[0].graphicalProperties.line.noFill = True
chart1.series[1].graphicalProperties.solidFill = RED_CORAL
chart1.series[1].graphicalProperties.line.noFill = True
chart1.width = 22
chart1.height = 13
ws_dash.add_chart(chart1, "B33")

# Margin % by month (line chart)
ws_dash['F18'].value = "% MARGEN BRUTO MENSUAL"
ws_dash['F18'].font = Font(name='Calibri', bold=True, size=11, color=GRAY_700)

ws_dash['F19'].value = "Mes"
ws_dash['G19'].value = "Margen %"
style_header(ws_dash['F19'], bg=GRAY_700)
style_header(ws_dash['G19'], bg=GRAY_700)

for i, m in enumerate(MONTHS_ES):
    r = 20 + i
    ws_dash.cell(row=r, column=6).value = m
    ws_dash.cell(row=r, column=6).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws_dash.cell(row=r, column=6).border = thin_border
    ws_dash.cell(row=r, column=7).value = f'=IFERROR((C{r}-D{r})/C{r},0)'
    ws_dash.cell(row=r, column=7).number_format = '0.0%'
    ws_dash.cell(row=r, column=7).border = thin_border

chart2 = LineChart()
chart2.title = "Evolucion del Margen Bruto"
chart2.style = 10
chart2.y_axis.title = "%"
chart2.y_axis.numFmt = '0%'
chart2.legend = None
chart2.width = 18
chart2.height = 13
d3 = Reference(ws_dash, min_col=7, min_row=19, max_row=31)
cats3 = Reference(ws_dash, min_col=6, min_row=20, max_row=31)
chart2.add_data(d3, titles_from_data=True)
chart2.set_categories(cats3)
chart2.series[0].graphicalProperties.line.solidFill = ORANGE
chart2.series[0].graphicalProperties.line.width = 28000
ws_dash.add_chart(chart2, "F33")

# ============================================================
# SHEET 2: INVENTARIO
# ============================================================
ws_inv = wb.create_sheet("Inventario")
ws_inv.sheet_properties.tabColor = PURPLE
ws_inv.sheet_view.showGridLines = False

inv_w = {1:4, 2:30, 3:20, 4:20, 5:20}
for c, w in inv_w.items():
    ws_inv.column_dimensions[get_column_letter(c)].width = w
fill_white(ws_inv, 1, 40, 1, 5)

ws_inv.merge_cells('A1:E1')
style_title(ws_inv['A1'], 16)
ws_inv['A1'].value = "INVENTARIO INICIAL Y FINAL"
ws_inv.row_dimensions[1].height = 35

ws_inv.merge_cells('A2:E2')
c = ws_inv['A2']
c.value = "Ingresa el valor del inventario al inicio y final del periodo para calcular el costo de ventas."
c.font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)

headers_inv = [("A","#"), ("B","CATEGORIA / PRODUCTO"), ("C","INV. INICIAL (Bs.)"),
               ("D","INV. FINAL (Bs.)"), ("E","VARIACION")]
ws_inv.row_dimensions[3].height = 28
for col_letter, label in headers_inv:
    style_header(ws_inv[f'{col_letter}3'], bg=DARK_BG, fg=GOLD, size=10)
    ws_inv[f'{col_letter}3'].value = label

# First row = TOTAL (sum of details below)
ws_inv.cell(row=4, column=1).value = ""
ws_inv.cell(row=4, column=2).value = "TOTAL INVENTARIO"
ws_inv.cell(row=4, column=2).font = Font(name='Calibri', bold=True, size=11, color=BLACK)
ws_inv.cell(row=4, column=2).border = thin_border
ws_inv.cell(row=4, column=3).value = '=SUM(C5:C34)'
ws_inv.cell(row=4, column=3).number_format = '#,##0.00'
style_output(ws_inv.cell(row=4, column=3), bold=True)
ws_inv.cell(row=4, column=4).value = '=SUM(D5:D34)'
ws_inv.cell(row=4, column=4).number_format = '#,##0.00'
style_output(ws_inv.cell(row=4, column=4), bold=True)
ws_inv.cell(row=4, column=5).value = '=D4-C4'
ws_inv.cell(row=4, column=5).number_format = '#,##0.00'
style_output(ws_inv.cell(row=4, column=5), bold=True)

for row in range(5, 5 + MAX_PRODUCTS):
    ws_inv.cell(row=row, column=1).value = f'=IF(B{row}="","",ROW()-4)'
    ws_inv.cell(row=row, column=1).font = Font(name='Calibri', size=10, color=GRAY_500)
    ws_inv.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    ws_inv.cell(row=row, column=1).border = thin_border

    c = ws_inv.cell(row=row, column=2)
    style_input(c); c.alignment = Alignment(horizontal='left', vertical='center')
    unlock_cell(c)

    c = ws_inv.cell(row=row, column=3)
    style_input(c); c.number_format = '#,##0.00'
    unlock_cell(c)

    c = ws_inv.cell(row=row, column=4)
    style_input(c); c.number_format = '#,##0.00'
    unlock_cell(c)

    c = ws_inv.cell(row=row, column=5)
    c.value = f'=IFERROR(D{row}-C{row},0)'
    c.number_format = '#,##0.00'
    style_output(c)

# Conditional formatting: negative variation = red
ws_inv.conditional_formatting.add(f'E5:E{4+MAX_PRODUCTS}',
    CellIsRule(operator='lessThan', formula=['0'],
              fill=PatternFill('solid', fgColor="FEE2E2"),
              font=Font(color=RED_ALERT, bold=True)))

# Sample data
sample_inv = [
    ("Materia Prima A", 25000, 22000),
    ("Materia Prima B", 15000, 18000),
    ("Productos en Proceso", 8000, 6000),
    ("Productos Terminados", 30000, 28000),
    ("Suministros", 5000, 4500),
]
for i, (name, ini, fin) in enumerate(sample_inv):
    ws_inv.cell(row=5+i, column=2).value = name
    ws_inv.cell(row=5+i, column=3).value = ini
    ws_inv.cell(row=5+i, column=4).value = fin

# ============================================================
# SHEET 3: COMPRAS
# ============================================================
ws_comp = wb.create_sheet("Compras")
ws_comp.sheet_properties.tabColor = TURQUOISE
ws_comp.sheet_view.showGridLines = False

comp_w = {1:14, 2:30, 3:20, 4:12, 5:18, 6:18}
for c, w in comp_w.items():
    ws_comp.column_dimensions[get_column_letter(c)].width = w

ws_comp.merge_cells('A1:F1')
style_title(ws_comp['A1'], 16)
ws_comp['A1'].value = "REGISTRO DE COMPRAS"
ws_comp['A1'].fill = PatternFill('solid', fgColor=WHITE)
ws_comp.row_dimensions[1].height = 35

ws_comp.merge_cells('A2:F2')
c = ws_comp['A2']
c.value = "Registra todas las compras de inventario/mercaderia del periodo."
c.font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)
c.fill = PatternFill('solid', fgColor=WHITE)

headers_comp = [("A","FECHA"), ("B","PROVEEDOR"), ("C","PRODUCTO/DETALLE"),
                ("D","CANTIDAD"), ("E","TOTAL COMPRA (Bs.)"), ("F","OBSERVACION")]
ws_comp.row_dimensions[3].height = 28
for col_letter, label in headers_comp:
    style_header(ws_comp[f'{col_letter}3'], bg=DARK_BG, fg=GOLD, size=10)
    ws_comp[f'{col_letter}3'].value = label

for row in range(4, 4 + MAX_ENTRIES):
    for col in range(1, 7):
        c = ws_comp.cell(row=row, column=col)
        style_input(c)
        unlock_cell(c)
        if col == 1: c.number_format = 'DD/MM/YYYY'
        elif col in [2, 3, 6]: c.alignment = Alignment(horizontal='left', vertical='center')
        elif col == 4: c.number_format = '#,##0'
        elif col == 5: c.number_format = '#,##0.00'

dv_pos = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
dv_pos.error = "El valor no puede ser negativo"
ws_comp.add_data_validation(dv_pos)
dv_pos.add('E4:E203')

# Sample purchases
from datetime import date
sample_comp = [
    (date(2025,1,5), "Proveedor ABC", "Materia Prima A", 100, 12500),
    (date(2025,1,10), "Distribuidora XYZ", "Materia Prima B", 50, 7500),
    (date(2025,1,15), "Suministros SRL", "Suministros varios", 1, 2000),
    (date(2025,2,1), "Proveedor ABC", "Materia Prima A", 80, 10000),
    (date(2025,2,5), "Distribuidora XYZ", "Materia Prima B", 60, 9000),
]
for i, (dt, prov, prod, qty, total) in enumerate(sample_comp):
    ws_comp.cell(row=4+i, column=1).value = dt
    ws_comp.cell(row=4+i, column=2).value = prov
    ws_comp.cell(row=4+i, column=3).value = prod
    ws_comp.cell(row=4+i, column=4).value = qty
    ws_comp.cell(row=4+i, column=5).value = total

# ============================================================
# SHEET 4: VENTAS (Sales with COGS)
# ============================================================
ws_ven = wb.create_sheet("Ventas")
ws_ven.sheet_properties.tabColor = ORANGE
ws_ven.sheet_view.showGridLines = False

ven_w = {1:14, 2:28, 3:12, 4:16, 5:16, 6:14, 7:16, 8:16, 9:14}
for c, w in ven_w.items():
    ws_ven.column_dimensions[get_column_letter(c)].width = w

ws_ven.merge_cells('A1:I1')
style_title(ws_ven['A1'], 16)
ws_ven['A1'].value = "VENTAS Y COSTO DE VENTAS"
ws_ven['A1'].fill = PatternFill('solid', fgColor=WHITE)
ws_ven.row_dimensions[1].height = 35

ws_ven.merge_cells('A2:I2')
c = ws_ven['A2']
c.value = "Registra cada venta con su costo unitario para calcular el margen bruto por producto."
c.font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)
c.fill = PatternFill('solid', fgColor=WHITE)

headers_ven = [("A","FECHA"), ("B","PRODUCTO"), ("C","CANTIDAD"),
               ("D","PRECIO UNIT."), ("E","COSTO UNIT."), ("F","MARGEN %"),
               ("G","TOTAL VENTA"), ("H","TOTAL COSTO"), ("I","UTILIDAD")]
ws_ven.row_dimensions[3].height = 28
for col_letter, label in headers_ven:
    style_header(ws_ven[f'{col_letter}3'], bg=DARK_BG, fg=GOLD, size=10)
    ws_ven[f'{col_letter}3'].value = label

for row in range(4, 4 + MAX_ENTRIES):
    # A: Date INPUT
    c = ws_ven.cell(row=row, column=1)
    style_input(c); c.number_format = 'DD/MM/YYYY'; unlock_cell(c)

    # B: Product INPUT
    c = ws_ven.cell(row=row, column=2)
    style_input(c); c.alignment = Alignment(horizontal='left', vertical='center'); unlock_cell(c)

    # C: Quantity INPUT
    c = ws_ven.cell(row=row, column=3)
    style_input(c); c.number_format = '#,##0'; unlock_cell(c)

    # D: Unit Price INPUT
    c = ws_ven.cell(row=row, column=4)
    style_input(c); c.number_format = '#,##0.00'; unlock_cell(c)

    # E: Unit Cost INPUT
    c = ws_ven.cell(row=row, column=5)
    style_input(c); c.number_format = '#,##0.00'; unlock_cell(c)

    # F: Margin % (CALCULATED)
    c = ws_ven.cell(row=row, column=6)
    c.value = f'=IFERROR((D{row}-E{row})/D{row},0)'
    c.number_format = '0.0%'
    style_output(c)

    # G: Total Sale (CALCULATED)
    c = ws_ven.cell(row=row, column=7)
    c.value = f'=IFERROR(C{row}*D{row},0)'
    c.number_format = '#,##0.00'
    style_output(c)

    # H: Total Cost (CALCULATED)
    c = ws_ven.cell(row=row, column=8)
    c.value = f'=IFERROR(C{row}*E{row},0)'
    c.number_format = '#,##0.00'
    style_output(c)

    # I: Profit (CALCULATED)
    c = ws_ven.cell(row=row, column=9)
    c.value = f'=G{row}-H{row}'
    c.number_format = '#,##0.00'
    style_output(c, bold=True)

# Conditional formatting: margin
ws_ven.conditional_formatting.add(f'F4:F{3+MAX_ENTRIES}',
    CellIsRule(operator='lessThan', formula=['0.2'],
              fill=PatternFill('solid', fgColor="FEE2E2"),
              font=Font(color=RED_ALERT, bold=True)))
ws_ven.conditional_formatting.add(f'F4:F{3+MAX_ENTRIES}',
    CellIsRule(operator='greaterThanOrEqual', formula=['0.4'],
              fill=PatternFill('solid', fgColor="D1FAE5"),
              font=Font(color=GREEN_OK, bold=True)))

# Sample sales
sample_ven = [
    (date(2025,1,8), "Producto Terminado A", 20, 150, 80),
    (date(2025,1,12), "Producto Terminado B", 15, 220, 120),
    (date(2025,1,15), "Producto Terminado A", 25, 150, 80),
    (date(2025,1,20), "Producto Terminado C", 10, 350, 200),
    (date(2025,1,25), "Producto Terminado B", 18, 220, 120),
    (date(2025,2,3), "Producto Terminado A", 30, 150, 80),
    (date(2025,2,8), "Producto Terminado C", 12, 350, 200),
]
for i, (dt, prod, qty, price, cost) in enumerate(sample_ven):
    ws_ven.cell(row=4+i, column=1).value = dt
    ws_ven.cell(row=4+i, column=2).value = prod
    ws_ven.cell(row=4+i, column=3).value = qty
    ws_ven.cell(row=4+i, column=4).value = price
    ws_ven.cell(row=4+i, column=5).value = cost

# ============================================================
# SHEET 5: CONFIG
# ============================================================
ws_conf = wb.create_sheet("Config")
ws_conf.sheet_properties.tabColor = GRAY_500
ws_conf.sheet_view.showGridLines = False
conf_w = {1:25, 2:45}
for c, w in conf_w.items():
    ws_conf.column_dimensions[get_column_letter(c)].width = w
fill_white(ws_conf, 1, 25, 1, 3)

ws_conf['A1'].value = "v1.0.0"
ws_conf['A1'].font = Font(name='Calibri', size=9, color=GRAY_400)
ws_conf.merge_cells('A3:B3')
style_title(ws_conf['A3'], 14)
ws_conf['A3'].value = "CONFIGURACION"

settings = [
    ("Producto", "Gestor de Costo de Ventas"),
    ("Version", "v1.0.0"),
    ("Marca", "No Somos Ignorantes"),
    ("Precio", "Bs. 59"),
    ("Proteccion", "nsi2024"),
    ("", ""),
    ("INSTRUCCIONES", ""),
    ("1.", "En 'Inventario', ingresa el valor del inventario inicial y final."),
    ("2.", "En 'Compras', registra todas las compras del periodo."),
    ("3.", "En 'Ventas', registra cada venta con su precio y costo unitario."),
    ("4.", "El Dashboard calcula automaticamente el COGS, margen bruto y rotacion."),
    ("5.", "Formula COGS: Inv. Inicial + Compras - Inv. Final = Costo de Ventas"),
    ("6.", "Para desbloquear: Revisar -> Desproteger hoja -> nsi2024"),
]
for i, (label, value) in enumerate(settings):
    r = 5 + i
    ws_conf.cell(row=r, column=1).value = label
    ws_conf.cell(row=r, column=1).font = Font(name='Calibri', size=10, bold=True, color=GRAY_700)
    ws_conf.cell(row=r, column=2).value = value
    ws_conf.cell(row=r, column=2).font = Font(name='Calibri', size=10, color=GRAY_600)

# ============================================================
# PROTECTION
# ============================================================
for ws in wb.worksheets:
    ws.protection.sheet = True
    ws.protection.password = "nsi2024"
    ws.protection.enable()

# Re-unlock all input cells
for row in range(5, 5+MAX_PRODUCTS):
    for col in [2, 3, 4]:
        unlock_cell(ws_inv.cell(row=row, column=col))

for row in range(4, 4+MAX_ENTRIES):
    for col in range(1, 7):
        unlock_cell(ws_comp.cell(row=row, column=col))

for row in range(4, 4+MAX_ENTRIES):
    for col in [1, 2, 3, 4, 5]:
        unlock_cell(ws_ven.cell(row=row, column=col))

# ============================================================
# SAVE & VERIFY
# ============================================================
output_path = r"D:\Landing-Page_marketplace\excel_products\Gestor_Costo_Ventas_NSI.xlsx"
wb.save(output_path)
print(f"[OK] Saved: {output_path}")

from openpyxl import load_workbook
wb2 = load_workbook(output_path)
formula_count = 0
for ws in wb2.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formula_count += 1
print(f"[OK] Sheets: {wb2.sheetnames}")
print(f"[OK] Total formulas: {formula_count}")
print(f"[OK] Charts: Dashboard={len(ws_dash._charts)}")
