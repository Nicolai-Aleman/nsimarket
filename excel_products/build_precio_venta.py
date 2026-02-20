"""
Build: Calculadora de Precio de Venta (Bs. 49)
No Somos Ignorantes v1.0
Sale price calculator: Main calculator, multiple pricing scenarios,
competitor comparison, break-even analysis.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

GOLD = "D4AF37"; DARK_BG = "1A1A2E"; ACCENT_BLUE = "0F3460"
WHITE = "FFFFFF"; BLACK = "000000"
LIGHT_YELLOW = "FFF9C4"; LIGHT_GREEN = "E8F5E9"
RED_CORAL = "E74C3C"; TURQUOISE = "16A085"; ORANGE = "E67E22"; PURPLE = "8E44AD"
LIGHT_GRAY = "BDC3C7"; GREEN_OK = "27AE60"; RED_ALERT = "E74C3C"
GRAY_300 = "D1D5DB"; GRAY_400 = "9CA3AF"; GRAY_500 = "6B7280"
GRAY_600 = "4B5563"; GRAY_700 = "374151"; GRAY_900 = "111827"

thin_border = Border(
    left=Side(style='thin', color=GRAY_300), right=Side(style='thin', color=GRAY_300),
    top=Side(style='thin', color=GRAY_300), bottom=Side(style='thin', color=GRAY_300))
gold_border = Border(
    left=Side(style='thin', color=GOLD), right=Side(style='thin', color=GOLD),
    top=Side(style='thin', color=GOLD), bottom=Side(style='thin', color=GOLD))

def style_title(cell, size=18):
    cell.font = Font(name='Calibri', bold=True, color=BLACK, size=size)
    cell.alignment = Alignment(horizontal='left', vertical='center')
def style_header(cell, bg=DARK_BG, fg=WHITE, size=10):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
def style_input(cell):
    cell.font = Font(name='Calibri', size=11, color="0000FF")
    cell.fill = PatternFill('solid', fgColor=LIGHT_YELLOW)
    cell.border = gold_border
    cell.alignment = Alignment(horizontal='center', vertical='center')
def style_output(cell, bold=False):
    cell.font = Font(name='Calibri', size=11, color=GRAY_900, bold=bold)
    cell.fill = PatternFill('solid', fgColor=LIGHT_GREEN)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='right', vertical='center')
def style_banner(cell, bg_color, fg=WHITE, size=10):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg_color)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
def style_label(cell, bold=False):
    cell.font = Font(name='Calibri', size=11, color=GRAY_700, bold=bold)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = thin_border
def fill_white(ws, r1, r2, c1, c2):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=WHITE)
def unlock_cell(cell):
    cell.protection = cell.protection.copy(locked=False)

MAX_PRODUCTS = 20
MAX_COMPETITORS = 10

# ============================================================
# SHEET 1: CALCULADORA (Main Calculator)
# ============================================================
ws = wb.active
ws.title = "Calculadora"
ws.sheet_properties.tabColor = RED_CORAL
ws.sheet_view.showGridLines = False

cw = {1:3, 2:30, 3:20, 4:3, 5:30, 6:20, 7:3}
for c, w in cw.items():
    ws.column_dimensions[get_column_letter(c)].width = w
fill_white(ws, 1, 55, 1, 7)

ws.merge_cells('B2:F2')
style_title(ws['B2'], 20)
ws['B2'].value = "CALCULADORA DE PRECIO DE VENTA"
ws.merge_cells('B3:F3')
ws['B3'].font = Font(name='Calibri', size=10, color=GRAY_500)
ws['B3'].value = "No Somos Ignorantes  |  v1.0  |  Define tu precio optimo"

ws.row_dimensions[4].height = 8

# ── LEFT COLUMN: COSTOS ──
ws.merge_cells('B5:C5')
style_banner(ws['B5'], RED_CORAL, WHITE, 11)
ws['B5'].value = "ESTRUCTURA DE COSTOS"

# Cost items
cost_items = [
    ("Costo de Materia Prima", "B7", "C7"),
    ("Mano de Obra Directa", "B8", "C8"),
    ("Costos Indirectos de Fabricacion", "B9", "C9"),
    ("Empaque y Embalaje", "B10", "C10"),
    ("Transporte / Envio", "B11", "C11"),
    ("Comision de Venta (%)", "B12", "C12"),
    ("Otros Costos Variables", "B13", "C13"),
]

for i, (label, label_cell, input_cell) in enumerate(cost_items):
    r = 7 + i
    style_label(ws[label_cell])
    ws[label_cell].value = label
    c = ws[input_cell]
    style_input(c)
    c.number_format = '#,##0.00'
    unlock_cell(c)

ws.row_dimensions[14].height = 4

# COSTO TOTAL
ws.merge_cells('B15:C15')
c = ws['B15']
style_banner(c, DARK_BG, GOLD, 11)
c.value = "COSTO TOTAL UNITARIO"
ws['C15'].font = Font(name='Calibri', bold=True, color=GOLD, size=11)

ws['B16'].value = "Costo Total (Bs.):"
style_label(ws['B16'], bold=True)
ws['C16'].value = '=SUM(C7:C13)'
ws['C16'].number_format = '#,##0.00'
style_output(ws['C16'], bold=True)

ws.row_dimensions[17].height = 8

# ── RIGHT COLUMN: PRECIO ──
ws.merge_cells('E5:F5')
style_banner(ws['E5'], TURQUOISE, WHITE, 11)
ws['E5'].value = "CALCULO DEL PRECIO"

# Method 1: Cost-plus (markup)
ws['E7'].value = "Margen deseado sobre costo (%):"
style_label(ws['E7'])
ws['F7'].value = 0.40
style_input(ws['F7'])
ws['F7'].number_format = '0%'
unlock_cell(ws['F7'])

ws['E8'].value = "Precio (Cost-Plus):"
style_label(ws['E8'], bold=True)
ws['F8'].value = '=IFERROR(C16*(1+F7),0)'
ws['F8'].number_format = '#,##0.00'
style_output(ws['F8'], bold=True)

ws.row_dimensions[9].height = 4

# Method 2: Target margin
ws['E10'].value = "Margen deseado sobre precio (%):"
style_label(ws['E10'])
ws['F10'].value = 0.30
style_input(ws['F10'])
ws['F10'].number_format = '0%'
unlock_cell(ws['F10'])

ws['E11'].value = "Precio (Target Margin):"
style_label(ws['E11'], bold=True)
ws['F11'].value = '=IFERROR(C16/(1-F10),0)'
ws['F11'].number_format = '#,##0.00'
style_output(ws['F11'], bold=True)

ws.row_dimensions[12].height = 4

# Method 3: Break-even
ws['E13'].value = "Costos Fijos Mensuales (Bs.):"
style_label(ws['E13'])
ws['F13'].value = 5000
style_input(ws['F13'])
ws['F13'].number_format = '#,##0.00'
unlock_cell(ws['F13'])

ws['E14'].value = "Unidades esperadas/mes:"
style_label(ws['E14'])
ws['F14'].value = 100
style_input(ws['F14'])
ws['F14'].number_format = '#,##0'
unlock_cell(ws['F14'])

ws['E15'].value = "Precio (Break-Even):"
style_label(ws['E15'], bold=True)
ws['F15'].value = '=IFERROR(F13/F14+C16,0)'
ws['F15'].number_format = '#,##0.00'
style_output(ws['F15'], bold=True)

ws.row_dimensions[17].height = 8

# ── RESUMEN DE PRECIOS ──
ws.merge_cells('B18:F18')
style_banner(ws['B18'], ORANGE, WHITE, 11)
ws['B18'].value = "RESUMEN DE METODOS DE PRECIO"

ws['B19'].value = "Metodo"
ws['C19'].value = "Precio"
ws['D19'].value = ""
ws['E19'].value = "Margen Bs."
ws['F19'].value = "Margen %"
style_header(ws['B19'], bg=GRAY_700)
style_header(ws['C19'], bg=GRAY_700)
style_header(ws['E19'], bg=GRAY_700)
style_header(ws['F19'], bg=GRAY_700)

methods = [
    ("Cost-Plus (Markup)", "=F8", "=F8-C16", "=IFERROR((F8-C16)/F8,0)"),
    ("Target Margin", "=F11", "=F11-C16", "=IFERROR((F11-C16)/F11,0)"),
    ("Break-Even + Margen", "=F15", "=F15-C16", "=IFERROR((F15-C16)/F15,0)"),
]
for i, (name, price_f, margin_f, pct_f) in enumerate(methods):
    r = 20 + i
    ws.cell(row=r, column=2).value = name
    ws.cell(row=r, column=2).font = Font(name='Calibri', size=11, color=GRAY_700)
    ws.cell(row=r, column=2).border = thin_border

    ws.cell(row=r, column=3).value = price_f
    ws.cell(row=r, column=3).number_format = '#,##0.00'
    style_output(ws.cell(row=r, column=3), bold=True)

    ws.cell(row=r, column=5).value = margin_f
    ws.cell(row=r, column=5).number_format = '#,##0.00'
    ws.cell(row=r, column=5).border = thin_border
    ws.cell(row=r, column=5).font = Font(name='Calibri', size=11, color=GRAY_700)

    ws.cell(row=r, column=6).value = pct_f
    ws.cell(row=r, column=6).number_format = '0.0%'
    ws.cell(row=r, column=6).border = thin_border
    ws.cell(row=r, column=6).font = Font(name='Calibri', size=11, color=GRAY_700)

ws.row_dimensions[23].height = 4

# Precio FINAL (user decides)
ws.merge_cells('B24:C24')
c = ws['B24']
style_banner(c, PURPLE, WHITE, 12)
c.value = "PRECIO FINAL ELEGIDO"
ws.merge_cells('E24:F24')
c = ws['E24']
style_input(c)
c.number_format = '#,##0.00'
c.font = Font(name='Calibri', size=18, bold=True, color="0000FF")
unlock_cell(c)

ws['B25'].value = "Ganancia por unidad:"
style_label(ws['B25'])
ws['C25'].value = '=IFERROR(E24-C16,0)'
ws['C25'].number_format = '#,##0.00'
style_output(ws['C25'], bold=True)

ws['E25'].value = "Margen final:"
style_label(ws['E25'])
ws['F25'].value = '=IFERROR((E24-C16)/E24,0)'
ws['F25'].number_format = '0.0%'
style_output(ws['F25'], bold=True)

ws['B26'].value = "Punto de equilibrio (unidades):"
style_label(ws['B26'])
ws['C26'].value = '=IFERROR(F13/(E24-C16),0)'
ws['C26'].number_format = '#,##0'
style_output(ws['C26'], bold=True)

# Sample data
ws['C7'].value = 25   # Materia prima
ws['C8'].value = 10   # Mano de obra
ws['C9'].value = 5    # Costos indirectos
ws['C10'].value = 3   # Empaque
ws['C11'].value = 2   # Transporte
ws['C12'].value = 0   # Comision
ws['C13'].value = 0   # Otros

# ============================================================
# SHEET 2: ESCENARIOS
# ============================================================
ws_esc = wb.create_sheet("Escenarios")
ws_esc.sheet_properties.tabColor = ORANGE
ws_esc.sheet_view.showGridLines = False

ew = {1:4, 2:30, 3:16, 4:16, 5:16, 6:16, 7:16, 8:16, 9:16}
for c, w in ew.items():
    ws_esc.column_dimensions[get_column_letter(c)].width = w
fill_white(ws_esc, 1, 35, 1, 9)

ws_esc.merge_cells('B1:I1')
style_title(ws_esc['B1'], 16)
ws_esc['B1'].value = "ESCENARIOS DE PRECIO"
ws_esc.row_dimensions[1].height = 35

ws_esc.merge_cells('B2:I2')
ws_esc['B2'].value = "Compara hasta 20 productos con diferentes estructuras de costos y precios."
ws_esc['B2'].font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)

headers_esc = [("B","PRODUCTO"), ("C","COSTO UNIT."), ("D","PRECIO VENTA"),
               ("E","MARGEN Bs."), ("F","MARGEN %"), ("G","VENTAS/MES"),
               ("H","INGRESO MENSUAL"), ("I","UTILIDAD MENSUAL")]
ws_esc.row_dimensions[3].height = 28
for col_letter, label in headers_esc:
    style_header(ws_esc[f'{col_letter}3'], bg=DARK_BG, fg=GOLD, size=9)
    ws_esc[f'{col_letter}3'].value = label

for row in range(4, 4 + MAX_PRODUCTS):
    # Product name (INPUT)
    c = ws_esc.cell(row=row, column=2)
    style_input(c); c.alignment = Alignment(horizontal='left', vertical='center'); unlock_cell(c)

    # Cost (INPUT)
    c = ws_esc.cell(row=row, column=3)
    style_input(c); c.number_format = '#,##0.00'; unlock_cell(c)

    # Price (INPUT)
    c = ws_esc.cell(row=row, column=4)
    style_input(c); c.number_format = '#,##0.00'; unlock_cell(c)

    # Margin Bs (CALC)
    c = ws_esc.cell(row=row, column=5)
    c.value = f'=IFERROR(D{row}-C{row},0)'
    c.number_format = '#,##0.00'
    style_output(c)

    # Margin % (CALC)
    c = ws_esc.cell(row=row, column=6)
    c.value = f'=IFERROR((D{row}-C{row})/D{row},0)'
    c.number_format = '0.0%'
    style_output(c)

    # Monthly sales (INPUT)
    c = ws_esc.cell(row=row, column=7)
    style_input(c); c.number_format = '#,##0'; unlock_cell(c)

    # Monthly revenue (CALC)
    c = ws_esc.cell(row=row, column=8)
    c.value = f'=IFERROR(D{row}*G{row},0)'
    c.number_format = '#,##0.00'
    style_output(c)

    # Monthly profit (CALC)
    c = ws_esc.cell(row=row, column=9)
    c.value = f'=IFERROR(E{row}*G{row},0)'
    c.number_format = '#,##0.00'
    style_output(c, bold=True)

# Total row
r_tot = 4 + MAX_PRODUCTS
ws_esc.cell(row=r_tot, column=2).value = "TOTAL"
ws_esc.cell(row=r_tot, column=2).font = Font(name='Calibri', bold=True, size=11, color=BLACK)
ws_esc.cell(row=r_tot, column=2).border = thin_border
for col in [8, 9]:
    c = ws_esc.cell(row=r_tot, column=col)
    c.value = f'=SUM({get_column_letter(col)}4:{get_column_letter(col)}{r_tot-1})'
    c.number_format = '#,##0.00'
    style_output(c, bold=True)

# Conditional formatting: low margin red, high margin green
ws_esc.conditional_formatting.add(f'F4:F{r_tot-1}',
    CellIsRule(operator='lessThan', formula=['0.2'],
              fill=PatternFill('solid', fgColor="FEE2E2"),
              font=Font(color=RED_ALERT, bold=True)))
ws_esc.conditional_formatting.add(f'F4:F{r_tot-1}',
    CellIsRule(operator='greaterThanOrEqual', formula=['0.4'],
              fill=PatternFill('solid', fgColor="D1FAE5"),
              font=Font(color=GREEN_OK, bold=True)))

# Data bars on profit
ws_esc.conditional_formatting.add(f'I4:I{r_tot-1}',
    DataBarRule(start_type='min', end_type='max', color=PURPLE))

# Bar chart: Profit by product
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Utilidad Mensual por Producto"
chart1.y_axis.title = "Bs."
chart1.legend = None
chart1.width = 22
chart1.height = 13
d1 = Reference(ws_esc, min_col=9, min_row=3, max_row=r_tot-1)
cats1 = Reference(ws_esc, min_col=2, min_row=4, max_row=r_tot-1)
chart1.add_data(d1, titles_from_data=True)
chart1.set_categories(cats1)
chart1.series[0].graphicalProperties.solidFill = PURPLE
chart1.series[0].graphicalProperties.line.noFill = True
ws_esc.add_chart(chart1, f"B{r_tot+2}")

# Sample scenario data
sample_esc = [
    ("Producto A - Basico", 45, 75, 80),
    ("Producto B - Estandar", 65, 120, 50),
    ("Producto C - Premium", 95, 200, 25),
    ("Servicio D - Consultoria", 30, 80, 15),
]
for i, (name, cost, price, qty) in enumerate(sample_esc):
    r = 4 + i
    ws_esc.cell(row=r, column=2).value = name
    ws_esc.cell(row=r, column=3).value = cost
    ws_esc.cell(row=r, column=4).value = price
    ws_esc.cell(row=r, column=7).value = qty

# ============================================================
# SHEET 3: COMPETENCIA
# ============================================================
ws_comp = wb.create_sheet("Competencia")
ws_comp.sheet_properties.tabColor = PURPLE
ws_comp.sheet_view.showGridLines = False

cpw = {1:4, 2:25, 3:18, 4:18, 5:16, 6:16, 7:16}
for c, w in cpw.items():
    ws_comp.column_dimensions[get_column_letter(c)].width = w
fill_white(ws_comp, 1, 30, 1, 7)

ws_comp.merge_cells('B1:G1')
style_title(ws_comp['B1'], 16)
ws_comp['B1'].value = "COMPARACION CON COMPETENCIA"
ws_comp.row_dimensions[1].height = 35

ws_comp.merge_cells('B2:G2')
ws_comp['B2'].value = "Compara tu precio con la competencia para posicionarte correctamente."
ws_comp['B2'].font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)

# Your price reference
ws_comp['B3'].value = "Tu Precio (Bs.):"
ws_comp['B3'].font = Font(name='Calibri', size=10, bold=True, color=GRAY_700)
ws_comp['C3'].value = '=Calculadora!E24'
ws_comp['C3'].number_format = '#,##0.00'
ws_comp['C3'].font = Font(name='Calibri', size=14, bold=True, color=ACCENT_BLUE)

headers_comp = [("B","COMPETIDOR"), ("C","PRODUCTO"), ("D","PRECIO (Bs.)"),
                ("E","DIFERENCIA"), ("F","POSICION"), ("G","OBSERVACIONES")]
ws_comp.row_dimensions[5].height = 28
for col_letter, label in headers_comp:
    style_header(ws_comp[f'{col_letter}5'], bg=DARK_BG, fg=GOLD, size=10)
    ws_comp[f'{col_letter}5'].value = label

for row in range(6, 6 + MAX_COMPETITORS):
    # Competidor (INPUT)
    c = ws_comp.cell(row=row, column=2)
    style_input(c); c.alignment = Alignment(horizontal='left', vertical='center'); unlock_cell(c)

    # Producto (INPUT)
    c = ws_comp.cell(row=row, column=3)
    style_input(c); c.alignment = Alignment(horizontal='left', vertical='center'); unlock_cell(c)

    # Precio (INPUT)
    c = ws_comp.cell(row=row, column=4)
    style_input(c); c.number_format = '#,##0.00'; unlock_cell(c)

    # Diferencia (CALC)
    c = ws_comp.cell(row=row, column=5)
    c.value = f'=IFERROR(C$3-D{row},0)'
    c.number_format = '+#,##0.00;-#,##0.00'
    style_output(c)

    # Posicion (CALC)
    c = ws_comp.cell(row=row, column=6)
    c.value = f'=IF(D{row}="","",IF(C$3<D{row},"DEBAJO",IF(C$3>D{row},"ARRIBA","IGUAL")))'
    c.font = Font(name='Calibri', size=10, bold=True)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border

    # Observaciones (INPUT)
    c = ws_comp.cell(row=row, column=7)
    style_input(c); c.alignment = Alignment(horizontal='left', vertical='center'); unlock_cell(c)

# Conditional formatting: Position
ws_comp.conditional_formatting.add(f'F6:F{5+MAX_COMPETITORS}',
    CellIsRule(operator='equal', formula=['"DEBAJO"'],
              fill=PatternFill('solid', fgColor="D1FAE5"),
              font=Font(color=GREEN_OK, bold=True)))
ws_comp.conditional_formatting.add(f'F6:F{5+MAX_COMPETITORS}',
    CellIsRule(operator='equal', formula=['"ARRIBA"'],
              fill=PatternFill('solid', fgColor="FEE2E2"),
              font=Font(color=RED_ALERT, bold=True)))
ws_comp.conditional_formatting.add(f'F6:F{5+MAX_COMPETITORS}',
    CellIsRule(operator='equal', formula=['"IGUAL"'],
              fill=PatternFill('solid', fgColor="FEF3C7"),
              font=Font(color="D97706", bold=True)))

# Sample competitors
sample_comp = [
    ("Competidor A", "Producto similar", 85, "Menor calidad"),
    ("Competidor B", "Producto premium", 120, "Marca reconocida"),
    ("Competidor C", "Producto basico", 50, "Sin servicio"),
]
for i, (comp, prod, price, obs) in enumerate(sample_comp):
    r = 6 + i
    ws_comp.cell(row=r, column=2).value = comp
    ws_comp.cell(row=r, column=3).value = prod
    ws_comp.cell(row=r, column=4).value = price
    ws_comp.cell(row=r, column=7).value = obs

# ============================================================
# SHEET 4: CONFIG
# ============================================================
ws_conf = wb.create_sheet("Config")
ws_conf.sheet_properties.tabColor = GRAY_500
ws_conf.sheet_view.showGridLines = False
cwid = {1:25, 2:45}
for c, w in cwid.items():
    ws_conf.column_dimensions[get_column_letter(c)].width = w
fill_white(ws_conf, 1, 25, 1, 3)
ws_conf['A1'].value = "v1.0.0"
ws_conf['A1'].font = Font(name='Calibri', size=9, color=GRAY_400)
ws_conf.merge_cells('A3:B3')
style_title(ws_conf['A3'], 14)
ws_conf['A3'].value = "CONFIGURACION"

settings = [
    ("Producto", "Calculadora de Precio de Venta"),
    ("Version", "v1.0.0"),
    ("Marca", "No Somos Ignorantes"),
    ("Precio", "Bs. 49"),
    ("Proteccion", "nsi2024"),
    ("", ""),
    ("INSTRUCCIONES", ""),
    ("1.", "En 'Calculadora', ingresa tus costos en las celdas amarillas."),
    ("2.", "Se calculan 3 metodos de precio: Cost-Plus, Target Margin, Break-Even."),
    ("3.", "Elige tu precio final en la celda morada grande."),
    ("4.", "En 'Escenarios', compara multiples productos/servicios."),
    ("5.", "En 'Competencia', analiza tu posicion vs competidores."),
    ("6.", "Para desbloquear: Revisar -> Desproteger hoja -> nsi2024"),
]
for i, (label, value) in enumerate(settings):
    r = 5 + i
    ws_conf.cell(row=r, column=1).value = label
    ws_conf.cell(row=r, column=1).font = Font(name='Calibri', size=10, bold=True, color=GRAY_700)
    ws_conf.cell(row=r, column=2).value = value
    ws_conf.cell(row=r, column=2).font = Font(name='Calibri', size=10, color=GRAY_600)

# ============================================================
# PROTECTION
# ============================================================
for sheet in wb.worksheets:
    sheet.protection.sheet = True
    sheet.protection.password = "nsi2024"
    sheet.protection.enable()

# Unlock Calculadora inputs
for r in range(7, 14):
    unlock_cell(ws.cell(row=r, column=3))
unlock_cell(ws['F7'])
unlock_cell(ws['F10'])
unlock_cell(ws['F13'])
unlock_cell(ws['F14'])
unlock_cell(ws['E24'])

# Unlock Escenarios
for row in range(4, 4+MAX_PRODUCTS):
    for col in [2, 3, 4, 7]:
        unlock_cell(ws_esc.cell(row=row, column=col))

# Unlock Competencia
for row in range(6, 6+MAX_COMPETITORS):
    for col in [2, 3, 4, 7]:
        unlock_cell(ws_comp.cell(row=row, column=col))

# ============================================================
# SAVE & VERIFY
# ============================================================
out = r"D:\Landing-Page_marketplace\excel_products\Calculadora_Precio_Venta_NSI.xlsx"
wb.save(out)
print(f"[OK] Saved: {out}")

from openpyxl import load_workbook
wb2 = load_workbook(out)
fc = sum(1 for s in wb2.worksheets for row in s.iter_rows() for cell in row
         if isinstance(cell.value, str) and cell.value.startswith('='))
print(f"[OK] Sheets: {wb2.sheetnames}")
print(f"[OK] Total formulas: {fc}")
print(f"[OK] Charts: Escenarios={len(ws_esc._charts)}")
