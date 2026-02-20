"""
Build: Analizador de Puntaje Crediticio (Bs. 39)
No Somos Ignorantes v1.0
Credit score analyzer: 5-factor model, payment tracker, what-if simulator, recommendations.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, RadarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule, DataBarRule, FormulaRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ============================================================
# COLORS
# ============================================================
GOLD = "D4AF37"
DARK_BG = "1A1A2E"
CARD_BG = "16213E"
ACCENT_BLUE = "0F3460"
WHITE = "FFFFFF"
BLACK = "000000"
LIGHT_YELLOW = "FFF9C4"
LIGHT_GREEN = "E8F5E9"
SUCCESS = "10B981"
WARNING = "F59E0B"
ERROR_RED = "EF4444"
INFO_BLUE = "3B82F6"
GRAY_100 = "F3F4F6"
GRAY_200 = "E5E7EB"
GRAY_300 = "D1D5DB"
GRAY_500 = "6B7280"
GRAY_700 = "374151"
GRAY_900 = "111827"

# Score-specific
SCORE_EXCELLENT = "047857"
SCORE_EXCELLENT_LIGHT = "A7F3D0"
SCORE_GOOD = "059669"
SCORE_GOOD_LIGHT = "D1FAE5"
SCORE_FAIR = "D97706"
SCORE_FAIR_LIGHT = "FEF3C7"
SCORE_POOR = "DC2626"
SCORE_POOR_LIGHT = "FEE2E2"
SCORE_VPOOR = "991B1B"

# Factor colors
F_HISTORY = "2563EB"
F_HISTORY_LIGHT = "DBEAFE"
F_UTIL = "7C3AED"
F_UTIL_LIGHT = "EDE9FE"
F_AGE = "059669"
F_AGE_LIGHT = "D1FAE5"
F_MIX = "D97706"
F_MIX_LIGHT = "FEF3C7"
F_INQUIRY = "DC2626"
F_INQUIRY_LIGHT = "FEE2E2"

FACTOR_COLORS = [F_HISTORY, F_UTIL, F_AGE, F_MIX, F_INQUIRY]
FACTOR_COLORS_LIGHT = [F_HISTORY_LIGHT, F_UTIL_LIGHT, F_AGE_LIGHT, F_MIX_LIGHT, F_INQUIRY_LIGHT]

# ============================================================
# STYLE HELPERS
# ============================================================
thin_border = Border(
    left=Side(style='thin', color=GRAY_300),
    right=Side(style='thin', color=GRAY_300),
    top=Side(style='thin', color=GRAY_300),
    bottom=Side(style='thin', color=GRAY_300)
)
gold_border = Border(
    left=Side(style='thin', color=GOLD),
    right=Side(style='thin', color=GOLD),
    top=Side(style='thin', color=GOLD),
    bottom=Side(style='thin', color=GOLD)
)

def style_header(cell, bg=DARK_BG, fg=WHITE, size=12):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')

def style_input(cell):
    cell.font = Font(name='Calibri', size=11, color="0000FF")
    cell.fill = PatternFill('solid', fgColor=LIGHT_YELLOW)
    cell.border = gold_border
    cell.alignment = Alignment(horizontal='right', vertical='center')

def style_output(cell, bold=False):
    cell.font = Font(name='Calibri', size=11, color=GRAY_900, bold=bold)
    cell.fill = PatternFill('solid', fgColor=LIGHT_GREEN)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='right', vertical='center')

def style_label(cell, bold=False):
    cell.font = Font(name='Calibri', size=11, color=GRAY_700, bold=bold)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def style_section(cell, bg, fg=WHITE):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=11)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def fmt_currency(cell):
    cell.number_format = '#,##0.00'

def fmt_pct(cell):
    cell.number_format = '0.0%'

def fmt_int(cell):
    cell.number_format = '#,##0'

def rh(ws, row, height):
    ws.row_dimensions[row].height = height

def fill_row(ws, row, cols, color):
    for c in cols:
        ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=color)

# Factor weights
WEIGHTS = [0.35, 0.30, 0.15, 0.10, 0.10]
FACTOR_NAMES = [
    "Historial de Pagos",
    "Uso de Credito",
    "Antiguedad de Cuentas",
    "Tipos de Credito",
    "Consultas Recientes"
]

# ============================================================
# SHEET 1: DASHBOARD
# ============================================================
ws = wb.active
ws.title = "Dashboard"
ws.sheet_properties.tabColor = GOLD

ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 32
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 20
ws.column_dimensions['G'].width = 2

ws.freeze_panes = 'B5'

# ── HEADER ──
rh(ws, 1, 8)
ws.merge_cells('B1:F1')
fill_row(ws, 1, range(2, 7), DARK_BG)

ws.merge_cells('B2:F2')
h = ws['B2']
h.value = "ANALIZADOR DE PUNTAJE CREDITICIO"
h.font = Font(name='Calibri', bold=True, color=GOLD, size=20)
h.fill = PatternFill('solid', fgColor=DARK_BG)
h.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws, 2, range(2, 7), DARK_BG)
rh(ws, 2, 42)

ws.merge_cells('B3:F3')
sub = ws['B3']
sub.value = "No Somos Ignorantes  |  v1.0  |  Conoce, entiende y mejora tu puntaje"
sub.font = Font(name='Calibri', color=GRAY_300, size=10, italic=True)
sub.fill = PatternFill('solid', fgColor=DARK_BG)
sub.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws, 3, range(2, 7), DARK_BG)
rh(ws, 3, 24)

rh(ws, 4, 8)

# ── SCORE DISPLAY ──
r = 5
rh(ws, r, 30)
ws.merge_cells(f'B{r}:F{r}')
sec = ws[f'B{r}']
sec.value = "TU PUNTAJE ESTIMADO"
style_section(sec, CARD_BG, fg=GOLD)
fill_row(ws, r, range(3, 7), CARD_BG)

# Big score number
rh(ws, 6, 60)
ws.merge_cells('B6:C6')
score_cell = ws['B6']
# Score formula: weighted sum * 5.5 + 300, capped at 300-850
score_cell.value = '=IFERROR(MAX(300,MIN(850,ROUND((C13*0.35+C14*0.30+C15*0.15+C16*0.10+C17*0.10)*5.5+300,0))),300)'
score_cell.font = Font(name='Calibri', bold=True, color=GOLD, size=48)
score_cell.fill = PatternFill('solid', fgColor=DARK_BG)
score_cell.alignment = Alignment(horizontal='center', vertical='center')
score_cell.border = Border(
    left=Side(style='medium', color=GOLD),
    right=Side(style='medium', color=GOLD),
    top=Side(style='medium', color=GOLD),
    bottom=Side(style='medium', color=GOLD))
ws.cell(row=6, column=3).fill = PatternFill('solid', fgColor=DARK_BG)

# Score classification
ws.merge_cells('D6:F6')
class_cell = ws['D6']
class_cell.value = '=IF(B6>=750,"EXCELENTE",IF(B6>=700,"MUY BUENO",IF(B6>=650,"BUENO",IF(B6>=550,"REGULAR",IF(B6>=450,"BAJO","MUY BAJO")))))'
class_cell.font = Font(name='Calibri', bold=True, size=24)
class_cell.alignment = Alignment(horizontal='center', vertical='center')
class_cell.border = thin_border

# Conditional formatting for classification
ws.conditional_formatting.add('D6:F6',
    FormulaRule(formula=['$B$6>=750'], fill=PatternFill('solid', fgColor=SCORE_EXCELLENT_LIGHT),
                font=Font(bold=True, color=SCORE_EXCELLENT, size=24)))
ws.conditional_formatting.add('D6:F6',
    FormulaRule(formula=['AND($B$6>=700,$B$6<750)'], fill=PatternFill('solid', fgColor=SCORE_GOOD_LIGHT),
                font=Font(bold=True, color=SCORE_GOOD, size=24)))
ws.conditional_formatting.add('D6:F6',
    FormulaRule(formula=['AND($B$6>=550,$B$6<700)'], fill=PatternFill('solid', fgColor=SCORE_FAIR_LIGHT),
                font=Font(bold=True, color=SCORE_FAIR, size=24)))
ws.conditional_formatting.add('D6:F6',
    FormulaRule(formula=['$B$6<550'], fill=PatternFill('solid', fgColor=SCORE_POOR_LIGHT),
                font=Font(bold=True, color=SCORE_POOR, size=24)))

# Score scale bar
rh(ws, 7, 28)
scale_labels = [("300", SCORE_VPOOR), ("450", SCORE_POOR), ("550", SCORE_FAIR),
                ("650", SCORE_GOOD), ("750", SCORE_EXCELLENT), ("850", SCORE_EXCELLENT)]
for i, (label, color) in enumerate(scale_labels):
    c = ws.cell(row=7, column=2 + i, value=label)
    c.font = Font(name='Calibri', bold=True, color=WHITE, size=9)
    c.fill = PatternFill('solid', fgColor=color)
    c.alignment = Alignment(horizontal='center', vertical='center')

# Next level info
rh(ws, 8, 30)
ws.merge_cells('B8:F8')
next_lvl = ws['B8']
next_lvl.value = '=IF(B6>=750,"Felicidades! Tienes un puntaje EXCELENTE. Mantelo!",IF(B6>=700,"Proximo nivel: EXCELENTE (750). Necesitas +"&TEXT(750-B6,"#,##0")&" puntos",IF(B6>=650,"Proximo nivel: MUY BUENO (700). Necesitas +"&TEXT(700-B6,"#,##0")&" puntos",IF(B6>=550,"Proximo nivel: BUENO (650). Necesitas +"&TEXT(650-B6,"#,##0")&" puntos","Proximo nivel: REGULAR (550). Necesitas +"&TEXT(550-B6,"#,##0")&" puntos"))))'
next_lvl.font = Font(name='Calibri', bold=True, color=INFO_BLUE, size=12)
next_lvl.alignment = Alignment(horizontal='center', vertical='center')
next_lvl.fill = PatternFill('solid', fgColor=GRAY_100)
next_lvl.border = thin_border

rh(ws, 9, 8)

# ── FACTOR INPUTS ──
r = 10
rh(ws, r, 30)
ws.merge_cells(f'B{r}:F{r}')
sec = ws[f'B{r}']
sec.value = "FACTORES QUE AFECTAN TU PUNTAJE (ingresa tu puntuacion 0-100)"
style_section(sec, ACCENT_BLUE)
fill_row(ws, r, range(3, 7), ACCENT_BLUE)

# Headers
r = 11
rh(ws, r, 26)
factor_headers = ['', 'FACTOR', 'TU PUNTAJE (0-100)', 'PESO (%)', 'CONTRIBUCION', 'ESTADO']
for i, hdr in enumerate(factor_headers):
    c = ws.cell(row=r, column=i + 1, value=hdr)
    if i >= 1:
        style_header(c, bg=CARD_BG, fg=GOLD, size=9)

# Detailed sub-labels
rh(ws, 12, 6)

# Data validation for factor scores
dv_score = DataValidation(type="whole", operator="between", formula1=0, formula2=100)
dv_score.error = "El puntaje debe estar entre 0 y 100"
dv_score.errorTitle = "Puntaje invalido"
dv_score.prompt = "Ingresa tu puntuacion estimada de 0 (peor) a 100 (mejor)"
dv_score.promptTitle = "Puntaje del Factor"
ws.add_data_validation(dv_score)

# Factor defaults (sample good-but-improvable profile)
factor_defaults = [85, 60, 75, 70, 45]
factor_descriptions = [
    "Pagos a tiempo en los ultimos 24 meses",
    "% de credito usado vs limite disponible",
    "Anios promedio de tus cuentas abiertas",
    "Variedad: tarjetas, prestamos, hipoteca",
    "Consultas al buro en ultimos 6 meses"
]

for i, (name, default, desc) in enumerate(zip(FACTOR_NAMES, factor_defaults, factor_descriptions)):
    row = 13 + i
    rh(ws, row, 34)
    bg = FACTOR_COLORS_LIGHT[i]

    # Factor name
    lbl = ws.cell(row=row, column=2, value=f"{name} ({int(WEIGHTS[i]*100)}%)")
    lbl.font = Font(name='Calibri', bold=True, color=FACTOR_COLORS[i], size=11)
    lbl.fill = PatternFill('solid', fgColor=bg)
    lbl.border = thin_border

    # Score input
    sc = ws.cell(row=row, column=3, value=default)
    style_input(sc)
    sc.font = Font(name='Calibri', size=14, bold=True, color="0000FF")
    fmt_int(sc)
    dv_score.add(f'C{row}')

    # Weight
    wt = ws.cell(row=row, column=4, value=WEIGHTS[i])
    style_output(wt)
    wt.fill = PatternFill('solid', fgColor=bg)
    fmt_pct(wt)

    # Contribution
    cont = ws.cell(row=row, column=5)
    cont.value = f'=ROUND(C{row}*D{row},1)'
    style_output(cont, bold=True)
    cont.fill = PatternFill('solid', fgColor=bg)
    fmt_int(cont)

    # Status
    stat = ws.cell(row=row, column=6)
    stat.value = f'=IF(C{row}>=80,"Excelente",IF(C{row}>=60,"Bueno",IF(C{row}>=40,"Mejorar","Critico")))'
    stat.alignment = Alignment(horizontal='center', vertical='center')
    stat.border = thin_border
    stat.font = Font(name='Calibri', bold=True, size=11)

# Conditional formatting for status
for row in range(13, 18):
    ws.conditional_formatting.add(f'F{row}',
        CellIsRule(operator='equal', formula=['"Excelente"'],
                  fill=PatternFill('solid', fgColor=SCORE_EXCELLENT_LIGHT),
                  font=Font(bold=True, color=SCORE_EXCELLENT)))
    ws.conditional_formatting.add(f'F{row}',
        CellIsRule(operator='equal', formula=['"Bueno"'],
                  fill=PatternFill('solid', fgColor=SCORE_GOOD_LIGHT),
                  font=Font(bold=True, color=SCORE_GOOD)))
    ws.conditional_formatting.add(f'F{row}',
        CellIsRule(operator='equal', formula=['"Mejorar"'],
                  fill=PatternFill('solid', fgColor=SCORE_FAIR_LIGHT),
                  font=Font(bold=True, color=SCORE_FAIR)))
    ws.conditional_formatting.add(f'F{row}',
        CellIsRule(operator='equal', formula=['"Critico"'],
                  fill=PatternFill('solid', fgColor=SCORE_POOR_LIGHT),
                  font=Font(bold=True, color=SCORE_POOR)))

# Data bars on score inputs
ws.conditional_formatting.add('C13:C17',
    DataBarRule(start_type='num', start_value=0, end_type='num', end_value=100,
                color=INFO_BLUE))

# Totals
rh(ws, 18, 28)
ws.cell(row=18, column=2, value="TOTAL PONDERADO").font = Font(name='Calibri', bold=True, color=WHITE, size=11)
ws.cell(row=18, column=2).fill = PatternFill('solid', fgColor=CARD_BG)
ws.cell(row=18, column=2).border = thin_border

ws.cell(row=18, column=3).value = '=ROUND(SUMPRODUCT(C13:C17,D13:D17)/SUM(D13:D17),0)'
ws.cell(row=18, column=3).font = Font(name='Calibri', bold=True, color=GOLD, size=14)
ws.cell(row=18, column=3).fill = PatternFill('solid', fgColor=CARD_BG)
ws.cell(row=18, column=3).alignment = Alignment(horizontal='center')
fmt_int(ws.cell(row=18, column=3))

ws.cell(row=18, column=4, value=1.0)
ws.cell(row=18, column=4).font = Font(name='Calibri', bold=True, color=GOLD, size=11)
ws.cell(row=18, column=4).fill = PatternFill('solid', fgColor=CARD_BG)
fmt_pct(ws.cell(row=18, column=4))

ws.cell(row=18, column=5).value = '=SUM(E13:E17)'
ws.cell(row=18, column=5).font = Font(name='Calibri', bold=True, color=GOLD, size=14)
ws.cell(row=18, column=5).fill = PatternFill('solid', fgColor=CARD_BG)
ws.cell(row=18, column=5).alignment = Alignment(horizontal='center')
fmt_int(ws.cell(row=18, column=5))

for col in [6]:
    ws.cell(row=18, column=col).fill = PatternFill('solid', fgColor=CARD_BG)

rh(ws, 19, 10)

# ── SCORE INTERPRETATION TABLE ──
r = 20
rh(ws, r, 28)
ws.merge_cells(f'B{r}:F{r}')
sec = ws[f'B{r}']
sec.value = "INTERPRETACION DEL PUNTAJE"
style_section(sec, GRAY_700)
fill_row(ws, r, range(3, 7), GRAY_700)

score_ranges = [
    ("750 - 850", "EXCELENTE", "Acceso a las mejores tasas y condiciones de credito", SCORE_EXCELLENT, SCORE_EXCELLENT_LIGHT),
    ("700 - 749", "MUY BUENO", "Buenas opciones de credito con tasas competitivas", SCORE_GOOD, SCORE_GOOD_LIGHT),
    ("650 - 699", "BUENO", "Acceso a credito estandar, tasas promedio del mercado", "0D9488", "CCFBF1"),
    ("550 - 649", "REGULAR", "Opciones limitadas, tasas mas altas, posibles garantias", SCORE_FAIR, SCORE_FAIR_LIGHT),
    ("450 - 549", "BAJO", "Dificultad para obtener credito, necesita mejorar", SCORE_POOR, SCORE_POOR_LIGHT),
    ("300 - 449", "MUY BAJO", "Requiere reconstruir historial. Busca asesoramiento", SCORE_VPOOR, "FECACA"),
]

for i, (rng, level, meaning, color, bg) in enumerate(score_ranges):
    row = 21 + i
    rh(ws, row, 28)

    r_c = ws.cell(row=row, column=2, value=rng)
    r_c.font = Font(name='Calibri', bold=True, size=10, color=color)
    r_c.fill = PatternFill('solid', fgColor=bg)
    r_c.border = thin_border
    r_c.alignment = Alignment(horizontal='center')

    l_c = ws.cell(row=row, column=3, value=level)
    l_c.font = Font(name='Calibri', bold=True, size=10, color=color)
    l_c.fill = PatternFill('solid', fgColor=bg)
    l_c.border = thin_border
    l_c.alignment = Alignment(horizontal='center')

    ws.merge_cells(f'D{row}:F{row}')
    m_c = ws.cell(row=row, column=4)
    m_c.value = meaning
    m_c.font = Font(name='Calibri', size=10, color=GRAY_700)
    m_c.fill = PatternFill('solid', fgColor=bg)
    m_c.border = thin_border
    m_c.alignment = Alignment(wrap_text=True, vertical='center')

# Highlight current range
for i in range(6):
    row = 21 + i
    ws.conditional_formatting.add(f'B{row}:F{row}',
        FormulaRule(
            formula=[f'AND($B$6>={[750,700,650,550,450,300][i]},$B$6<{[851,750,700,650,550,450][i]})'],
            fill=PatternFill('solid', fgColor=GOLD),
            font=Font(bold=True, color=DARK_BG)))

rh(ws, 27, 10)

# ── CHARTS ──
# Radar chart for factor scores
radar = RadarChart()
radar.type = "filled"
radar.title = "Perfil de Puntaje Crediticio"
radar.style = 10
radar.width = 16
radar.height = 14

# Helper data for radar
ws.cell(row=60, column=2, value="Factor")
ws.cell(row=60, column=3, value="Tu Puntaje")
ws.cell(row=60, column=4, value="Ideal")
for i, name in enumerate(FACTOR_NAMES):
    ws.cell(row=61+i, column=2, value=name)
    ws.cell(row=61+i, column=3).value = f'=C{13+i}'
    ws.cell(row=61+i, column=4, value=100)

data_radar = Reference(ws, min_col=3, min_row=60, max_row=65)
ideal_radar = Reference(ws, min_col=4, min_row=60, max_row=65)
cats_radar = Reference(ws, min_col=2, min_row=61, max_row=65)
radar.add_data(data_radar, titles_from_data=True)
radar.add_data(ideal_radar, titles_from_data=True)
radar.set_categories(cats_radar)
radar.series[0].graphicalProperties.solidFill = INFO_BLUE
radar.series[1].graphicalProperties.solidFill = GRAY_300
ws.add_chart(radar, "B28")

# Bar chart: factor contributions
bar_factors = BarChart()
bar_factors.type = "col"
bar_factors.title = "Contribucion de Cada Factor a tu Puntaje"
bar_factors.y_axis.title = "Puntos"
bar_factors.style = 10
bar_factors.width = 16
bar_factors.height = 14

ws.cell(row=68, column=2, value="Factor")
ws.cell(row=68, column=3, value="Contribucion")
for i, name in enumerate(FACTOR_NAMES):
    ws.cell(row=69+i, column=2, value=name)
    ws.cell(row=69+i, column=3).value = f'=E{13+i}'

data_bar = Reference(ws, min_col=3, min_row=68, max_row=73)
cats_bar = Reference(ws, min_col=2, min_row=69, max_row=73)
bar_factors.add_data(data_bar, titles_from_data=True)
bar_factors.set_categories(cats_bar)

for i in range(5):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = FACTOR_COLORS[i]
    bar_factors.series[0].data_points.append(pt)

bar_factors.dataLabels = DataLabelList()
bar_factors.dataLabels.showVal = True
ws.add_chart(bar_factors, "D28")

# Pie chart: weight distribution
pie = PieChart()
pie.title = "Peso de Cada Factor (%)"
pie.style = 10
pie.width = 16
pie.height = 12

ws.cell(row=76, column=2, value="Factor")
ws.cell(row=76, column=3, value="Peso")
for i, name in enumerate(FACTOR_NAMES):
    ws.cell(row=77+i, column=2, value=name)
    ws.cell(row=77+i, column=3, value=WEIGHTS[i])

pie_data = Reference(ws, min_col=3, min_row=76, max_row=81)
pie_cats = Reference(ws, min_col=2, min_row=77, max_row=81)
pie.add_data(pie_data, titles_from_data=True)
pie.set_categories(pie_cats)

pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
pie.dataLabels.showCatName = True

for i in range(5):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = FACTOR_COLORS[i]
    pie.series[0].data_points.append(pt)

ws.add_chart(pie, "B44")


# ============================================================
# SHEET 2: HISTORIAL (Payment history tracker)
# ============================================================
ws2 = wb.create_sheet("Historial")
ws2.sheet_properties.tabColor = F_HISTORY

ws2.column_dimensions['A'].width = 2
ws2.column_dimensions['B'].width = 24
ws2.column_dimensions['C'].width = 16
ws2.column_dimensions['D'].width = 16
ws2.column_dimensions['E'].width = 16
ws2.column_dimensions['F'].width = 16
ws2.column_dimensions['G'].width = 16
ws2.column_dimensions['H'].width = 2

ws2.freeze_panes = 'B6'

# Header
ws2.merge_cells('B1:G1')
fill_row(ws2, 1, range(2, 8), DARK_BG)
rh(ws2, 1, 8)

ws2.merge_cells('B2:G2')
h = ws2['B2']
h.value = "HISTORIAL DE PAGOS"
h.font = Font(name='Calibri', bold=True, color=GOLD, size=18)
h.fill = PatternFill('solid', fgColor=DARK_BG)
h.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws2, 2, range(2, 8), DARK_BG)
rh(ws2, 2, 38)

ws2.merge_cells('B3:G3')
sub = ws2['B3']
sub.value = "Registra tus pagos de los ultimos 24 meses para calcular tu puntuacion"
sub.font = Font(name='Calibri', color=GRAY_300, size=10, italic=True)
sub.fill = PatternFill('solid', fgColor=DARK_BG)
sub.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws2, 3, range(2, 8), DARK_BG)
rh(ws2, 3, 24)

rh(ws2, 4, 8)

# ── PAYMENT HISTORY INPUTS ──
r = 5
ws2.merge_cells(f'B{r}:G{r}')
sec = ws2[f'B{r}']
sec.value = "RESUMEN DE PAGOS (ultimos 24 meses)"
style_section(sec, F_HISTORY)
fill_row(ws2, r, range(3, 8), F_HISTORY)
rh(ws2, r, 28)

# Headers
r = 6
rh(ws2, r, 26)
hist_headers = ['', 'METRICA', 'CANTIDAD', 'IMPACTO', 'PUNTOS PERDIDOS']
for i, hdr in enumerate(hist_headers):
    c = ws2.cell(row=r, column=i + 1, value=hdr)
    if i >= 1:
        style_header(c, bg=CARD_BG, fg=GOLD, size=9)

# Payment history metrics
hist_metrics = [
    (7, "Total de pagos realizados", 24, "Base del historial"),
    (8, "Pagos puntuales (a tiempo)", 20, "Factor positivo principal"),
    (9, "Pagos atrasados 1-30 dias", 3, "-10 puntos cada uno"),
    (10, "Pagos atrasados 31-60 dias", 1, "-20 puntos cada uno"),
    (11, "Pagos atrasados 61-90 dias", 0, "-40 puntos cada uno"),
    (12, "Pagos atrasados 90+ dias", 0, "-50 puntos cada uno"),
    (13, "Cuentas en cobranza", 0, "-30 puntos cada una"),
]

dv_count = DataValidation(type="whole", operator="greaterThanOrEqual", formula1=0)
dv_count.error = "La cantidad no puede ser negativa"
dv_count.errorTitle = "Cantidad invalida"
ws2.add_data_validation(dv_count)

for (row, label, default, impact) in hist_metrics:
    rh(ws2, row, 28)

    lbl = ws2.cell(row=row, column=2, value=label)
    style_label(lbl, bold=True)
    lbl.fill = PatternFill('solid', fgColor=F_HISTORY_LIGHT)
    lbl.border = thin_border

    val = ws2.cell(row=row, column=3, value=default)
    style_input(val)
    fmt_int(val)
    dv_count.add(f'C{row}')

    imp = ws2.cell(row=row, column=4, value=impact)
    imp.font = Font(name='Calibri', size=9, color=GRAY_500, italic=True)
    imp.fill = PatternFill('solid', fgColor=F_HISTORY_LIGHT)
    imp.border = thin_border

    # Points lost
    pts = ws2.cell(row=row, column=5)
    if row == 7 or row == 8:
        pts.value = 0
    elif row == 9:
        pts.value = f'=C{row}*10'
    elif row == 10:
        pts.value = f'=C{row}*20'
    elif row == 11:
        pts.value = f'=C{row}*40'
    elif row == 12:
        pts.value = f'=C{row}*50'
    elif row == 13:
        pts.value = f'=C{row}*30'
    style_output(pts)
    fmt_int(pts)

# Calculated score
rh(ws2, 14, 6)
rh(ws2, 15, 32)
ws2.merge_cells('B15:C15')
lbl = ws2['B15']
lbl.value = "PUNTAJE HISTORIAL DE PAGOS"
lbl.font = Font(name='Calibri', bold=True, color=WHITE, size=12)
lbl.fill = PatternFill('solid', fgColor=F_HISTORY)
lbl.alignment = Alignment(horizontal='center', vertical='center')
ws2.cell(row=15, column=3).fill = PatternFill('solid', fgColor=F_HISTORY)

ws2.merge_cells('D15:E15')
score_hist = ws2['D15']
score_hist.value = '=MAX(0,MIN(100,100-SUM(E9:E13)))'
score_hist.font = Font(name='Calibri', bold=True, color=WHITE, size=20)
score_hist.fill = PatternFill('solid', fgColor=F_HISTORY)
score_hist.alignment = Alignment(horizontal='center', vertical='center')
fmt_int(score_hist)

rh(ws2, 16, 6)
ws2.merge_cells('B16:E16')
tip = ws2['B16']
tip.value = "Usa este puntaje en el Dashboard (factor: Historial de Pagos)"
tip.font = Font(name='Calibri', size=10, color=INFO_BLUE, italic=True)
tip.alignment = Alignment(horizontal='center')

rh(ws2, 17, 10)

# ── CREDIT UTILIZATION CALCULATOR ──
r = 18
ws2.merge_cells(f'B{r}:G{r}')
sec = ws2[f'B{r}']
sec.value = "CALCULADORA DE USO DE CREDITO"
style_section(sec, F_UTIL)
fill_row(ws2, r, range(3, 8), F_UTIL)
rh(ws2, r, 28)

# Credit cards/lines input
r = 19
rh(ws2, r, 26)
util_headers = ['', 'TARJETA/LINEA', 'LIMITE (Bs.)', 'SALDO USADO (Bs.)', 'USO (%)', 'ESTADO']
for i, hdr in enumerate(util_headers):
    c = ws2.cell(row=r, column=i + 1, value=hdr)
    if i >= 1:
        style_header(c, bg=CARD_BG, fg=GOLD, size=9)

sample_cards = [
    ("Visa Clasica", 15000, 4500),
    ("Mastercard Gold", 30000, 12000),
    ("Linea de credito BNB", 50000, 5000),
]

NUM_CARDS = 8
dv_limit = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1=0)
dv_limit.error = "No puede ser negativo"
ws2.add_data_validation(dv_limit)

for i in range(NUM_CARDS):
    row = 20 + i
    rh(ws2, row, 26)

    # Name
    n = ws2.cell(row=row, column=2)
    if i < len(sample_cards):
        n.value = sample_cards[i][0]
    style_label(n)
    n.fill = PatternFill('solid', fgColor=LIGHT_YELLOW)
    n.border = gold_border
    n.font = Font(name='Calibri', size=11, color="0000FF")

    # Limit
    lim = ws2.cell(row=row, column=3)
    if i < len(sample_cards):
        lim.value = sample_cards[i][1]
    style_input(lim)
    fmt_currency(lim)
    dv_limit.add(f'C{row}')

    # Balance
    bal = ws2.cell(row=row, column=4)
    if i < len(sample_cards):
        bal.value = sample_cards[i][2]
    style_input(bal)
    fmt_currency(bal)
    dv_limit.add(f'D{row}')

    # Usage %
    use = ws2.cell(row=row, column=5)
    use.value = f'=IFERROR(IF(C{row}>0,D{row}/C{row},0),0)'
    style_output(use)
    fmt_pct(use)

    # Status
    stat = ws2.cell(row=row, column=6)
    stat.value = f'=IF(C{row}=0,"",IF(E{row}<0.1,"Optimo",IF(E{row}<0.3,"Bueno",IF(E{row}<0.5,"Cuidado",IF(E{row}<0.75,"Alto","Critico")))))'
    stat.alignment = Alignment(horizontal='center')
    stat.border = thin_border
    stat.font = Font(name='Calibri', bold=True)

# Status formatting
for row in range(20, 20 + NUM_CARDS):
    ws2.conditional_formatting.add(f'F{row}',
        CellIsRule(operator='equal', formula=['"Optimo"'],
                  fill=PatternFill('solid', fgColor=SCORE_EXCELLENT_LIGHT),
                  font=Font(bold=True, color=SCORE_EXCELLENT)))
    ws2.conditional_formatting.add(f'F{row}',
        CellIsRule(operator='equal', formula=['"Bueno"'],
                  fill=PatternFill('solid', fgColor=SCORE_GOOD_LIGHT),
                  font=Font(bold=True, color=SCORE_GOOD)))
    ws2.conditional_formatting.add(f'F{row}',
        CellIsRule(operator='equal', formula=['"Cuidado"'],
                  fill=PatternFill('solid', fgColor=SCORE_FAIR_LIGHT),
                  font=Font(bold=True, color=SCORE_FAIR)))
    ws2.conditional_formatting.add(f'F{row}',
        CellIsRule(operator='equal', formula=['"Alto"'],
                  fill=PatternFill('solid', fgColor=SCORE_POOR_LIGHT),
                  font=Font(bold=True, color=SCORE_POOR)))
    ws2.conditional_formatting.add(f'F{row}',
        CellIsRule(operator='equal', formula=['"Critico"'],
                  fill=PatternFill('solid', fgColor="FECACA"),
                  font=Font(bold=True, color=SCORE_VPOOR)))

# Color scale on usage
ws2.conditional_formatting.add(f'E20:E{20+NUM_CARDS-1}',
    ColorScaleRule(start_type='num', start_value=0, start_color=SCORE_EXCELLENT_LIGHT,
                   mid_type='num', mid_value=0.3, mid_color=SCORE_FAIR_LIGHT,
                   end_type='num', end_value=1, end_color=SCORE_POOR_LIGHT))

# Totals
tot_row = 20 + NUM_CARDS
rh(ws2, tot_row, 28)
ws2.cell(row=tot_row, column=2, value="TOTAL").font = Font(name='Calibri', bold=True, color=WHITE, size=11)
ws2.cell(row=tot_row, column=2).fill = PatternFill('solid', fgColor=F_UTIL)
ws2.cell(row=tot_row, column=2).alignment = Alignment(horizontal='center')

for col in [3, 4]:
    c = ws2.cell(row=tot_row, column=col)
    c.value = f'=SUM({get_column_letter(col)}20:{get_column_letter(col)}{20+NUM_CARDS-1})'
    c.font = Font(name='Calibri', bold=True, color=WHITE, size=11)
    c.fill = PatternFill('solid', fgColor=F_UTIL)
    fmt_currency(c)

avg_use = ws2.cell(row=tot_row, column=5)
avg_use.value = f'=IFERROR(D{tot_row}/C{tot_row},0)'
avg_use.font = Font(name='Calibri', bold=True, color=WHITE, size=11)
avg_use.fill = PatternFill('solid', fgColor=F_UTIL)
fmt_pct(avg_use)

for col in [6]:
    ws2.cell(row=tot_row, column=col).fill = PatternFill('solid', fgColor=F_UTIL)

# Credit utilization score
rh(ws2, tot_row + 1, 6)
rh(ws2, tot_row + 2, 32)
ws2.merge_cells(f'B{tot_row+2}:C{tot_row+2}')
lbl = ws2[f'B{tot_row+2}']
lbl.value = "PUNTAJE USO DE CREDITO"
lbl.font = Font(name='Calibri', bold=True, color=WHITE, size=12)
lbl.fill = PatternFill('solid', fgColor=F_UTIL)
lbl.alignment = Alignment(horizontal='center', vertical='center')
ws2.cell(row=tot_row+2, column=3).fill = PatternFill('solid', fgColor=F_UTIL)

ws2.merge_cells(f'D{tot_row+2}:E{tot_row+2}')
score_util = ws2[f'D{tot_row+2}']
score_util.value = f'=IF(E{tot_row}<0.1,100,IF(E{tot_row}<0.3,90,IF(E{tot_row}<0.5,70,IF(E{tot_row}<0.75,50,30))))'
score_util.font = Font(name='Calibri', bold=True, color=WHITE, size=20)
score_util.fill = PatternFill('solid', fgColor=F_UTIL)
score_util.alignment = Alignment(horizontal='center', vertical='center')
fmt_int(score_util)

ws2.merge_cells(f'B{tot_row+3}:E{tot_row+3}')
tip2 = ws2[f'B{tot_row+3}']
tip2.value = "Usa este puntaje en el Dashboard (factor: Uso de Credito)"
tip2.font = Font(name='Calibri', size=10, color=INFO_BLUE, italic=True)
tip2.alignment = Alignment(horizontal='center')


# ============================================================
# SHEET 3: SIMULADOR (What-if scenarios)
# ============================================================
ws3 = wb.create_sheet("Simulador")
ws3.sheet_properties.tabColor = WARNING

ws3.column_dimensions['A'].width = 2
ws3.column_dimensions['B'].width = 32
ws3.column_dimensions['C'].width = 18
ws3.column_dimensions['D'].width = 18
ws3.column_dimensions['E'].width = 18
ws3.column_dimensions['F'].width = 20
ws3.column_dimensions['G'].width = 2

ws3.freeze_panes = 'B5'

# Header
ws3.merge_cells('B1:F1')
fill_row(ws3, 1, range(2, 7), DARK_BG)
rh(ws3, 1, 8)

ws3.merge_cells('B2:F2')
h = ws3['B2']
h.value = "SIMULADOR: QUE PASA SI...?"
h.font = Font(name='Calibri', bold=True, color=GOLD, size=18)
h.fill = PatternFill('solid', fgColor=DARK_BG)
h.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws3, 2, range(2, 7), DARK_BG)
rh(ws3, 2, 40)

ws3.merge_cells('B3:F3')
sub = ws3['B3']
sub.value = "Simula diferentes escenarios y ve como afectan tu puntaje"
sub.font = Font(name='Calibri', color=GRAY_300, size=10, italic=True)
sub.fill = PatternFill('solid', fgColor=DARK_BG)
sub.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws3, 3, range(2, 7), DARK_BG)
rh(ws3, 3, 24)

rh(ws3, 4, 8)

# ── SCENARIO TABLE ──
r = 5
ws3.merge_cells(f'B{r}:F{r}')
sec = ws3[f'B{r}']
sec.value = "ESCENARIOS DE MEJORA"
style_section(sec, WARNING)
fill_row(ws3, r, range(3, 7), WARNING)
rh(ws3, r, 28)

r = 6
rh(ws3, r, 26)
sim_headers = ['', 'ESCENARIO', 'PUNTAJE ACTUAL', 'PUNTAJE NUEVO', 'CAMBIO', 'IMPACTO']
for i, hdr in enumerate(sim_headers):
    c = ws3.cell(row=r, column=i + 1, value=hdr)
    if i >= 1:
        style_header(c, bg=CARD_BG, fg=GOLD, size=9)

# Scenarios with formulas
scenarios = [
    (7, "Pago todas las deudas atrasadas",
     '=Dashboard!B6',
     '=IFERROR(MAX(300,MIN(850,ROUND((100*0.35+Dashboard!C14*0.30+Dashboard!C15*0.15+Dashboard!C16*0.10+Dashboard!C17*0.10)*5.5+300,0))),300)'),
    (8, "Reduzco uso de credito al 10%",
     '=Dashboard!B6',
     '=IFERROR(MAX(300,MIN(850,ROUND((Dashboard!C13*0.35+100*0.30+Dashboard!C15*0.15+Dashboard!C16*0.10+Dashboard!C17*0.10)*5.5+300,0))),300)'),
    (9, "Mantengo cuentas por 5+ anios",
     '=Dashboard!B6',
     '=IFERROR(MAX(300,MIN(850,ROUND((Dashboard!C13*0.35+Dashboard!C14*0.30+90*0.15+Dashboard!C16*0.10+Dashboard!C17*0.10)*5.5+300,0))),300)'),
    (10, "Agrego un tipo de credito nuevo",
     '=Dashboard!B6',
     '=IFERROR(MAX(300,MIN(850,ROUND((Dashboard!C13*0.35+Dashboard!C14*0.30+Dashboard!C15*0.15+85*0.10+Dashboard!C17*0.10)*5.5+300,0))),300)'),
    (11, "No hago mas consultas por 6 meses",
     '=Dashboard!B6',
     '=IFERROR(MAX(300,MIN(850,ROUND((Dashboard!C13*0.35+Dashboard!C14*0.30+Dashboard!C15*0.15+Dashboard!C16*0.10+80*0.10)*5.5+300,0))),300)'),
    (12, "ESCENARIO IDEAL (todo perfecto)",
     '=Dashboard!B6',
     '=IFERROR(MAX(300,MIN(850,ROUND((100*0.35+100*0.30+100*0.15+100*0.10+100*0.10)*5.5+300,0))),300)'),
    (13, "PEOR CASO (todo critico)",
     '=Dashboard!B6',
     '=IFERROR(MAX(300,MIN(850,ROUND((10*0.35+10*0.30+10*0.15+10*0.10+10*0.10)*5.5+300,0))),300)'),
]

for (row, label, current, new_score) in scenarios:
    rh(ws3, row, 30)
    bg = GRAY_100 if (row - 7) % 2 == 0 else WHITE

    lbl = ws3.cell(row=row, column=2, value=label)
    style_label(lbl, bold=True)
    lbl.fill = PatternFill('solid', fgColor=bg)
    lbl.border = thin_border

    cur = ws3.cell(row=row, column=3)
    cur.value = current
    style_output(cur)
    cur.fill = PatternFill('solid', fgColor=bg)
    fmt_int(cur)

    new = ws3.cell(row=row, column=4)
    new.value = new_score
    style_output(new, bold=True)
    new.fill = PatternFill('solid', fgColor=bg)
    fmt_int(new)

    # Change
    chg = ws3.cell(row=row, column=5)
    chg.value = f'=D{row}-C{row}'
    chg.font = Font(name='Calibri', bold=True, size=12)
    chg.alignment = Alignment(horizontal='center', vertical='center')
    chg.border = thin_border
    chg.fill = PatternFill('solid', fgColor=bg)
    fmt_int(chg)

    # Impact description
    imp = ws3.cell(row=row, column=6)
    imp.value = f'=IF(E{row}>50,"ALTO IMPACTO",IF(E{row}>20,"IMPACTO MEDIO",IF(E{row}>0,"IMPACTO BAJO",IF(E{row}=0,"SIN CAMBIO","NEGATIVO"))))'
    imp.alignment = Alignment(horizontal='center', vertical='center')
    imp.border = thin_border
    imp.font = Font(name='Calibri', bold=True)

# Conditional formatting
for row in range(7, 14):
    ws3.conditional_formatting.add(f'E{row}',
        CellIsRule(operator='greaterThan', formula=['0'],
                  fill=PatternFill('solid', fgColor=SCORE_GOOD_LIGHT),
                  font=Font(bold=True, color=SCORE_GOOD, size=14)))
    ws3.conditional_formatting.add(f'E{row}',
        CellIsRule(operator='lessThan', formula=['0'],
                  fill=PatternFill('solid', fgColor=SCORE_POOR_LIGHT),
                  font=Font(bold=True, color=SCORE_POOR, size=14)))

    ws3.conditional_formatting.add(f'F{row}',
        CellIsRule(operator='equal', formula=['"ALTO IMPACTO"'],
                  fill=PatternFill('solid', fgColor=SCORE_EXCELLENT_LIGHT),
                  font=Font(bold=True, color=SCORE_EXCELLENT)))
    ws3.conditional_formatting.add(f'F{row}',
        CellIsRule(operator='equal', formula=['"NEGATIVO"'],
                  fill=PatternFill('solid', fgColor=SCORE_POOR_LIGHT),
                  font=Font(bold=True, color=SCORE_POOR)))

rh(ws3, 14, 10)

# ── CUSTOM SIMULATOR ──
r = 15
ws3.merge_cells(f'B{r}:F{r}')
sec = ws3[f'B{r}']
sec.value = "SIMULADOR PERSONALIZADO"
style_section(sec, INFO_BLUE)
fill_row(ws3, r, range(3, 7), INFO_BLUE)
rh(ws3, r, 28)

ws3.merge_cells(f'B16:F16')
ws3['B16'].value = "Ingresa los puntajes hipoteticos para cada factor y ve el resultado"
ws3['B16'].font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)
ws3['B16'].alignment = Alignment(horizontal='center')
rh(ws3, 16, 22)

r = 17
rh(ws3, r, 24)
cust_headers = ['', 'FACTOR', 'ACTUAL', 'SIMULADO', 'DIFERENCIA']
for i, hdr in enumerate(cust_headers):
    c = ws3.cell(row=r, column=i + 1, value=hdr)
    if i >= 1:
        style_header(c, bg=CARD_BG, fg=GOLD, size=9)

dv_sim = DataValidation(type="whole", operator="between", formula1=0, formula2=100)
dv_sim.error = "Valor entre 0 y 100"
dv_sim.errorTitle = "Invalido"
ws3.add_data_validation(dv_sim)

for i, name in enumerate(FACTOR_NAMES):
    row = 18 + i
    rh(ws3, row, 28)
    bg = FACTOR_COLORS_LIGHT[i]

    lbl = ws3.cell(row=row, column=2, value=name)
    lbl.font = Font(name='Calibri', bold=True, color=FACTOR_COLORS[i], size=10)
    lbl.fill = PatternFill('solid', fgColor=bg)
    lbl.border = thin_border

    actual = ws3.cell(row=row, column=3)
    actual.value = f'=Dashboard!C{13+i}'
    style_output(actual)
    actual.fill = PatternFill('solid', fgColor=bg)
    fmt_int(actual)

    sim = ws3.cell(row=row, column=4)
    sim.value = factor_defaults[i] + 10  # slightly improved default
    style_input(sim)
    fmt_int(sim)
    dv_sim.add(f'D{row}')

    diff = ws3.cell(row=row, column=5)
    diff.value = f'=D{row}-C{row}'
    style_output(diff)
    fmt_int(diff)

# Simulated score result
rh(ws3, 23, 6)
rh(ws3, 24, 38)
ws3.merge_cells('B24:C24')
lbl = ws3['B24']
lbl.value = "PUNTAJE SIMULADO"
lbl.font = Font(name='Calibri', bold=True, color=GOLD, size=14)
lbl.fill = PatternFill('solid', fgColor=DARK_BG)
lbl.alignment = Alignment(horizontal='center', vertical='center')
ws3.cell(row=24, column=3).fill = PatternFill('solid', fgColor=DARK_BG)

ws3.merge_cells('D24:E24')
sim_score = ws3['D24']
sim_score.value = '=IFERROR(MAX(300,MIN(850,ROUND((D18*0.35+D19*0.30+D20*0.15+D21*0.10+D22*0.10)*5.5+300,0))),300)'
sim_score.font = Font(name='Calibri', bold=True, color=GOLD, size=28)
sim_score.fill = PatternFill('solid', fgColor=DARK_BG)
sim_score.alignment = Alignment(horizontal='center', vertical='center')
fmt_int(sim_score)

rh(ws3, 25, 28)
ws3.merge_cells('B25:C25')
ws3['B25'].value = "VS ACTUAL:"
ws3['B25'].font = Font(name='Calibri', bold=True, size=11, color=GRAY_500)
ws3['B25'].alignment = Alignment(horizontal='center')

ws3.merge_cells('D25:E25')
diff_score = ws3['D25']
diff_score.value = '=D24-Dashboard!B6'
diff_score.font = Font(name='Calibri', bold=True, size=16)
diff_score.alignment = Alignment(horizontal='center', vertical='center')
fmt_int(diff_score)

ws3.conditional_formatting.add('D25:E25',
    CellIsRule(operator='greaterThan', formula=['0'],
              fill=PatternFill('solid', fgColor=SCORE_GOOD_LIGHT),
              font=Font(bold=True, color=SCORE_GOOD, size=16)))
ws3.conditional_formatting.add('D25:E25',
    CellIsRule(operator='lessThan', formula=['0'],
              fill=PatternFill('solid', fgColor=SCORE_POOR_LIGHT),
              font=Font(bold=True, color=SCORE_POOR, size=16)))

# Bar chart: actual vs simulated
bar_sim = BarChart()
bar_sim.type = "col"
bar_sim.grouping = "clustered"
bar_sim.title = "Actual vs Simulado por Factor"
bar_sim.y_axis.title = "Puntaje"
bar_sim.style = 10
bar_sim.width = 22
bar_sim.height = 14

actual_ref = Reference(ws3, min_col=3, min_row=17, max_row=22)
sim_ref = Reference(ws3, min_col=4, min_row=17, max_row=22)
cats_sim = Reference(ws3, min_col=2, min_row=18, max_row=22)
bar_sim.add_data(actual_ref, titles_from_data=True)
bar_sim.add_data(sim_ref, titles_from_data=True)
bar_sim.set_categories(cats_sim)
bar_sim.series[0].graphicalProperties.solidFill = GRAY_500
bar_sim.series[1].graphicalProperties.solidFill = INFO_BLUE
ws3.add_chart(bar_sim, "B27")


# ============================================================
# SHEET 4: RECOMENDACIONES
# ============================================================
ws4 = wb.create_sheet("Recomendaciones")
ws4.sheet_properties.tabColor = SUCCESS

ws4.column_dimensions['A'].width = 2
ws4.column_dimensions['B'].width = 20
ws4.column_dimensions['C'].width = 50
ws4.column_dimensions['D'].width = 16
ws4.column_dimensions['E'].width = 2

ws4.freeze_panes = 'B5'

# Header
ws4.merge_cells('B1:D1')
fill_row(ws4, 1, range(2, 5), DARK_BG)
rh(ws4, 1, 8)

ws4.merge_cells('B2:D2')
h = ws4['B2']
h.value = "RECOMENDACIONES PERSONALIZADAS"
h.font = Font(name='Calibri', bold=True, color=GOLD, size=18)
h.fill = PatternFill('solid', fgColor=DARK_BG)
h.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws4, 2, range(2, 5), DARK_BG)
rh(ws4, 2, 38)

ws4.merge_cells('B3:D3')
sub = ws4['B3']
sub.value = "Acciones concretas basadas en tu perfil actual"
sub.font = Font(name='Calibri', color=GRAY_300, size=10, italic=True)
sub.fill = PatternFill('solid', fgColor=DARK_BG)
sub.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws4, 3, range(2, 5), DARK_BG)
rh(ws4, 3, 24)

rh(ws4, 4, 8)

# Dynamic recommendations per factor
factor_recs = [
    ("Historial de Pagos", F_HISTORY, F_HISTORY_LIGHT, 13, [
        ("Configura pagos automaticos para TODAS tus cuentas", "Si tienes atrasos, paga al menos el minimo HOY"),
        ("Negocia eliminacion de atrasos menores con el banco", "Establece alertas de pago 5 dias antes del vencimiento"),
        ("Un solo pago atrasado puede bajar 50+ puntos", "Si no puedes pagar todo, paga el minimo para evitar reporte"),
    ]),
    ("Uso de Credito", F_UTIL, F_UTIL_LIGHT, 14, [
        ("Mantene tu uso por debajo del 30% del limite", "Si usas mas del 50%, paga antes de la fecha de corte"),
        ("Pide aumento de limite (sin usarlo) para bajar el %", "Distribuye gastos en varias tarjetas en vez de una"),
        ("El punto optimo es usar menos del 10% del limite", "Paga 2 veces al mes para mantener el saldo bajo"),
    ]),
    ("Antiguedad", F_AGE, F_AGE_LIGHT, 15, [
        ("NO cierres tu tarjeta mas antigua, mantela activa", "Usa tu tarjeta vieja al menos 1 vez al mes"),
        ("La antiguedad promedio ideal es 5+ anios", "Abre cuentas nuevas solo cuando sea necesario"),
        ("Cada cuenta nueva baja tu promedio de antiguedad", "Si necesitas una nueva tarjeta, conserva las viejas"),
    ]),
    ("Tipos de Credito", F_MIX, F_MIX_LIGHT, 16, [
        ("Ideal: 1 tarjeta + 1 prestamo + 1 linea de credito", "No abras creditos solo por diversificar (ten un proposito)"),
        ("Los creditos hipotecarios suman mucho al mix", "Un prestamo personal pequeno puede ayudar al mix"),
        ("Las tiendas con credito propio tambien cuentan", "La variedad demuestra que manejas diferentes obligaciones"),
    ]),
    ("Consultas Recientes", F_INQUIRY, F_INQUIRY_LIGHT, 17, [
        ("Evita solicitar multiples creditos en poco tiempo", "Cada consulta baja tu puntaje 5-10 puntos temporalmente"),
        ("Agrupa solicitudes del mismo tipo en 2 semanas", "Las consultas se borran del reporte despues de 2 anios"),
        ("Consultar tu PROPIO reporte no afecta tu puntaje", "Evita pre-aprobaciones que generan consultas duras"),
    ]),
]

row_offset = 5
for factor_name, color, bg, dash_row, recs in factor_recs:
    # Section header
    ws4.merge_cells(f'B{row_offset}:C{row_offset}')
    sec = ws4[f'B{row_offset}']
    sec.value = f'{factor_name} - Tu puntaje:'
    style_section(sec, color)
    ws4.cell(row=row_offset, column=3).fill = PatternFill('solid', fgColor=color)
    rh(ws4, row_offset, 28)

    # Dynamic score display
    score_d = ws4.cell(row=row_offset, column=4)
    score_d.value = f'=Dashboard!C{dash_row}'
    score_d.font = Font(name='Calibri', bold=True, color=WHITE, size=14)
    score_d.fill = PatternFill('solid', fgColor=color)
    score_d.alignment = Alignment(horizontal='center')

    row_offset += 1

    # Dynamic recommendation based on score
    for j, (good_tip, bad_tip) in enumerate(recs):
        rh(ws4, row_offset, 30)
        ws4.cell(row=row_offset, column=2, value=f"Accion {j+1}").font = Font(name='Calibri', bold=True, size=10, color=color)
        ws4.cell(row=row_offset, column=2).fill = PatternFill('solid', fgColor=bg)
        ws4.cell(row=row_offset, column=2).border = thin_border
        ws4.cell(row=row_offset, column=2).alignment = Alignment(horizontal='center')

        ws4.merge_cells(f'C{row_offset}:D{row_offset}')
        rec = ws4.cell(row=row_offset, column=3)
        rec.value = f'=IF(Dashboard!C{dash_row}>=70,"{good_tip}","{bad_tip}")'
        rec.font = Font(name='Calibri', size=10, color=GRAY_700)
        rec.fill = PatternFill('solid', fgColor=bg)
        rec.border = thin_border
        rec.alignment = Alignment(wrap_text=True, vertical='center')

        row_offset += 1

    rh(ws4, row_offset, 6)
    row_offset += 1


# ============================================================
# SHEET 5: CONFIG
# ============================================================
ws5 = wb.create_sheet("Config")
ws5.sheet_properties.tabColor = GRAY_500

ws5.column_dimensions['A'].width = 25
ws5.column_dimensions['B'].width = 55

ws5.merge_cells('A1:B1')
h = ws5['A1']
h.value = "CONFIGURACION"
h.font = Font(name='Calibri', bold=True, color=GOLD, size=16)
h.fill = PatternFill('solid', fgColor=DARK_BG)
h.alignment = Alignment(horizontal='center', vertical='center')
ws5.cell(row=1, column=2).fill = PatternFill('solid', fgColor=DARK_BG)
rh(ws5, 1, 36)

config_data = [
    ("Producto", "Analizador de Puntaje Crediticio"),
    ("Version", "v1.0.0"),
    ("Marca", "No Somos Ignorantes"),
    ("Precio", "Bs. 39"),
    ("Modelo", "Aproximacion buros bolivianos (educativo)"),
    ("Rango puntaje", "300 - 850"),
    ("Factores", "5 (Historial, Uso, Antiguedad, Mix, Consultas)"),
    ("Proteccion", "nsi2024"),
    ("Contacto", "nosomosignorantes@gmail.com"),
]

for i, (key, val) in enumerate(config_data):
    row = 3 + i
    k = ws5.cell(row=row, column=1, value=key)
    k.font = Font(name='Calibri', bold=True, size=11, color=GRAY_700)
    k.border = thin_border
    v = ws5.cell(row=row, column=2, value=val)
    v.font = Font(name='Calibri', size=11, color=GRAY_900)
    v.border = thin_border

r = 14
ws5.merge_cells(f'A{r}:B{r}')
inst = ws5[f'A{r}']
inst.value = "INSTRUCCIONES DE USO"
inst.font = Font(name='Calibri', bold=True, color=GOLD, size=14)
inst.fill = PatternFill('solid', fgColor=DARK_BG)
inst.alignment = Alignment(horizontal='center')
ws5.cell(row=r, column=2).fill = PatternFill('solid', fgColor=DARK_BG)
rh(ws5, r, 32)

instructions = [
    "1. Ve al Dashboard e ingresa tu puntuacion estimada (0-100) para cada factor.",
    "2. Usa la hoja 'Historial' para calcular tu puntaje exacto de pagos y uso de credito.",
    "3. Copia los puntajes calculados al Dashboard para mayor precision.",
    "4. Ve al 'Simulador' para explorar que pasaria si mejoras ciertos factores.",
    "5. Revisa 'Recomendaciones' para acciones concretas basadas en tu perfil.",
    "",
    "COMO ESTIMAR CADA FACTOR (0-100):",
    "- Historial: 100 = nunca atrasado, -10 por cada atraso <30d, -20 por >30d",
    "- Uso credito: 100 = <10% usado, 90 = <30%, 70 = <50%, 50 = <75%, 30 = >75%",
    "- Antiguedad: 100 = >10 anios promedio, 80 = 5-10, 60 = 3-5, 40 = 1-3, 20 = <1",
    "- Mix: 100 = 4+ tipos (hipoteca, auto, tarjeta, personal), 50 = solo tarjetas",
    "- Consultas: 100 = 0 en 6 meses, 80 = 1-2, 60 = 3-4, 40 = 5+",
    "",
    "NOTA IMPORTANTE:",
    "Este es un modelo educativo basado en la estructura de buros de credito.",
    "NO es tu puntaje oficial. Para tu puntaje real, consulta tu buro de credito.",
]

for i, line in enumerate(instructions):
    row = 15 + i
    ws5.merge_cells(f'A{row}:B{row}')
    c = ws5[f'A{row}']
    c.value = line
    c.font = Font(name='Calibri', size=10, color=GRAY_700)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    rh(ws5, row, 20)


# ============================================================
# SHEET PROTECTION
# ============================================================
ws.protection.sheet = True
ws.protection.password = "nsi2024"
ws.protection.enable()
# Unlock factor score inputs
for row in range(13, 18):
    ws.cell(row=row, column=3).protection = ws.cell(row=row, column=3).protection.copy(locked=False)

ws2.protection.sheet = True
ws2.protection.password = "nsi2024"
ws2.protection.enable()
# Unlock history inputs
for row in range(7, 14):
    ws2.cell(row=row, column=3).protection = ws2.cell(row=row, column=3).protection.copy(locked=False)
# Unlock card inputs
for i in range(NUM_CARDS):
    row = 20 + i
    for col in [2, 3, 4]:
        ws2.cell(row=row, column=col).protection = ws2.cell(row=row, column=col).protection.copy(locked=False)

ws3.protection.sheet = True
ws3.protection.password = "nsi2024"
ws3.protection.enable()
# Unlock simulator inputs
for row in range(18, 23):
    ws3.cell(row=row, column=4).protection = ws3.cell(row=row, column=4).protection.copy(locked=False)

for ws_name in ["Recomendaciones", "Config"]:
    sheet = wb[ws_name]
    sheet.protection.sheet = True
    sheet.protection.password = "nsi2024"
    sheet.protection.enable()


# ============================================================
# SAVE
# ============================================================
OUTPUT = "D:/Landing-Page_marketplace/excel_products/Analizador_Puntaje_Crediticio_NSI.xlsx"
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print("Sheets:", wb.sheetnames)
print("Done!")
