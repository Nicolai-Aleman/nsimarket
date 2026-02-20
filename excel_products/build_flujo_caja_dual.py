"""
Build: Flujo de Caja Dual Bs/USD (Bs. 99)
No Somos Ignorantes v1.0
Dual currency cashflow: Dashboard, BOB flow, USD flow, exchange rates,
projections. Anchor product for entrepreneurs.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

GOLD = "D4AF37"; DARK_BG = "1A1A2E"; ACCENT_BLUE = "0F3460"
WHITE = "FFFFFF"; BLACK = "000000"
LIGHT_YELLOW = "FFF9C4"; LIGHT_GREEN = "E8F5E9"
RED_CORAL = "E74C3C"; TURQUOISE = "16A085"; ORANGE = "E67E22"; PURPLE = "8E44AD"
LIGHT_GRAY = "BDC3C7"; GREEN_OK = "27AE60"; RED_ALERT = "E74C3C"
GRAY_300 = "D1D5DB"; GRAY_400 = "9CA3AF"; GRAY_500 = "6B7280"
GRAY_600 = "4B5563"; GRAY_700 = "374151"; GRAY_900 = "111827"

thin_border = Border(
    left=Side(style='thin', color=GRAY_300), right=Side(style='thin', color=GRAY_300),
    top=Side(style='thin', color=GRAY_300), bottom=Side(style='thin', color=GRAY_300))
gold_border = Border(
    left=Side(style='thin', color=GOLD), right=Side(style='thin', color=GOLD),
    top=Side(style='thin', color=GOLD), bottom=Side(style='thin', color=GOLD))

def style_title(cell, size=18):
    cell.font = Font(name='Calibri', bold=True, color=BLACK, size=size)
    cell.alignment = Alignment(horizontal='left', vertical='center')
def style_header(cell, bg=DARK_BG, fg=WHITE, size=10):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
def style_input(cell):
    cell.font = Font(name='Calibri', size=11, color="0000FF")
    cell.fill = PatternFill('solid', fgColor=LIGHT_YELLOW)
    cell.border = gold_border
    cell.alignment = Alignment(horizontal='center', vertical='center')
def style_output(cell, bold=False):
    cell.font = Font(name='Calibri', size=11, color=GRAY_900, bold=bold)
    cell.fill = PatternFill('solid', fgColor=LIGHT_GREEN)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='right', vertical='center')
def style_banner(cell, bg_color, fg=WHITE, size=10):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg_color)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
def fill_white(ws, r1, r2, c1, c2):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=WHITE)
def unlock_cell(cell):
    cell.protection = cell.protection.copy(locked=False)

MONTHS_ES = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
MAX_ENTRIES = 200

# ============================================================
# SHEET 1: DASHBOARD
# ============================================================
ws = wb.active
ws.title = "Dashboard"
ws.sheet_properties.tabColor = RED_CORAL
ws.sheet_view.showGridLines = False

dw = {1:3, 2:22, 3:16, 4:16, 5:3, 6:22, 7:16, 8:16, 9:3}
for c, w in dw.items():
    ws.column_dimensions[get_column_letter(c)].width = w
fill_white(ws, 1, 55, 1, 9)

ws.merge_cells('B2:H2')
style_title(ws['B2'], 20)
ws['B2'].value = "FLUJO DE CAJA DUAL (Bs / USD)"
ws.merge_cells('B3:H3')
ws['B3'].font = Font(name='Calibri', size=10, color=GRAY_500)
ws['B3'].value = "No Somos Ignorantes  |  v1.0  |  Control de efectivo en dos monedas"

# Exchange rate
ws['B4'].value = "Tipo de Cambio (Bs/USD):"
ws['B4'].font = Font(name='Calibri', size=10, bold=True, color=GRAY_700)
ws['D4'].value = 6.96
style_input(ws['D4'])
ws['D4'].number_format = '#,##0.00'
unlock_cell(ws['D4'])

ws.row_dimensions[5].height = 6

# KPI 1: Saldo BOB
ws.merge_cells('B6:D6')
style_banner(ws['B6'], RED_CORAL, WHITE, 10)
ws['B6'].value = "SALDO EN BOLIVIANOS (Bs.)"
ws.merge_cells('B7:D7')
c = ws['B7']
c.font = Font(name='Calibri', bold=True, color=BLACK, size=24)
c.alignment = Alignment(horizontal='right', vertical='center')
c.value = ('=IFERROR(SUMPRODUCT((Flujo_BOB!D4:D203="Ingreso")*Flujo_BOB!E4:E203)'
           '-SUMPRODUCT((Flujo_BOB!D4:D203="Egreso")*Flujo_BOB!E4:E203),0)')
c.number_format = '#,##0.00'

# KPI 2: Saldo USD
ws.merge_cells('F6:H6')
style_banner(ws['F6'], TURQUOISE, WHITE, 10)
ws['F6'].value = "SALDO EN DOLARES (USD)"
ws.merge_cells('F7:H7')
c = ws['F7']
c.font = Font(name='Calibri', bold=True, color=BLACK, size=24)
c.alignment = Alignment(horizontal='right', vertical='center')
c.value = ('=IFERROR(SUMPRODUCT((Flujo_USD!D4:D203="Ingreso")*Flujo_USD!E4:E203)'
           '-SUMPRODUCT((Flujo_USD!D4:D203="Egreso")*Flujo_USD!E4:E203),0)')
c.number_format = '#,##0.00'

ws.row_dimensions[8].height = 4

# KPI 3: Saldo Combinado en Bs
ws.merge_cells('B9:D9')
style_banner(ws['B9'], ORANGE, WHITE, 10)
ws['B9'].value = "SALDO TOTAL COMBINADO (Bs.)"
ws.merge_cells('B10:D10')
c = ws['B10']
c.font = Font(name='Calibri', bold=True, color=BLACK, size=24)
c.alignment = Alignment(horizontal='right', vertical='center')
c.value = '=B7+(F7*D4)'
c.number_format = '#,##0.00'

# KPI 4: Flujo Neto del Mes
ws.merge_cells('F9:H9')
style_banner(ws['F9'], PURPLE, WHITE, 10)
ws['F9'].value = "FLUJO NETO DEL MES (Bs.)"
ws.merge_cells('F10:H10')
c = ws['F10']
c.font = Font(name='Calibri', bold=True, color=BLACK, size=24)
c.alignment = Alignment(horizontal='right', vertical='center')
# Net flow this month = BOB inflows - BOB outflows + (USD inflows - USD outflows) * TC
c.value = ('=IFERROR('
           'SUMPRODUCT((MONTH(Flujo_BOB!A4:A203)=MONTH(TODAY()))*(YEAR(Flujo_BOB!A4:A203)=YEAR(TODAY()))*(Flujo_BOB!D4:D203="Ingreso")*Flujo_BOB!E4:E203)'
           '-SUMPRODUCT((MONTH(Flujo_BOB!A4:A203)=MONTH(TODAY()))*(YEAR(Flujo_BOB!A4:A203)=YEAR(TODAY()))*(Flujo_BOB!D4:D203="Egreso")*Flujo_BOB!E4:E203)'
           '+('
           'SUMPRODUCT((MONTH(Flujo_USD!A4:A203)=MONTH(TODAY()))*(YEAR(Flujo_USD!A4:A203)=YEAR(TODAY()))*(Flujo_USD!D4:D203="Ingreso")*Flujo_USD!E4:E203)'
           '-SUMPRODUCT((MONTH(Flujo_USD!A4:A203)=MONTH(TODAY()))*(YEAR(Flujo_USD!A4:A203)=YEAR(TODAY()))*(Flujo_USD!D4:D203="Egreso")*Flujo_USD!E4:E203)'
           ')*D4,0)')
c.number_format = '#,##0.00'

ws.row_dimensions[11].height = 8

# Monthly trend table
ws['B12'].value = "FLUJO MENSUAL COMBINADO (Bs.)"
ws['B12'].font = Font(name='Calibri', bold=True, size=11, color=GRAY_700)

ws['B13'].value = "Mes"
ws['C13'].value = "Ingresos Bs."
ws['D13'].value = "Egresos Bs."
style_header(ws['B13'], bg=GRAY_700)
style_header(ws['C13'], bg=GRAY_700)
style_header(ws['D13'], bg=GRAY_700)

for i, m in enumerate(MONTHS_ES):
    r = 14 + i
    ws.cell(row=r, column=2).value = m
    ws.cell(row=r, column=2).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws.cell(row=r, column=2).border = thin_border

    # Ingresos BOB + USD converted
    ws.cell(row=r, column=3).value = (
        f'=IFERROR(SUMPRODUCT((MONTH(Flujo_BOB!A4:A203)={i+1})*(YEAR(Flujo_BOB!A4:A203)=YEAR(TODAY()))*(Flujo_BOB!D4:D203="Ingreso")*Flujo_BOB!E4:E203)'
        f'+SUMPRODUCT((MONTH(Flujo_USD!A4:A203)={i+1})*(YEAR(Flujo_USD!A4:A203)=YEAR(TODAY()))*(Flujo_USD!D4:D203="Ingreso")*Flujo_USD!E4:E203)*D4,0)')
    ws.cell(row=r, column=3).number_format = '#,##0'
    ws.cell(row=r, column=3).border = thin_border

    # Egresos BOB + USD converted
    ws.cell(row=r, column=4).value = (
        f'=IFERROR(SUMPRODUCT((MONTH(Flujo_BOB!A4:A203)={i+1})*(YEAR(Flujo_BOB!A4:A203)=YEAR(TODAY()))*(Flujo_BOB!D4:D203="Egreso")*Flujo_BOB!E4:E203)'
        f'+SUMPRODUCT((MONTH(Flujo_USD!A4:A203)={i+1})*(YEAR(Flujo_USD!A4:A203)=YEAR(TODAY()))*(Flujo_USD!D4:D203="Egreso")*Flujo_USD!E4:E203)*D4,0)')
    ws.cell(row=r, column=4).number_format = '#,##0'
    ws.cell(row=r, column=4).border = thin_border

# Chart: Monthly cashflow
chart1 = BarChart()
chart1.type = "col"
chart1.grouping = "clustered"
chart1.style = 10
chart1.title = "Flujo de Caja Mensual Combinado"
chart1.y_axis.title = "Bs."
chart1.legend.position = 'b'
d1 = Reference(ws, min_col=3, min_row=13, max_row=25)
d2 = Reference(ws, min_col=4, min_row=13, max_row=25)
cats = Reference(ws, min_col=2, min_row=14, max_row=25)
chart1.add_data(d1, titles_from_data=True)
chart1.add_data(d2, titles_from_data=True)
chart1.set_categories(cats)
chart1.series[0].graphicalProperties.solidFill = GREEN_OK
chart1.series[0].graphicalProperties.line.noFill = True
chart1.series[1].graphicalProperties.solidFill = RED_CORAL
chart1.series[1].graphicalProperties.line.noFill = True
chart1.width = 22
chart1.height = 13
ws.add_chart(chart1, "B27")

# Saldo acumulado line chart
ws['F12'].value = "SALDO ACUMULADO"
ws['F12'].font = Font(name='Calibri', bold=True, size=11, color=GRAY_700)
ws['F13'].value = "Mes"
ws['G13'].value = "Saldo Bs."
ws['H13'].value = "Saldo USD"
style_header(ws['F13'], bg=GRAY_700)
style_header(ws['G13'], bg=GRAY_700)
style_header(ws['H13'], bg=GRAY_700)

for i, m in enumerate(MONTHS_ES):
    r = 14 + i
    ws.cell(row=r, column=6).value = m
    ws.cell(row=r, column=6).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws.cell(row=r, column=6).border = thin_border

    # Cumulative BOB balance
    months_formula = "+".join([
        f'SUMPRODUCT((MONTH(Flujo_BOB!A4:A203)={j+1})*(YEAR(Flujo_BOB!A4:A203)=YEAR(TODAY()))'
        f'*IF(Flujo_BOB!D4:D203="Ingreso",1,-1)*Flujo_BOB!E4:E203)' for j in range(i+1)])
    ws.cell(row=r, column=7).value = f'=IFERROR({months_formula},0)'
    ws.cell(row=r, column=7).number_format = '#,##0'
    ws.cell(row=r, column=7).border = thin_border

    months_formula_usd = "+".join([
        f'SUMPRODUCT((MONTH(Flujo_USD!A4:A203)={j+1})*(YEAR(Flujo_USD!A4:A203)=YEAR(TODAY()))'
        f'*IF(Flujo_USD!D4:D203="Ingreso",1,-1)*Flujo_USD!E4:E203)' for j in range(i+1)])
    ws.cell(row=r, column=8).value = f'=IFERROR({months_formula_usd},0)'
    ws.cell(row=r, column=8).number_format = '#,##0'
    ws.cell(row=r, column=8).border = thin_border

chart2 = LineChart()
chart2.title = "Saldo Acumulado por Moneda"
chart2.style = 10
chart2.y_axis.title = "Monto"
chart2.legend.position = 'b'
chart2.width = 18
chart2.height = 13
d3 = Reference(ws, min_col=7, min_row=13, max_row=25)
d4 = Reference(ws, min_col=8, min_row=13, max_row=25)
cats2 = Reference(ws, min_col=6, min_row=14, max_row=25)
chart2.add_data(d3, titles_from_data=True)
chart2.add_data(d4, titles_from_data=True)
chart2.set_categories(cats2)
chart2.series[0].graphicalProperties.line.solidFill = TURQUOISE
chart2.series[0].graphicalProperties.line.width = 25000
chart2.series[1].graphicalProperties.line.solidFill = ORANGE
chart2.series[1].graphicalProperties.line.width = 25000
ws.add_chart(chart2, "F27")

# ============================================================
# Helper to build flow sheets (BOB and USD)
# ============================================================
def build_flow_sheet(ws_flow, currency, tab_color, symbol):
    ws_flow.sheet_properties.tabColor = tab_color
    ws_flow.sheet_view.showGridLines = False

    fw = {1:14, 2:28, 3:20, 4:12, 5:18, 6:18, 7:28}
    for c, w in fw.items():
        ws_flow.column_dimensions[get_column_letter(c)].width = w

    ws_flow.merge_cells('A1:G1')
    style_title(ws_flow['A1'], 16)
    ws_flow['A1'].value = f"FLUJO DE CAJA EN {currency}"
    ws_flow['A1'].fill = PatternFill('solid', fgColor=WHITE)
    ws_flow.row_dimensions[1].height = 35

    ws_flow.merge_cells('A2:G2')
    ws_flow['A2'].value = f"Registra todos los ingresos y egresos en {currency}."
    ws_flow['A2'].font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)
    ws_flow['A2'].fill = PatternFill('solid', fgColor=WHITE)

    headers = [("A","FECHA"), ("B","CONCEPTO"), ("C","CATEGORIA"),
               ("D","TIPO"), ("E",f"MONTO ({symbol})"), ("F","SALDO ACUMULADO"),
               ("G","OBSERVACION")]
    ws_flow.row_dimensions[3].height = 28
    for col_letter, label in headers:
        style_header(ws_flow[f'{col_letter}3'], bg=DARK_BG, fg=GOLD, size=10)
        ws_flow[f'{col_letter}3'].value = label

    for row in range(4, 4 + MAX_ENTRIES):
        # Date
        c = ws_flow.cell(row=row, column=1)
        style_input(c); c.number_format = 'DD/MM/YYYY'; unlock_cell(c)

        # Concepto
        c = ws_flow.cell(row=row, column=2)
        style_input(c); c.alignment = Alignment(horizontal='left', vertical='center'); unlock_cell(c)

        # Categoria
        c = ws_flow.cell(row=row, column=3)
        style_input(c); unlock_cell(c)

        # Tipo
        c = ws_flow.cell(row=row, column=4)
        style_input(c); unlock_cell(c)

        # Monto
        c = ws_flow.cell(row=row, column=5)
        style_input(c); c.number_format = '#,##0.00'; unlock_cell(c)

        # Saldo acumulado (running balance)
        c = ws_flow.cell(row=row, column=6)
        if row == 4:
            c.value = '=IF(D4="Ingreso",E4,IF(D4="Egreso",-E4,0))'
        else:
            c.value = f'=IF(A{row}="",F{row-1},F{row-1}+IF(D{row}="Ingreso",E{row},IF(D{row}="Egreso",-E{row},0)))'
        c.number_format = '#,##0.00'
        style_output(c)

        # Observacion
        c = ws_flow.cell(row=row, column=7)
        style_input(c); c.alignment = Alignment(horizontal='left', vertical='center'); unlock_cell(c)

    # Tipo dropdown
    dv_tipo = DataValidation(type="list", formula1='"Ingreso,Egreso"', allow_blank=True)
    dv_tipo.prompt = "Ingreso = entrada, Egreso = salida"
    dv_tipo.promptTitle = "Tipo"
    ws_flow.add_data_validation(dv_tipo)
    dv_tipo.add('D4:D203')

    # Category dropdown
    cats = '"Ventas,Servicios,Alquiler,Sueldos,Compras,Impuestos,Prestamos,Otros"'
    dv_cat = DataValidation(type="list", formula1=cats, allow_blank=True)
    dv_cat.prompt = "Categoria del movimiento"
    dv_cat.promptTitle = "Categoria"
    ws_flow.add_data_validation(dv_cat)
    dv_cat.add('C4:C203')

    # Amount >= 0
    dv_pos = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    dv_pos.error = "El monto no puede ser negativo"
    ws_flow.add_data_validation(dv_pos)
    dv_pos.add('E4:E203')

    # Conditional formatting: negative balance = red
    ws_flow.conditional_formatting.add('F4:F203',
        CellIsRule(operator='lessThan', formula=['0'],
                  fill=PatternFill('solid', fgColor="FEE2E2"),
                  font=Font(color=RED_ALERT, bold=True)))

    # Tipo colors
    ws_flow.conditional_formatting.add('D4:D203',
        CellIsRule(operator='equal', formula=['"Ingreso"'],
                  fill=PatternFill('solid', fgColor="D1FAE5"),
                  font=Font(color=GREEN_OK, bold=True)))
    ws_flow.conditional_formatting.add('D4:D203',
        CellIsRule(operator='equal', formula=['"Egreso"'],
                  fill=PatternFill('solid', fgColor="FEE2E2"),
                  font=Font(color=RED_ALERT, bold=True)))

# Build BOB and USD sheets
ws_bob = wb.create_sheet("Flujo_BOB")
build_flow_sheet(ws_bob, "BOLIVIANOS", TURQUOISE, "Bs.")

ws_usd = wb.create_sheet("Flujo_USD")
build_flow_sheet(ws_usd, "DOLARES", ORANGE, "USD")

# Sample data BOB
from datetime import date
sample_bob = [
    (date(2025,1,2), "Ventas enero semana 1", "Ventas", "Ingreso", 15000, ""),
    (date(2025,1,5), "Alquiler local", "Alquiler", "Egreso", 3500, ""),
    (date(2025,1,5), "Sueldos enero", "Sueldos", "Egreso", 8000, "3 empleados"),
    (date(2025,1,10), "Ventas enero semana 2", "Ventas", "Ingreso", 12000, ""),
    (date(2025,1,15), "Servicios basicos", "Servicios", "Egreso", 700, ""),
    (date(2025,1,20), "Ventas enero semana 3", "Ventas", "Ingreso", 18000, ""),
    (date(2025,1,25), "Compra mercaderia", "Compras", "Egreso", 10000, ""),
    (date(2025,2,1), "Ventas febrero semana 1", "Ventas", "Ingreso", 16000, ""),
    (date(2025,2,5), "Alquiler local", "Alquiler", "Egreso", 3500, ""),
]
for i, (dt, concept, cat, tipo, amount, obs) in enumerate(sample_bob):
    r = 4 + i
    ws_bob.cell(row=r, column=1).value = dt
    ws_bob.cell(row=r, column=2).value = concept
    ws_bob.cell(row=r, column=3).value = cat
    ws_bob.cell(row=r, column=4).value = tipo
    ws_bob.cell(row=r, column=5).value = amount
    ws_bob.cell(row=r, column=7).value = obs

# Sample data USD
sample_usd = [
    (date(2025,1,3), "Pago cliente exportacion", "Ventas", "Ingreso", 2000, ""),
    (date(2025,1,10), "Compra insumos importados", "Compras", "Egreso", 800, ""),
    (date(2025,1,20), "Comision plataforma", "Servicios", "Egreso", 150, "Stripe"),
    (date(2025,2,3), "Pago cliente exportacion", "Ventas", "Ingreso", 2500, ""),
]
for i, (dt, concept, cat, tipo, amount, obs) in enumerate(sample_usd):
    r = 4 + i
    ws_usd.cell(row=r, column=1).value = dt
    ws_usd.cell(row=r, column=2).value = concept
    ws_usd.cell(row=r, column=3).value = cat
    ws_usd.cell(row=r, column=4).value = tipo
    ws_usd.cell(row=r, column=5).value = amount
    ws_usd.cell(row=r, column=7).value = obs

# ============================================================
# SHEET 4: TIPO CAMBIO (Exchange Rate History)
# ============================================================
ws_tc = wb.create_sheet("TipoCambio")
ws_tc.sheet_properties.tabColor = PURPLE
ws_tc.sheet_view.showGridLines = False

tc_w = {1:14, 2:14, 3:14, 4:14, 5:14}
for c, w in tc_w.items():
    ws_tc.column_dimensions[get_column_letter(c)].width = w
fill_white(ws_tc, 1, 40, 1, 5)

ws_tc.merge_cells('A1:E1')
style_title(ws_tc['A1'], 16)
ws_tc['A1'].value = "HISTORIAL TIPO DE CAMBIO"
ws_tc.row_dimensions[1].height = 35

ws_tc.merge_cells('A2:E2')
ws_tc['A2'].value = "Registra el tipo de cambio mensual Bs/USD para referencia."
ws_tc['A2'].font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)

headers_tc = [("A","MES"), ("B","COMPRA"), ("C","VENTA"), ("D","PROMEDIO"), ("E","VARIACION")]
ws_tc.row_dimensions[3].height = 28
for col_letter, label in headers_tc:
    style_header(ws_tc[f'{col_letter}3'], bg=DARK_BG, fg=GOLD, size=10)
    ws_tc[f'{col_letter}3'].value = label

for i, m in enumerate(MONTHS_ES):
    r = 4 + i
    ws_tc.cell(row=r, column=1).value = m
    ws_tc.cell(row=r, column=1).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws_tc.cell(row=r, column=1).border = thin_border

    c = ws_tc.cell(row=r, column=2)
    style_input(c); c.number_format = '#,##0.00'; unlock_cell(c)

    c = ws_tc.cell(row=r, column=3)
    style_input(c); c.number_format = '#,##0.00'; unlock_cell(c)

    c = ws_tc.cell(row=r, column=4)
    c.value = f'=IFERROR((B{r}+C{r})/2,0)'
    c.number_format = '#,##0.00'
    style_output(c)

    c = ws_tc.cell(row=r, column=5)
    if i == 0:
        c.value = '=0'
    else:
        c.value = f'=IFERROR((D{r}-D{r-1})/D{r-1},0)'
    c.number_format = '+0.00%;-0.00%'
    style_output(c)

# Sample exchange rates
for i in range(12):
    ws_tc.cell(row=4+i, column=2).value = 6.86
    ws_tc.cell(row=4+i, column=3).value = 6.96

# ============================================================
# SHEET 5: PROYECCION
# ============================================================
ws_proj = wb.create_sheet("Proyeccion")
ws_proj.sheet_properties.tabColor = "3498DB"
ws_proj.sheet_view.showGridLines = False

pj_w = {1:4, 2:22, 3:16, 4:16, 5:16}
for c, w in pj_w.items():
    ws_proj.column_dimensions[get_column_letter(c)].width = w
fill_white(ws_proj, 1, 30, 1, 5)

ws_proj.merge_cells('B1:E1')
style_title(ws_proj['B1'], 16)
ws_proj['B1'].value = "PROYECCION DE FLUJO DE CAJA"
ws_proj.row_dimensions[1].height = 35

ws_proj.merge_cells('B2:E2')
ws_proj['B2'].value = "Ingresa tus estimaciones de ingresos y egresos para los proximos meses."
ws_proj['B2'].font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)

headers_proj = [("B","MES"), ("C","INGRESOS PROY. (Bs.)"), ("D","EGRESOS PROY. (Bs.)"), ("E","SALDO PROY.")]
ws_proj.row_dimensions[4].height = 28
for col_letter, label in headers_proj:
    style_header(ws_proj[f'{col_letter}4'], bg=DARK_BG, fg=GOLD, size=10)
    ws_proj[f'{col_letter}4'].value = label

for i, m in enumerate(MONTHS_ES):
    r = 5 + i
    ws_proj.cell(row=r, column=2).value = m
    ws_proj.cell(row=r, column=2).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws_proj.cell(row=r, column=2).border = thin_border

    c = ws_proj.cell(row=r, column=3)
    style_input(c); c.number_format = '#,##0.00'; unlock_cell(c)

    c = ws_proj.cell(row=r, column=4)
    style_input(c); c.number_format = '#,##0.00'; unlock_cell(c)

    c = ws_proj.cell(row=r, column=5)
    if i == 0:
        c.value = f'=C{r}-D{r}'
    else:
        c.value = f'=E{r-1}+C{r}-D{r}'
    c.number_format = '#,##0.00'
    style_output(c, bold=True)

# Conditional formatting: negative projected balance
ws_proj.conditional_formatting.add('E5:E16',
    CellIsRule(operator='lessThan', formula=['0'],
              fill=PatternFill('solid', fgColor="FEE2E2"),
              font=Font(color=RED_ALERT, bold=True)))

# ============================================================
# SHEET 6: CONFIG
# ============================================================
ws_conf = wb.create_sheet("Config")
ws_conf.sheet_properties.tabColor = GRAY_500
ws_conf.sheet_view.showGridLines = False
cw = {1:25, 2:50}
for c, w in cw.items():
    ws_conf.column_dimensions[get_column_letter(c)].width = w
fill_white(ws_conf, 1, 25, 1, 3)
ws_conf['A1'].value = "v1.0.0"
ws_conf['A1'].font = Font(name='Calibri', size=9, color=GRAY_400)
ws_conf.merge_cells('A3:B3')
style_title(ws_conf['A3'], 14)
ws_conf['A3'].value = "CONFIGURACION"

settings = [
    ("Producto", "Flujo de Caja Dual Bs/USD"),
    ("Version", "v1.0.0"),
    ("Marca", "No Somos Ignorantes"),
    ("Precio", "Bs. 99"),
    ("Proteccion", "nsi2024"),
    ("", ""),
    ("INSTRUCCIONES", ""),
    ("1.", "Ingresa el tipo de cambio actual en el Dashboard (celda D4)."),
    ("2.", "En 'Flujo_BOB', registra ingresos/egresos en Bolivianos."),
    ("3.", "En 'Flujo_USD', registra ingresos/egresos en Dolares."),
    ("4.", "El Dashboard combina ambas monedas automaticamente."),
    ("5.", "En 'TipoCambio', registra el TC mensual de referencia."),
    ("6.", "En 'Proyeccion', estima flujos futuros."),
    ("7.", "Para desbloquear: Revisar -> Desproteger hoja -> nsi2024"),
]
for i, (label, value) in enumerate(settings):
    r = 5 + i
    ws_conf.cell(row=r, column=1).value = label
    ws_conf.cell(row=r, column=1).font = Font(name='Calibri', size=10, bold=True, color=GRAY_700)
    ws_conf.cell(row=r, column=2).value = value
    ws_conf.cell(row=r, column=2).font = Font(name='Calibri', size=10, color=GRAY_600)

# ============================================================
# PROTECTION
# ============================================================
for sheet in wb.worksheets:
    sheet.protection.sheet = True
    sheet.protection.password = "nsi2024"
    sheet.protection.enable()

# Unlock Dashboard inputs
unlock_cell(ws['D4'])

# BOB and USD flows already unlocked in build_flow_sheet
# Re-unlock to be safe
for ws_flow in [ws_bob, ws_usd]:
    for row in range(4, 4+MAX_ENTRIES):
        for col in [1, 2, 3, 4, 5, 7]:
            unlock_cell(ws_flow.cell(row=row, column=col))

# TipoCambio
for i in range(12):
    unlock_cell(ws_tc.cell(row=4+i, column=2))
    unlock_cell(ws_tc.cell(row=4+i, column=3))

# Proyeccion
for i in range(12):
    unlock_cell(ws_proj.cell(row=5+i, column=3))
    unlock_cell(ws_proj.cell(row=5+i, column=4))

# ============================================================
# SAVE & VERIFY
# ============================================================
out = r"D:\Landing-Page_marketplace\excel_products\Flujo_Caja_Dual_NSI.xlsx"
wb.save(out)
print(f"[OK] Saved: {out}")

from openpyxl import load_workbook
wb2 = load_workbook(out)
fc = sum(1 for s in wb2.worksheets for row in s.iter_rows() for cell in row
         if isinstance(cell.value, str) and cell.value.startswith('='))
print(f"[OK] Sheets: {wb2.sheetnames}")
print(f"[OK] Total formulas: {fc}")
print(f"[OK] Charts: Dashboard={len(ws._charts)}")
