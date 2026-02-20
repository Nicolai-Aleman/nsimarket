"""
Build: Estados Financieros Completos (Bs. 199)
No Somos Ignorantes v1.0
PREMIUM: Full financial statements — Balance Sheet, Income Statement,
Cash Flow Statement, Chart of Accounts, Journal Entries, Financial Ratios.
Executive dashboard design.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule, DataBarRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

GOLD = "D4AF37"; DARK_BG = "1A1A2E"; ACCENT_BLUE = "0F3460"
WHITE = "FFFFFF"; BLACK = "000000"
LIGHT_YELLOW = "FFF9C4"; LIGHT_GREEN = "E8F5E9"
RED_CORAL = "E74C3C"; TURQUOISE = "16A085"; ORANGE = "E67E22"; PURPLE = "8E44AD"
LIGHT_GRAY = "BDC3C7"; GREEN_OK = "27AE60"; RED_ALERT = "E74C3C"
GRAY_300 = "D1D5DB"; GRAY_400 = "9CA3AF"; GRAY_500 = "6B7280"
GRAY_600 = "4B5563"; GRAY_700 = "374151"; GRAY_800 = "1F2937"; GRAY_900 = "111827"

thin_border = Border(
    left=Side(style='thin', color=GRAY_300), right=Side(style='thin', color=GRAY_300),
    top=Side(style='thin', color=GRAY_300), bottom=Side(style='thin', color=GRAY_300))
gold_border = Border(
    left=Side(style='thin', color=GOLD), right=Side(style='thin', color=GOLD),
    top=Side(style='thin', color=GOLD), bottom=Side(style='thin', color=GOLD))
bottom_double = Border(bottom=Side(style='double', color=BLACK))

def style_title(cell, size=18):
    cell.font = Font(name='Calibri', bold=True, color=BLACK, size=size)
    cell.alignment = Alignment(horizontal='left', vertical='center')
def style_header(cell, bg=DARK_BG, fg=WHITE, size=10):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
def style_input(cell):
    cell.font = Font(name='Calibri', size=11, color="0000FF")
    cell.fill = PatternFill('solid', fgColor=LIGHT_YELLOW)
    cell.border = gold_border
    cell.alignment = Alignment(horizontal='right', vertical='center')
def style_output(cell, bold=False):
    cell.font = Font(name='Calibri', size=11, color=GRAY_900, bold=bold)
    cell.fill = PatternFill('solid', fgColor=LIGHT_GREEN)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='right', vertical='center')
def style_banner(cell, bg_color, fg=WHITE, size=10):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg_color)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
def style_label(cell, bold=False, indent=0):
    cell.font = Font(name='Calibri', size=11, color=GRAY_700, bold=bold)
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=indent)
    cell.border = thin_border
def style_subtotal(cell):
    cell.font = Font(name='Calibri', size=11, color=BLACK, bold=True)
    cell.border = Border(top=Side(style='thin', color=BLACK), bottom=Side(style='thin', color=BLACK))
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.number_format = '#,##0.00'
def style_total(cell):
    cell.font = Font(name='Calibri', size=12, color=BLACK, bold=True)
    cell.border = Border(top=Side(style='thin', color=BLACK), bottom=Side(style='double', color=BLACK))
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.number_format = '#,##0.00'
def fill_white(ws, r1, r2, c1, c2):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=WHITE)
def unlock_cell(cell):
    cell.protection = cell.protection.copy(locked=False)

# ============================================================
# SHEET 1: DASHBOARD
# ============================================================
ws = wb.active
ws.title = "Dashboard"
ws.sheet_properties.tabColor = RED_CORAL
ws.sheet_view.showGridLines = False

dw = {1:3, 2:22, 3:16, 4:16, 5:3, 6:22, 7:16, 8:16, 9:3}
for c, w in dw.items():
    ws.column_dimensions[get_column_letter(c)].width = w
fill_white(ws, 1, 50, 1, 9)

ws.merge_cells('B2:H2')
style_title(ws['B2'], 20)
ws['B2'].value = "ESTADOS FINANCIEROS - PANEL EJECUTIVO"
ws.merge_cells('B3:H3')
ws['B3'].font = Font(name='Calibri', size=10, color=GRAY_500)
ws['B3'].value = "No Somos Ignorantes  |  v1.0  |  Balance + Resultados + Flujo de Caja"

ws['B4'].value = "Periodo:"
ws['B4'].font = Font(name='Calibri', size=10, bold=True, color=GRAY_700)
ws['C4'].value = "Enero - Diciembre 2025"
style_input(ws['C4'])
ws['C4'].alignment = Alignment(horizontal='left', vertical='center')
unlock_cell(ws['C4'])

ws.row_dimensions[5].height = 6

# KPI 1: Total Activos
ws.merge_cells('B6:D6')
style_banner(ws['B6'], RED_CORAL, WHITE, 10)
ws['B6'].value = "TOTAL ACTIVOS (Bs.)"
ws.merge_cells('B7:D7')
c = ws['B7']
c.font = Font(name='Calibri', bold=True, color=BLACK, size=24)
c.alignment = Alignment(horizontal='right', vertical='center')
c.value = '=BalanceGeneral!C20'
c.number_format = '#,##0.00'

# KPI 2: Utilidad Neta
ws.merge_cells('F6:H6')
style_banner(ws['F6'], TURQUOISE, WHITE, 10)
ws['F6'].value = "UTILIDAD NETA (Bs.)"
ws.merge_cells('F7:H7')
c = ws['F7']
c.font = Font(name='Calibri', bold=True, color=BLACK, size=24)
c.alignment = Alignment(horizontal='right', vertical='center')
c.value = '=EstadoResultados!C27'
c.number_format = '#,##0.00'

ws.row_dimensions[8].height = 4

# KPI 3: ROE
ws.merge_cells('B9:D9')
style_banner(ws['B9'], ORANGE, WHITE, 10)
ws['B9'].value = "ROE (Retorno sobre Patrimonio)"
ws.merge_cells('B10:D10')
c = ws['B10']
c.font = Font(name='Calibri', bold=True, color=BLACK, size=24)
c.alignment = Alignment(horizontal='right', vertical='center')
c.value = '=Razones!C5'
c.number_format = '0.0%'

# KPI 4: Razon Corriente
ws.merge_cells('F9:H9')
style_banner(ws['F9'], PURPLE, WHITE, 10)
ws['F9'].value = "RAZON CORRIENTE"
ws.merge_cells('F10:H10')
c = ws['F10']
c.font = Font(name='Calibri', bold=True, color=BLACK, size=24)
c.alignment = Alignment(horizontal='right', vertical='center')
c.value = '=Razones!C7'
c.number_format = '#,##0.00"x"'

ws.row_dimensions[11].height = 6

# Verification banner
ws.merge_cells('B12:H12')
c = ws['B12']
c.value = '=IF(BalanceGeneral!C20=BalanceGeneral!C33,"BALANCE CUADRA: Activos = Pasivos + Patrimonio","ERROR: EL BALANCE NO CUADRA")'
c.font = Font(name='Calibri', bold=True, size=12, color=WHITE)
c.fill = PatternFill('solid', fgColor=GREEN_OK)
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = thin_border

ws.row_dimensions[13].height = 8

# Chart data: Asset composition
ws['B14'].value = "COMPOSICION DE ACTIVOS"
ws['B14'].font = Font(name='Calibri', bold=True, size=11, color=GRAY_700)
ws['B15'].value = "Tipo"; ws['C15'].value = "Monto"
style_header(ws['B15'], bg=GRAY_700); style_header(ws['C15'], bg=GRAY_700)

ws['B16'].value = "Activo Corriente"; ws['C16'].value = '=BalanceGeneral!C11'
ws['B17'].value = "Activo No Corriente"; ws['C17'].value = '=BalanceGeneral!C19'
for r in [16, 17]:
    ws.cell(row=r, column=2).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws.cell(row=r, column=2).border = thin_border
    ws.cell(row=r, column=3).number_format = '#,##0.00'
    ws.cell(row=r, column=3).border = thin_border

chart1 = PieChart()
chart1.title = "Composicion de Activos"
chart1.style = 10
d1 = Reference(ws, min_col=3, min_row=15, max_row=17)
cats1 = Reference(ws, min_col=2, min_row=16, max_row=17)
chart1.add_data(d1, titles_from_data=True)
chart1.set_categories(cats1)
chart1.width = 16; chart1.height = 12
for idx, color in enumerate([TURQUOISE, ORANGE]):
    pt = DataPoint(idx=idx)
    pt.graphicalProperties.solidFill = color
    chart1.series[0].data_points.append(pt)
chart1.dataLabels = DataLabelList()
chart1.dataLabels.showPercent = True
ws.add_chart(chart1, "B19")

# Profitability chart
ws['F14'].value = "ESTADO DE RESULTADOS"
ws['F14'].font = Font(name='Calibri', bold=True, size=11, color=GRAY_700)
ws['F15'].value = "Concepto"; ws['G15'].value = "Monto"
style_header(ws['F15'], bg=GRAY_700); style_header(ws['G15'], bg=GRAY_700)

items_er = [("Ingresos", '=EstadoResultados!C6'), ("Costo Ventas", '=EstadoResultados!C10'),
            ("Gastos Op.", '=EstadoResultados!C19'), ("Utilidad Neta", '=EstadoResultados!C27')]
for i, (lbl, formula) in enumerate(items_er):
    r = 16 + i
    ws.cell(row=r, column=6).value = lbl
    ws.cell(row=r, column=6).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws.cell(row=r, column=6).border = thin_border
    ws.cell(row=r, column=7).value = formula
    ws.cell(row=r, column=7).number_format = '#,##0.00'
    ws.cell(row=r, column=7).border = thin_border

chart2 = BarChart()
chart2.type = "col"
chart2.style = 10
chart2.title = "Resumen de Resultados"
chart2.y_axis.title = "Bs."
chart2.legend = None
chart2.width = 16; chart2.height = 12
d2 = Reference(ws, min_col=7, min_row=15, max_row=19)
cats2 = Reference(ws, min_col=6, min_row=16, max_row=19)
chart2.add_data(d2, titles_from_data=True)
chart2.set_categories(cats2)
bar_colors = [TURQUOISE, RED_CORAL, ORANGE, PURPLE]
for idx, color in enumerate(bar_colors):
    pt = DataPoint(idx=idx)
    pt.graphicalProperties.solidFill = color
    chart2.series[0].data_points.append(pt)
ws.add_chart(chart2, "F19")

# ============================================================
# SHEET 2: PLAN DE CUENTAS
# ============================================================
ws_pc = wb.create_sheet("PlanCuentas")
ws_pc.sheet_properties.tabColor = TURQUOISE
ws_pc.sheet_view.showGridLines = False

pcw = {1:12, 2:35, 3:18, 4:16}
for c, w in pcw.items():
    ws_pc.column_dimensions[get_column_letter(c)].width = w

ws_pc.merge_cells('A1:D1')
style_title(ws_pc['A1'], 16)
ws_pc['A1'].value = "PLAN DE CUENTAS"
ws_pc['A1'].fill = PatternFill('solid', fgColor=WHITE)
ws_pc.row_dimensions[1].height = 35

headers_pc = [("A","CODIGO"), ("B","CUENTA"), ("C","TIPO"), ("D","NATURALEZA")]
ws_pc.row_dimensions[3].height = 28
for col_letter, label in headers_pc:
    style_header(ws_pc[f'{col_letter}3'], bg=DARK_BG, fg=GOLD, size=10)
    ws_pc[f'{col_letter}3'].value = label

# Standard chart of accounts
accounts = [
    # Activos
    ("1000", "ACTIVOS", "Titulo", ""),
    ("1100", "Activo Corriente", "Subtitulo", ""),
    ("1110", "Caja y Bancos", "Activo", "Deudora"),
    ("1120", "Cuentas por Cobrar", "Activo", "Deudora"),
    ("1130", "Inventarios", "Activo", "Deudora"),
    ("1140", "Anticipos y Prepagados", "Activo", "Deudora"),
    ("1200", "Activo No Corriente", "Subtitulo", ""),
    ("1210", "Propiedad, Planta y Equipo", "Activo", "Deudora"),
    ("1220", "Depreciacion Acumulada", "Activo", "Acreedora"),
    ("1230", "Intangibles", "Activo", "Deudora"),
    # Pasivos
    ("2000", "PASIVOS", "Titulo", ""),
    ("2100", "Pasivo Corriente", "Subtitulo", ""),
    ("2110", "Cuentas por Pagar", "Pasivo", "Acreedora"),
    ("2120", "Sueldos por Pagar", "Pasivo", "Acreedora"),
    ("2130", "Impuestos por Pagar", "Pasivo", "Acreedora"),
    ("2140", "Deudas a Corto Plazo", "Pasivo", "Acreedora"),
    ("2200", "Pasivo No Corriente", "Subtitulo", ""),
    ("2210", "Deudas a Largo Plazo", "Pasivo", "Acreedora"),
    # Patrimonio
    ("3000", "PATRIMONIO", "Titulo", ""),
    ("3100", "Capital Social", "Patrimonio", "Acreedora"),
    ("3200", "Reservas", "Patrimonio", "Acreedora"),
    ("3300", "Utilidades Retenidas", "Patrimonio", "Acreedora"),
    ("3400", "Resultado del Ejercicio", "Patrimonio", "Acreedora"),
    # Ingresos
    ("4000", "INGRESOS", "Titulo", ""),
    ("4100", "Ventas", "Ingreso", "Acreedora"),
    ("4200", "Otros Ingresos", "Ingreso", "Acreedora"),
    ("4300", "Ingresos Financieros", "Ingreso", "Acreedora"),
    # Costos
    ("5000", "COSTOS", "Titulo", ""),
    ("5100", "Costo de Ventas", "Costo", "Deudora"),
    # Gastos
    ("6000", "GASTOS OPERATIVOS", "Titulo", ""),
    ("6100", "Sueldos y Salarios", "Gasto", "Deudora"),
    ("6200", "Alquiler", "Gasto", "Deudora"),
    ("6300", "Servicios Basicos", "Gasto", "Deudora"),
    ("6400", "Depreciacion", "Gasto", "Deudora"),
    ("6500", "Publicidad y Marketing", "Gasto", "Deudora"),
    ("6600", "Otros Gastos Operativos", "Gasto", "Deudora"),
    ("7000", "GASTOS FINANCIEROS", "Titulo", ""),
    ("7100", "Intereses Pagados", "Gasto", "Deudora"),
    ("8000", "IMPUESTOS", "Titulo", ""),
    ("8100", "Impuesto sobre Utilidades", "Gasto", "Deudora"),
]

for i, (code, name, tipo, nature) in enumerate(accounts):
    r = 4 + i
    ws_pc.cell(row=r, column=1).value = code
    ws_pc.cell(row=r, column=2).value = name
    ws_pc.cell(row=r, column=3).value = tipo
    ws_pc.cell(row=r, column=4).value = nature

    is_title = tipo in ["Titulo", "Subtitulo"]
    ws_pc.cell(row=r, column=1).font = Font(name='Calibri', size=10,
        bold=is_title, color=DARK_BG if is_title else GRAY_700)
    ws_pc.cell(row=r, column=2).font = Font(name='Calibri', size=10,
        bold=is_title, color=DARK_BG if is_title else GRAY_700)
    ws_pc.cell(row=r, column=2).alignment = Alignment(
        horizontal='left', vertical='center', indent=0 if is_title else 1)
    ws_pc.cell(row=r, column=3).font = Font(name='Calibri', size=10, color=GRAY_500)
    ws_pc.cell(row=r, column=4).font = Font(name='Calibri', size=10, color=GRAY_500)
    for col in range(1, 5):
        ws_pc.cell(row=r, column=col).border = thin_border

# ============================================================
# SHEET 3: BALANCE GENERAL
# ============================================================
ws_bg = wb.create_sheet("BalanceGeneral")
ws_bg.sheet_properties.tabColor = RED_CORAL
ws_bg.sheet_view.showGridLines = False

bgw = {1:4, 2:35, 3:20, 4:4}
for c, w in bgw.items():
    ws_bg.column_dimensions[get_column_letter(c)].width = w
fill_white(ws_bg, 1, 40, 1, 4)

ws_bg.merge_cells('B1:C1')
style_title(ws_bg['B1'], 16)
ws_bg['B1'].value = "ESTADO DE SITUACION FINANCIERA"
ws_bg.row_dimensions[1].height = 30

ws_bg.merge_cells('B2:C2')
ws_bg['B2'].value = '="Al "&Dashboard!C4'
ws_bg['B2'].font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)

# ACTIVOS
ws_bg.merge_cells('B4:C4')
style_banner(ws_bg['B4'], DARK_BG, GOLD, 12)
ws_bg['B4'].value = "ACTIVOS"

ws_bg['B5'].value = "Activo Corriente"
ws_bg['B5'].font = Font(name='Calibri', size=11, bold=True, color=GRAY_800)

# Asset items (INPUT)
asset_current = [
    (6, "Caja y Bancos", 25000),
    (7, "Cuentas por Cobrar", 15000),
    (8, "Inventarios", 30000),
    (9, "Anticipos y Prepagados", 5000),
]
for r, label, sample in asset_current:
    style_label(ws_bg.cell(row=r, column=2), indent=1)
    ws_bg.cell(row=r, column=2).value = label
    c = ws_bg.cell(row=r, column=3)
    style_input(c); c.number_format = '#,##0.00'; c.value = sample; unlock_cell(c)

# Subtotal Activo Corriente (row 11)
r = 11
ws_bg.cell(row=10, column=2).value = ""
ws_bg.cell(row=r, column=2).value = "Total Activo Corriente"
ws_bg.cell(row=r, column=2).font = Font(name='Calibri', size=11, bold=True, color=BLACK)
ws_bg.cell(row=r, column=2).border = thin_border
c = ws_bg.cell(row=r, column=3)
c.value = '=SUM(C6:C9)'
style_subtotal(c)

# Non-current assets
ws_bg['B13'].value = "Activo No Corriente"
ws_bg['B13'].font = Font(name='Calibri', size=11, bold=True, color=GRAY_800)

asset_noncurrent = [
    (14, "Propiedad, Planta y Equipo", 150000),
    (15, "(-) Depreciacion Acumulada", -30000),
    (16, "Intangibles", 10000),
]
for r, label, sample in asset_noncurrent:
    style_label(ws_bg.cell(row=r, column=2), indent=1)
    ws_bg.cell(row=r, column=2).value = label
    c = ws_bg.cell(row=r, column=3)
    style_input(c); c.number_format = '#,##0.00'; c.value = sample; unlock_cell(c)

# Subtotal Activo No Corriente
ws_bg.cell(row=17, column=2).value = ""
r = 19
ws_bg['B18'].value = ""
ws_bg.cell(row=r, column=2).value = "Total Activo No Corriente"
ws_bg.cell(row=r, column=2).font = Font(name='Calibri', size=11, bold=True, color=BLACK)
ws_bg.cell(row=r, column=2).border = thin_border
c = ws_bg.cell(row=r, column=3)
c.value = '=SUM(C14:C16)'
style_subtotal(c)

# TOTAL ACTIVOS
r = 20
ws_bg.cell(row=r, column=2).value = "TOTAL ACTIVOS"
ws_bg.cell(row=r, column=2).font = Font(name='Calibri', size=12, bold=True, color=BLACK)
ws_bg.cell(row=r, column=2).border = thin_border
c = ws_bg.cell(row=r, column=3)
c.value = '=C11+C19'
style_total(c)

ws_bg.row_dimensions[21].height = 8

# PASIVOS
ws_bg.merge_cells('B22:C22')
style_banner(ws_bg['B22'], DARK_BG, GOLD, 12)
ws_bg['B22'].value = "PASIVOS Y PATRIMONIO"

ws_bg['B23'].value = "Pasivo Corriente"
ws_bg['B23'].font = Font(name='Calibri', size=11, bold=True, color=GRAY_800)

liability_current = [
    (24, "Cuentas por Pagar", 20000),
    (25, "Sueldos por Pagar", 5000),
    (26, "Impuestos por Pagar", 3000),
    (27, "Deudas a Corto Plazo", 15000),
]
for r, label, sample in liability_current:
    style_label(ws_bg.cell(row=r, column=2), indent=1)
    ws_bg.cell(row=r, column=2).value = label
    c = ws_bg.cell(row=r, column=3)
    style_input(c); c.number_format = '#,##0.00'; c.value = sample; unlock_cell(c)

# Total Pasivo Corriente
r = 28
ws_bg.cell(row=r, column=2).value = "Total Pasivo Corriente"
ws_bg.cell(row=r, column=2).font = Font(name='Calibri', size=11, bold=True, color=BLACK)
ws_bg.cell(row=r, column=2).border = thin_border
c = ws_bg.cell(row=r, column=3)
c.value = '=SUM(C24:C27)'
style_subtotal(c)

# Non-current liabilities
ws_bg['B29'].value = "Pasivo No Corriente"
ws_bg['B29'].font = Font(name='Calibri', size=11, bold=True, color=GRAY_800)

ws_bg.cell(row=30, column=2).value = "Deudas a Largo Plazo"
style_label(ws_bg.cell(row=30, column=2), indent=1)
c = ws_bg.cell(row=30, column=3)
style_input(c); c.number_format = '#,##0.00'; c.value = 50000; unlock_cell(c)

# TOTAL PASIVOS
r = 31
ws_bg.cell(row=r, column=2).value = "TOTAL PASIVOS"
ws_bg.cell(row=r, column=2).font = Font(name='Calibri', size=11, bold=True, color=BLACK)
ws_bg.cell(row=r, column=2).border = thin_border
c = ws_bg.cell(row=r, column=3)
c.value = '=C28+C30'
style_subtotal(c)

# PATRIMONIO
ws_bg['B32'].value = "Patrimonio"
ws_bg['B32'].font = Font(name='Calibri', size=11, bold=True, color=GRAY_800)

equity_items = [
    (33, "Capital Social", 80000),
    (34, "Reservas", 5000),
    (35, "Utilidades Retenidas", 17000),
    (36, "Resultado del Ejercicio", None),  # Linked to Income Statement
]
for r, label, sample in equity_items:
    style_label(ws_bg.cell(row=r, column=2), indent=1)
    ws_bg.cell(row=r, column=2).value = label
    c = ws_bg.cell(row=r, column=3)
    if sample is not None:
        style_input(c); c.number_format = '#,##0.00'; c.value = sample; unlock_cell(c)
    else:
        c.value = '=EstadoResultados!C27'
        c.number_format = '#,##0.00'
        style_output(c, bold=True)

# TOTAL PATRIMONIO
r = 37
ws_bg.cell(row=r, column=2).value = "TOTAL PATRIMONIO"
ws_bg.cell(row=r, column=2).font = Font(name='Calibri', size=11, bold=True, color=BLACK)
ws_bg.cell(row=r, column=2).border = thin_border
c = ws_bg.cell(row=r, column=3)
c.value = '=SUM(C33:C36)'
style_subtotal(c)

# TOTAL PASIVOS + PATRIMONIO (should = Total Activos)
# Redefine row 33 to be Total P+P
# Actually let's put it at row 33 but we already used that. Let me use row 38.
# Wait, I have equity items at 33-36. Let me fix the total at row 38.

# Fix: The row numbering. Let me recalculate.
# Row 31 = TOTAL PASIVOS, row 32 = Patrimonio header, rows 33-36 = equity items
# Row 37 = TOTAL PATRIMONIO. OK that's fine.

# But Dashboard references C33 for Total P+P. Let me fix.
# Actually let me just not reference C33 in Dashboard. Let me use the correct rows.

# TOTAL PASIVOS + PATRIMONIO
r = 38
ws_bg.cell(row=r, column=2).value = "TOTAL PASIVOS + PATRIMONIO"
ws_bg.cell(row=r, column=2).font = Font(name='Calibri', size=12, bold=True, color=BLACK)
ws_bg.cell(row=r, column=2).border = thin_border
c = ws_bg.cell(row=r, column=3)
c.value = '=C31+C37'
style_total(c)

# Fix Dashboard reference: C33 should be C38
# We'll fix after

# Verification
r = 39
ws_bg.merge_cells('B39:C39')
c = ws_bg['B39']
c.value = '=IF(ABS(C20-C38)<0.01,"CUADRA: Activos = Pasivos + Patrimonio","ERROR: NO CUADRA")'
c.font = Font(name='Calibri', bold=True, size=11, color=WHITE)
c.fill = PatternFill('solid', fgColor=GREEN_OK)
c.alignment = Alignment(horizontal='center', vertical='center')

# Conditional formatting for verification
ws_bg.conditional_formatting.add('B39:C39',
    FormulaRule(formula=['ABS(C20-C38)>=0.01'],
               fill=PatternFill('solid', fgColor=RED_ALERT),
               font=Font(color=WHITE, bold=True)))

# Fix Dashboard reference
ws['B12'].value = '=IF(ABS(BalanceGeneral!C20-BalanceGeneral!C38)<0.01,"BALANCE CUADRA: Activos = Pasivos + Patrimonio","ERROR: EL BALANCE NO CUADRA")'

# FormulaRule already imported at top

# ============================================================
# SHEET 4: ESTADO DE RESULTADOS
# ============================================================
ws_er = wb.create_sheet("EstadoResultados")
ws_er.sheet_properties.tabColor = ORANGE
ws_er.sheet_view.showGridLines = False

erw = {1:4, 2:35, 3:20, 4:4}
for c, w in erw.items():
    ws_er.column_dimensions[get_column_letter(c)].width = w
fill_white(ws_er, 1, 35, 1, 4)

ws_er.merge_cells('B1:C1')
style_title(ws_er['B1'], 16)
ws_er['B1'].value = "ESTADO DE RESULTADOS"
ws_er.row_dimensions[1].height = 30

ws_er.merge_cells('B2:C2')
ws_er['B2'].value = '="Periodo: "&Dashboard!C4'
ws_er['B2'].font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)

# INGRESOS
ws_er.merge_cells('B4:C4')
style_banner(ws_er['B4'], TURQUOISE, WHITE, 11)
ws_er['B4'].value = "INGRESOS"

income_items = [
    (5, "Ventas Netas", 350000),
    (6, "Otros Ingresos", 8000),
]
for r, label, sample in income_items:
    style_label(ws_er.cell(row=r, column=2), indent=1)
    ws_er.cell(row=r, column=2).value = label
    c = ws_er.cell(row=r, column=3)
    style_input(c); c.number_format = '#,##0.00'; c.value = sample; unlock_cell(c)

r = 7
ws_er.cell(row=r, column=2).value = "TOTAL INGRESOS"
ws_er.cell(row=r, column=2).font = Font(name='Calibri', size=11, bold=True, color=BLACK)
ws_er.cell(row=r, column=2).border = thin_border
c = ws_er.cell(row=r, column=3)
c.value = '=SUM(C5:C6)'
style_subtotal(c)

ws_er.row_dimensions[8].height = 4

# COSTO DE VENTAS
ws_er.merge_cells('B9:C9')
style_banner(ws_er['B9'], RED_CORAL, WHITE, 11)
ws_er['B9'].value = "COSTO DE VENTAS"

ws_er.cell(row=10, column=2).value = "Costo de Ventas"
style_label(ws_er.cell(row=10, column=2), indent=1)
c = ws_er.cell(row=10, column=3)
style_input(c); c.number_format = '#,##0.00'; c.value = 180000; unlock_cell(c)

# UTILIDAD BRUTA
r = 12
ws_er.cell(row=r, column=2).value = "UTILIDAD BRUTA"
ws_er.cell(row=r, column=2).font = Font(name='Calibri', size=11, bold=True, color=BLACK)
ws_er.cell(row=r, column=2).border = thin_border
c = ws_er.cell(row=r, column=3)
c.value = '=C7-C10'
style_subtotal(c)

ws_er.row_dimensions[13].height = 4

# GASTOS OPERATIVOS
ws_er.merge_cells('B14:C14')
style_banner(ws_er['B14'], ORANGE, WHITE, 11)
ws_er['B14'].value = "GASTOS OPERATIVOS"

opex_items = [
    (15, "Sueldos y Salarios", 80000),
    (16, "Alquiler", 36000),
    (17, "Servicios Basicos", 12000),
    (18, "Depreciacion", 15000),
    (19, "Publicidad y Marketing", 8000),
    (20, "Otros Gastos Operativos", 5000),
]
for r, label, sample in opex_items:
    style_label(ws_er.cell(row=r, column=2), indent=1)
    ws_er.cell(row=r, column=2).value = label
    c = ws_er.cell(row=r, column=3)
    style_input(c); c.number_format = '#,##0.00'; c.value = sample; unlock_cell(c)

r = 21
ws_er.cell(row=r, column=2).value = "TOTAL GASTOS OPERATIVOS"
ws_er.cell(row=r, column=2).font = Font(name='Calibri', size=11, bold=True, color=BLACK)
ws_er.cell(row=r, column=2).border = thin_border
c = ws_er.cell(row=r, column=3)
c.value = '=SUM(C15:C20)'
style_subtotal(c)

# UTILIDAD OPERATIVA
r = 22
ws_er.cell(row=r, column=2).value = "UTILIDAD OPERATIVA (EBIT)"
ws_er.cell(row=r, column=2).font = Font(name='Calibri', size=11, bold=True, color=BLACK)
ws_er.cell(row=r, column=2).border = thin_border
c = ws_er.cell(row=r, column=3)
c.value = '=C12-C21'
style_subtotal(c)

ws_er.row_dimensions[23].height = 4

# GASTOS FINANCIEROS
ws_er.cell(row=24, column=2).value = "(-) Gastos Financieros"
style_label(ws_er.cell(row=24, column=2), indent=1)
c = ws_er.cell(row=24, column=3)
style_input(c); c.number_format = '#,##0.00'; c.value = 6000; unlock_cell(c)

# UTILIDAD ANTES DE IMPUESTOS
r = 25
ws_er.cell(row=r, column=2).value = "UTILIDAD ANTES DE IMPUESTOS"
ws_er.cell(row=r, column=2).font = Font(name='Calibri', size=11, bold=True, color=BLACK)
ws_er.cell(row=r, column=2).border = thin_border
c = ws_er.cell(row=r, column=3)
c.value = '=C22-C24'
style_subtotal(c)

# IMPUESTOS
ws_er.cell(row=26, column=2).value = "(-) Impuesto sobre Utilidades (25%)"
style_label(ws_er.cell(row=26, column=2), indent=1)
c = ws_er.cell(row=26, column=3)
c.value = '=IFERROR(IF(C25>0,C25*0.25,0),0)'
c.number_format = '#,##0.00'
style_output(c)

# UTILIDAD NETA
r = 27
ws_er.cell(row=r, column=2).value = "UTILIDAD NETA"
ws_er.cell(row=r, column=2).font = Font(name='Calibri', size=12, bold=True, color=BLACK)
ws_er.cell(row=r, column=2).border = thin_border
c = ws_er.cell(row=r, column=3)
c.value = '=C25-C26'
style_total(c)

# Fix Dashboard reference for income statement
ws['C16'].value = '=BalanceGeneral!C11'  # Already correct
ws['C17'].value = '=BalanceGeneral!C19'  # Already correct

# Fix Dashboard ER references to use C7 for total ingresos and C21 for opex
# Row 16 of dashboard = Ingresos
ws.cell(row=16, column=7).value = '=EstadoResultados!C7'
ws.cell(row=17, column=7).value = '=EstadoResultados!C10'
ws.cell(row=18, column=7).value = '=EstadoResultados!C21'
ws.cell(row=19, column=7).value = '=EstadoResultados!C27'

# Fix labels
ws.cell(row=16, column=6).value = "Ingresos"
ws.cell(row=17, column=6).value = "Costo Ventas"
ws.cell(row=18, column=6).value = "Gastos Op."
ws.cell(row=19, column=6).value = "Utilidad Neta"

# ============================================================
# SHEET 5: FLUJO DE CAJA
# ============================================================
ws_fc = wb.create_sheet("FlujoCaja")
ws_fc.sheet_properties.tabColor = PURPLE
ws_fc.sheet_view.showGridLines = False

fcw = {1:4, 2:35, 3:20, 4:4}
for c, w in fcw.items():
    ws_fc.column_dimensions[get_column_letter(c)].width = w
fill_white(ws_fc, 1, 35, 1, 4)

ws_fc.merge_cells('B1:C1')
style_title(ws_fc['B1'], 16)
ws_fc['B1'].value = "ESTADO DE FLUJO DE EFECTIVO"
ws_fc.row_dimensions[1].height = 30

ws_fc.merge_cells('B2:C2')
ws_fc['B2'].value = '="Periodo: "&Dashboard!C4'
ws_fc['B2'].font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)

# OPERACIONES
ws_fc.merge_cells('B4:C4')
style_banner(ws_fc['B4'], TURQUOISE, WHITE, 11)
ws_fc['B4'].value = "ACTIVIDADES DE OPERACION"

fc_ops = [
    (5, "Utilidad Neta", None, '=EstadoResultados!C27'),
    (6, "(+) Depreciacion", 15000, None),
    (7, "(+/-) Cambio en Ctas por Cobrar", -3000, None),
    (8, "(+/-) Cambio en Inventarios", 2000, None),
    (9, "(+/-) Cambio en Ctas por Pagar", 4000, None),
]
for r, label, sample, formula in fc_ops:
    style_label(ws_fc.cell(row=r, column=2), indent=1)
    ws_fc.cell(row=r, column=2).value = label
    c = ws_fc.cell(row=r, column=3)
    if formula:
        c.value = formula
        c.number_format = '#,##0.00'
        style_output(c, bold=True)
    else:
        style_input(c); c.number_format = '#,##0.00'; c.value = sample; unlock_cell(c)

r = 10
ws_fc.cell(row=r, column=2).value = "Flujo Neto de Operaciones"
ws_fc.cell(row=r, column=2).font = Font(name='Calibri', size=11, bold=True, color=BLACK)
ws_fc.cell(row=r, column=2).border = thin_border
c = ws_fc.cell(row=r, column=3)
c.value = '=SUM(C5:C9)'
style_subtotal(c)

ws_fc.row_dimensions[11].height = 4

# INVERSION
ws_fc.merge_cells('B12:C12')
style_banner(ws_fc['B12'], ORANGE, WHITE, 11)
ws_fc['B12'].value = "ACTIVIDADES DE INVERSION"

fc_inv = [
    (13, "Compra de Activos Fijos", -20000),
    (14, "Venta de Activos", 5000),
]
for r, label, sample in fc_inv:
    style_label(ws_fc.cell(row=r, column=2), indent=1)
    ws_fc.cell(row=r, column=2).value = label
    c = ws_fc.cell(row=r, column=3)
    style_input(c); c.number_format = '#,##0.00'; c.value = sample; unlock_cell(c)

r = 15
ws_fc.cell(row=r, column=2).value = "Flujo Neto de Inversion"
ws_fc.cell(row=r, column=2).font = Font(name='Calibri', size=11, bold=True, color=BLACK)
ws_fc.cell(row=r, column=2).border = thin_border
c = ws_fc.cell(row=r, column=3)
c.value = '=SUM(C13:C14)'
style_subtotal(c)

ws_fc.row_dimensions[16].height = 4

# FINANCIAMIENTO
ws_fc.merge_cells('B17:C17')
style_banner(ws_fc['B17'], PURPLE, WHITE, 11)
ws_fc['B17'].value = "ACTIVIDADES DE FINANCIAMIENTO"

fc_fin = [
    (18, "Nuevos Prestamos", 10000),
    (19, "Pago de Prestamos", -8000),
    (20, "Aportes de Capital", 0),
    (21, "Distribucion de Dividendos", -5000),
]
for r, label, sample in fc_fin:
    style_label(ws_fc.cell(row=r, column=2), indent=1)
    ws_fc.cell(row=r, column=2).value = label
    c = ws_fc.cell(row=r, column=3)
    style_input(c); c.number_format = '#,##0.00'; c.value = sample; unlock_cell(c)

r = 22
ws_fc.cell(row=r, column=2).value = "Flujo Neto de Financiamiento"
ws_fc.cell(row=r, column=2).font = Font(name='Calibri', size=11, bold=True, color=BLACK)
ws_fc.cell(row=r, column=2).border = thin_border
c = ws_fc.cell(row=r, column=3)
c.value = '=SUM(C18:C21)'
style_subtotal(c)

ws_fc.row_dimensions[23].height = 8

# TOTALS
r = 24
ws_fc.cell(row=r, column=2).value = "AUMENTO/DISMINUCION NETO DE EFECTIVO"
ws_fc.cell(row=r, column=2).font = Font(name='Calibri', size=11, bold=True, color=BLACK)
ws_fc.cell(row=r, column=2).border = thin_border
c = ws_fc.cell(row=r, column=3)
c.value = '=C10+C15+C22'
style_subtotal(c)

ws_fc.cell(row=25, column=2).value = "Efectivo al Inicio del Periodo"
style_label(ws_fc.cell(row=25, column=2))
c = ws_fc.cell(row=25, column=3)
style_input(c); c.number_format = '#,##0.00'; c.value = 20000; unlock_cell(c)

r = 26
ws_fc.cell(row=r, column=2).value = "EFECTIVO AL FINAL DEL PERIODO"
ws_fc.cell(row=r, column=2).font = Font(name='Calibri', size=12, bold=True, color=BLACK)
ws_fc.cell(row=r, column=2).border = thin_border
c = ws_fc.cell(row=r, column=3)
c.value = '=C24+C25'
style_total(c)

# ============================================================
# SHEET 6: RAZONES FINANCIERAS
# ============================================================
ws_rz = wb.create_sheet("Razones")
ws_rz.sheet_properties.tabColor = "3498DB"
ws_rz.sheet_view.showGridLines = False

rzw = {1:4, 2:30, 3:16, 4:30, 5:4}
for c, w in rzw.items():
    ws_rz.column_dimensions[get_column_letter(c)].width = w
fill_white(ws_rz, 1, 30, 1, 5)

ws_rz.merge_cells('B1:D1')
style_title(ws_rz['B1'], 16)
ws_rz['B1'].value = "RAZONES FINANCIERAS"
ws_rz.row_dimensions[1].height = 35

ws_rz.merge_cells('B2:D2')
ws_rz['B2'].value = "Indicadores calculados automaticamente desde el Balance y Estado de Resultados."
ws_rz['B2'].font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)

# Headers
ws_rz['B3'].value = "INDICADOR"
ws_rz['C3'].value = "VALOR"
ws_rz['D3'].value = "INTERPRETACION"
style_header(ws_rz['B3'], bg=DARK_BG, fg=GOLD, size=10)
style_header(ws_rz['C3'], bg=DARK_BG, fg=GOLD, size=10)
style_header(ws_rz['D3'], bg=DARK_BG, fg=GOLD, size=10)

# RENTABILIDAD section
ws_rz.merge_cells('B4:D4')
style_banner(ws_rz['B4'], TURQUOISE, WHITE, 10)
ws_rz['B4'].value = "RENTABILIDAD"

ratios = [
    # (row, name, formula, format, interpretation)
    (5, "ROE (Retorno sobre Patrimonio)",
     '=IFERROR(EstadoResultados!C27/BalanceGeneral!C37,0)', '0.0%',
     '=IF(C5>0.15,"Excelente",IF(C5>0.08,"Bueno","Bajo"))'),
    (6, "ROA (Retorno sobre Activos)",
     '=IFERROR(EstadoResultados!C27/BalanceGeneral!C20,0)', '0.0%',
     '=IF(C6>0.10,"Excelente",IF(C6>0.05,"Bueno","Bajo"))'),
]

# LIQUIDEZ section
ws_rz.merge_cells(f'B{7}:D{7}')
style_banner(ws_rz[f'B{7}'], ORANGE, WHITE, 10)
ws_rz[f'B{7}'].value = "LIQUIDEZ"

ratios += [
    (8, "Razon Corriente",  # Fix: was 7 in KPI ref, need to match
     '=IFERROR(BalanceGeneral!C11/BalanceGeneral!C28,0)', '#,##0.00',
     '=IF(C8>2,"Holgada",IF(C8>1,"Adecuada","Riesgo"))'),
    (9, "Prueba Acida",
     '=IFERROR((BalanceGeneral!C11-BalanceGeneral!C8)/BalanceGeneral!C28,0)', '#,##0.00',
     '=IF(C9>1,"Buena",IF(C9>0.5,"Aceptable","Riesgo"))'),
    (10, "Capital de Trabajo (Bs.)",
     '=BalanceGeneral!C11-BalanceGeneral!C28', '#,##0.00',
     '=IF(C10>0,"Positivo - Liquidez OK","Negativo - Riesgo")'),
]

# ENDEUDAMIENTO section
ws_rz.merge_cells(f'B{11}:D{11}')
style_banner(ws_rz[f'B{11}'], RED_CORAL, WHITE, 10)
ws_rz[f'B{11}'].value = "ENDEUDAMIENTO"

ratios += [
    (12, "Razon de Endeudamiento",
     '=IFERROR(BalanceGeneral!C31/BalanceGeneral!C20,0)', '0.0%',
     '=IF(C12<0.4,"Conservador",IF(C12<0.6,"Moderado","Alto riesgo"))'),
    (13, "Apalancamiento (Deuda/Patrimonio)",
     '=IFERROR(BalanceGeneral!C31/BalanceGeneral!C37,0)', '#,##0.00"x"',
     '=IF(C13<1,"Bajo apalancamiento",IF(C13<2,"Moderado","Alto"))'),
]

# EFICIENCIA section
ws_rz.merge_cells(f'B{14}:D{14}')
style_banner(ws_rz[f'B{14}'], PURPLE, WHITE, 10)
ws_rz[f'B{14}'].value = "EFICIENCIA"

ratios += [
    (15, "Margen Bruto",
     '=IFERROR(EstadoResultados!C12/EstadoResultados!C7,0)', '0.0%',
     '=IF(C15>0.4,"Excelente",IF(C15>0.25,"Bueno","Bajo"))'),
    (16, "Margen Operativo",
     '=IFERROR(EstadoResultados!C22/EstadoResultados!C7,0)', '0.0%',
     '=IF(C16>0.15,"Excelente",IF(C16>0.08,"Bueno","Bajo"))'),
    (17, "Margen Neto",
     '=IFERROR(EstadoResultados!C27/EstadoResultados!C7,0)', '0.0%',
     '=IF(C17>0.10,"Excelente",IF(C17>0.05,"Bueno","Bajo"))'),
]

for r, name, formula, fmt, interp in ratios:
    ws_rz.cell(row=r, column=2).value = name
    ws_rz.cell(row=r, column=2).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws_rz.cell(row=r, column=2).border = thin_border

    c = ws_rz.cell(row=r, column=3)
    c.value = formula
    c.number_format = fmt
    style_output(c, bold=True)

    ws_rz.cell(row=r, column=4).value = interp
    ws_rz.cell(row=r, column=4).font = Font(name='Calibri', size=10, color=GRAY_600)
    ws_rz.cell(row=r, column=4).border = thin_border

# Fix Dashboard KPI reference for Razon Corriente (was C7, should be C8)
ws['F10'].value = '=Razones!C8'

# Conditional formatting for interpretation
for r_range in ['D5:D6', 'D8:D10', 'D12:D13', 'D15:D17']:
    ws_rz.conditional_formatting.add(r_range,
        CellIsRule(operator='containsText', formula=['"Excelente"'],
                  fill=PatternFill('solid', fgColor="D1FAE5"),
                  font=Font(color=GREEN_OK, bold=True)))
    ws_rz.conditional_formatting.add(r_range,
        CellIsRule(operator='containsText', formula=['"Riesgo"'],
                  fill=PatternFill('solid', fgColor="FEE2E2"),
                  font=Font(color=RED_ALERT, bold=True)))

# ============================================================
# SHEET 7: CONFIG
# ============================================================
ws_conf = wb.create_sheet("Config")
ws_conf.sheet_properties.tabColor = GRAY_500
ws_conf.sheet_view.showGridLines = False
cwid = {1:25, 2:55}
for c, w in cwid.items():
    ws_conf.column_dimensions[get_column_letter(c)].width = w
fill_white(ws_conf, 1, 28, 1, 3)
ws_conf['A1'].value = "v1.0.0"
ws_conf['A1'].font = Font(name='Calibri', size=9, color=GRAY_400)
ws_conf.merge_cells('A3:B3')
style_title(ws_conf['A3'], 14)
ws_conf['A3'].value = "CONFIGURACION"

settings = [
    ("Producto", "Estados Financieros Completos"),
    ("Version", "v1.0.0"),
    ("Marca", "No Somos Ignorantes"),
    ("Precio", "Bs. 199 (Premium)"),
    ("Proteccion", "nsi2024"),
    ("", ""),
    ("INSTRUCCIONES", ""),
    ("1.", "Ingresa el periodo en el Dashboard."),
    ("2.", "En 'BalanceGeneral', ingresa los saldos de cada cuenta (celdas amarillas)."),
    ("3.", "En 'EstadoResultados', ingresa ingresos, costos y gastos."),
    ("4.", "En 'FlujoCaja', ingresa los movimientos de efectivo."),
    ("5.", "La Utilidad del Ejercicio se vincula automaticamente al Balance."),
    ("6.", "Las Razones Financieras se calculan automaticamente."),
    ("7.", "El Dashboard muestra KPIs, verificacion de cuadre y graficos."),
    ("8.", "Para desbloquear: Revisar -> Desproteger hoja -> nsi2024"),
    ("", ""),
    ("NOTA", "El Plan de Cuentas es referencial. Ajusta segun tu empresa."),
]
for i, (label, value) in enumerate(settings):
    r = 5 + i
    ws_conf.cell(row=r, column=1).value = label
    ws_conf.cell(row=r, column=1).font = Font(name='Calibri', size=10, bold=True, color=GRAY_700)
    ws_conf.cell(row=r, column=2).value = value
    ws_conf.cell(row=r, column=2).font = Font(name='Calibri', size=10, color=GRAY_600)

# ============================================================
# PROTECTION
# ============================================================
for sheet in wb.worksheets:
    sheet.protection.sheet = True
    sheet.protection.password = "nsi2024"
    sheet.protection.enable()

# Unlock Dashboard
unlock_cell(ws['C4'])

# Unlock Balance General
for r in [6, 7, 8, 9, 14, 15, 16, 24, 25, 26, 27, 30, 33, 34, 35]:
    unlock_cell(ws_bg.cell(row=r, column=3))

# Unlock Estado de Resultados
for r in [5, 6, 10, 15, 16, 17, 18, 19, 20, 24]:
    unlock_cell(ws_er.cell(row=r, column=3))

# Unlock Flujo de Caja
for r in [6, 7, 8, 9, 13, 14, 18, 19, 20, 21, 25]:
    unlock_cell(ws_fc.cell(row=r, column=3))

# ============================================================
# SAVE & VERIFY
# ============================================================
out = r"D:\Landing-Page_marketplace\excel_products\Estados_Financieros_Completos_NSI.xlsx"
wb.save(out)
print(f"[OK] Saved: {out}")

from openpyxl import load_workbook
wb2 = load_workbook(out)
fc_count = sum(1 for s in wb2.worksheets for row in s.iter_rows() for cell in row
               if isinstance(cell.value, str) and cell.value.startswith('='))
print(f"[OK] Sheets: {wb2.sheetnames}")
print(f"[OK] Total formulas: {fc_count}")
print(f"[OK] Charts: Dashboard={len(ws._charts)}")
print(f"[OK] Financial ratios: 10 indicators")
print(f"[OK] Chart of accounts: {len(accounts)} entries")
