"""
Build: Amortizador de Deudas PRO (Bs. 69)
No Somos Ignorantes v1.0
3 amortization methods (French, German, American), full comparison, prepayment simulator.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, AreaChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule, DataBarRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ============================================================
# COLOR PALETTE
# ============================================================
GOLD = "D4AF37"
DARK_BG = "1A1A2E"
CARD_BG = "16213E"
ACCENT_BLUE = "0F3460"
WHITE = "FFFFFF"
BLACK = "000000"
LIGHT_YELLOW = "FFF9C4"
LIGHT_GREEN = "E8F5E9"
SUCCESS = "10B981"
WARNING = "F59E0B"
ERROR_RED = "EF4444"
INFO_BLUE = "3B82F6"
GRAY_100 = "F3F4F6"
GRAY_200 = "E5E7EB"
GRAY_300 = "D1D5DB"
GRAY_500 = "6B7280"
GRAY_700 = "374151"
GRAY_900 = "111827"

# Method-specific colors
FRENCH_BLUE = "2563EB"
FRENCH_BLUE_LIGHT = "DBEAFE"
GERMAN_GREEN = "059669"
GERMAN_GREEN_LIGHT = "D1FAE5"
AMERICAN_ORANGE = "D97706"
AMERICAN_ORANGE_LIGHT = "FEF3C7"
WINNER_TEAL = "0D9488"
WINNER_TEAL_LIGHT = "CCFBF1"
PURPLE = "7C3AED"
PURPLE_LIGHT = "EDE9FE"

# ============================================================
# STYLE HELPERS
# ============================================================
thin_border = Border(
    left=Side(style='thin', color=GRAY_300),
    right=Side(style='thin', color=GRAY_300),
    top=Side(style='thin', color=GRAY_300),
    bottom=Side(style='thin', color=GRAY_300)
)
gold_border = Border(
    left=Side(style='thin', color=GOLD),
    right=Side(style='thin', color=GOLD),
    top=Side(style='thin', color=GOLD),
    bottom=Side(style='thin', color=GOLD)
)
thick_bottom = Border(bottom=Side(style='medium', color=GRAY_700))

def style_header(cell, bg=DARK_BG, fg=WHITE, size=12):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')

def style_input(cell):
    cell.font = Font(name='Calibri', size=11, color="0000FF")
    cell.fill = PatternFill('solid', fgColor=LIGHT_YELLOW)
    cell.border = gold_border
    cell.alignment = Alignment(horizontal='right', vertical='center')

def style_output(cell, bold=False):
    cell.font = Font(name='Calibri', size=11, color=GRAY_900, bold=bold)
    cell.fill = PatternFill('solid', fgColor=LIGHT_GREEN)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='right', vertical='center')

def style_label(cell, bold=False):
    cell.font = Font(name='Calibri', size=11, color=GRAY_700, bold=bold)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def style_section(cell, bg, fg=WHITE):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=11)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def fmt_currency(cell):
    cell.number_format = '#,##0.00'

def fmt_pct(cell):
    cell.number_format = '0.00%'

def fmt_int(cell):
    cell.number_format = '#,##0'

def rh(ws, row, height):
    ws.row_dimensions[row].height = height

def fill_row(ws, row, cols, color):
    for c in cols:
        ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=color)

MAX_MONTHS = 360

# ============================================================
# SHEET 1: CALCULADORA (Main input + summary)
# ============================================================
ws = wb.active
ws.title = "Calculadora"
ws.sheet_properties.tabColor = GOLD

ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 32
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 22
ws.column_dimensions['F'].width = 22
ws.column_dimensions['G'].width = 22
ws.column_dimensions['H'].width = 2

ws.freeze_panes = 'B6'

# ── HEADER ──
rh(ws, 1, 8)
ws.merge_cells('B1:G1')
fill_row(ws, 1, range(2, 8), DARK_BG)

ws.merge_cells('B2:G2')
h = ws['B2']
h.value = "AMORTIZADOR DE DEUDAS PRO"
h.font = Font(name='Calibri', bold=True, color=GOLD, size=22)
h.fill = PatternFill('solid', fgColor=DARK_BG)
h.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws, 2, range(2, 8), DARK_BG)
rh(ws, 2, 48)

ws.merge_cells('B3:G3')
sub = ws['B3']
sub.value = "No Somos Ignorantes  |  v1.0  |  Frances - Aleman - Americano en un solo lugar"
sub.font = Font(name='Calibri', color=GRAY_300, size=10, italic=True)
sub.fill = PatternFill('solid', fgColor=DARK_BG)
sub.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws, 3, range(2, 8), DARK_BG)
rh(ws, 3, 24)

rh(ws, 4, 8)

# ── LOAN INPUT SECTION ──
r = 5
rh(ws, r, 30)
ws.merge_cells(f'B{r}:G{r}')
sec = ws[f'B{r}']
sec.value = "DATOS DEL PRESTAMO"
style_section(sec, ACCENT_BLUE)
fill_row(ws, r, range(3, 8), ACCENT_BLUE)

inputs_data = [
    (6, "Capital del prestamo (Bs.)", 200000, True, False),
    (7, "Tasa de interes anual (%)", 0.08, False, True),
    (8, "Plazo (meses)", 120, False, False),
    (9, "Fecha de inicio", None, False, False),
]

# Data validations
dv_capital = DataValidation(type="decimal", operator="between", formula1=1000, formula2=10000000)
dv_capital.error = "El capital debe estar entre Bs. 1,000 y Bs. 10,000,000"
dv_capital.errorTitle = "Capital invalido"
dv_capital.prompt = "Ingresa el monto total del prestamo"
dv_capital.promptTitle = "Capital"
ws.add_data_validation(dv_capital)

dv_rate = DataValidation(type="decimal", operator="between", formula1=0, formula2=1)
dv_rate.error = "La tasa debe estar entre 0% y 100%"
dv_rate.errorTitle = "Tasa invalida"
dv_rate.prompt = "Ingresa la tasa anual como decimal (ej: 0.08 = 8%)"
dv_rate.promptTitle = "Tasa Anual"
ws.add_data_validation(dv_rate)

dv_term = DataValidation(type="whole", operator="between", formula1=1, formula2=360)
dv_term.error = "El plazo debe estar entre 1 y 360 meses"
dv_term.errorTitle = "Plazo invalido"
dv_term.prompt = "Ingresa el plazo en meses (ej: 120 = 10 anios)"
dv_term.promptTitle = "Plazo"
ws.add_data_validation(dv_term)

for (row, label, default, is_cur, is_pct) in inputs_data:
    rh(ws, row, 32)
    lbl = ws.cell(row=row, column=2, value=label)
    style_label(lbl, bold=True)
    lbl.fill = PatternFill('solid', fgColor=LIGHT_YELLOW)
    lbl.border = thin_border

    ws.merge_cells(f'C{row}:D{row}')
    val = ws.cell(row=row, column=3)
    if row == 9:
        from datetime import date
        val.value = date(2025, 1, 1)
        style_input(val)
        val.number_format = 'DD/MM/YYYY'
    else:
        val.value = default
        style_input(val)
        if is_cur:
            fmt_currency(val)
            val.font = Font(name='Calibri', size=14, bold=True, color="0000FF")
        elif is_pct:
            fmt_pct(val)
            val.font = Font(name='Calibri', size=14, bold=True, color="0000FF")
        else:
            fmt_int(val)
            val.font = Font(name='Calibri', size=14, bold=True, color="0000FF")

    # Helper info
    if row == 8:
        info = ws.cell(row=row, column=5)
        info.value = '=IFERROR(ROUND(C8/12,1)&" anios","")'
        info.font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)

dv_capital.add('C6')
dv_rate.add('C7')
dv_term.add('C8')

rh(ws, 10, 10)

# ── RESULTS: 3 METHODS SIDE BY SIDE ──
r = 11
rh(ws, r, 30)
ws.merge_cells(f'B{r}:G{r}')
sec = ws[f'B{r}']
sec.value = "COMPARACION DE METODOS"
style_section(sec, CARD_BG, fg=GOLD)
fill_row(ws, r, range(3, 8), CARD_BG)

# Column headers for comparison
r = 12
rh(ws, r, 28)
method_headers = ['', 'METRICA', 'FRANCES', 'ALEMAN', 'AMERICANO', 'MEJOR']
for i, hdr in enumerate(method_headers):
    c = ws.cell(row=r, column=i + 1, value=hdr)
    if i >= 1:
        style_header(c, bg=CARD_BG, fg=GOLD, size=10)

# Method color indicator row
r = 13
rh(ws, r, 6)
for col, color in [(3, FRENCH_BLUE), (4, GERMAN_GREEN), (5, AMERICAN_ORANGE)]:
    ws.cell(row=r, column=col).fill = PatternFill('solid', fgColor=color)

# Comparison metrics
metrics = [
    (14, "Cuota del primer mes (Bs.)",
     '=IFERROR(-PMT($C$7/12,$C$8,$C$6),0)',
     '=IFERROR($C$6/$C$8+$C$6*$C$7/12,0)',
     '=IFERROR($C$6*$C$7/12,0)',
     True),
    (15, "Cuota del ultimo mes (Bs.)",
     '=IFERROR(-PMT($C$7/12,$C$8,$C$6),0)',
     '=IFERROR($C$6/$C$8+($C$6/$C$8)*$C$7/12,0)',
     '=IFERROR($C$6+$C$6*$C$7/12,0)',
     True),
    (16, "Cuota promedio (Bs.)",
     '=IFERROR(-PMT($C$7/12,$C$8,$C$6),0)',
     '=IFERROR(($C$6/$C$8+$C$6*$C$7/12+$C$6/$C$8+($C$6/$C$8)*$C$7/12)/2,0)',
     '=IFERROR(($C$6*$C$7/12*($C$8-1)+$C$6+$C$6*$C$7/12)/$C$8,0)',
     True),
    (17, "Total a pagar (Bs.)",
     '=IFERROR(-PMT($C$7/12,$C$8,$C$6)*$C$8,0)',
     '=IFERROR($C$6+$C$6*$C$7/12*($C$8+1)/2,0)',
     '=IFERROR($C$6+$C$6*$C$7/12*$C$8,0)',
     True),
    (18, "Total intereses (Bs.)",
     '=IFERROR(C17-$C$6,0)',
     '=IFERROR(D17-$C$6,0)',
     '=IFERROR(E17-$C$6,0)',
     True),
    (19, "Relacion intereses/capital (%)",
     '=IFERROR(C18/$C$6,0)',
     '=IFERROR(D18/$C$6,0)',
     '=IFERROR(E18/$C$6,0)',
     False),
]

for (row, label, french, german, american, is_cur) in metrics:
    rh(ws, row, 30)
    lbl = ws.cell(row=row, column=2, value=label)
    style_label(lbl, bold=True)
    lbl.border = thin_border

    # French
    fc = ws.cell(row=row, column=3)
    fc.value = french
    style_output(fc, bold=True)
    fc.fill = PatternFill('solid', fgColor=FRENCH_BLUE_LIGHT)
    if is_cur:
        fmt_currency(fc)
    else:
        fmt_pct(fc)

    # German
    gc = ws.cell(row=row, column=4)
    gc.value = german
    style_output(gc, bold=True)
    gc.fill = PatternFill('solid', fgColor=GERMAN_GREEN_LIGHT)
    if is_cur:
        fmt_currency(gc)
    else:
        fmt_pct(gc)

    # American
    ac = ws.cell(row=row, column=5)
    ac.value = american
    style_output(ac, bold=True)
    ac.fill = PatternFill('solid', fgColor=AMERICAN_ORANGE_LIGHT)
    if is_cur:
        fmt_currency(ac)
    else:
        fmt_pct(ac)

    # Best (lowest for total/interest rows, fixed cuota advantage for French)
    best_c = ws.cell(row=row, column=6)
    if row in [17, 18]:  # total and interest: lower is better
        best_c.value = f'=IF(MIN(C{row},D{row},E{row})=C{row},"Frances",IF(MIN(C{row},D{row},E{row})=D{row},"Aleman","Americano"))'
    elif row == 14:  # first cuota: lower is better
        best_c.value = f'=IF(MIN(C{row},D{row},E{row})=C{row},"Frances",IF(MIN(C{row},D{row},E{row})=D{row},"Aleman","Americano"))'
    elif row == 19:
        best_c.value = f'=IF(MIN(C{row},D{row},E{row})=C{row},"Frances",IF(MIN(C{row},D{row},E{row})=D{row},"Aleman","Americano"))'
    else:
        best_c.value = f'=IF(MIN(C{row},D{row},E{row})=C{row},"Frances",IF(MIN(C{row},D{row},E{row})=D{row},"Aleman","Americano"))'
    best_c.font = Font(name='Calibri', bold=True, size=11, color=WINNER_TEAL)
    best_c.fill = PatternFill('solid', fgColor=WINNER_TEAL_LIGHT)
    best_c.alignment = Alignment(horizontal='center', vertical='center')
    best_c.border = thin_border

# Conditional formatting: highlight winner cells
for row in range(14, 20):
    for col in [3, 4, 5]:
        cell_ref = f'{get_column_letter(col)}{row}'
        if row in [17, 18, 19]:
            ws.conditional_formatting.add(cell_ref,
                FormulaRule(formula=[f'{cell_ref}=MIN($C{row},$D{row},$E{row})'],
                            fill=PatternFill('solid', fgColor=WINNER_TEAL_LIGHT),
                            font=Font(bold=True, color=WINNER_TEAL)))

rh(ws, 20, 10)

# ── WINNER BANNER ──
r = 21
rh(ws, r, 38)
ws.merge_cells(f'B{r}:G{r}')
winner = ws[f'B{r}']
winner.value = '=IFERROR("MENOR COSTO TOTAL: "&F17&" - Pagas Bs. "&TEXT(MIN(C17,D17,E17),"#,##0.00")&" | Ahorras Bs. "&TEXT(MAX(C17,D17,E17)-MIN(C17,D17,E17),"#,##0.00")&" vs el mas caro","Ingresa datos del prestamo")'
winner.font = Font(name='Calibri', bold=True, color=WINNER_TEAL, size=13)
winner.fill = PatternFill('solid', fgColor=WINNER_TEAL_LIGHT)
winner.alignment = Alignment(horizontal='center', vertical='center')
winner.border = Border(
    left=Side(style='medium', color=WINNER_TEAL),
    right=Side(style='medium', color=WINNER_TEAL),
    top=Side(style='medium', color=WINNER_TEAL),
    bottom=Side(style='medium', color=WINNER_TEAL))
fill_row(ws, r, range(3, 8), WINNER_TEAL_LIGHT)

rh(ws, 22, 10)

# ── METHOD DESCRIPTIONS ──
r = 23
rh(ws, r, 30)
ws.merge_cells(f'B{r}:G{r}')
sec = ws[f'B{r}']
sec.value = "COMO FUNCIONA CADA METODO"
style_section(sec, PURPLE)
fill_row(ws, r, range(3, 8), PURPLE)

descs = [
    (24, "FRANCES (Cuota Fija)", "La cuota es siempre igual. Al inicio pagas mas interes, al final mas capital. El mas comun en Bolivia.", FRENCH_BLUE, FRENCH_BLUE_LIGHT),
    (25, "ALEMAN (Amortizacion Fija)", "El capital que pagas es fijo cada mes. La cuota baja con el tiempo porque los intereses disminuyen.", GERMAN_GREEN, GERMAN_GREEN_LIGHT),
    (26, "AMERICANO (Solo Interes)", "Solo pagas intereses cada mes. Al final devuelves TODO el capital de golpe (pago balloon).", AMERICAN_ORANGE, AMERICAN_ORANGE_LIGHT),
]

for (row, title, desc, color, bg) in descs:
    rh(ws, row, 42)
    t = ws.cell(row=row, column=2, value=title)
    t.font = Font(name='Calibri', bold=True, color=color, size=11)
    t.fill = PatternFill('solid', fgColor=bg)
    t.border = thin_border
    t.alignment = Alignment(vertical='center')

    ws.merge_cells(f'C{row}:G{row}')
    d = ws.cell(row=row, column=3)
    d.value = desc
    d.font = Font(name='Calibri', size=10, color=GRAY_700)
    d.fill = PatternFill('solid', fgColor=bg)
    d.border = thin_border
    d.alignment = Alignment(wrap_text=True, vertical='center')

rh(ws, 27, 10)

# ── CHART: Total a Pagar comparison ──
bar1 = BarChart()
bar1.type = "col"
bar1.title = "Total a Pagar por Metodo (Bs.)"
bar1.y_axis.title = "Total (Bs.)"
bar1.style = 10
bar1.width = 20
bar1.height = 14

# Helper data for chart
ws.cell(row=70, column=2, value="Metodo")
ws.cell(row=70, column=3, value="Total a Pagar")
ws.cell(row=70, column=4, value="Total Intereses")
methods_names = ["Frances", "Aleman", "Americano"]
for i, name in enumerate(methods_names):
    ws.cell(row=71+i, column=2, value=name)
    ws.cell(row=71+i, column=3).value = f'={get_column_letter(3+i)}17'
    ws.cell(row=71+i, column=4).value = f'={get_column_letter(3+i)}18'
    fmt_currency(ws.cell(row=71+i, column=3))
    fmt_currency(ws.cell(row=71+i, column=4))

data1 = Reference(ws, min_col=3, min_row=70, max_row=73)
cats1 = Reference(ws, min_col=2, min_row=71, max_row=73)
bar1.add_data(data1, titles_from_data=True)
bar1.set_categories(cats1)

colors_bar = [FRENCH_BLUE, GERMAN_GREEN, AMERICAN_ORANGE]
for i in range(3):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = colors_bar[i]
    bar1.series[0].data_points.append(pt)

bar1.dataLabels = DataLabelList()
bar1.dataLabels.showVal = True
bar1.dataLabels.numFmt = '#,##0'
ws.add_chart(bar1, "B28")

# ── CHART: Stacked Capital vs Interest ──
bar2 = BarChart()
bar2.type = "col"
bar2.grouping = "stacked"
bar2.title = "Desglose: Capital vs Intereses"
bar2.y_axis.title = "Monto (Bs.)"
bar2.style = 10
bar2.width = 20
bar2.height = 14

# Helper for stacked
ws.cell(row=76, column=2, value="Metodo")
ws.cell(row=76, column=3, value="Capital")
ws.cell(row=76, column=4, value="Intereses")
for i, name in enumerate(methods_names):
    ws.cell(row=77+i, column=2, value=name)
    ws.cell(row=77+i, column=3).value = '=$C$6'
    ws.cell(row=77+i, column=4).value = f'={get_column_letter(3+i)}18'
    fmt_currency(ws.cell(row=77+i, column=3))
    fmt_currency(ws.cell(row=77+i, column=4))

cap_ref = Reference(ws, min_col=3, min_row=76, max_row=79)
int_ref = Reference(ws, min_col=4, min_row=76, max_row=79)
cats2 = Reference(ws, min_col=2, min_row=77, max_row=79)
bar2.add_data(cap_ref, titles_from_data=True)
bar2.add_data(int_ref, titles_from_data=True)
bar2.set_categories(cats2)
bar2.series[0].graphicalProperties.solidFill = SUCCESS
bar2.series[1].graphicalProperties.solidFill = ERROR_RED
ws.add_chart(bar2, "B44")


# ============================================================
# HELPER: Build amortization table sheet
# ============================================================
def build_amort_sheet(name, tab_color, method, method_label):
    ws_t = wb.create_sheet(name)
    ws_t.sheet_properties.tabColor = tab_color

    ws_t.column_dimensions['A'].width = 8
    ws_t.column_dimensions['B'].width = 14
    ws_t.column_dimensions['C'].width = 18
    ws_t.column_dimensions['D'].width = 18
    ws_t.column_dimensions['E'].width = 18
    ws_t.column_dimensions['F'].width = 18
    ws_t.column_dimensions['G'].width = 18
    ws_t.column_dimensions['H'].width = 14

    ws_t.freeze_panes = 'A5'

    # Header
    ws_t.merge_cells('A1:H1')
    h = ws_t['A1']
    h.value = f"TABLA DE AMORTIZACION - {method_label}"
    h.font = Font(name='Calibri', bold=True, color=GOLD, size=16)
    h.fill = PatternFill('solid', fgColor=DARK_BG)
    h.alignment = Alignment(horizontal='center', vertical='center')
    fill_row(ws_t, 1, range(1, 9), DARK_BG)
    rh(ws_t, 1, 38)

    # Summary row
    rh(ws_t, 2, 24)
    ws_t.cell(row=2, column=1, value="Capital:").font = Font(name='Calibri', bold=True, size=9)
    c2b = ws_t.cell(row=2, column=2)
    c2b.value = '=Calculadora!C6'
    c2b.font = Font(name='Calibri', bold=True, size=9, color=ACCENT_BLUE)
    c2b.number_format = '#,##0.00'

    ws_t.cell(row=2, column=3, value="Tasa:").font = Font(name='Calibri', bold=True, size=9)
    c2d = ws_t.cell(row=2, column=4)
    c2d.value = '=Calculadora!C7'
    c2d.font = Font(name='Calibri', bold=True, size=9, color=ACCENT_BLUE)
    c2d.number_format = '0.00%'

    ws_t.cell(row=2, column=5, value="Plazo:").font = Font(name='Calibri', bold=True, size=9)
    c2f = ws_t.cell(row=2, column=6)
    c2f.value = '=Calculadora!C8&" meses"'
    c2f.font = Font(name='Calibri', bold=True, size=9, color=ACCENT_BLUE)

    # Total interest and total paid
    rh(ws_t, 3, 24)
    ws_t.cell(row=3, column=1, value="Total Pagado:").font = Font(name='Calibri', bold=True, size=9)
    tp = ws_t.cell(row=3, column=2)
    tp.value = f'=IFERROR(SUM(D5:D{5+MAX_MONTHS-1}),0)'
    tp.font = Font(name='Calibri', bold=True, size=9, color=tab_color)
    tp.number_format = '#,##0.00'

    ws_t.cell(row=3, column=3, value="Total Interes:").font = Font(name='Calibri', bold=True, size=9)
    ti = ws_t.cell(row=3, column=4)
    ti.value = f'=IFERROR(SUM(E5:E{5+MAX_MONTHS-1}),0)'
    ti.font = Font(name='Calibri', bold=True, size=9, color=ERROR_RED)
    ti.number_format = '#,##0.00'

    ws_t.cell(row=3, column=5, value="Total Capital:").font = Font(name='Calibri', bold=True, size=9)
    tc = ws_t.cell(row=3, column=6)
    tc.value = f'=IFERROR(SUM(F5:F{5+MAX_MONTHS-1}),0)'
    tc.font = Font(name='Calibri', bold=True, size=9, color=SUCCESS)
    tc.number_format = '#,##0.00'

    # Column headers
    r = 4
    rh(ws_t, r, 26)
    headers = ['#', 'FECHA', 'SALDO INICIAL', 'CUOTA', 'INTERES', 'CAPITAL', 'SALDO FINAL', '% PAGADO']
    for i, hdr in enumerate(headers):
        c = ws_t.cell(row=r, column=i + 1, value=hdr)
        style_header(c, bg=CARD_BG, fg=GOLD, size=9)

    # Amortization rows
    for m in range(MAX_MONTHS):
        row = 5 + m
        rh(ws_t, row, 18)

        # Month number
        ws_t.cell(row=row, column=1, value=m + 1).alignment = Alignment(horizontal='center')
        ws_t.cell(row=row, column=1).font = Font(name='Calibri', size=8, color=GRAY_700)
        ws_t.cell(row=row, column=1).border = thin_border

        # Date
        dt = ws_t.cell(row=row, column=2)
        if m == 0:
            dt.value = f'=Calculadora!$C$9'
        else:
            dt.value = f'=IFERROR(IF(C{row}<=0.01,"",EDATE(B{row-1},1)),"")'
        dt.number_format = 'MMM-YYYY'
        dt.font = Font(name='Calibri', size=8, color=GRAY_700)
        dt.alignment = Alignment(horizontal='center')
        dt.border = thin_border

        # Starting balance
        sb = ws_t.cell(row=row, column=3)
        if m == 0:
            sb.value = '=IF(Calculadora!$C$6>0,Calculadora!$C$6,0)'
        else:
            sb.value = f'=IF(G{row-1}<=0.01,0,G{row-1})'
        style_output(sb)
        fmt_currency(sb)
        sb.font = Font(name='Calibri', size=8, color=GRAY_900)

        # Interest
        ie = ws_t.cell(row=row, column=5)
        ie.value = f'=IF(C{row}<=0.01,0,ROUND(C{row}*Calculadora!$C$7/12,2))'
        style_output(ie)
        fmt_currency(ie)
        ie.font = Font(name='Calibri', size=8, color=GRAY_900)
        ie.fill = PatternFill('solid', fgColor="FEE2E2")

        # Method-specific: Cuota and Capital
        cuota = ws_t.cell(row=row, column=4)
        capital = ws_t.cell(row=row, column=6)

        if method == "french":
            cuota.value = f'=IF(C{row}<=0.01,0,IFERROR(-PMT(Calculadora!$C$7/12,Calculadora!$C$8,Calculadora!$C$6),0))'
            capital.value = f'=IF(C{row}<=0.01,0,D{row}-E{row})'
        elif method == "german":
            capital.value = f'=IF(C{row}<=0.01,0,ROUND(Calculadora!$C$6/Calculadora!$C$8,2))'
            cuota.value = f'=IF(C{row}<=0.01,0,F{row}+E{row})'
        elif method == "american":
            if m == 0:
                # Not the last month: interest only
                cuota.value = f'=IF(C{row}<=0.01,0,IF(A{row}<Calculadora!$C$8,E{row},E{row}+C{row}))'
            else:
                cuota.value = f'=IF(C{row}<=0.01,0,IF(A{row}<Calculadora!$C$8,E{row},E{row}+C{row}))'
            capital.value = f'=IF(C{row}<=0.01,0,D{row}-E{row})'

        style_output(cuota)
        fmt_currency(cuota)
        cuota.font = Font(name='Calibri', size=8, color=GRAY_900)

        style_output(capital)
        fmt_currency(capital)
        capital.font = Font(name='Calibri', size=8, color=GRAY_900)
        capital.fill = PatternFill('solid', fgColor=LIGHT_GREEN)

        # Ending balance
        eb = ws_t.cell(row=row, column=7)
        eb.value = f'=IF(C{row}<=0.01,0,MAX(ROUND(C{row}-F{row},2),0))'
        style_output(eb)
        fmt_currency(eb)
        eb.font = Font(name='Calibri', size=8, color=GRAY_900)

        # % paid
        pp = ws_t.cell(row=row, column=8)
        pp.value = f'=IFERROR(IF(Calculadora!$C$6>0,1-(G{row}/Calculadora!$C$6),0),0)'
        style_output(pp)
        fmt_pct(pp)
        pp.font = Font(name='Calibri', size=8, color=GRAY_900)

    # Highlight zero balance rows
    ws_t.conditional_formatting.add(f'G5:G{5+MAX_MONTHS-1}',
        CellIsRule(operator='lessThanOrEqual', formula=['0.01'],
                  fill=PatternFill('solid', fgColor=WINNER_TEAL_LIGHT),
                  font=Font(bold=True, color=WINNER_TEAL)))

    # Data bar for progress
    ws_t.conditional_formatting.add(f'H5:H{5+MAX_MONTHS-1}',
        DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1,
                    color=tab_color))

    # Totals
    tot_row = 5 + MAX_MONTHS
    rh(ws_t, tot_row, 26)
    ws_t.cell(row=tot_row, column=1, value="TOTAL").font = Font(name='Calibri', bold=True, color=WHITE, size=9)
    ws_t.cell(row=tot_row, column=1).fill = PatternFill('solid', fgColor=tab_color)
    ws_t.cell(row=tot_row, column=1).alignment = Alignment(horizontal='center')
    for col in range(2, 9):
        ws_t.cell(row=tot_row, column=col).fill = PatternFill('solid', fgColor=tab_color)
    for col in [4, 5, 6]:
        c = ws_t.cell(row=tot_row, column=col)
        c.value = f'=SUM({get_column_letter(col)}5:{get_column_letter(col)}{5+MAX_MONTHS-1})'
        c.font = Font(name='Calibri', bold=True, color=WHITE, size=9)
        fmt_currency(c)

    # Charts
    chart_row = tot_row + 3

    # Line: balance over time
    line = LineChart()
    line.title = f"Evolucion del Saldo - {method_label}"
    line.y_axis.title = "Saldo (Bs.)"
    line.x_axis.title = "Mes"
    line.style = 10
    line.width = 26
    line.height = 14

    bal_ref = Reference(ws_t, min_col=7, min_row=4, max_row=5+MAX_MONTHS-1)
    m_cats = Reference(ws_t, min_col=1, min_row=5, max_row=5+MAX_MONTHS-1)
    line.add_data(bal_ref, titles_from_data=True)
    line.set_categories(m_cats)
    s = line.series[0]
    s.graphicalProperties.line.solidFill = tab_color
    s.graphicalProperties.line.width = 20000
    ws_t.add_chart(line, f"A{chart_row}")

    # Stacked area: capital vs interest
    bar_ci = BarChart()
    bar_ci.type = "col"
    bar_ci.grouping = "stacked"
    bar_ci.title = f"Capital vs Interes por Mes - {method_label}"
    bar_ci.y_axis.title = "Monto (Bs.)"
    bar_ci.style = 10
    bar_ci.width = 26
    bar_ci.height = 14

    # Show first N months only for readability
    show_months = min(MAX_MONTHS, 120)
    cap_d = Reference(ws_t, min_col=6, min_row=4, max_row=5+show_months-1)
    int_d = Reference(ws_t, min_col=5, min_row=4, max_row=5+show_months-1)
    m_cats2 = Reference(ws_t, min_col=1, min_row=5, max_row=5+show_months-1)
    bar_ci.add_data(cap_d, titles_from_data=True)
    bar_ci.add_data(int_d, titles_from_data=True)
    bar_ci.set_categories(m_cats2)
    bar_ci.series[0].graphicalProperties.solidFill = SUCCESS
    bar_ci.series[1].graphicalProperties.solidFill = ERROR_RED
    ws_t.add_chart(bar_ci, f"A{chart_row + 16}")

    return ws_t


# Build the 3 amortization sheets
ws_fr = build_amort_sheet("Tabla_Frances", FRENCH_BLUE, "french", "METODO FRANCES (Cuota Fija)")
ws_ge = build_amort_sheet("Tabla_Aleman", GERMAN_GREEN, "german", "METODO ALEMAN (Amortizacion Fija)")
ws_am = build_amort_sheet("Tabla_Americano", AMERICAN_ORANGE, "american", "METODO AMERICANO (Interes + Balloon)")


# ============================================================
# SHEET 5: COMPARACION (Visual comparison of all 3 methods)
# ============================================================
ws_comp = wb.create_sheet("Comparacion")
ws_comp.sheet_properties.tabColor = WINNER_TEAL

ws_comp.column_dimensions['A'].width = 2
ws_comp.column_dimensions['B'].width = 16
ws_comp.column_dimensions['C'].width = 18
ws_comp.column_dimensions['D'].width = 18
ws_comp.column_dimensions['E'].width = 18
ws_comp.column_dimensions['F'].width = 2

ws_comp.freeze_panes = 'B5'

# Header
ws_comp.merge_cells('B1:E1')
fill_row(ws_comp, 1, range(2, 6), DARK_BG)
rh(ws_comp, 1, 8)

ws_comp.merge_cells('B2:E2')
h = ws_comp['B2']
h.value = "COMPARACION VISUAL DE METODOS"
h.font = Font(name='Calibri', bold=True, color=GOLD, size=18)
h.fill = PatternFill('solid', fgColor=DARK_BG)
h.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws_comp, 2, range(2, 6), DARK_BG)
rh(ws_comp, 2, 40)

ws_comp.merge_cells('B3:E3')
sub = ws_comp['B3']
sub.value = "Evolucion de cuotas y saldos en los 3 metodos"
sub.font = Font(name='Calibri', color=GRAY_300, size=10, italic=True)
sub.fill = PatternFill('solid', fgColor=DARK_BG)
sub.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws_comp, 3, range(2, 6), DARK_BG)
rh(ws_comp, 3, 24)

rh(ws_comp, 4, 8)

# ── CUOTA EVOLUTION TABLE (yearly summary) ──
r = 5
ws_comp.merge_cells(f'B{r}:E{r}')
sec = ws_comp[f'B{r}']
sec.value = "CUOTA POR ANIO (Bs.)"
style_section(sec, CARD_BG, fg=GOLD)
fill_row(ws_comp, r, range(3, 6), CARD_BG)
rh(ws_comp, r, 28)

r = 6
rh(ws_comp, r, 24)
for i, hdr in enumerate(['', 'ANIO', 'FRANCES', 'ALEMAN', 'AMERICANO']):
    c = ws_comp.cell(row=r, column=i + 1, value=hdr)
    if i >= 1:
        style_header(c, bg=CARD_BG, fg=GOLD, size=9)

max_years = 30
for y in range(max_years):
    row = 7 + y
    rh(ws_comp, row, 20)

    # Year
    ws_comp.cell(row=row, column=2, value=y+1).alignment = Alignment(horizontal='center')
    ws_comp.cell(row=row, column=2).font = Font(name='Calibri', size=9)
    ws_comp.cell(row=row, column=2).border = thin_border

    month_row = 5 + y * 12  # first month of that year in amort tables

    # French (fixed cuota)
    fc = ws_comp.cell(row=row, column=3)
    fc.value = f'=IFERROR(IF(Tabla_Frances!C{month_row}>0.01,Tabla_Frances!D{month_row},""),"")'
    style_output(fc)
    fmt_currency(fc)
    fc.font = Font(name='Calibri', size=9, color=FRENCH_BLUE)

    # German (decreasing)
    gc = ws_comp.cell(row=row, column=4)
    gc.value = f'=IFERROR(IF(Tabla_Aleman!C{month_row}>0.01,Tabla_Aleman!D{month_row},""),"")'
    style_output(gc)
    fmt_currency(gc)
    gc.font = Font(name='Calibri', size=9, color=GERMAN_GREEN)

    # American
    ac = ws_comp.cell(row=row, column=5)
    ac.value = f'=IFERROR(IF(Tabla_Americano!C{month_row}>0.01,Tabla_Americano!D{month_row},""),"")'
    style_output(ac)
    fmt_currency(ac)
    ac.font = Font(name='Calibri', size=9, color=AMERICAN_ORANGE)

# Line chart: Cuota evolution
line_cuota = LineChart()
line_cuota.title = "Evolucion de la Cuota Mensual por Metodo"
line_cuota.y_axis.title = "Cuota (Bs.)"
line_cuota.x_axis.title = "Anio"
line_cuota.style = 10
line_cuota.width = 22
line_cuota.height = 14

fr_ref = Reference(ws_comp, min_col=3, min_row=6, max_row=7+max_years-1)
ge_ref = Reference(ws_comp, min_col=4, min_row=6, max_row=7+max_years-1)
am_ref = Reference(ws_comp, min_col=5, min_row=6, max_row=7+max_years-1)
yr_cats = Reference(ws_comp, min_col=2, min_row=7, max_row=7+max_years-1)

line_cuota.add_data(fr_ref, titles_from_data=True)
line_cuota.add_data(ge_ref, titles_from_data=True)
line_cuota.add_data(am_ref, titles_from_data=True)
line_cuota.set_categories(yr_cats)

line_cuota.series[0].graphicalProperties.line.solidFill = FRENCH_BLUE
line_cuota.series[0].graphicalProperties.line.width = 25000
line_cuota.series[1].graphicalProperties.line.solidFill = GERMAN_GREEN
line_cuota.series[1].graphicalProperties.line.width = 25000
line_cuota.series[2].graphicalProperties.line.solidFill = AMERICAN_ORANGE
line_cuota.series[2].graphicalProperties.line.width = 25000

ws_comp.add_chart(line_cuota, f"B{7+max_years+1}")

# ── SALDO EVOLUTION TABLE (yearly) ──
bal_start = 7 + max_years + 18
ws_comp.merge_cells(f'B{bal_start}:E{bal_start}')
sec2 = ws_comp[f'B{bal_start}']
sec2.value = "SALDO PENDIENTE POR ANIO (Bs.)"
style_section(sec2, CARD_BG, fg=GOLD)
fill_row(ws_comp, bal_start, range(3, 6), CARD_BG)
rh(ws_comp, bal_start, 28)

bal_start += 1
rh(ws_comp, bal_start, 24)
for i, hdr in enumerate(['', 'ANIO', 'FRANCES', 'ALEMAN', 'AMERICANO']):
    c = ws_comp.cell(row=bal_start, column=i + 1, value=hdr)
    if i >= 1:
        style_header(c, bg=CARD_BG, fg=GOLD, size=9)

for y in range(max_years):
    row = bal_start + 1 + y
    rh(ws_comp, row, 20)

    ws_comp.cell(row=row, column=2, value=y+1).alignment = Alignment(horizontal='center')
    ws_comp.cell(row=row, column=2).font = Font(name='Calibri', size=9)
    ws_comp.cell(row=row, column=2).border = thin_border

    month_row = 5 + (y + 1) * 12 - 1  # last month of that year

    fc = ws_comp.cell(row=row, column=3)
    fc.value = f'=IFERROR(Tabla_Frances!G{month_row},"")'
    style_output(fc)
    fmt_currency(fc)
    fc.font = Font(name='Calibri', size=9, color=FRENCH_BLUE)

    gc = ws_comp.cell(row=row, column=4)
    gc.value = f'=IFERROR(Tabla_Aleman!G{month_row},"")'
    style_output(gc)
    fmt_currency(gc)
    gc.font = Font(name='Calibri', size=9, color=GERMAN_GREEN)

    ac = ws_comp.cell(row=row, column=5)
    ac.value = f'=IFERROR(Tabla_Americano!G{month_row},"")'
    style_output(ac)
    fmt_currency(ac)
    ac.font = Font(name='Calibri', size=9, color=AMERICAN_ORANGE)

# Line chart: Balance evolution
line_bal = LineChart()
line_bal.title = "Evolucion del Saldo Pendiente por Metodo"
line_bal.y_axis.title = "Saldo (Bs.)"
line_bal.x_axis.title = "Anio"
line_bal.style = 10
line_bal.width = 22
line_bal.height = 14

fr_bal = Reference(ws_comp, min_col=3, min_row=bal_start, max_row=bal_start+max_years)
ge_bal = Reference(ws_comp, min_col=4, min_row=bal_start, max_row=bal_start+max_years)
am_bal = Reference(ws_comp, min_col=5, min_row=bal_start, max_row=bal_start+max_years)
yr_cats2 = Reference(ws_comp, min_col=2, min_row=bal_start+1, max_row=bal_start+max_years)

line_bal.add_data(fr_bal, titles_from_data=True)
line_bal.add_data(ge_bal, titles_from_data=True)
line_bal.add_data(am_bal, titles_from_data=True)
line_bal.set_categories(yr_cats2)

line_bal.series[0].graphicalProperties.line.solidFill = FRENCH_BLUE
line_bal.series[0].graphicalProperties.line.width = 25000
line_bal.series[1].graphicalProperties.line.solidFill = GERMAN_GREEN
line_bal.series[1].graphicalProperties.line.width = 25000
line_bal.series[2].graphicalProperties.line.solidFill = AMERICAN_ORANGE
line_bal.series[2].graphicalProperties.line.width = 25000

ws_comp.add_chart(line_bal, f"B{bal_start+max_years+2}")


# ============================================================
# SHEET 6: CONFIG
# ============================================================
ws_cfg = wb.create_sheet("Config")
ws_cfg.sheet_properties.tabColor = GRAY_500

ws_cfg.column_dimensions['A'].width = 25
ws_cfg.column_dimensions['B'].width = 55

ws_cfg.merge_cells('A1:B1')
h = ws_cfg['A1']
h.value = "CONFIGURACION"
h.font = Font(name='Calibri', bold=True, color=GOLD, size=16)
h.fill = PatternFill('solid', fgColor=DARK_BG)
h.alignment = Alignment(horizontal='center', vertical='center')
ws_cfg.cell(row=1, column=2).fill = PatternFill('solid', fgColor=DARK_BG)
rh(ws_cfg, 1, 36)

config_data = [
    ("Producto", "Amortizador de Deudas PRO"),
    ("Version", "v1.0.0"),
    ("Marca", "No Somos Ignorantes"),
    ("Precio", "Bs. 69"),
    ("Metodos", "Frances, Aleman, Americano"),
    ("Max plazo", "360 meses (30 anios)"),
    ("Proteccion", "nsi2024"),
    ("Contacto", "nosomosignorantes@gmail.com"),
]

for i, (key, val) in enumerate(config_data):
    row = 3 + i
    k = ws_cfg.cell(row=row, column=1, value=key)
    k.font = Font(name='Calibri', bold=True, size=11, color=GRAY_700)
    k.border = thin_border
    v = ws_cfg.cell(row=row, column=2, value=val)
    v.font = Font(name='Calibri', size=11, color=GRAY_900)
    v.border = thin_border

r = 13
ws_cfg.merge_cells(f'A{r}:B{r}')
inst = ws_cfg[f'A{r}']
inst.value = "INSTRUCCIONES DE USO"
inst.font = Font(name='Calibri', bold=True, color=GOLD, size=14)
inst.fill = PatternFill('solid', fgColor=DARK_BG)
inst.alignment = Alignment(horizontal='center')
ws_cfg.cell(row=r, column=2).fill = PatternFill('solid', fgColor=DARK_BG)
rh(ws_cfg, r, 32)

instructions = [
    "1. En la hoja 'Calculadora' ingresa: capital, tasa anual, plazo y fecha de inicio.",
    "2. La tabla de comparacion muestra los 3 metodos lado a lado automaticamente.",
    "3. Revisa cada hoja de tabla (Frances, Aleman, Americano) para el detalle mes a mes.",
    "4. La hoja 'Comparacion' muestra graficos de evolucion de cuotas y saldos.",
    "",
    "METODO FRANCES (el mas comun en Bolivia):",
    "- Cuota FIJA todos los meses.",
    "- Al inicio pagas mas interes, al final mas capital.",
    "- Ideal si quieres planificar un presupuesto fijo.",
    "",
    "METODO ALEMAN:",
    "- Amortizacion de capital FIJA (siempre pagas lo mismo de capital).",
    "- La cuota BAJA cada mes porque los intereses disminuyen.",
    "- Pagas MENOS intereses en total vs Frances.",
    "",
    "METODO AMERICANO:",
    "- Solo pagas intereses cada mes (cuota baja).",
    "- Al final del plazo devuelves TODO el capital de golpe.",
    "- Riesgoso: requiere disciplina para ahorrar el capital.",
    "- Pagas MAS intereses que cualquier otro metodo.",
]

for i, line in enumerate(instructions):
    row = 14 + i
    ws_cfg.merge_cells(f'A{row}:B{row}')
    c = ws_cfg[f'A{row}']
    c.value = line
    c.font = Font(name='Calibri', size=10, color=GRAY_700)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    rh(ws_cfg, row, 20)


# ============================================================
# SHEET PROTECTION
# ============================================================
# Calculadora: protect outputs, allow inputs
ws.protection.sheet = True
ws.protection.password = "nsi2024"
ws.protection.enable()
for row in [6, 7, 8, 9]:
    ws.cell(row=row, column=3).protection = ws.cell(row=row, column=3).protection.copy(locked=False)

# Amortization tables: fully protected
for ws_name in ["Tabla_Frances", "Tabla_Aleman", "Tabla_Americano", "Comparacion", "Config"]:
    sheet = wb[ws_name]
    sheet.protection.sheet = True
    sheet.protection.password = "nsi2024"
    sheet.protection.enable()

# ============================================================
# SAVE
# ============================================================
OUTPUT = "D:/Landing-Page_marketplace/excel_products/Amortizador_Deudas_PRO_NSI.xlsx"
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print("Sheets:", wb.sheetnames)
print("Done!")
