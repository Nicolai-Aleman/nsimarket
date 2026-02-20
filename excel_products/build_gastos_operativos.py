"""
Build: Control de Gastos Operativos (Bs. 59)
No Somos Ignorantes v1.0
Operating expenses tracker: Dashboard with KPIs, expense entry by category,
budget vs actual comparison, 12-month trend analysis.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint, SeriesLabel
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ============================================================
# COLORS
# ============================================================
GOLD = "D4AF37"
DARK_BG = "1A1A2E"
CARD_BG = "16213E"
ACCENT_BLUE = "0F3460"
WHITE = "FFFFFF"
BLACK = "000000"
LIGHT_YELLOW = "FFF9C4"
LIGHT_GREEN = "E8F5E9"
RED_CORAL = "E74C3C"
TURQUOISE = "16A085"
ORANGE = "E67E22"
PURPLE = "8E44AD"
LIGHT_GRAY = "BDC3C7"
GREEN_OK = "27AE60"
YELLOW_WARN = "F1C40F"
RED_ALERT = "E74C3C"
GRAY_100 = "F3F4F6"
GRAY_200 = "E5E7EB"
GRAY_300 = "D1D5DB"
GRAY_400 = "9CA3AF"
GRAY_500 = "6B7280"
GRAY_600 = "4B5563"
GRAY_700 = "374151"
GRAY_800 = "1F2937"
GRAY_900 = "111827"

# ============================================================
# STYLE HELPERS
# ============================================================
thin_border = Border(
    left=Side(style='thin', color=GRAY_300),
    right=Side(style='thin', color=GRAY_300),
    top=Side(style='thin', color=GRAY_300),
    bottom=Side(style='thin', color=GRAY_300)
)
gold_border = Border(
    left=Side(style='thin', color=GOLD),
    right=Side(style='thin', color=GOLD),
    top=Side(style='thin', color=GOLD),
    bottom=Side(style='thin', color=GOLD)
)

def style_title(cell, size=18):
    cell.font = Font(name='Calibri', bold=True, color=BLACK, size=size)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def style_kpi_number(cell, size=24):
    cell.font = Font(name='Calibri', bold=True, color=BLACK, size=size)
    cell.alignment = Alignment(horizontal='right', vertical='center')

def style_kpi_label(cell):
    cell.font = Font(name='Calibri', size=9, color="808080")
    cell.alignment = Alignment(horizontal='left', vertical='center')

def style_header(cell, bg=DARK_BG, fg=WHITE, size=10):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

def style_input(cell):
    cell.font = Font(name='Calibri', size=11, color="0000FF")
    cell.fill = PatternFill('solid', fgColor=LIGHT_YELLOW)
    cell.border = gold_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

def style_output(cell, bold=False):
    cell.font = Font(name='Calibri', size=11, color=GRAY_900, bold=bold)
    cell.fill = PatternFill('solid', fgColor=LIGHT_GREEN)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='right', vertical='center')

def style_banner(cell, bg_color, fg=WHITE, size=10):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg_color)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

def fill_white(ws, r1, r2, c1, c2):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=WHITE)

def unlock_cell(cell):
    cell.protection = cell.protection.copy(locked=False)

MONTHS_ES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
             "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

EXPENSE_CATEGORIES = [
    "Alquiler", "Servicios (luz/agua/gas)", "Internet/Telefono",
    "Sueldos y Salarios", "Seguro Social", "Material de Oficina",
    "Publicidad/Marketing", "Transporte/Combustible",
    "Mantenimiento", "Seguros", "Impuestos/Patentes",
    "Otros Gastos"
]

# ============================================================
# SHEET 1: DASHBOARD
# ============================================================
ws_dash = wb.active
ws_dash.title = "Dashboard"
ws_dash.sheet_properties.tabColor = RED_CORAL
ws_dash.sheet_view.showGridLines = False

dash_widths = {1: 3, 2: 22, 3: 16, 4: 16, 5: 3, 6: 22, 7: 16, 8: 16, 9: 3}
for c, w in dash_widths.items():
    ws_dash.column_dimensions[get_column_letter(c)].width = w
fill_white(ws_dash, 1, 55, 1, 9)

# Title
ws_dash.merge_cells('B2:H2')
c = ws_dash['B2']
style_title(c, 20)
c.value = "CONTROL DE GASTOS OPERATIVOS"
ws_dash.merge_cells('B3:H3')
c = ws_dash['B3']
c.font = Font(name='Calibri', size=10, color=GRAY_500)
c.value = "No Somos Ignorantes  |  v1.0  |  Resumen del periodo actual"

# Year selector
ws_dash['B4'].value = "Ano:"
ws_dash['B4'].font = Font(name='Calibri', size=10, bold=True, color=GRAY_700)
ws_dash['C4'].value = 2025
style_input(ws_dash['C4'])
unlock_cell(ws_dash['C4'])

ws_dash.row_dimensions[5].height = 6

# ── KPI 1: Total Gastos del Mes ──
ws_dash.merge_cells('B6:D6')
c = ws_dash['B6']
style_banner(c, RED_CORAL, WHITE, 10)
c.value = "GASTOS DEL MES ACTUAL (Bs.)"
ws_dash.merge_cells('B7:D7')
c = ws_dash['B7']
style_kpi_number(c, 26)
c.value = '=IFERROR(SUMPRODUCT((MONTH(Gastos!A4:A503)=MONTH(TODAY()))*(YEAR(Gastos!A4:A503)=YEAR(TODAY()))*Gastos!D4:D503),0)'
c.number_format = '#,##0.00'

ws_dash['B8'].value = "Presupuesto mes:"
style_kpi_label(ws_dash['B8'])
ws_dash['C8'].value = '=IFERROR(INDEX(Presupuesto!C5:N5,1,MONTH(TODAY())),0)'
ws_dash['C8'].number_format = '#,##0'
ws_dash['C8'].font = Font(name='Calibri', size=11, color=GRAY_700)

ws_dash['B9'].value = "Desviacion:"
style_kpi_label(ws_dash['B9'])
ws_dash.merge_cells('C9:D9')
c = ws_dash['C9']
c.value = '=IFERROR(B7-C8,0)'
c.number_format = '#,##0.00'
c.font = Font(name='Calibri', bold=True, size=14, color=BLACK)

# ── KPI 2: Gastos Acumulados Ano ──
ws_dash.merge_cells('F6:H6')
c = ws_dash['F6']
style_banner(c, TURQUOISE, WHITE, 10)
c.value = "GASTOS ACUMULADOS ANO (Bs.)"
ws_dash.merge_cells('F7:H7')
c = ws_dash['F7']
style_kpi_number(c, 26)
c.value = '=IFERROR(SUMPRODUCT((YEAR(Gastos!A4:A503)=C4)*Gastos!D4:D503),0)'
c.number_format = '#,##0.00'

ws_dash['F8'].value = "Presupuesto anual:"
style_kpi_label(ws_dash['F8'])
ws_dash['G8'].value = '=IFERROR(SUM(Presupuesto!C5:N5),0)'
ws_dash['G8'].number_format = '#,##0'
ws_dash['G8'].font = Font(name='Calibri', size=11, color=GRAY_700)

ws_dash['F9'].value = "% ejecutado:"
style_kpi_label(ws_dash['F9'])
ws_dash.merge_cells('G9:H9')
c = ws_dash['G9']
c.value = '=IFERROR(F7/G8,0)'
c.number_format = '0.0%'
c.font = Font(name='Calibri', bold=True, size=14, color=BLACK)

ws_dash.row_dimensions[10].height = 6

# ── KPI 3: Gasto Promedio Mensual ──
ws_dash.merge_cells('B11:D11')
c = ws_dash['B11']
style_banner(c, ORANGE, WHITE, 10)
c.value = "PROMEDIO MENSUAL (Bs.)"
ws_dash.merge_cells('B12:D12')
c = ws_dash['B12']
style_kpi_number(c, 26)
c.value = '=IFERROR(F7/MONTH(TODAY()),0)'
c.number_format = '#,##0.00'

# ── KPI 4: Categoria Mayor Gasto ──
ws_dash.merge_cells('F11:H11')
c = ws_dash['F11']
style_banner(c, PURPLE, WHITE, 10)
c.value = "CATEGORIA MAYOR GASTO"
ws_dash.merge_cells('F12:H12')
c = ws_dash['F12']
c.font = Font(name='Calibri', bold=True, size=16, color=BLACK)
c.alignment = Alignment(horizontal='center', vertical='center')
c.value = '=IFERROR(INDEX(Gastos!C4:C503,MATCH(MAX(Gastos!D4:D503),Gastos!D4:D503,0)),"--")'

ws_dash.row_dimensions[13].height = 6

# ── Monthly Trend Table (for chart) ──
ws_dash['B14'].value = "TENDENCIA MENSUAL"
ws_dash['B14'].font = Font(name='Calibri', bold=True, size=11, color=GRAY_700)

ws_dash['B15'].value = "Mes"
ws_dash['C15'].value = "Gasto Real"
ws_dash['D15'].value = "Presupuesto"
style_header(ws_dash['B15'], bg=GRAY_700)
style_header(ws_dash['C15'], bg=GRAY_700)
style_header(ws_dash['D15'], bg=GRAY_700)

for i, m in enumerate(MONTHS_ES):
    r = 16 + i
    ws_dash.cell(row=r, column=2).value = m
    ws_dash.cell(row=r, column=2).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws_dash.cell(row=r, column=2).border = thin_border

    ws_dash.cell(row=r, column=3).value = (
        f'=IFERROR(SUMPRODUCT((MONTH(Gastos!A4:A503)={i+1})'
        f'*(YEAR(Gastos!A4:A503)=C4)*Gastos!D4:D503),0)'
    )
    ws_dash.cell(row=r, column=3).number_format = '#,##0.00'
    ws_dash.cell(row=r, column=3).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws_dash.cell(row=r, column=3).border = thin_border

    # Budget from Presupuesto sheet
    col_letter_pres = get_column_letter(3 + i)  # C=Ene, D=Feb...N=Dic
    ws_dash.cell(row=r, column=4).value = f'=IFERROR(Presupuesto!{col_letter_pres}5,0)'
    ws_dash.cell(row=r, column=4).number_format = '#,##0.00'
    ws_dash.cell(row=r, column=4).font = Font(name='Calibri', size=10, color=LIGHT_GRAY)
    ws_dash.cell(row=r, column=4).border = thin_border

# Chart: Monthly trend (Actual vs Budget)
chart1 = BarChart()
chart1.type = "col"
chart1.grouping = "clustered"
chart1.style = 10
chart1.title = "Gastos Mensuales: Real vs Presupuesto"
chart1.y_axis.title = "Bs."
chart1.legend.position = 'b'

data_real = Reference(ws_dash, min_col=3, min_row=15, max_row=27)
data_pres = Reference(ws_dash, min_col=4, min_row=15, max_row=27)
cats_m = Reference(ws_dash, min_col=2, min_row=16, max_row=27)

chart1.add_data(data_real, titles_from_data=True)
chart1.add_data(data_pres, titles_from_data=True)
chart1.set_categories(cats_m)

chart1.series[0].graphicalProperties.solidFill = RED_CORAL
chart1.series[0].graphicalProperties.line.noFill = True
chart1.series[1].graphicalProperties.solidFill = LIGHT_GRAY
chart1.series[1].graphicalProperties.line.noFill = True

chart1.width = 22
chart1.height = 13
ws_dash.add_chart(chart1, "B29")

# ── Pie chart: Gastos por Categoria ──
ws_dash['F14'].value = "GASTOS POR CATEGORIA"
ws_dash['F14'].font = Font(name='Calibri', bold=True, size=11, color=GRAY_700)

ws_dash['F15'].value = "Categoria"
ws_dash['G15'].value = "Total (Bs.)"
style_header(ws_dash['F15'], bg=GRAY_700)
style_header(ws_dash['G15'], bg=GRAY_700)

for i, cat in enumerate(EXPENSE_CATEGORIES):
    r = 16 + i
    ws_dash.cell(row=r, column=6).value = cat
    ws_dash.cell(row=r, column=6).font = Font(name='Calibri', size=9, color=GRAY_700)
    ws_dash.cell(row=r, column=6).border = thin_border
    ws_dash.cell(row=r, column=7).value = (
        f'=IFERROR(SUMPRODUCT((Gastos!C4:C503=F{r})'
        f'*(YEAR(Gastos!A4:A503)=C4)*Gastos!D4:D503),0)'
    )
    ws_dash.cell(row=r, column=7).number_format = '#,##0.00'
    ws_dash.cell(row=r, column=7).font = Font(name='Calibri', size=9, color=GRAY_700)
    ws_dash.cell(row=r, column=7).border = thin_border

chart2 = PieChart()
chart2.title = "Distribucion de Gastos"
chart2.style = 10
data2 = Reference(ws_dash, min_col=7, min_row=15, max_row=27)
cats2 = Reference(ws_dash, min_col=6, min_row=16, max_row=27)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.width = 18
chart2.height = 13

pie_colors = [RED_CORAL, TURQUOISE, ORANGE, PURPLE, "3498DB", "1ABC9C",
              "F39C12", "2ECC71", "E74C3C", "9B59B6", LIGHT_GRAY, "34495E"]
for idx, color in enumerate(pie_colors):
    pt = DataPoint(idx=idx)
    pt.graphicalProperties.solidFill = color
    chart2.series[0].data_points.append(pt)

chart2.dataLabels = DataLabelList()
chart2.dataLabels.showPercent = True
chart2.dataLabels.showVal = False
ws_dash.add_chart(chart2, "F29")

# ============================================================
# SHEET 2: GASTOS (Expense Entry)
# ============================================================
ws_gas = wb.create_sheet("Gastos")
ws_gas.sheet_properties.tabColor = ORANGE
ws_gas.sheet_view.showGridLines = False

gas_widths = {1: 14, 2: 30, 3: 26, 4: 16, 5: 16, 6: 30}
for c, w in gas_widths.items():
    ws_gas.column_dimensions[get_column_letter(c)].width = w

ws_gas.merge_cells('A1:F1')
c = ws_gas['A1']
style_title(c, 16)
c.value = "REGISTRO DE GASTOS OPERATIVOS"
c.fill = PatternFill('solid', fgColor=WHITE)
ws_gas.row_dimensions[1].height = 35

ws_gas.merge_cells('A2:F2')
c = ws_gas['A2']
c.value = "Registra cada gasto con fecha, descripcion y categoria. Las celdas amarillas son editables."
c.font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)
c.fill = PatternFill('solid', fgColor=WHITE)

headers_gas = [
    ("A", "FECHA"),
    ("B", "DESCRIPCION"),
    ("C", "CATEGORIA"),
    ("D", "MONTO (Bs.)"),
    ("E", "METODO PAGO"),
    ("F", "OBSERVACION")
]
ws_gas.row_dimensions[3].height = 28
for col_letter, label in headers_gas:
    cell = ws_gas[f'{col_letter}3']
    style_header(cell, bg=DARK_BG, fg=GOLD, size=10)
    cell.value = label

# 500 rows for expenses
for row in range(4, 504):
    for col in range(1, 7):
        c = ws_gas.cell(row=row, column=col)
        style_input(c)
        unlock_cell(c)
        if col == 1:
            c.number_format = 'DD/MM/YYYY'
        elif col == 2 or col == 6:
            c.alignment = Alignment(horizontal='left', vertical='center')
        elif col == 4:
            c.number_format = '#,##0.00'

# Data validation: Category dropdown
cat_list = ",".join(EXPENSE_CATEGORIES)
dv_cat = DataValidation(type="list", formula1=f'"{cat_list}"', allow_blank=True)
dv_cat.prompt = "Selecciona la categoria del gasto"
dv_cat.promptTitle = "Categoria"
ws_gas.add_data_validation(dv_cat)
dv_cat.add('C4:C503')

# Data validation: Payment method
dv_pay = DataValidation(type="list", formula1='"Efectivo,Transferencia,Cheque,Tarjeta,Otro"', allow_blank=True)
dv_pay.prompt = "Selecciona el metodo de pago"
dv_pay.promptTitle = "Metodo de Pago"
ws_gas.add_data_validation(dv_pay)
dv_pay.add('E4:E503')

# Data validation: Amount >= 0
dv_pos = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
dv_pos.prompt = "Ingresa el monto del gasto"
dv_pos.promptTitle = "Monto"
dv_pos.error = "El monto no puede ser negativo"
dv_pos.errorTitle = "Monto invalido"
ws_gas.add_data_validation(dv_pos)
dv_pos.add('D4:D503')

# Sample data
from datetime import date
sample_expenses = [
    (date(2025, 1, 1), "Alquiler local enero", "Alquiler", 3500, "Transferencia", ""),
    (date(2025, 1, 5), "Luz y agua enero", "Servicios (luz/agua/gas)", 450, "Efectivo", ""),
    (date(2025, 1, 5), "Internet fibra optica", "Internet/Telefono", 250, "Transferencia", "Plan empresarial"),
    (date(2025, 1, 10), "Sueldos enero", "Sueldos y Salarios", 8000, "Transferencia", "3 empleados"),
    (date(2025, 1, 15), "Papeleria y toner", "Material de Oficina", 320, "Efectivo", ""),
    (date(2025, 1, 15), "Publicidad Facebook", "Publicidad/Marketing", 500, "Tarjeta", "Campana enero"),
    (date(2025, 1, 20), "Gasolina vehiculo", "Transporte/Combustible", 400, "Efectivo", ""),
    (date(2025, 1, 25), "Reparacion AC", "Mantenimiento", 600, "Cheque", "Aire acondicionado"),
    (date(2025, 2, 1), "Alquiler local febrero", "Alquiler", 3500, "Transferencia", ""),
    (date(2025, 2, 5), "Luz y agua febrero", "Servicios (luz/agua/gas)", 480, "Efectivo", ""),
]
for i, (dt, desc, cat, amount, pay, obs) in enumerate(sample_expenses):
    r = 4 + i
    ws_gas.cell(row=r, column=1).value = dt
    ws_gas.cell(row=r, column=2).value = desc
    ws_gas.cell(row=r, column=3).value = cat
    ws_gas.cell(row=r, column=4).value = amount
    ws_gas.cell(row=r, column=5).value = pay
    ws_gas.cell(row=r, column=6).value = obs

# Conditional formatting: Data bars on amounts
ws_gas.conditional_formatting.add('D4:D503',
    DataBarRule(start_type='min', end_type='max', color=RED_CORAL))

# ============================================================
# SHEET 3: PRESUPUESTO (Budget vs Actual)
# ============================================================
ws_pres = wb.create_sheet("Presupuesto")
ws_pres.sheet_properties.tabColor = TURQUOISE
ws_pres.sheet_view.showGridLines = False

pres_widths = {1: 4, 2: 28}
for c, w in pres_widths.items():
    ws_pres.column_dimensions[get_column_letter(c)].width = w
for i in range(12):
    ws_pres.column_dimensions[get_column_letter(3 + i)].width = 14
ws_pres.column_dimensions[get_column_letter(15)].width = 16  # O = Total

fill_white(ws_pres, 1, 60, 1, 16)

# Title
ws_pres.merge_cells('B1:O1')
c = ws_pres['B1']
style_title(c, 16)
c.value = "PRESUPUESTO vs REAL POR CATEGORIA"
ws_pres.row_dimensions[1].height = 35

# Year selector
ws_pres['B2'].value = "Ano:"
ws_pres['B2'].font = Font(name='Calibri', size=10, bold=True, color=GRAY_700)
ws_pres['C2'].value = '=Dashboard!C4'
ws_pres['C2'].font = Font(name='Calibri', size=11, bold=True, color=ACCENT_BLUE)

# ── Section 1: PRESUPUESTO (Budget) ──
ws_pres['A4'].value = ""
ws_pres.merge_cells('B4:O4')
c = ws_pres['B4']
style_banner(c, TURQUOISE, WHITE, 11)
c.value = "PRESUPUESTO MENSUAL (Bs.)"

# Month headers
ws_pres['B5'].value = "CATEGORIA"
style_header(ws_pres['B5'], bg=DARK_BG, fg=GOLD, size=9)
for i, m in enumerate(MONTHS_ES):
    cell = ws_pres.cell(row=5, column=3 + i)
    style_header(cell, bg=DARK_BG, fg=GOLD, size=9)
    cell.value = m.upper()
ws_pres.cell(row=5, column=15).value = "TOTAL"
style_header(ws_pres.cell(row=5, column=15), bg=DARK_BG, fg=GOLD, size=9)

# Budget rows (INPUT)
for i, cat in enumerate(EXPENSE_CATEGORIES):
    r = 6 + i
    ws_pres.cell(row=r, column=2).value = cat
    ws_pres.cell(row=r, column=2).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws_pres.cell(row=r, column=2).border = thin_border

    for m in range(12):
        c = ws_pres.cell(row=r, column=3 + m)
        style_input(c)
        c.number_format = '#,##0'
        unlock_cell(c)

    # Total
    c = ws_pres.cell(row=r, column=15)
    c.value = f'=SUM(C{r}:N{r})'
    c.number_format = '#,##0'
    style_output(c, bold=True)

# Total row
r_tot_budget = 6 + len(EXPENSE_CATEGORIES)
ws_pres.cell(row=r_tot_budget, column=2).value = "TOTAL PRESUPUESTO"
ws_pres.cell(row=r_tot_budget, column=2).font = Font(name='Calibri', size=10, bold=True, color=BLACK)
ws_pres.cell(row=r_tot_budget, column=2).border = thin_border
for m in range(12):
    c = ws_pres.cell(row=r_tot_budget, column=3 + m)
    c.value = f'=SUM({get_column_letter(3+m)}6:{get_column_letter(3+m)}{r_tot_budget-1})'
    c.number_format = '#,##0'
    style_output(c, bold=True)
c = ws_pres.cell(row=r_tot_budget, column=15)
c.value = f'=SUM(O6:O{r_tot_budget-1})'
c.number_format = '#,##0'
style_output(c, bold=True)

# Row for total budget per month (referenced by Dashboard) — this is row 5 total = r_tot_budget
# Actually Dashboard references row 5 for totals — let me add a clean total in row 5
# We'll fix Dashboard to point to r_tot_budget

# ── Section 2: GASTO REAL ──
r_real_start = r_tot_budget + 2
ws_pres.merge_cells(f'B{r_real_start}:O{r_real_start}')
c = ws_pres[f'B{r_real_start}']
style_banner(c, RED_CORAL, WHITE, 11)
c.value = "GASTO REAL MENSUAL (Bs.)"

r_real_header = r_real_start + 1
ws_pres.cell(row=r_real_header, column=2).value = "CATEGORIA"
style_header(ws_pres.cell(row=r_real_header, column=2), bg=DARK_BG, fg=GOLD, size=9)
for i, m in enumerate(MONTHS_ES):
    cell = ws_pres.cell(row=r_real_header, column=3 + i)
    style_header(cell, bg=DARK_BG, fg=GOLD, size=9)
    cell.value = m.upper()
ws_pres.cell(row=r_real_header, column=15).value = "TOTAL"
style_header(ws_pres.cell(row=r_real_header, column=15), bg=DARK_BG, fg=GOLD, size=9)

for i, cat in enumerate(EXPENSE_CATEGORIES):
    r = r_real_header + 1 + i
    ws_pres.cell(row=r, column=2).value = cat
    ws_pres.cell(row=r, column=2).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws_pres.cell(row=r, column=2).border = thin_border

    for m in range(12):
        c = ws_pres.cell(row=r, column=3 + m)
        c.value = (
            f'=IFERROR(SUMPRODUCT((Gastos!C4:C503="{cat}")'
            f'*(MONTH(Gastos!A4:A503)={m+1})'
            f'*(YEAR(Gastos!A4:A503)=C2)*Gastos!D4:D503),0)'
        )
        c.number_format = '#,##0'
        c.font = Font(name='Calibri', size=10, color=GRAY_700)
        c.border = thin_border

    c = ws_pres.cell(row=r, column=15)
    c.value = f'=SUM(C{r}:N{r})'
    c.number_format = '#,##0'
    style_output(c)

r_tot_real = r_real_header + 1 + len(EXPENSE_CATEGORIES)
ws_pres.cell(row=r_tot_real, column=2).value = "TOTAL REAL"
ws_pres.cell(row=r_tot_real, column=2).font = Font(name='Calibri', size=10, bold=True, color=BLACK)
ws_pres.cell(row=r_tot_real, column=2).border = thin_border
for m in range(12):
    c = ws_pres.cell(row=r_tot_real, column=3 + m)
    c.value = f'=SUM({get_column_letter(3+m)}{r_real_header+1}:{get_column_letter(3+m)}{r_tot_real-1})'
    c.number_format = '#,##0'
    style_output(c, bold=True)
c = ws_pres.cell(row=r_tot_real, column=15)
c.value = f'=SUM(O{r_real_header+1}:O{r_tot_real-1})'
c.number_format = '#,##0'
style_output(c, bold=True)

# ── Section 3: DESVIACION (Variance) ──
r_var_start = r_tot_real + 2
ws_pres.merge_cells(f'B{r_var_start}:O{r_var_start}')
c = ws_pres[f'B{r_var_start}']
style_banner(c, ORANGE, WHITE, 11)
c.value = "DESVIACION (Presupuesto - Real)"

r_var_header = r_var_start + 1
ws_pres.cell(row=r_var_header, column=2).value = "CATEGORIA"
style_header(ws_pres.cell(row=r_var_header, column=2), bg=DARK_BG, fg=GOLD, size=9)
for i, m in enumerate(MONTHS_ES):
    cell = ws_pres.cell(row=r_var_header, column=3 + i)
    style_header(cell, bg=DARK_BG, fg=GOLD, size=9)
    cell.value = m.upper()
ws_pres.cell(row=r_var_header, column=15).value = "TOTAL"
style_header(ws_pres.cell(row=r_var_header, column=15), bg=DARK_BG, fg=GOLD, size=9)

for i, cat in enumerate(EXPENSE_CATEGORIES):
    r = r_var_header + 1 + i
    r_budget = 6 + i
    r_actual = r_real_header + 1 + i
    ws_pres.cell(row=r, column=2).value = cat
    ws_pres.cell(row=r, column=2).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws_pres.cell(row=r, column=2).border = thin_border

    for m in range(12):
        col = get_column_letter(3 + m)
        c = ws_pres.cell(row=r, column=3 + m)
        c.value = f'={col}{r_budget}-{col}{r_actual}'
        c.number_format = '#,##0'
        c.font = Font(name='Calibri', size=10, color=GRAY_700)
        c.border = thin_border

    c = ws_pres.cell(row=r, column=15)
    c.value = f'=SUM(C{r}:N{r})'
    c.number_format = '#,##0'
    c.font = Font(name='Calibri', size=10, bold=True, color=GRAY_700)
    c.border = thin_border

# Conditional formatting: Negative variance = red, positive = green
var_range = f'C{r_var_header+1}:N{r_var_header+len(EXPENSE_CATEGORIES)}'
ws_pres.conditional_formatting.add(var_range,
    CellIsRule(operator='lessThan', formula=['0'],
              fill=PatternFill('solid', fgColor="FEE2E2"),
              font=Font(color=RED_ALERT, bold=True)))
ws_pres.conditional_formatting.add(var_range,
    CellIsRule(operator='greaterThanOrEqual', formula=['0'],
              fill=PatternFill('solid', fgColor="D1FAE5"),
              font=Font(color=GREEN_OK)))

# Fix Dashboard references to point to the correct total rows
# Dashboard!C8 should reference the budget total for current month
# We need to update Dashboard formulas

# Sample budget data
budget_data = {
    "Alquiler": 3500,
    "Servicios (luz/agua/gas)": 500,
    "Internet/Telefono": 250,
    "Sueldos y Salarios": 8000,
    "Seguro Social": 1200,
    "Material de Oficina": 400,
    "Publicidad/Marketing": 600,
    "Transporte/Combustible": 500,
    "Mantenimiento": 300,
    "Seguros": 200,
    "Impuestos/Patentes": 150,
    "Otros Gastos": 200,
}
for i, cat in enumerate(EXPENSE_CATEGORIES):
    r = 6 + i
    monthly = budget_data.get(cat, 0)
    for m in range(12):
        ws_pres.cell(row=r, column=3 + m).value = monthly

# ============================================================
# SHEET 4: TENDENCIAS (Historical Trends)
# ============================================================
ws_trend = wb.create_sheet("Tendencias")
ws_trend.sheet_properties.tabColor = PURPLE
ws_trend.sheet_view.showGridLines = False

trend_widths = {1: 4, 2: 28}
for c, w in trend_widths.items():
    ws_trend.column_dimensions[get_column_letter(c)].width = w
for i in range(12):
    ws_trend.column_dimensions[get_column_letter(3 + i)].width = 14

fill_white(ws_trend, 1, 40, 1, 15)

ws_trend.merge_cells('B1:N1')
c = ws_trend['B1']
style_title(c, 16)
c.value = "TENDENCIAS DE GASTOS"
ws_trend.row_dimensions[1].height = 35

ws_trend['B2'].value = "Evolucion mensual por categoria. Datos calculados automaticamente desde el registro de gastos."
ws_trend['B2'].font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)

# Headers
ws_trend['B4'].value = "CATEGORIA"
style_header(ws_trend['B4'], bg=DARK_BG, fg=GOLD, size=9)
for i, m in enumerate(MONTHS_ES):
    cell = ws_trend.cell(row=4, column=3 + i)
    style_header(cell, bg=DARK_BG, fg=GOLD, size=9)
    cell.value = m.upper()

for i, cat in enumerate(EXPENSE_CATEGORIES):
    r = 5 + i
    ws_trend.cell(row=r, column=2).value = cat
    ws_trend.cell(row=r, column=2).font = Font(name='Calibri', size=10, color=GRAY_700)
    ws_trend.cell(row=r, column=2).border = thin_border

    for m in range(12):
        c = ws_trend.cell(row=r, column=3 + m)
        c.value = (
            f'=IFERROR(SUMPRODUCT((Gastos!C4:C503="{cat}")'
            f'*(MONTH(Gastos!A4:A503)={m+1})'
            f'*(YEAR(Gastos!A4:A503)=Dashboard!C4)*Gastos!D4:D503),0)'
        )
        c.number_format = '#,##0'
        c.font = Font(name='Calibri', size=10, color=GRAY_700)
        c.border = thin_border

# Line chart: top 4 expense categories
chart3 = LineChart()
chart3.title = "Tendencia Mensual - Top Categorias"
chart3.style = 10
chart3.y_axis.title = "Bs."
chart3.x_axis.title = None
chart3.legend.position = 'b'
chart3.width = 24
chart3.height = 14

# Add data for first 4 categories (most common big expenses)
for i in range(4):
    data_ref = Reference(ws_trend, min_col=3, max_col=14, min_row=5+i)
    chart3.add_data(data_ref, from_rows=True)
    chart3.series[i].tx = SeriesLabel(v=ws_trend.cell(row=5+i, column=2).value)

cats_months = Reference(ws_trend, min_col=3, max_col=14, min_row=4)
chart3.set_categories(cats_months)

line_colors = [RED_CORAL, TURQUOISE, ORANGE, PURPLE]
for i, color in enumerate(line_colors):
    chart3.series[i].graphicalProperties.line.solidFill = color
    chart3.series[i].graphicalProperties.line.width = 25000

ws_trend.add_chart(chart3, "B18")

# ============================================================
# SHEET 5: CONFIG
# ============================================================
ws_conf = wb.create_sheet("Config")
ws_conf.sheet_properties.tabColor = GRAY_500
ws_conf.sheet_view.showGridLines = False

conf_widths = {1: 25, 2: 40}
for c, w in conf_widths.items():
    ws_conf.column_dimensions[get_column_letter(c)].width = w

fill_white(ws_conf, 1, 25, 1, 3)

ws_conf['A1'].value = "v1.0.0"
ws_conf['A1'].font = Font(name='Calibri', size=9, color=GRAY_400)

ws_conf.merge_cells('A3:B3')
style_title(ws_conf['A3'], 14)
ws_conf['A3'].value = "CONFIGURACION"

settings = [
    ("Producto", "Control de Gastos Operativos"),
    ("Version", "v1.0.0"),
    ("Marca", "No Somos Ignorantes"),
    ("Precio", "Bs. 59"),
    ("Proteccion", "nsi2024"),
    ("", ""),
    ("INSTRUCCIONES", ""),
    ("1.", "Las celdas AMARILLAS son editables."),
    ("2.", "En 'Gastos', registra cada gasto con fecha, descripcion y categoria."),
    ("3.", "En 'Presupuesto', define tu presupuesto mensual por categoria."),
    ("4.", "El Dashboard muestra KPIs, graficos de tendencia y comparacion real vs presupuesto."),
    ("5.", "La hoja 'Tendencias' muestra la evolucion mensual por categoria."),
    ("6.", "Para desbloquear: Revisar -> Desproteger hoja -> nsi2024"),
]
for i, (label, value) in enumerate(settings):
    r = 5 + i
    ws_conf.cell(row=r, column=1).value = label
    ws_conf.cell(row=r, column=1).font = Font(name='Calibri', size=10, bold=True, color=GRAY_700)
    ws_conf.cell(row=r, column=2).value = value
    ws_conf.cell(row=r, column=2).font = Font(name='Calibri', size=10, color=GRAY_600)

# ============================================================
# Fix Dashboard formulas to reference correct rows
# ============================================================
# Dashboard C8 (budget for current month) should reference the budget total row
ws_dash['C8'].value = f'=IFERROR(INDEX(Presupuesto!C{r_tot_budget}:N{r_tot_budget},1,MONTH(TODAY())),0)'

# Dashboard G8 (annual budget) should reference budget total
ws_dash['G8'].value = f'=IFERROR(Presupuesto!O{r_tot_budget},0)'

# Also fix Dashboard monthly trend to reference correct budget totals
for i in range(12):
    r = 16 + i
    col_letter_pres = get_column_letter(3 + i)
    ws_dash.cell(row=r, column=4).value = f'=IFERROR(Presupuesto!{col_letter_pres}{r_tot_budget},0)'

# ============================================================
# SHEET PROTECTION
# ============================================================
for ws in wb.worksheets:
    ws.protection.sheet = True
    ws.protection.password = "nsi2024"
    ws.protection.enable()

# Re-unlock all input cells
# Gastos: all columns
for row in range(4, 504):
    for col in range(1, 7):
        unlock_cell(ws_gas.cell(row=row, column=col))

# Presupuesto: budget cells only (rows 6 to 6+11, cols C-N)
for i in range(len(EXPENSE_CATEGORIES)):
    for m in range(12):
        unlock_cell(ws_pres.cell(row=6 + i, column=3 + m))

# Dashboard: year selector
unlock_cell(ws_dash['C4'])

# ============================================================
# SAVE
# ============================================================
output_path = r"D:\Landing-Page_marketplace\excel_products\Control_Gastos_Operativos_NSI.xlsx"
wb.save(output_path)
print(f"[OK] Saved: {output_path}")

# Verification
from openpyxl import load_workbook
wb2 = load_workbook(output_path)
formula_count = 0
for ws in wb2.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formula_count += 1

print(f"[OK] Sheets: {wb2.sheetnames}")
print(f"[OK] Total formulas: {formula_count}")
print(f"[OK] Charts: Dashboard={len(ws_dash._charts)}, Tendencias={len(ws_trend._charts)}")
print(f"[OK] Categories: {len(EXPENSE_CATEGORIES)}")
print(f"[OK] Budget total row: {r_tot_budget}, Real total row: {r_tot_real}")
