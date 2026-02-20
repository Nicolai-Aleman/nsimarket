"""
Build: Presupuesto 50/30/20 - FREE Lead Magnet
No Somos Ignorantes v1.0
Professional Excel with charts, conditional formatting, data validation, protection.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers, NamedStyle
)
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

wb = Workbook()

# ============================================================
# COLOR PALETTE
# ============================================================
GOLD = "D4AF37"
DARK_BG = "1A1A2E"
CARD_BG = "16213E"
ACCENT_BLUE = "0F3460"
WHITE = "FFFFFF"
BLACK = "000000"
LIGHT_YELLOW = "FFF9C4"
LIGHT_GREEN = "E8F5E9"
SUCCESS = "10B981"
WARNING = "F59E0B"
ERROR = "EF4444"
INFO_BLUE = "3B82F6"
GRAY_100 = "F3F4F6"
GRAY_200 = "E5E7EB"
GRAY_300 = "D1D5DB"
GRAY_500 = "6B7280"
GRAY_700 = "374151"
GRAY_900 = "111827"
NECESIDADES_COLOR = "3B82F6"   # Blue
DESEOS_COLOR = "F59E0B"        # Amber
AHORRO_COLOR = "10B981"        # Green
NECESIDADES_LIGHT = "DBEAFE"
DESEOS_LIGHT = "FEF3C7"
AHORRO_LIGHT = "D1FAE5"

# ============================================================
# STYLE HELPERS
# ============================================================
thin_border = Border(
    left=Side(style='thin', color=GRAY_300),
    right=Side(style='thin', color=GRAY_300),
    top=Side(style='thin', color=GRAY_300),
    bottom=Side(style='thin', color=GRAY_300)
)

gold_border = Border(
    left=Side(style='thin', color=GOLD),
    right=Side(style='thin', color=GOLD),
    top=Side(style='thin', color=GOLD),
    bottom=Side(style='thin', color=GOLD)
)

thick_bottom = Border(bottom=Side(style='medium', color=GRAY_700))

def style_header(cell, bg=DARK_BG, fg=WHITE, size=12):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')

def style_subheader(cell, bg=ACCENT_BLUE, fg=WHITE):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=11)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def style_input(cell):
    cell.font = Font(name='Calibri', size=11, color="0000FF")
    cell.fill = PatternFill('solid', fgColor=LIGHT_YELLOW)
    cell.border = gold_border
    cell.alignment = Alignment(horizontal='right', vertical='center')

def style_output(cell, bold=False):
    cell.font = Font(name='Calibri', size=11, color=GRAY_900, bold=bold)
    cell.fill = PatternFill('solid', fgColor=LIGHT_GREEN)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='right', vertical='center')

def style_label(cell, bold=False, indent=0):
    cell.font = Font(name='Calibri', size=11, color=GRAY_700, bold=bold)
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=indent)

def style_category_header(cell, color, text_color=WHITE):
    cell.font = Font(name='Calibri', bold=True, color=text_color, size=11)
    cell.fill = PatternFill('solid', fgColor=color)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def apply_currency_format(cell):
    cell.number_format = '#,##0.00'

def apply_pct_format(cell):
    cell.number_format = '0.0%'

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height


# ============================================================
# SHEET 1: PRESUPUESTO (Main Budget)
# ============================================================
ws = wb.active
ws.title = "Presupuesto"
ws.sheet_properties.tabColor = GOLD

# Column widths
ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 22
ws.column_dimensions['G'].width = 2

# Freeze panes
ws.freeze_panes = 'B7'

# ── HEADER BANNER ──
ws.merge_cells('B1:F1')
set_row_height(ws, 1, 8)

ws.merge_cells('B2:F2')
h = ws['B2']
h.value = "PRESUPUESTO 50/30/20"
h.font = Font(name='Calibri', bold=True, color=GOLD, size=20)
h.fill = PatternFill('solid', fgColor=DARK_BG)
h.alignment = Alignment(horizontal='center', vertical='center')
for col in range(2, 7):
    c = ws.cell(row=2, column=col)
    c.fill = PatternFill('solid', fgColor=DARK_BG)
set_row_height(ws, 2, 42)

ws.merge_cells('B3:F3')
sub = ws['B3']
sub.value = "No Somos Ignorantes  |  v1.0  |  Tu primer paso hacia la libertad financiera"
sub.font = Font(name='Calibri', color=GRAY_300, size=10, italic=True)
sub.fill = PatternFill('solid', fgColor=DARK_BG)
sub.alignment = Alignment(horizontal='center', vertical='center')
for col in range(2, 7):
    c = ws.cell(row=3, column=col)
    c.fill = PatternFill('solid', fgColor=DARK_BG)
set_row_height(ws, 3, 24)

# Spacer row
set_row_height(ws, 4, 8)

# ── INCOME INPUT ──
set_row_height(ws, 5, 32)
ws.merge_cells('B5:C5')
lbl = ws['B5']
lbl.value = "TU INGRESO MENSUAL (Bs.)"
lbl.font = Font(name='Calibri', bold=True, color=GRAY_900, size=13)
lbl.alignment = Alignment(horizontal='left', vertical='center')

ws.merge_cells('D5:E5')
inc = ws['D5']
inc.value = 5000
style_input(inc)
inc.font = Font(name='Calibri', size=14, bold=True, color="0000FF")
apply_currency_format(inc)
# Data validation: positive number
dv_income = DataValidation(type="decimal", operator="greaterThan", formula1=0)
dv_income.error = "Ingresa un monto mayor a 0"
dv_income.errorTitle = "Monto inválido"
dv_income.prompt = "Ingresa tu ingreso mensual en Bolivianos"
dv_income.promptTitle = "Ingreso Mensual"
ws.add_data_validation(dv_income)
dv_income.add('D5')

# Spacer
set_row_height(ws, 6, 6)

# ── COLUMN HEADERS ──
row = 7
set_row_height(ws, row, 28)
headers = ['', 'CATEGORÍA', 'MONTO (Bs.)', '% DEL INGRESO', 'LÍMITE (Bs.)', 'ESTADO']
for i, hdr in enumerate(headers):
    c = ws.cell(row=row, column=i+1, value=hdr)
    if i >= 1:
        style_header(c, bg=CARD_BG, fg=GOLD, size=10)
        c.border = Border(bottom=Side(style='medium', color=GOLD))

# ============================================================
# 50% NECESIDADES SECTION
# ============================================================
row = 8
set_row_height(ws, row, 30)
ws.merge_cells(f'B{row}:F{row}')
sec = ws[f'B{row}']
sec.value = "50% NECESIDADES — Lo que necesitas para vivir"
style_category_header(sec, NECESIDADES_COLOR)
for col in range(3, 7):
    ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor=NECESIDADES_COLOR)

necesidades_items = [
    "Vivienda (alquiler/cuota)", "Servicios básicos (luz, agua, gas)",
    "Alimentación y mercado", "Transporte",
    "Seguros (salud, vida)", "Educación obligatoria",
    "Medicinas / Salud", "Teléfono / Internet básico"
]

nec_start = 9
for i, item in enumerate(necesidades_items):
    r = nec_start + i
    set_row_height(ws, r, 26)
    # Label
    lbl = ws.cell(row=r, column=2, value=item)
    style_label(lbl, indent=1)
    lbl.fill = PatternFill('solid', fgColor=NECESIDADES_LIGHT)
    lbl.border = thin_border
    # Input amount
    amt = ws.cell(row=r, column=3)
    if i == 0:
        amt.value = 1200
    elif i == 1:
        amt.value = 200
    elif i == 2:
        amt.value = 800
    elif i == 3:
        amt.value = 300
    else:
        amt.value = 0
    style_input(amt)
    apply_currency_format(amt)
    # % of income
    pct = ws.cell(row=r, column=4)
    pct.value = f'=IF($D$5>0,C{r}/$D$5,0)'
    style_output(pct)
    apply_pct_format(pct)
    pct.fill = PatternFill('solid', fgColor=NECESIDADES_LIGHT)
    pct.border = thin_border

# Necesidades SUBTOTAL row
nec_end = nec_start + len(necesidades_items) - 1
nec_total_row = nec_end + 1
set_row_height(ws, nec_total_row, 30)
lbl = ws.cell(row=nec_total_row, column=2, value="TOTAL NECESIDADES")
lbl.font = Font(name='Calibri', bold=True, color=WHITE, size=11)
lbl.fill = PatternFill('solid', fgColor=NECESIDADES_COLOR)
lbl.border = thick_bottom

total_c = ws.cell(row=nec_total_row, column=3)
total_c.value = f'=SUM(C{nec_start}:C{nec_end})'
total_c.font = Font(name='Calibri', bold=True, color=WHITE, size=11)
total_c.fill = PatternFill('solid', fgColor=NECESIDADES_COLOR)
apply_currency_format(total_c)
total_c.border = thick_bottom

pct_c = ws.cell(row=nec_total_row, column=4)
pct_c.value = f'=IF($D$5>0,C{nec_total_row}/$D$5,0)'
pct_c.font = Font(name='Calibri', bold=True, color=WHITE, size=11)
pct_c.fill = PatternFill('solid', fgColor=NECESIDADES_COLOR)
apply_pct_format(pct_c)
pct_c.border = thick_bottom

limit_c = ws.cell(row=nec_total_row, column=5)
limit_c.value = '=$D$5*0.5'
limit_c.font = Font(name='Calibri', bold=True, color=WHITE, size=11)
limit_c.fill = PatternFill('solid', fgColor=NECESIDADES_COLOR)
apply_currency_format(limit_c)
limit_c.border = thick_bottom

status_c = ws.cell(row=nec_total_row, column=6)
status_c.value = f'=IF(C{nec_total_row}<=E{nec_total_row},"DENTRO DEL LIMITE","EXCEDIDO")'
status_c.font = Font(name='Calibri', bold=True, color=WHITE, size=11)
status_c.fill = PatternFill('solid', fgColor=NECESIDADES_COLOR)
status_c.alignment = Alignment(horizontal='center', vertical='center')
status_c.border = thick_bottom

# Remaining row
nec_remaining = nec_total_row + 1
set_row_height(ws, nec_remaining, 22)
ws.cell(row=nec_remaining, column=2, value="Disponible restante").font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)
rem = ws.cell(row=nec_remaining, column=5)
rem.value = f'=E{nec_total_row}-C{nec_total_row}'
rem.font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)
apply_currency_format(rem)

# Spacer
spacer1 = nec_remaining + 1
set_row_height(ws, spacer1, 8)

# ============================================================
# 30% DESEOS SECTION
# ============================================================
des_header = spacer1 + 1
set_row_height(ws, des_header, 30)
ws.merge_cells(f'B{des_header}:F{des_header}')
sec = ws[f'B{des_header}']
sec.value = "30% DESEOS — Lo que te hace feliz"
style_category_header(sec, DESEOS_COLOR, text_color=GRAY_900)
for col in range(3, 7):
    ws.cell(row=des_header, column=col).fill = PatternFill('solid', fgColor=DESEOS_COLOR)

deseos_items = [
    "Entretenimiento / Salidas", "Restaurantes / Comida fuera",
    "Compras personales (ropa, etc.)", "Streaming / Suscripciones",
    "Hobbies / Deportes", "Regalos",
    "Viajes / Vacaciones", "Otros deseos"
]

des_start = des_header + 1
for i, item in enumerate(deseos_items):
    r = des_start + i
    set_row_height(ws, r, 26)
    lbl = ws.cell(row=r, column=2, value=item)
    style_label(lbl, indent=1)
    lbl.fill = PatternFill('solid', fgColor=DESEOS_LIGHT)
    lbl.border = thin_border
    amt = ws.cell(row=r, column=3)
    if i == 0:
        amt.value = 300
    elif i == 1:
        amt.value = 200
    elif i == 2:
        amt.value = 150
    elif i == 3:
        amt.value = 50
    else:
        amt.value = 0
    style_input(amt)
    apply_currency_format(amt)
    pct = ws.cell(row=r, column=4)
    pct.value = f'=IF($D$5>0,C{r}/$D$5,0)'
    style_output(pct)
    apply_pct_format(pct)
    pct.fill = PatternFill('solid', fgColor=DESEOS_LIGHT)
    pct.border = thin_border

des_end = des_start + len(deseos_items) - 1
des_total_row = des_end + 1
set_row_height(ws, des_total_row, 30)
lbl = ws.cell(row=des_total_row, column=2, value="TOTAL DESEOS")
lbl.font = Font(name='Calibri', bold=True, color=GRAY_900, size=11)
lbl.fill = PatternFill('solid', fgColor=DESEOS_COLOR)
lbl.border = thick_bottom

total_c = ws.cell(row=des_total_row, column=3)
total_c.value = f'=SUM(C{des_start}:C{des_end})'
total_c.font = Font(name='Calibri', bold=True, color=GRAY_900, size=11)
total_c.fill = PatternFill('solid', fgColor=DESEOS_COLOR)
apply_currency_format(total_c)
total_c.border = thick_bottom

pct_c = ws.cell(row=des_total_row, column=4)
pct_c.value = f'=IF($D$5>0,C{des_total_row}/$D$5,0)'
pct_c.font = Font(name='Calibri', bold=True, color=GRAY_900, size=11)
pct_c.fill = PatternFill('solid', fgColor=DESEOS_COLOR)
apply_pct_format(pct_c)
pct_c.border = thick_bottom

limit_c = ws.cell(row=des_total_row, column=5)
limit_c.value = '=$D$5*0.3'
limit_c.font = Font(name='Calibri', bold=True, color=GRAY_900, size=11)
limit_c.fill = PatternFill('solid', fgColor=DESEOS_COLOR)
apply_currency_format(limit_c)
limit_c.border = thick_bottom

status_c = ws.cell(row=des_total_row, column=6)
status_c.value = f'=IF(C{des_total_row}<=E{des_total_row},"DENTRO DEL LIMITE","EXCEDIDO")'
status_c.font = Font(name='Calibri', bold=True, color=GRAY_900, size=11)
status_c.fill = PatternFill('solid', fgColor=DESEOS_COLOR)
status_c.alignment = Alignment(horizontal='center', vertical='center')
status_c.border = thick_bottom

des_remaining = des_total_row + 1
set_row_height(ws, des_remaining, 22)
ws.cell(row=des_remaining, column=2, value="Disponible restante").font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)
rem = ws.cell(row=des_remaining, column=5)
rem.value = f'=E{des_total_row}-C{des_total_row}'
rem.font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)
apply_currency_format(rem)

spacer2 = des_remaining + 1
set_row_height(ws, spacer2, 8)

# ============================================================
# 20% AHORRO SECTION
# ============================================================
aho_header = spacer2 + 1
set_row_height(ws, aho_header, 30)
ws.merge_cells(f'B{aho_header}:F{aho_header}')
sec = ws[f'B{aho_header}']
sec.value = "20% AHORRO E INVERSIONES — Tu futuro"
style_category_header(sec, AHORRO_COLOR)
for col in range(3, 7):
    ws.cell(row=aho_header, column=col).fill = PatternFill('solid', fgColor=AHORRO_COLOR)

ahorro_items = [
    "Fondo de emergencia", "Inversiones / Ahorro largo plazo",
    "Metas de ahorro (viaje, auto, etc.)", "Pago extra de deudas",
    "Educación / Cursos", "Retiro / Jubilación"
]

aho_start = aho_header + 1
for i, item in enumerate(ahorro_items):
    r = aho_start + i
    set_row_height(ws, r, 26)
    lbl = ws.cell(row=r, column=2, value=item)
    style_label(lbl, indent=1)
    lbl.fill = PatternFill('solid', fgColor=AHORRO_LIGHT)
    lbl.border = thin_border
    amt = ws.cell(row=r, column=3)
    if i == 0:
        amt.value = 500
    elif i == 1:
        amt.value = 300
    elif i == 2:
        amt.value = 200
    else:
        amt.value = 0
    style_input(amt)
    apply_currency_format(amt)
    pct = ws.cell(row=r, column=4)
    pct.value = f'=IF($D$5>0,C{r}/$D$5,0)'
    style_output(pct)
    apply_pct_format(pct)
    pct.fill = PatternFill('solid', fgColor=AHORRO_LIGHT)
    pct.border = thin_border

aho_end = aho_start + len(ahorro_items) - 1
aho_total_row = aho_end + 1
set_row_height(ws, aho_total_row, 30)
lbl = ws.cell(row=aho_total_row, column=2, value="TOTAL AHORRO")
lbl.font = Font(name='Calibri', bold=True, color=WHITE, size=11)
lbl.fill = PatternFill('solid', fgColor=AHORRO_COLOR)
lbl.border = thick_bottom

total_c = ws.cell(row=aho_total_row, column=3)
total_c.value = f'=SUM(C{aho_start}:C{aho_end})'
total_c.font = Font(name='Calibri', bold=True, color=WHITE, size=11)
total_c.fill = PatternFill('solid', fgColor=AHORRO_COLOR)
apply_currency_format(total_c)
total_c.border = thick_bottom

pct_c = ws.cell(row=aho_total_row, column=4)
pct_c.value = f'=IF($D$5>0,C{aho_total_row}/$D$5,0)'
pct_c.font = Font(name='Calibri', bold=True, color=WHITE, size=11)
pct_c.fill = PatternFill('solid', fgColor=AHORRO_COLOR)
apply_pct_format(pct_c)
pct_c.border = thick_bottom

limit_c = ws.cell(row=aho_total_row, column=5)
limit_c.value = '=$D$5*0.2'
limit_c.font = Font(name='Calibri', bold=True, color=WHITE, size=11)
limit_c.fill = PatternFill('solid', fgColor=AHORRO_COLOR)
apply_currency_format(limit_c)
limit_c.border = thick_bottom

status_c = ws.cell(row=aho_total_row, column=6)
status_c.value = f'=IF(C{aho_total_row}>=E{aho_total_row},"META ALCANZADA","AUN FALTA")'
status_c.font = Font(name='Calibri', bold=True, color=WHITE, size=11)
status_c.fill = PatternFill('solid', fgColor=AHORRO_COLOR)
status_c.alignment = Alignment(horizontal='center', vertical='center')
status_c.border = thick_bottom

aho_remaining = aho_total_row + 1
set_row_height(ws, aho_remaining, 22)
ws.cell(row=aho_remaining, column=2, value="Te falta ahorrar").font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)
rem = ws.cell(row=aho_remaining, column=5)
rem.value = f'=MAX(0,E{aho_total_row}-C{aho_total_row})'
rem.font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)
apply_currency_format(rem)

# ============================================================
# GRAND TOTAL SECTION
# ============================================================
spacer3 = aho_remaining + 1
set_row_height(ws, spacer3, 12)

gt_row = spacer3 + 1
set_row_height(ws, gt_row, 36)
ws.merge_cells(f'B{gt_row}:B{gt_row}')
lbl = ws.cell(row=gt_row, column=2, value="TOTAL GASTOS + AHORRO")
lbl.font = Font(name='Calibri', bold=True, color=WHITE, size=13)
lbl.fill = PatternFill('solid', fgColor=GRAY_900)
lbl.border = Border(bottom=Side(style='thick', color=GOLD))

gt_total = ws.cell(row=gt_row, column=3)
gt_total.value = f'=C{nec_total_row}+C{des_total_row}+C{aho_total_row}'
gt_total.font = Font(name='Calibri', bold=True, color=GOLD, size=13)
gt_total.fill = PatternFill('solid', fgColor=GRAY_900)
apply_currency_format(gt_total)
gt_total.border = Border(bottom=Side(style='thick', color=GOLD))

gt_pct = ws.cell(row=gt_row, column=4)
gt_pct.value = f'=IF($D$5>0,C{gt_row}/$D$5,0)'
gt_pct.font = Font(name='Calibri', bold=True, color=GOLD, size=13)
gt_pct.fill = PatternFill('solid', fgColor=GRAY_900)
apply_pct_format(gt_pct)
gt_pct.border = Border(bottom=Side(style='thick', color=GOLD))

gt_limit = ws.cell(row=gt_row, column=5)
gt_limit.value = '=$D$5'
gt_limit.font = Font(name='Calibri', bold=True, color=GOLD, size=13)
gt_limit.fill = PatternFill('solid', fgColor=GRAY_900)
apply_currency_format(gt_limit)
gt_limit.border = Border(bottom=Side(style='thick', color=GOLD))

gt_status = ws.cell(row=gt_row, column=6)
gt_status.value = f'=IF(C{gt_row}<=E{gt_row},"PRESUPUESTO OK","GASTAS DE MAS")'
gt_status.font = Font(name='Calibri', bold=True, color=GOLD, size=13)
gt_status.fill = PatternFill('solid', fgColor=GRAY_900)
gt_status.alignment = Alignment(horizontal='center', vertical='center')
gt_status.border = Border(bottom=Side(style='thick', color=GOLD))

# Sobrante / Deficit row
sob_row = gt_row + 1
set_row_height(ws, sob_row, 28)
ws.cell(row=sob_row, column=2, value="SOBRANTE / (DEFICIT)").font = Font(name='Calibri', bold=True, color=GRAY_700, size=11)
sob = ws.cell(row=sob_row, column=3)
sob.value = f'=$D$5-C{gt_row}'
sob.font = Font(name='Calibri', bold=True, size=12)
apply_currency_format(sob)

# ── CONDITIONAL FORMATTING ──
# Status columns: green for OK, red for exceeded
for status_row in [nec_total_row, des_total_row]:
    ws.conditional_formatting.add(f'F{status_row}',
        CellIsRule(operator='equal', formula=['"DENTRO DEL LIMITE"'],
                   fill=PatternFill('solid', fgColor='C6EFCE'),
                   font=Font(color='006100', bold=True)))
    ws.conditional_formatting.add(f'F{status_row}',
        CellIsRule(operator='equal', formula=['"EXCEDIDO"'],
                   fill=PatternFill('solid', fgColor='FFC7CE'),
                   font=Font(color='9C0006', bold=True)))

ws.conditional_formatting.add(f'F{aho_total_row}',
    CellIsRule(operator='equal', formula=['"META ALCANZADA"'],
               fill=PatternFill('solid', fgColor='C6EFCE'),
               font=Font(color='006100', bold=True)))
ws.conditional_formatting.add(f'F{aho_total_row}',
    CellIsRule(operator='equal', formula=['"AUN FALTA"'],
               fill=PatternFill('solid', fgColor='FFC7CE'),
               font=Font(color='9C0006', bold=True)))

# Grand total status
ws.conditional_formatting.add(f'F{gt_row}',
    CellIsRule(operator='equal', formula=['"PRESUPUESTO OK"'],
               font=Font(color='10B981', bold=True, size=13)))
ws.conditional_formatting.add(f'F{gt_row}',
    CellIsRule(operator='equal', formula=['"GASTAS DE MAS"'],
               font=Font(color='EF4444', bold=True, size=13)))

# Sobrante: green if positive, red if negative
ws.conditional_formatting.add(f'C{sob_row}',
    CellIsRule(operator='greaterThanOrEqual', formula=['0'],
               font=Font(color='006100', bold=True, size=12)))
ws.conditional_formatting.add(f'C{sob_row}',
    CellIsRule(operator='lessThan', formula=['0'],
               font=Font(color='9C0006', bold=True, size=12)))

# Data validation for all input cells
dv_amounts = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1=0)
dv_amounts.error = "Ingresa un monto de 0 o mayor"
dv_amounts.errorTitle = "Monto inválido"
ws.add_data_validation(dv_amounts)
for r in range(nec_start, nec_end + 1):
    dv_amounts.add(f'C{r}')
for r in range(des_start, des_end + 1):
    dv_amounts.add(f'C{r}')
for r in range(aho_start, aho_end + 1):
    dv_amounts.add(f'C{r}')


# ============================================================
# SHEET 2: DASHBOARD (Charts and Summary)
# ============================================================
ws2 = wb.create_sheet("Dashboard")
ws2.sheet_properties.tabColor = ACCENT_BLUE

ws2.column_dimensions['A'].width = 2
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 16
ws2.column_dimensions['D'].width = 16
ws2.column_dimensions['E'].width = 16
ws2.column_dimensions['F'].width = 16
ws2.column_dimensions['G'].width = 16
ws2.column_dimensions['H'].width = 2

# Header
ws2.merge_cells('B1:G1')
set_row_height(ws2, 1, 8)
ws2.merge_cells('B2:G2')
h = ws2['B2']
h.value = "DASHBOARD - RESUMEN VISUAL"
h.font = Font(name='Calibri', bold=True, color=GOLD, size=18)
h.fill = PatternFill('solid', fgColor=DARK_BG)
h.alignment = Alignment(horizontal='center', vertical='center')
for col in range(2, 8):
    ws2.cell(row=2, column=col).fill = PatternFill('solid', fgColor=DARK_BG)
set_row_height(ws2, 2, 40)

ws2.merge_cells('B3:G3')
sub = ws2['B3']
sub.value = "Tus finanzas de un vistazo"
sub.font = Font(name='Calibri', color=GRAY_300, size=10, italic=True)
sub.fill = PatternFill('solid', fgColor=DARK_BG)
sub.alignment = Alignment(horizontal='center', vertical='center')
for col in range(2, 8):
    ws2.cell(row=3, column=col).fill = PatternFill('solid', fgColor=DARK_BG)
set_row_height(ws2, 3, 24)

set_row_height(ws2, 4, 8)

# ── KPI Cards ──
# Card: Income
set_row_height(ws2, 5, 20)
set_row_height(ws2, 6, 28)
ws2.cell(row=5, column=2, value="INGRESO MENSUAL").font = Font(name='Calibri', size=9, color=GRAY_500, bold=True)
kpi1 = ws2.cell(row=6, column=2)
kpi1.value = "=Presupuesto!D5"
kpi1.font = Font(name='Calibri', size=16, bold=True, color=GRAY_900)
apply_currency_format(kpi1)

# Card: Total spent
ws2.cell(row=5, column=3, value="TOTAL ASIGNADO").font = Font(name='Calibri', size=9, color=GRAY_500, bold=True)
kpi2 = ws2.cell(row=6, column=3)
kpi2.value = f"=Presupuesto!C{gt_row}"
kpi2.font = Font(name='Calibri', size=16, bold=True, color=GRAY_900)
apply_currency_format(kpi2)

# Card: Sobrante
ws2.cell(row=5, column=4, value="SOBRANTE").font = Font(name='Calibri', size=9, color=GRAY_500, bold=True)
kpi3 = ws2.cell(row=6, column=4)
kpi3.value = f"=Presupuesto!C{sob_row}"
kpi3.font = Font(name='Calibri', size=16, bold=True, color=SUCCESS)
apply_currency_format(kpi3)

# Card: % Used
ws2.cell(row=5, column=5, value="% UTILIZADO").font = Font(name='Calibri', size=9, color=GRAY_500, bold=True)
kpi4 = ws2.cell(row=6, column=5)
kpi4.value = f"=Presupuesto!D{gt_row}"
kpi4.font = Font(name='Calibri', size=16, bold=True, color=INFO_BLUE)
apply_pct_format(kpi4)

# Card: Status
ws2.cell(row=5, column=6, value="ESTADO").font = Font(name='Calibri', size=9, color=GRAY_500, bold=True)
kpi5 = ws2.cell(row=6, column=6)
kpi5.value = f"=Presupuesto!F{gt_row}"
kpi5.font = Font(name='Calibri', size=14, bold=True, color=GOLD)
kpi5.alignment = Alignment(horizontal='center')

set_row_height(ws2, 7, 8)

# ── PIE CHART DATA (hidden helper) ──
ws2.cell(row=8, column=2, value="Categoría").font = Font(name='Calibri', bold=True, color=GRAY_700, size=10)
ws2.cell(row=8, column=3, value="Monto").font = Font(name='Calibri', bold=True, color=GRAY_700, size=10)
ws2.cell(row=8, column=4, value="Límite ideal").font = Font(name='Calibri', bold=True, color=GRAY_700, size=10)
ws2.cell(row=8, column=5, value="Diferencia").font = Font(name='Calibri', bold=True, color=GRAY_700, size=10)

ws2.cell(row=9, column=2, value="Necesidades (50%)").font = Font(name='Calibri', color=NECESIDADES_COLOR, bold=True)
ws2.cell(row=9, column=3).value = f"=Presupuesto!C{nec_total_row}"
apply_currency_format(ws2.cell(row=9, column=3))
ws2.cell(row=9, column=4).value = "=Presupuesto!D5*0.5"
apply_currency_format(ws2.cell(row=9, column=4))
ws2.cell(row=9, column=5).value = "=D9-C9"
apply_currency_format(ws2.cell(row=9, column=5))

ws2.cell(row=10, column=2, value="Deseos (30%)").font = Font(name='Calibri', color=DESEOS_COLOR, bold=True)
ws2.cell(row=10, column=3).value = f"=Presupuesto!C{des_total_row}"
apply_currency_format(ws2.cell(row=10, column=3))
ws2.cell(row=10, column=4).value = "=Presupuesto!D5*0.3"
apply_currency_format(ws2.cell(row=10, column=4))
ws2.cell(row=10, column=5).value = "=D10-C10"
apply_currency_format(ws2.cell(row=10, column=5))

ws2.cell(row=11, column=2, value="Ahorro (20%)").font = Font(name='Calibri', color=AHORRO_COLOR, bold=True)
ws2.cell(row=11, column=3).value = f"=Presupuesto!C{aho_total_row}"
apply_currency_format(ws2.cell(row=11, column=3))
ws2.cell(row=11, column=4).value = "=Presupuesto!D5*0.2"
apply_currency_format(ws2.cell(row=11, column=4))
ws2.cell(row=11, column=5).value = "=D11-C11"
apply_currency_format(ws2.cell(row=11, column=5))

# Conditional formatting for difference column
ws2.conditional_formatting.add('E9:E11',
    CellIsRule(operator='greaterThanOrEqual', formula=['0'],
               font=Font(color='006100', bold=True),
               fill=PatternFill('solid', fgColor='C6EFCE')))
ws2.conditional_formatting.add('E9:E11',
    CellIsRule(operator='lessThan', formula=['0'],
               font=Font(color='9C0006', bold=True),
               fill=PatternFill('solid', fgColor='FFC7CE')))

# ── PIE CHART: Distribution ──
pie = PieChart()
pie.title = "Distribución de tu Presupuesto"
pie.style = 10
pie.width = 16
pie.height = 12

data = Reference(ws2, min_col=3, min_row=8, max_row=11)
cats = Reference(ws2, min_col=2, min_row=9, max_row=11)
pie.add_data(data, titles_from_data=True)
pie.set_categories(cats)

# Colors for slices
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
slice_colors = [NECESIDADES_COLOR, DESEOS_COLOR, AHORRO_COLOR]
for i, color in enumerate(slice_colors):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = color
    pie.series[0].data_points.append(pt)

pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
pie.dataLabels.showVal = True
pie.dataLabels.showCatName = True
pie.dataLabels.numFmt = '#,##0'

ws2.add_chart(pie, "B13")

# ── BAR CHART: Actual vs Ideal ──
bar = BarChart()
bar.type = "col"
bar.title = "Real vs Ideal (50/30/20)"
bar.style = 10
bar.width = 16
bar.height = 12
bar.y_axis.title = "Bs."

data_actual = Reference(ws2, min_col=3, min_row=8, max_row=11)
data_ideal = Reference(ws2, min_col=4, min_row=8, max_row=11)
cats = Reference(ws2, min_col=2, min_row=9, max_row=11)

bar.add_data(data_actual, titles_from_data=True)
bar.add_data(data_ideal, titles_from_data=True)
bar.set_categories(cats)
bar.shape = 4

bar.series[0].graphicalProperties.solidFill = GOLD
bar.series[1].graphicalProperties.solidFill = GRAY_300

ws2.add_chart(bar, "B30")


# ============================================================
# SHEET 3: INSTRUCCIONES
# ============================================================
ws3 = wb.create_sheet("Instrucciones")
ws3.sheet_properties.tabColor = SUCCESS

ws3.column_dimensions['A'].width = 2
ws3.column_dimensions['B'].width = 80
ws3.column_dimensions['C'].width = 2

ws3.merge_cells('B2:B2')
h = ws3['B2']
h.value = "COMO USAR TU PRESUPUESTO 50/30/20"
h.font = Font(name='Calibri', bold=True, color=GOLD, size=16)
h.fill = PatternFill('solid', fgColor=DARK_BG)
h.alignment = Alignment(horizontal='center', vertical='center')
set_row_height(ws3, 2, 40)

instructions = [
    "",
    "PASO 1: INGRESA TU INGRESO MENSUAL",
    "   Ve a la hoja 'Presupuesto' y escribe tu ingreso neto (lo que recibes",
    "   despues de impuestos) en la celda amarilla grande.",
    "",
    "PASO 2: LLENA TUS NECESIDADES (50%)",
    "   Escribe cuanto gastas en cada categoria de necesidades.",
    "   Estas son cosas que NECESITAS para vivir: vivienda, comida, transporte.",
    "   La regla dice que no deberia ser mas del 50% de tu ingreso.",
    "",
    "PASO 3: LLENA TUS DESEOS (30%)",
    "   Escribe cuanto gastas en cosas que te gustan pero no son esenciales.",
    "   Restaurantes, entretenimiento, compras, suscripciones.",
    "   Maximo 30% de tu ingreso.",
    "",
    "PASO 4: LLENA TUS AHORROS (20%)",
    "   Escribe cuanto AHORRAS e INVIERTES cada mes.",
    "   Fondo de emergencia, inversiones, metas de ahorro.",
    "   Minimo 20% de tu ingreso.",
    "",
    "PASO 5: REVISA TU DASHBOARD",
    "   Ve a la hoja 'Dashboard' para ver tus graficos y resumen visual.",
    "   Ahi veras si estas dentro de los limites recomendados.",
    "",
    "REGLA DE ORO:",
    "   Si NECESIDADES > 50%: Busca reducir gastos fijos (negociar alquiler,",
    "   cambiar de plan de telefono, cocinar mas en casa).",
    "   Si DESEOS > 30%: Identifica que puedes eliminar o reducir.",
    "   Si AHORRO < 20%: Prioriza el ahorro, pagalo primero.",
    "",
    "CELDAS AMARILLAS = Tu escribes ahi",
    "CELDAS VERDES = Se calculan solas (no las toques!)",
    "",
    "---",
    "",
    "Quieres llevar tus finanzas al siguiente nivel?",
    "Visita nosomosignorantes.com para herramientas avanzadas:",
    "  - Bola de Nieve: Elimina TODAS tus deudas sistematicamente",
    "  - Amortizador PRO: Calcula tus prestamos con precision",
    "  - Y mucho mas!",
    "",
    "Siguenos en YouTube: @NoSomosIgnorantes",
]

for i, line in enumerate(instructions):
    r = 4 + i
    c = ws3.cell(row=r, column=2, value=line)
    if line.startswith("PASO") or line.startswith("REGLA") or line.startswith("CELDAS"):
        c.font = Font(name='Calibri', bold=True, color=GOLD, size=12)
    elif line.startswith("Quieres") or line.startswith("Visita") or line.startswith("Siguenos"):
        c.font = Font(name='Calibri', bold=True, color=NECESIDADES_COLOR, size=11)
    elif line.startswith("  -"):
        c.font = Font(name='Calibri', color=ACCENT_BLUE, size=11)
    else:
        c.font = Font(name='Calibri', color=GRAY_700, size=11)


# ============================================================
# SHEET 4: CONFIG
# ============================================================
ws4 = wb.create_sheet("Config")
ws4.sheet_properties.tabColor = GRAY_500
ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 40

ws4['A1'] = "Version"
ws4['B1'] = "v1.0.0"
ws4['A2'] = "Producto"
ws4['B2'] = "Presupuesto 50/30/20"
ws4['A3'] = "Marca"
ws4['B3'] = "No Somos Ignorantes"
ws4['A4'] = "Sitio web"
ws4['B4'] = "nosomosignorantes.com"
ws4['A5'] = "YouTube"
ws4['B5'] = "@NoSomosIgnorantes"
ws4['A6'] = "Precio"
ws4['B6'] = "GRATIS (Lead Magnet)"
ws4['A7'] = "Proteccion"
ws4['B7'] = "Celdas de formula protegidas"
ws4['A8'] = "Fecha"
ws4['B8'] = "Febrero 2026"

for r in range(1, 9):
    ws4.cell(row=r, column=1).font = Font(name='Calibri', bold=True, color=GRAY_700, size=10)
    ws4.cell(row=r, column=2).font = Font(name='Calibri', color=GRAY_900, size=10)


# ============================================================
# SHEET PROTECTION
# ============================================================
# Unlock input cells, lock everything else
for ws_name in ['Presupuesto']:
    sheet = wb[ws_name]
    for row in sheet.iter_rows():
        for cell in row:
            cell.protection = cell.protection.copy(locked=True)

    # Unlock input cells (income)
    sheet['D5'].protection = sheet['D5'].protection.copy(locked=False)
    # Unlock all category amount inputs
    for r in range(nec_start, nec_end + 1):
        sheet.cell(row=r, column=3).protection = sheet.cell(row=r, column=3).protection.copy(locked=False)
    for r in range(des_start, des_end + 1):
        sheet.cell(row=r, column=3).protection = sheet.cell(row=r, column=3).protection.copy(locked=False)
    for r in range(aho_start, aho_end + 1):
        sheet.cell(row=r, column=3).protection = sheet.cell(row=r, column=3).protection.copy(locked=False)

    sheet.protection.sheet = True
    sheet.protection.password = 'nsi2024'

# Protect Dashboard (read-only)
wb['Dashboard'].protection.sheet = True
wb['Dashboard'].protection.password = 'nsi2024'

# ============================================================
# PRINT SETTINGS
# ============================================================
for sheet in wb.sheetnames:
    s = wb[sheet]
    s.page_setup.orientation = 'portrait'
    s.page_setup.paperSize = s.PAPERSIZE_LETTER
    s.page_setup.fitToWidth = 1
    s.page_setup.fitToHeight = 0
    s.sheet_properties.pageSetUpPr.fitToPage = True

# ============================================================
# SAVE
# ============================================================
output_path = r"D:\Landing-Page_marketplace\excel_products\Presupuesto_50_30_20_NSI.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
print(f"Necesidades rows: {nec_start}-{nec_end}, total row: {nec_total_row}")
print(f"Deseos rows: {des_start}-{des_end}, total row: {des_total_row}")
print(f"Ahorro rows: {aho_start}-{aho_end}, total row: {aho_total_row}")
print(f"Grand total row: {gt_row}, Sobrante row: {sob_row}")
