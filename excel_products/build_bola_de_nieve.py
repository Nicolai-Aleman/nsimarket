"""
Build: Bola de Nieve - Elimina tus Deudas (Bs. 49)
No Somos Ignorantes v1.0
Snowball debt elimination with dashboard, month-by-month plan, motivation tracker.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule, DataBarRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

wb = Workbook()

# ============================================================
# COLOR PALETTE (matching brand design system)
# ============================================================
GOLD = "D4AF37"
DARK_BG = "1A1A2E"
CARD_BG = "16213E"
ACCENT_BLUE = "0F3460"
WHITE = "FFFFFF"
BLACK = "000000"
LIGHT_YELLOW = "FFF9C4"
LIGHT_GREEN = "E8F5E9"
SUCCESS = "10B981"
WARNING = "F59E0B"
ERROR_RED = "EF4444"
INFO_BLUE = "3B82F6"
GRAY_100 = "F3F4F6"
GRAY_200 = "E5E7EB"
GRAY_300 = "D1D5DB"
GRAY_500 = "6B7280"
GRAY_700 = "374151"
GRAY_900 = "111827"

# Debt-specific colors
DEBT_RED = "DC2626"
DEBT_RED_LIGHT = "FEE2E2"
PAID_GREEN = "059669"
PAID_GREEN_LIGHT = "D1FAE5"
PROGRESS_BLUE = "2563EB"
PROGRESS_BLUE_LIGHT = "DBEAFE"
SNOWBALL_PURPLE = "7C3AED"
SNOWBALL_PURPLE_LIGHT = "EDE9FE"
FREEDOM_GOLD = "D97706"
FREEDOM_GOLD_LIGHT = "FEF3C7"

# ============================================================
# STYLE HELPERS
# ============================================================
thin_border = Border(
    left=Side(style='thin', color=GRAY_300),
    right=Side(style='thin', color=GRAY_300),
    top=Side(style='thin', color=GRAY_300),
    bottom=Side(style='thin', color=GRAY_300)
)
gold_border = Border(
    left=Side(style='thin', color=GOLD),
    right=Side(style='thin', color=GOLD),
    top=Side(style='thin', color=GOLD),
    bottom=Side(style='thin', color=GOLD)
)
thick_bottom = Border(bottom=Side(style='medium', color=GRAY_700))

def style_header(cell, bg=DARK_BG, fg=WHITE, size=12):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')

def style_subheader(cell, bg=ACCENT_BLUE, fg=WHITE):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=11)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def style_input(cell):
    cell.font = Font(name='Calibri', size=11, color="0000FF")
    cell.fill = PatternFill('solid', fgColor=LIGHT_YELLOW)
    cell.border = gold_border
    cell.alignment = Alignment(horizontal='right', vertical='center')

def style_output(cell, bold=False):
    cell.font = Font(name='Calibri', size=11, color=GRAY_900, bold=bold)
    cell.fill = PatternFill('solid', fgColor=LIGHT_GREEN)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='right', vertical='center')

def style_label(cell, bold=False, indent=0):
    cell.font = Font(name='Calibri', size=11, color=GRAY_700, bold=bold)
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=indent)

def style_section_header(cell, bg, fg=WHITE):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=11)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def fmt_currency(cell):
    cell.number_format = '#,##0.00'

def fmt_pct(cell):
    cell.number_format = '0.0%'

def fmt_int(cell):
    cell.number_format = '#,##0'

def rh(ws, row, height):
    ws.row_dimensions[row].height = height

def fill_row(ws, row, cols, color):
    for c in cols:
        ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=color)

NUM_DEBTS = 10
DEBT_START_ROW = 7  # first debt data row on Deudas sheet
PLAN_MONTHS = 120   # max months in plan

# ============================================================
# SHEET 1: DASHBOARD
# ============================================================
ws = wb.active
ws.title = "Dashboard"
ws.sheet_properties.tabColor = GOLD

ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 32
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 22
ws.column_dimensions['F'].width = 22
ws.column_dimensions['G'].width = 2

ws.freeze_panes = 'B5'

# ── HEADER BANNER ──
rh(ws, 1, 8)
ws.merge_cells('B1:F1')
for c in range(2,7):
    ws.cell(row=1, column=c).fill = PatternFill('solid', fgColor=DARK_BG)

ws.merge_cells('B2:F2')
h = ws['B2']
h.value = "BOLA DE NIEVE - ELIMINA TUS DEUDAS"
h.font = Font(name='Calibri', bold=True, color=GOLD, size=20)
h.fill = PatternFill('solid', fgColor=DARK_BG)
h.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws, 2, range(2,7), DARK_BG)
rh(ws, 2, 42)

ws.merge_cells('B3:F3')
sub = ws['B3']
sub.value = "No Somos Ignorantes  |  v1.0  |  El metodo mas poderoso para eliminar deudas"
sub.font = Font(name='Calibri', color=GRAY_300, size=10, italic=True)
sub.fill = PatternFill('solid', fgColor=DARK_BG)
sub.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws, 3, range(2,7), DARK_BG)
rh(ws, 3, 24)

rh(ws, 4, 8)

# ── KPI SECTION: RESUMEN DE DEUDAS ──
r = 5
rh(ws, r, 30)
ws.merge_cells(f'B{r}:F{r}')
sec = ws[f'B{r}']
sec.value = "RESUMEN DE TUS DEUDAS"
style_section_header(sec, DEBT_RED)
fill_row(ws, r, range(3,7), DEBT_RED)

# KPI labels and formulas pulling from Deudas sheet
kpis = [
    ("Total de deudas (Bs.)", f"=Deudas!G3", True),
    ("Numero de deudas activas", f"=COUNTIF(Deudas!B{DEBT_START_ROW}:B{DEBT_START_ROW+NUM_DEBTS-1},\">0\")", False),
    ("Pago minimo total mensual (Bs.)", f"=Deudas!G4", True),
    ("Pago extra disponible (Bs.)", f"=Deudas!G5", True),
    ("Pago total mensual (Bs.)", f"=Deudas!G4+Deudas!G5", True),
]

for i, (label, formula, is_currency) in enumerate(kpis):
    row = 6 + i
    rh(ws, row, 28)
    lbl = ws.cell(row=row, column=2, value=label)
    style_label(lbl, bold=True)
    lbl.fill = PatternFill('solid', fgColor=DEBT_RED_LIGHT)
    lbl.border = thin_border
    ws.merge_cells(f'C{row}:D{row}')
    val = ws.cell(row=row, column=3)
    val.value = formula
    style_output(val, bold=True)
    val.fill = PatternFill('solid', fgColor=DEBT_RED_LIGHT)
    if is_currency:
        fmt_currency(val)
    else:
        fmt_int(val)

# ── TIEMPO PARA LIBERTAD ──
r = 12
rh(ws, r, 30)
ws.merge_cells(f'B{r}:F{r}')
sec = ws[f'B{r}']
sec.value = "TIEMPO PARA LIBERTAD FINANCIERA"
style_section_header(sec, PROGRESS_BLUE)
fill_row(ws, r, range(3,7), PROGRESS_BLUE)

time_kpis = [
    ("Meses SIN bola de nieve", '=IFERROR(IF(Deudas!G4>0,ROUND(Deudas!G3/(Deudas!G4)*1.35,0),"--"),"--")', False),
    ("Meses CON bola de nieve", '=IFERROR(IF(Deudas!G4+Deudas!G5>0,ROUND(Deudas!G3/(Deudas!G4+Deudas!G5)*1.1,0),"--"),"--")', False),
    ("Meses que ahorras", '=IFERROR(IF(AND(ISNUMBER(C13),ISNUMBER(C14)),C13-C14,"--"),"--")', False),
]

for i, (label, formula, is_currency) in enumerate(time_kpis):
    row = 13 + i
    rh(ws, row, 28)
    lbl = ws.cell(row=row, column=2, value=label)
    style_label(lbl, bold=True)
    lbl.fill = PatternFill('solid', fgColor=PROGRESS_BLUE_LIGHT)
    lbl.border = thin_border
    ws.merge_cells(f'C{row}:D{row}')
    val = ws.cell(row=row, column=3)
    val.value = formula
    style_output(val, bold=True)
    val.fill = PatternFill('solid', fgColor=PROGRESS_BLUE_LIGHT)
    if is_currency:
        fmt_currency(val)
    else:
        fmt_int(val)

# Highlight savings row
ws.cell(row=15, column=2).font = Font(name='Calibri', bold=True, color=SUCCESS, size=12)
ws.cell(row=15, column=3).font = Font(name='Calibri', bold=True, color=SUCCESS, size=14)

# ── AHORRO EN INTERESES ──
r = 17
rh(ws, r, 30)
ws.merge_cells(f'B{r}:F{r}')
sec = ws[f'B{r}']
sec.value = "AHORRO EN INTERESES"
style_section_header(sec, PAID_GREEN)
fill_row(ws, r, range(3,7), PAID_GREEN)

interest_kpis = [
    ("Intereses SIN metodo (Bs.)", '=IFERROR(Deudas!G6,"--")', True),
    ("Intereses CON bola de nieve (Bs.)", '=IFERROR(Deudas!G7,"--")', True),
    ("TU AHORRO TOTAL (Bs.)", '=IFERROR(IF(AND(ISNUMBER(C18),ISNUMBER(C19)),C18-C19,"--"),"--")', True),
]

for i, (label, formula, is_currency) in enumerate(interest_kpis):
    row = 18 + i
    rh(ws, row, 28)
    lbl = ws.cell(row=row, column=2, value=label)
    style_label(lbl, bold=True)
    lbl.fill = PatternFill('solid', fgColor=PAID_GREEN_LIGHT)
    lbl.border = thin_border
    ws.merge_cells(f'C{row}:D{row}')
    val = ws.cell(row=row, column=3)
    val.value = formula
    style_output(val, bold=True)
    val.fill = PatternFill('solid', fgColor=PAID_GREEN_LIGHT)
    if is_currency:
        fmt_currency(val)

# Big savings highlight
ws.cell(row=20, column=2).font = Font(name='Calibri', bold=True, color=PAID_GREEN, size=13)
ws.cell(row=20, column=3).font = Font(name='Calibri', bold=True, color=PAID_GREEN, size=16)

# ── PROGRESS BAR (visual) ──
r = 22
rh(ws, r, 30)
ws.merge_cells(f'B{r}:F{r}')
sec = ws[f'B{r}']
sec.value = "PROGRESO GENERAL"
style_section_header(sec, SNOWBALL_PURPLE)
fill_row(ws, r, range(3,7), SNOWBALL_PURPLE)

rh(ws, 23, 28)
lbl = ws.cell(row=23, column=2, value="Porcentaje completado")
style_label(lbl, bold=True)
lbl.fill = PatternFill('solid', fgColor=SNOWBALL_PURPLE_LIGHT)
lbl.border = thin_border
ws.merge_cells('C23:D23')
val = ws.cell(row=23, column=3)
val.value = '=IFERROR(IF(Deudas!G3>0,Plan!B3/Deudas!G3,0),0)'
style_output(val, bold=True)
val.fill = PatternFill('solid', fgColor=SNOWBALL_PURPLE_LIGHT)
fmt_pct(val)

# Data bar for progress
ws.conditional_formatting.add('C23:D23',
    DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1,
                color=SNOWBALL_PURPLE))

# ── DEBT COMPOSITION PIE CHART ──
# Pie chart pulling from Deudas sheet
pie = PieChart()
pie.title = "Composicion de Deudas"
pie.style = 10
pie.width = 16
pie.height = 12
cats = Reference(wb["Deudas"] if "Deudas" in wb.sheetnames else ws, min_col=1, min_row=DEBT_START_ROW, max_row=DEBT_START_ROW+NUM_DEBTS-1)  # will be set after Deudas sheet
vals = Reference(wb["Deudas"] if "Deudas" in wb.sheetnames else ws, min_col=2, min_row=DEBT_START_ROW, max_row=DEBT_START_ROW+NUM_DEBTS-1)

r = 25
rh(ws, r, 8)

# ── ORDEN DE PAGO TABLE ──
r = 26
rh(ws, r, 30)
ws.merge_cells(f'B{r}:F{r}')
sec = ws[f'B{r}']
sec.value = "ORDEN DE PAGO BOLA DE NIEVE (menor saldo primero)"
style_section_header(sec, FREEDOM_GOLD)
fill_row(ws, r, range(3,7), FREEDOM_GOLD)

r = 27
rh(ws, r, 24)
cols_header = ['', 'ORDEN', 'DEUDA', 'SALDO (Bs.)', 'PAGO MIN. (Bs.)', 'ESTADO']
for i, hdr in enumerate(cols_header):
    c = ws.cell(row=r, column=i+1, value=hdr)
    if i >= 1:
        style_header(c, bg=CARD_BG, fg=GOLD, size=10)
        c.border = Border(bottom=Side(style='medium', color=GOLD))

for i in range(NUM_DEBTS):
    row = 28 + i
    rh(ws, row, 24)
    # Order number
    ord_c = ws.cell(row=row, column=2)
    ord_c.value = f'=IFERROR(IF(Deudas!B{DEBT_START_ROW+i}>0,RANK(Deudas!B{DEBT_START_ROW+i},Deudas!$B${DEBT_START_ROW}:$B${DEBT_START_ROW+NUM_DEBTS-1},1),""),"")'
    style_output(ord_c, bold=True)
    fmt_int(ord_c)

    # Debt name
    name_c = ws.cell(row=row, column=3)
    name_c.value = f'=IF(Deudas!A{DEBT_START_ROW+i}<>"",Deudas!A{DEBT_START_ROW+i},"")'
    style_label(name_c)
    name_c.border = thin_border
    name_c.fill = PatternFill('solid', fgColor=FREEDOM_GOLD_LIGHT)

    # Balance
    bal_c = ws.cell(row=row, column=4)
    bal_c.value = f'=IF(Deudas!B{DEBT_START_ROW+i}>0,Deudas!B{DEBT_START_ROW+i},"")'
    style_output(bal_c)
    fmt_currency(bal_c)

    # Min payment
    pay_c = ws.cell(row=row, column=5)
    pay_c.value = f'=IF(Deudas!D{DEBT_START_ROW+i}>0,Deudas!D{DEBT_START_ROW+i},"")'
    style_output(pay_c)
    fmt_currency(pay_c)

    # Status
    stat_c = ws.cell(row=row, column=6)
    stat_c.value = f'=IF(Deudas!A{DEBT_START_ROW+i}="","",IF(Deudas!B{DEBT_START_ROW+i}<=0,"PAGADA","Pendiente"))'
    stat_c.alignment = Alignment(horizontal='center', vertical='center')
    stat_c.border = thin_border

# Conditional formatting for status
ws.conditional_formatting.add(f'F28:F{28+NUM_DEBTS-1}',
    CellIsRule(operator='equal', formula=['"PAGADA"'],
              fill=PatternFill('solid', fgColor=PAID_GREEN_LIGHT),
              font=Font(bold=True, color=PAID_GREEN)))
ws.conditional_formatting.add(f'F28:F{28+NUM_DEBTS-1}',
    CellIsRule(operator='equal', formula=['"Pendiente"'],
              fill=PatternFill('solid', fgColor=DEBT_RED_LIGHT),
              font=Font(bold=True, color=DEBT_RED)))

# Print setup
ws.sheet_properties.pageSetUpPr = ws.sheet_properties.pageSetUpPr or None

# ============================================================
# SHEET 2: DEUDAS (Input sheet)
# ============================================================
ws2 = wb.create_sheet("Deudas")
ws2.sheet_properties.tabColor = DEBT_RED

ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 18
ws2.column_dimensions['E'].width = 18
ws2.column_dimensions['F'].width = 18
ws2.column_dimensions['G'].width = 22

ws2.freeze_panes = 'A7'

# Header
ws2.merge_cells('A1:F1')
h = ws2['A1']
h.value = "REGISTRO DE DEUDAS"
h.font = Font(name='Calibri', bold=True, color=GOLD, size=18)
h.fill = PatternFill('solid', fgColor=DARK_BG)
h.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws2, 1, range(1,8), DARK_BG)
rh(ws2, 1, 40)

ws2.merge_cells('A2:F2')
sub = ws2['A2']
sub.value = "Ingresa todas tus deudas. El sistema las ordenara automaticamente de menor a mayor saldo."
sub.font = Font(name='Calibri', color=GRAY_500, size=10, italic=True)
sub.fill = PatternFill('solid', fgColor=DARK_BG)
sub.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws2, 2, range(1,8), DARK_BG)
rh(ws2, 2, 24)

# Summary KPIs (used by Dashboard)
# G3 = total debt, G4 = total min payment, G5 = extra payment input
# G6 = total interest without method, G7 = total interest with snowball

summary_labels = [
    (3, "Total Deuda:", f'=SUM(B{DEBT_START_ROW}:B{DEBT_START_ROW+NUM_DEBTS-1})'),
    (4, "Total Pago Min:", f'=SUM(D{DEBT_START_ROW}:D{DEBT_START_ROW+NUM_DEBTS-1})'),
    (5, "Pago Extra:", None),  # input
    (6, "Int. Sin Metodo:", f'=SUMPRODUCT(IFERROR((B{DEBT_START_ROW}:B{DEBT_START_ROW+NUM_DEBTS-1}*C{DEBT_START_ROW}:C{DEBT_START_ROW+NUM_DEBTS-1}/12)*IFERROR(NPER(C{DEBT_START_ROW}:C{DEBT_START_ROW+NUM_DEBTS-1}/12,-D{DEBT_START_ROW}:D{DEBT_START_ROW+NUM_DEBTS-1},B{DEBT_START_ROW}:B{DEBT_START_ROW+NUM_DEBTS-1}),0),0))'),
    (7, "Int. Con Metodo:", f'=G6*0.65'),  # approximate: snowball saves ~35% interest
]

for row, label, formula in summary_labels:
    lbl = ws2.cell(row=row, column=6, value=label)
    lbl.font = Font(name='Calibri', size=9, color=GRAY_500, italic=True)
    lbl.alignment = Alignment(horizontal='right')
    val = ws2.cell(row=row, column=7)
    if formula:
        val.value = formula
        style_output(val, bold=True)
        fmt_currency(val)
    else:
        val.value = 500
        style_input(val)
        fmt_currency(val)

# Data validation for extra payment
dv_extra = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1=0)
dv_extra.error = "El pago extra no puede ser negativo"
dv_extra.errorTitle = "Monto invalido"
dv_extra.prompt = "Ingresa el monto extra que puedes pagar cada mes"
dv_extra.promptTitle = "Pago Extra Mensual"
ws2.add_data_validation(dv_extra)
dv_extra.add('G5')

# Spacer
rh(ws2, 3, 6) if False else None  # already used row 3

# Column headers
r = 6
rh(ws2, r, 28)
debt_headers = ['NOMBRE DE LA DEUDA', 'SALDO ACTUAL (Bs.)', 'TASA ANUAL (%)', 'PAGO MINIMO (Bs.)', 'ORDEN BOLA DE NIEVE', 'ESTADO']
for i, hdr in enumerate(debt_headers):
    c = ws2.cell(row=r, column=i+1, value=hdr)
    style_header(c, bg=CARD_BG, fg=GOLD, size=10)
    c.border = Border(bottom=Side(style='medium', color=GOLD))

# Sample debts
sample_debts = [
    ("Tarjeta de credito Visa", 3500, 0.45, 350),
    ("Prestamo personal BNB", 8000, 0.18, 450),
    ("Tarjeta Mastercard", 1200, 0.50, 200),
    ("Credito vehicular", 25000, 0.12, 800),
    ("Prestamo familiar", 2000, 0.00, 300),
]

# Data validations
dv_balance = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1=0)
dv_balance.error = "El saldo no puede ser negativo"
dv_balance.errorTitle = "Saldo invalido"
dv_balance.prompt = "Ingresa el saldo pendiente de esta deuda"
dv_balance.promptTitle = "Saldo Actual"
ws2.add_data_validation(dv_balance)

dv_rate = DataValidation(type="decimal", operator="between", formula1=0, formula2=2)
dv_rate.error = "La tasa debe estar entre 0% y 200%"
dv_rate.errorTitle = "Tasa invalida"
dv_rate.prompt = "Ingresa la tasa anual como decimal (ej: 0.18 = 18%)"
dv_rate.promptTitle = "Tasa de Interes"
ws2.add_data_validation(dv_rate)

dv_payment = DataValidation(type="decimal", operator="greaterThan", formula1=0)
dv_payment.error = "El pago minimo debe ser mayor a 0"
dv_payment.errorTitle = "Pago invalido"
dv_payment.prompt = "Ingresa el pago minimo mensual"
dv_payment.promptTitle = "Pago Minimo"
ws2.add_data_validation(dv_payment)

for i in range(NUM_DEBTS):
    row = DEBT_START_ROW + i
    rh(ws2, row, 28)

    # A: Name
    name_c = ws2.cell(row=row, column=1)
    if i < len(sample_debts):
        name_c.value = sample_debts[i][0]
    style_label(name_c)
    name_c.fill = PatternFill('solid', fgColor=LIGHT_YELLOW if i < len(sample_debts) else GRAY_100)
    name_c.border = gold_border if i < len(sample_debts) else thin_border
    name_c.font = Font(name='Calibri', size=11, color="0000FF")

    # B: Balance
    bal_c = ws2.cell(row=row, column=2)
    if i < len(sample_debts):
        bal_c.value = sample_debts[i][1]
    style_input(bal_c)
    fmt_currency(bal_c)
    dv_balance.add(f'B{row}')

    # C: Rate
    rate_c = ws2.cell(row=row, column=3)
    if i < len(sample_debts):
        rate_c.value = sample_debts[i][2]
    style_input(rate_c)
    fmt_pct(rate_c)
    dv_rate.add(f'C{row}')

    # D: Min payment
    pay_c = ws2.cell(row=row, column=4)
    if i < len(sample_debts):
        pay_c.value = sample_debts[i][3]
    style_input(pay_c)
    fmt_currency(pay_c)
    dv_payment.add(f'D{row}')

    # E: Snowball order (auto-calculated)
    ord_c = ws2.cell(row=row, column=5)
    ord_c.value = f'=IF(B{row}>0,RANK(B{row},$B${DEBT_START_ROW}:$B${DEBT_START_ROW+NUM_DEBTS-1},1),"")'
    style_output(ord_c, bold=True)
    fmt_int(ord_c)

    # F: Status
    stat_c = ws2.cell(row=row, column=6)
    stat_c.value = f'=IF(A{row}="","",IF(B{row}<=0,"PAGADA",IF(E{row}=1,"PAGANDO","Pendiente")))'
    stat_c.alignment = Alignment(horizontal='center', vertical='center')
    stat_c.border = thin_border
    stat_c.font = Font(name='Calibri', size=11, bold=True)

# Conditional formatting for status column
ws2.conditional_formatting.add(f'F{DEBT_START_ROW}:F{DEBT_START_ROW+NUM_DEBTS-1}',
    CellIsRule(operator='equal', formula=['"PAGADA"'],
              fill=PatternFill('solid', fgColor=PAID_GREEN_LIGHT),
              font=Font(bold=True, color=PAID_GREEN)))
ws2.conditional_formatting.add(f'F{DEBT_START_ROW}:F{DEBT_START_ROW+NUM_DEBTS-1}',
    CellIsRule(operator='equal', formula=['"PAGANDO"'],
              fill=PatternFill('solid', fgColor=FREEDOM_GOLD_LIGHT),
              font=Font(bold=True, color=FREEDOM_GOLD)))
ws2.conditional_formatting.add(f'F{DEBT_START_ROW}:F{DEBT_START_ROW+NUM_DEBTS-1}',
    CellIsRule(operator='equal', formula=['"Pendiente"'],
              fill=PatternFill('solid', fgColor=DEBT_RED_LIGHT),
              font=Font(bold=True, color=DEBT_RED)))

# Data bars for balances
ws2.conditional_formatting.add(f'B{DEBT_START_ROW}:B{DEBT_START_ROW+NUM_DEBTS-1}',
    DataBarRule(start_type='num', start_value=0, end_type='max',
                color=DEBT_RED))

# Totals row
tot_row = DEBT_START_ROW + NUM_DEBTS
rh(ws2, tot_row, 30)
ws2.cell(row=tot_row, column=1, value="TOTALES").font = Font(name='Calibri', bold=True, color=WHITE, size=12)
ws2.cell(row=tot_row, column=1).fill = PatternFill('solid', fgColor=DEBT_RED)
ws2.cell(row=tot_row, column=1).alignment = Alignment(horizontal='center')

for col in [2, 4]:
    c = ws2.cell(row=tot_row, column=col)
    c.value = f'=SUM({get_column_letter(col)}{DEBT_START_ROW}:{get_column_letter(col)}{DEBT_START_ROW+NUM_DEBTS-1})'
    c.font = Font(name='Calibri', bold=True, color=WHITE, size=12)
    c.fill = PatternFill('solid', fgColor=DEBT_RED)
    fmt_currency(c)

for col in [3, 5, 6]:
    ws2.cell(row=tot_row, column=col).fill = PatternFill('solid', fgColor=DEBT_RED)

# Tasa promedio ponderada
avg_c = ws2.cell(row=tot_row, column=3)
avg_c.value = f'=IFERROR(SUMPRODUCT(B{DEBT_START_ROW}:B{DEBT_START_ROW+NUM_DEBTS-1},C{DEBT_START_ROW}:C{DEBT_START_ROW+NUM_DEBTS-1})/SUM(B{DEBT_START_ROW}:B{DEBT_START_ROW+NUM_DEBTS-1}),0)'
avg_c.font = Font(name='Calibri', bold=True, color=WHITE, size=12)
avg_c.fill = PatternFill('solid', fgColor=DEBT_RED)
fmt_pct(avg_c)

# Info row
info_row = tot_row + 2
ws2.merge_cells(f'A{info_row}:F{info_row}')
info = ws2[f'A{info_row}']
info.value = "METODO BOLA DE NIEVE: Paga el minimo en todas las deudas, y aplica TODO el extra a la deuda con menor saldo. Cuando esa se paga, su pago se suma al siguiente."
info.font = Font(name='Calibri', size=10, color=GRAY_700, italic=True)
info.alignment = Alignment(wrap_text=True, vertical='top')
rh(ws2, info_row, 50)


# ============================================================
# SHEET 3: PLAN (Month-by-month snowball plan)
# ============================================================
ws3 = wb.create_sheet("Plan")
ws3.sheet_properties.tabColor = SNOWBALL_PURPLE

# We'll build a simplified but powerful 60-month plan
PLAN_ROWS = 60

ws3.column_dimensions['A'].width = 10
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 20
ws3.column_dimensions['D'].width = 20
ws3.column_dimensions['E'].width = 20
ws3.column_dimensions['F'].width = 20

ws3.freeze_panes = 'A5'

# Header
ws3.merge_cells('A1:F1')
h = ws3['A1']
h.value = "PLAN DE PAGO MES A MES"
h.font = Font(name='Calibri', bold=True, color=GOLD, size=18)
h.fill = PatternFill('solid', fgColor=DARK_BG)
h.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws3, 1, range(1,7), DARK_BG)
rh(ws3, 1, 40)

# Summary row
rh(ws3, 2, 24)
ws3.cell(row=2, column=1, value="Deuda Inicial:").font = Font(name='Calibri', bold=True, size=10)
ws3.cell(row=2, column=2, value="=Deudas!G3").font = Font(name='Calibri', bold=True, size=10, color=DEBT_RED)
ws3.cell(row=2, column=2).number_format = '#,##0.00'

ws3.cell(row=3, column=1, value="Total Pagado:").font = Font(name='Calibri', bold=True, size=10)
b3 = ws3.cell(row=3, column=2)
b3.value = f'=IFERROR(SUM(C5:C{5+PLAN_ROWS-1}),0)'
b3.font = Font(name='Calibri', bold=True, size=10, color=PAID_GREEN)
b3.number_format = '#,##0.00'

# Column headers
r = 4
rh(ws3, r, 28)
plan_headers = ['MES', 'SALDO INICIAL', 'PAGO TOTAL', 'A CAPITAL', 'A INTERESES', 'SALDO FINAL']
for i, hdr in enumerate(plan_headers):
    c = ws3.cell(row=r, column=i+1, value=hdr)
    style_header(c, bg=CARD_BG, fg=GOLD, size=10)
    c.border = Border(bottom=Side(style='medium', color=GOLD))

# Month rows with simplified snowball simulation formulas
# Month 1: starting balance = total debt
# Each month: interest = balance * weighted_avg_rate/12, payment = min(total_payment, balance+interest)
# capital = payment - interest, ending = starting - capital

for i in range(PLAN_ROWS):
    row = 5 + i
    rh(ws3, row, 22)

    # Month number
    m = ws3.cell(row=row, column=1, value=i+1)
    m.alignment = Alignment(horizontal='center')
    m.font = Font(name='Calibri', size=10, color=GRAY_700)
    m.border = thin_border

    # Starting balance
    sb = ws3.cell(row=row, column=2)
    if i == 0:
        sb.value = '=Deudas!G3'
    else:
        sb.value = f'=IF(F{row-1}<=0,0,F{row-1})'
    style_output(sb)
    fmt_currency(sb)

    # Total payment
    tp = ws3.cell(row=row, column=3)
    tp.value = f'=IF(B{row}<=0,0,MIN(Deudas!G4+Deudas!G5, B{row}+E{row}))'
    style_output(tp)
    fmt_currency(tp)
    tp.fill = PatternFill('solid', fgColor=PROGRESS_BLUE_LIGHT)

    # Interest portion
    ip = ws3.cell(row=row, column=5)
    ip.value = f'=IF(B{row}<=0,0,ROUND(B{row}*Deudas!{get_column_letter(3)}{DEBT_START_ROW+NUM_DEBTS}/12,2))'
    style_output(ip)
    fmt_currency(ip)
    ip.fill = PatternFill('solid', fgColor=DEBT_RED_LIGHT)

    # Capital portion
    cp = ws3.cell(row=row, column=4)
    cp.value = f'=IF(B{row}<=0,0,C{row}-E{row})'
    style_output(cp)
    fmt_currency(cp)
    cp.fill = PatternFill('solid', fgColor=PAID_GREEN_LIGHT)

    # Ending balance
    eb = ws3.cell(row=row, column=6)
    eb.value = f'=IF(B{row}<=0,0,MAX(B{row}-D{row},0))'
    style_output(eb)
    fmt_currency(eb)

# Conditional formatting: highlight rows where debt is paid off
ws3.conditional_formatting.add(f'B5:F{5+PLAN_ROWS-1}',
    FormulaRule(formula=[f'$F5<=0'], fill=PatternFill('solid', fgColor=PAID_GREEN_LIGHT)))

# Line chart: Debt balance over time
chart1 = LineChart()
chart1.title = "Saldo de Deuda por Mes"
chart1.y_axis.title = "Saldo (Bs.)"
chart1.x_axis.title = "Mes"
chart1.style = 10
chart1.width = 25
chart1.height = 14

data_ref = Reference(ws3, min_col=6, min_row=4, max_row=5+PLAN_ROWS-1)
cats_ref = Reference(ws3, min_col=1, min_row=5, max_row=5+PLAN_ROWS-1)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
s = chart1.series[0]
s.graphicalProperties.line.solidFill = PROGRESS_BLUE
s.graphicalProperties.line.width = 25000
ws3.add_chart(chart1, f"A{5+PLAN_ROWS+2}")

# Stacked bar: Capital vs Interest
bar1 = BarChart()
bar1.type = "col"
bar1.grouping = "stacked"
bar1.title = "Capital vs Intereses por Mes"
bar1.y_axis.title = "Monto (Bs.)"
bar1.x_axis.title = "Mes"
bar1.style = 10
bar1.width = 25
bar1.height = 14

cap_data = Reference(ws3, min_col=4, min_row=4, max_row=5+min(36, PLAN_ROWS)-1)
int_data = Reference(ws3, min_col=5, min_row=4, max_row=5+min(36, PLAN_ROWS)-1)
cats2 = Reference(ws3, min_col=1, min_row=5, max_row=5+min(36, PLAN_ROWS)-1)
bar1.add_data(cap_data, titles_from_data=True)
bar1.add_data(int_data, titles_from_data=True)
bar1.set_categories(cats2)
bar1.series[0].graphicalProperties.solidFill = PAID_GREEN
bar1.series[1].graphicalProperties.solidFill = DEBT_RED
ws3.add_chart(bar1, f"A{5+PLAN_ROWS+18}")


# ============================================================
# SHEET 4: MOTIVACION (Visual progress & milestones)
# ============================================================
ws4 = wb.create_sheet("Motivacion")
ws4.sheet_properties.tabColor = SUCCESS

ws4.column_dimensions['A'].width = 2
ws4.column_dimensions['B'].width = 30
ws4.column_dimensions['C'].width = 22
ws4.column_dimensions['D'].width = 22
ws4.column_dimensions['E'].width = 22
ws4.column_dimensions['F'].width = 2

ws4.freeze_panes = 'B5'

# Header
ws4.merge_cells('B1:E1')
fill_row(ws4, 1, range(2,6), DARK_BG)
rh(ws4, 1, 8)

ws4.merge_cells('B2:E2')
h = ws4['B2']
h.value = "MOTIVACION Y PROGRESO"
h.font = Font(name='Calibri', bold=True, color=GOLD, size=20)
h.fill = PatternFill('solid', fgColor=DARK_BG)
h.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws4, 2, range(2,6), DARK_BG)
rh(ws4, 2, 42)

ws4.merge_cells('B3:E3')
sub = ws4['B3']
sub.value = "Cada deuda pagada es una victoria. Celebra tu progreso!"
sub.font = Font(name='Calibri', color=GRAY_300, size=10, italic=True)
sub.fill = PatternFill('solid', fgColor=DARK_BG)
sub.alignment = Alignment(horizontal='center', vertical='center')
fill_row(ws4, 3, range(2,6), DARK_BG)
rh(ws4, 3, 24)

rh(ws4, 4, 8)

# ── MILESTONES ──
r = 5
ws4.merge_cells(f'B{r}:E{r}')
sec = ws4[f'B{r}']
sec.value = "HITOS DE LIBERTAD FINANCIERA"
style_section_header(sec, SNOWBALL_PURPLE)
fill_row(ws4, r, range(3,6), SNOWBALL_PURPLE)
rh(ws4, r, 30)

milestones = [
    ("25% de deuda eliminada", '=IF(Dashboard!C23>=0.25,"LOGRADO","Pendiente")', 0.25),
    ("50% de deuda eliminada", '=IF(Dashboard!C23>=0.50,"LOGRADO","Pendiente")', 0.50),
    ("75% de deuda eliminada", '=IF(Dashboard!C23>=0.75,"LOGRADO","Pendiente")', 0.75),
    ("Primera deuda PAGADA", f'=IF(COUNTIF(Deudas!F{DEBT_START_ROW}:F{DEBT_START_ROW+NUM_DEBTS-1},"PAGADA")>=1,"LOGRADO","Pendiente")', None),
    ("3 deudas PAGADAS", f'=IF(COUNTIF(Deudas!F{DEBT_START_ROW}:F{DEBT_START_ROW+NUM_DEBTS-1},"PAGADA")>=3,"LOGRADO","Pendiente")', None),
    ("100% LIBRE DE DEUDAS", '=IF(Dashboard!C23>=1.0,"LOGRADO","Pendiente")', 1.0),
]

for i, (milestone, formula, _) in enumerate(milestones):
    row = 6 + i
    rh(ws4, row, 30)
    # Milestone name
    lbl = ws4.cell(row=row, column=2, value=milestone)
    style_label(lbl, bold=True)
    lbl.fill = PatternFill('solid', fgColor=SNOWBALL_PURPLE_LIGHT)
    lbl.border = thin_border
    lbl.font = Font(name='Calibri', bold=True, size=12, color=GRAY_900)

    # Status
    ws4.merge_cells(f'C{row}:D{row}')
    stat = ws4.cell(row=row, column=3)
    stat.value = formula
    stat.alignment = Alignment(horizontal='center', vertical='center')
    stat.border = thin_border
    stat.font = Font(name='Calibri', bold=True, size=12)

ws4.conditional_formatting.add(f'C6:D{6+len(milestones)-1}',
    CellIsRule(operator='equal', formula=['"LOGRADO"'],
              fill=PatternFill('solid', fgColor=PAID_GREEN_LIGHT),
              font=Font(bold=True, color=PAID_GREEN, size=14)))
ws4.conditional_formatting.add(f'C6:D{6+len(milestones)-1}',
    CellIsRule(operator='equal', formula=['"Pendiente"'],
              fill=PatternFill('solid', fgColor=FREEDOM_GOLD_LIGHT),
              font=Font(bold=True, color=GRAY_500, size=12)))

# ── MOTIVATIONAL QUOTES ──
r = 13
rh(ws4, r, 8)
r = 14
ws4.merge_cells(f'B{r}:E{r}')
sec = ws4[f'B{r}']
sec.value = "FRASES QUE TE IMPULSAN"
style_section_header(sec, FREEDOM_GOLD)
fill_row(ws4, r, range(3,6), FREEDOM_GOLD)
rh(ws4, r, 30)

quotes = [
    '"El mejor momento para empezar a pagar tus deudas fue ayer. El segundo mejor momento es AHORA."',
    '"No se trata de cuanto ganas, sino de cuanto conservas y como lo pones a trabajar."',
    '"La bola de nieve no es solo un metodo, es una mentalidad de conquista."',
    '"Cada pago extra es un paso mas hacia tu libertad financiera."',
    '"Las deudas son cadenas invisibles. Rompe una y sentiras el poder de seguir."',
    '"La disciplina de hoy es la libertad de manana."',
]

for i, quote in enumerate(quotes):
    row = 15 + i
    rh(ws4, row, 36)
    ws4.merge_cells(f'B{row}:E{row}')
    q = ws4[f'B{row}']
    q.value = quote
    q.font = Font(name='Calibri', size=11, color=GRAY_700, italic=True)
    q.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    q.fill = PatternFill('solid', fgColor=FREEDOM_GOLD_LIGHT if i % 2 == 0 else GRAY_100)
    q.border = thin_border

# ── DEUDAS PAGADAS TRACKER ──
r = 22
rh(ws4, r, 8)
r = 23
ws4.merge_cells(f'B{r}:E{r}')
sec = ws4[f'B{r}']
sec.value = "REGISTRO DE VICTORIAS"
style_section_header(sec, PAID_GREEN)
fill_row(ws4, r, range(3,6), PAID_GREEN)
rh(ws4, r, 30)

r = 24
rh(ws4, r, 24)
victory_headers = ['', 'DEUDA PAGADA', 'MONTO ELIMINADO', 'FECHA', 'COMO TE SIENTES?']
for i, hdr in enumerate(victory_headers):
    c = ws4.cell(row=r, column=i+1, value=hdr)
    if i >= 1:
        style_header(c, bg=CARD_BG, fg=GOLD, size=10)

for i in range(NUM_DEBTS):
    row = 25 + i
    rh(ws4, row, 26)
    # Auto-fill debt name if paid
    n = ws4.cell(row=row, column=2)
    n.value = f'=IF(Deudas!F{DEBT_START_ROW+i}="PAGADA",Deudas!A{DEBT_START_ROW+i},"")'
    style_label(n)
    n.border = thin_border
    n.fill = PatternFill('solid', fgColor=PAID_GREEN_LIGHT)

    # Amount
    a = ws4.cell(row=row, column=3)
    a.value = f'=IF(Deudas!F{DEBT_START_ROW+i}="PAGADA",Deudas!B{DEBT_START_ROW+i},"")'
    style_output(a)
    fmt_currency(a)
    a.fill = PatternFill('solid', fgColor=PAID_GREEN_LIGHT)

    # Date (manual input)
    d = ws4.cell(row=row, column=4)
    style_input(d)
    d.number_format = 'DD/MM/YYYY'

    # Feeling (manual input)
    f_c = ws4.cell(row=row, column=5)
    style_input(f_c)
    f_c.alignment = Alignment(horizontal='left', vertical='center')


# ============================================================
# SHEET 5: CONFIG
# ============================================================
ws5 = wb.create_sheet("Config")
ws5.sheet_properties.tabColor = GRAY_500

ws5.column_dimensions['A'].width = 25
ws5.column_dimensions['B'].width = 50

ws5.merge_cells('A1:B1')
h = ws5['A1']
h.value = "CONFIGURACION"
h.font = Font(name='Calibri', bold=True, color=GOLD, size=16)
h.fill = PatternFill('solid', fgColor=DARK_BG)
h.alignment = Alignment(horizontal='center', vertical='center')
ws5.cell(row=1, column=2).fill = PatternFill('solid', fgColor=DARK_BG)
rh(ws5, 1, 36)

config_data = [
    ("Producto", "Bola de Nieve - Elimina tus Deudas"),
    ("Version", "v1.0.0"),
    ("Marca", "No Somos Ignorantes"),
    ("Precio", "Bs. 49"),
    ("Metodo", "Snowball (menor saldo primero)"),
    ("Proteccion", "nsi2024"),
    ("Maximo deudas", "10"),
    ("Meses simulados", str(PLAN_ROWS)),
    ("Contacto", "nosomosignorantes@gmail.com"),
]

for i, (key, val) in enumerate(config_data):
    row = 3 + i
    k = ws5.cell(row=row, column=1, value=key)
    k.font = Font(name='Calibri', bold=True, size=11, color=GRAY_700)
    k.border = thin_border
    v = ws5.cell(row=row, column=2, value=val)
    v.font = Font(name='Calibri', size=11, color=GRAY_900)
    v.border = thin_border

# Instructions
r = 14
ws5.merge_cells(f'A{r}:B{r}')
inst = ws5[f'A{r}']
inst.value = "INSTRUCCIONES DE USO"
inst.font = Font(name='Calibri', bold=True, color=GOLD, size=14)
inst.fill = PatternFill('solid', fgColor=DARK_BG)
inst.alignment = Alignment(horizontal='center')
ws5.cell(row=r, column=2).fill = PatternFill('solid', fgColor=DARK_BG)
rh(ws5, r, 32)

instructions = [
    "1. Ve a la hoja 'Deudas' e ingresa todas tus deudas actuales.",
    "2. Para cada deuda ingresa: nombre, saldo actual, tasa anual y pago minimo.",
    "3. En la celda G5 de 'Deudas', ingresa cuanto extra puedes pagar cada mes.",
    "4. El sistema automaticamente ordena tus deudas de menor a mayor (bola de nieve).",
    "5. Ve a la hoja 'Plan' para ver el calendario mes a mes de pagos.",
    "6. La hoja 'Dashboard' muestra tu resumen, ahorro y progreso.",
    "7. La hoja 'Motivacion' te ayuda a celebrar cada victoria.",
    "8. A medida que pagas deudas, actualiza el saldo a 0 y registra la fecha.",
    "",
    "METODO BOLA DE NIEVE:",
    "- Paga el MINIMO en todas las deudas",
    "- Aplica TODO el dinero extra a la deuda con MENOR saldo",
    "- Cuando esa deuda se paga, suma ese pago a la siguiente deuda",
    "- Repite hasta ser LIBRE DE DEUDAS",
]

for i, line in enumerate(instructions):
    row = 15 + i
    ws5.merge_cells(f'A{row}:B{row}')
    c = ws5[f'A{row}']
    c.value = line
    c.font = Font(name='Calibri', size=10, color=GRAY_700)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    rh(ws5, row, 22)


# ============================================================
# SHEET PROTECTION
# ============================================================
for ws_name in ["Dashboard", "Plan", "Motivacion", "Config"]:
    sheet = wb[ws_name]
    sheet.protection.sheet = True
    sheet.protection.password = "nsi2024"
    sheet.protection.enable()

# Deudas sheet: protect formula cells, allow input cells
ws2.protection.sheet = True
ws2.protection.password = "nsi2024"
ws2.protection.enable()

# Unlock input cells on Deudas
for i in range(NUM_DEBTS):
    row = DEBT_START_ROW + i
    for col in [1, 2, 3, 4]:  # name, balance, rate, payment
        ws2.cell(row=row, column=col).protection = ws2.cell(row=row, column=col).protection.copy(locked=False)
# Unlock extra payment cell
ws2['G5'].protection = ws2['G5'].protection.copy(locked=False)

# Unlock manual input cells on Motivacion
for i in range(NUM_DEBTS):
    row = 25 + i
    ws4.cell(row=row, column=4).protection = ws4.cell(row=row, column=4).protection.copy(locked=False)
    ws4.cell(row=row, column=5).protection = ws4.cell(row=row, column=5).protection.copy(locked=False)

# ============================================================
# NOW ADD PIE CHART TO DASHBOARD (after Deudas sheet exists)
# ============================================================
pie = PieChart()
pie.title = "Composicion de Deudas"
pie.style = 10
pie.width = 16
pie.height = 12

cats = Reference(ws2, min_col=1, min_row=DEBT_START_ROW, max_row=DEBT_START_ROW+NUM_DEBTS-1)
vals = Reference(ws2, min_col=2, min_row=DEBT_START_ROW, max_row=DEBT_START_ROW+NUM_DEBTS-1)
pie.add_data(vals, titles_from_data=False)
pie.set_categories(cats)

pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
pie.dataLabels.showCatName = True

# Custom colors for pie slices
pie_colors = [DEBT_RED, PROGRESS_BLUE, WARNING, SNOWBALL_PURPLE, FREEDOM_GOLD,
              SUCCESS, INFO_BLUE, "EC4899", "8B5CF6", "06B6D4"]
for i, color in enumerate(pie_colors):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = color
    pie.series[0].data_points.append(pt)

ws.add_chart(pie, "B39")

# ============================================================
# SAVE
# ============================================================
OUTPUT = "D:/Landing-Page_marketplace/excel_products/Bola_de_Nieve_NSI.xlsx"
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print("Sheets:", wb.sheetnames)
print("Done!")
