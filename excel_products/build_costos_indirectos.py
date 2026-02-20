"""
Build: Calculadora de Costos Indirectos (Bs. 69)
No Somos Ignorantes v1.0
Indirect cost allocation calculator: Dashboard, cost pools, allocation bases,
product cost buildup with full absorption costing.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# Colors
GOLD = "D4AF37"; DARK_BG = "1A1A2E"; ACCENT_BLUE = "0F3460"
WHITE = "FFFFFF"; BLACK = "000000"
LIGHT_YELLOW = "FFF9C4"; LIGHT_GREEN = "E8F5E9"
RED_CORAL = "E74C3C"; TURQUOISE = "16A085"; ORANGE = "E67E22"; PURPLE = "8E44AD"
LIGHT_GRAY = "BDC3C7"; GREEN_OK = "27AE60"; YELLOW_WARN = "F1C40F"; RED_ALERT = "E74C3C"
GRAY_300 = "D1D5DB"; GRAY_400 = "9CA3AF"; GRAY_500 = "6B7280"
GRAY_600 = "4B5563"; GRAY_700 = "374151"; GRAY_900 = "111827"

thin_border = Border(
    left=Side(style='thin', color=GRAY_300), right=Side(style='thin', color=GRAY_300),
    top=Side(style='thin', color=GRAY_300), bottom=Side(style='thin', color=GRAY_300))
gold_border = Border(
    left=Side(style='thin', color=GOLD), right=Side(style='thin', color=GOLD),
    top=Side(style='thin', color=GOLD), bottom=Side(style='thin', color=GOLD))

def style_title(cell, size=18):
    cell.font = Font(name='Calibri', bold=True, color=BLACK, size=size)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def style_header(cell, bg=DARK_BG, fg=WHITE, size=10):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

def style_input(cell):
    cell.font = Font(name='Calibri', size=11, color="0000FF")
    cell.fill = PatternFill('solid', fgColor=LIGHT_YELLOW)
    cell.border = gold_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

def style_output(cell, bold=False):
    cell.font = Font(name='Calibri', size=11, color=GRAY_900, bold=bold)
    cell.fill = PatternFill('solid', fgColor=LIGHT_GREEN)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='right', vertical='center')

def style_banner(cell, bg_color, fg=WHITE, size=10):
    cell.font = Font(name='Calibri', bold=True, color=fg, size=size)
    cell.fill = PatternFill('solid', fgColor=bg_color)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

def fill_white(ws, r1, r2, c1, c2):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=WHITE)

def unlock_cell(cell):
    cell.protection = cell.protection.copy(locked=False)

MAX_COSTS = 20  # Up to 20 indirect cost items
MAX_PRODUCTS = 15  # Up to 15 products

# ============================================================
# SHEET 1: DASHBOARD
# ============================================================
ws_dash = wb.active
ws_dash.title = "Dashboard"
ws_dash.sheet_properties.tabColor = RED_CORAL
ws_dash.sheet_view.showGridLines = False

dash_widths = {1: 3, 2: 24, 3: 18, 4: 18, 5: 3, 6: 24, 7: 18, 8: 18, 9: 3}
for c, w in dash_widths.items():
    ws_dash.column_dimensions[get_column_letter(c)].width = w
fill_white(ws_dash, 1, 55, 1, 9)

ws_dash.merge_cells('B2:H2')
c = ws_dash['B2']
style_title(c, 20)
c.value = "COSTOS INDIRECTOS - PANEL EJECUTIVO"

ws_dash.merge_cells('B3:H3')
c = ws_dash['B3']
c.font = Font(name='Calibri', size=10, color=GRAY_500)
c.value = "No Somos Ignorantes  |  v1.0  |  Asignacion de costos por producto"

ws_dash.row_dimensions[4].height = 6

# KPI 1: Total Costos Indirectos
ws_dash.merge_cells('B5:D5')
style_banner(ws_dash['B5'], RED_CORAL, WHITE, 10)
ws_dash['B5'].value = "TOTAL COSTOS INDIRECTOS (Bs.)"
ws_dash.merge_cells('B6:D6')
c = ws_dash['B6']
c.font = Font(name='Calibri', bold=True, color=BLACK, size=26)
c.alignment = Alignment(horizontal='right', vertical='center')
c.value = '=IFERROR(SUM(CostosIndirectos!D4:D23),0)'
c.number_format = '#,##0.00'

# KPI 2: Total Costos Directos
ws_dash.merge_cells('F5:H5')
style_banner(ws_dash['F5'], TURQUOISE, WHITE, 10)
ws_dash['F5'].value = "TOTAL COSTOS DIRECTOS (Bs.)"
ws_dash.merge_cells('F6:H6')
c = ws_dash['F6']
c.font = Font(name='Calibri', bold=True, color=BLACK, size=26)
c.alignment = Alignment(horizontal='right', vertical='center')
c.value = '=IFERROR(SUM(Productos!D4:D18),0)'
c.number_format = '#,##0.00'

ws_dash.row_dimensions[7].height = 4

ws_dash['B8'].value = "% Costos Indirectos:"
ws_dash['B8'].font = Font(name='Calibri', size=9, color="808080")
ws_dash.merge_cells('C8:D8')
c = ws_dash['C8']
c.value = '=IFERROR(B6/(B6+F6),0)'
c.number_format = '0.0%'
c.font = Font(name='Calibri', bold=True, size=14, color=BLACK)

ws_dash['F8'].value = "% Costos Directos:"
ws_dash['F8'].font = Font(name='Calibri', size=9, color="808080")
ws_dash.merge_cells('G8:H8')
c = ws_dash['G8']
c.value = '=IFERROR(F6/(B6+F6),0)'
c.number_format = '0.0%'
c.font = Font(name='Calibri', bold=True, size=14, color=BLACK)

ws_dash.row_dimensions[9].height = 6

# KPI 3: Tasa de Asignacion
ws_dash.merge_cells('B10:D10')
style_banner(ws_dash['B10'], ORANGE, WHITE, 10)
ws_dash['B10'].value = "TASA DE ASIGNACION"
ws_dash.merge_cells('B11:D11')
c = ws_dash['B11']
c.font = Font(name='Calibri', bold=True, color=BLACK, size=22)
c.alignment = Alignment(horizontal='right', vertical='center')
c.value = '=IFERROR(B6/SUM(BaseAsignacion!D4:D18),0)'
c.number_format = '#,##0.00'

ws_dash['B12'].value = "Bs. por unidad de base"
ws_dash['B12'].font = Font(name='Calibri', size=9, color="808080")

# KPI 4: Costo Total Promedio
ws_dash.merge_cells('F10:H10')
style_banner(ws_dash['F10'], PURPLE, WHITE, 10)
ws_dash['F10'].value = "COSTO TOTAL PROMEDIO (Bs.)"
ws_dash.merge_cells('F11:H11')
c = ws_dash['F11']
c.font = Font(name='Calibri', bold=True, color=BLACK, size=22)
c.alignment = Alignment(horizontal='right', vertical='center')
c.value = '=IFERROR(AVERAGE(Productos!H4:H18),0)'
c.number_format = '#,##0.00'

ws_dash['F12'].value = "Costo directo + indirecto por producto"
ws_dash['F12'].font = Font(name='Calibri', size=9, color="808080")

ws_dash.row_dimensions[13].height = 8

# ── Chart data: Costos por producto ──
ws_dash['B14'].value = "COSTO TOTAL POR PRODUCTO"
ws_dash['B14'].font = Font(name='Calibri', bold=True, size=11, color=GRAY_700)

ws_dash['B15'].value = "Producto"
ws_dash['C15'].value = "Costo Directo"
ws_dash['D15'].value = "Costo Indirecto"
style_header(ws_dash['B15'], bg=GRAY_700)
style_header(ws_dash['C15'], bg=GRAY_700)
style_header(ws_dash['D15'], bg=GRAY_700)

for i in range(MAX_PRODUCTS):
    r = 16 + i
    ws_dash.cell(row=r, column=2).value = f'=IFERROR(Productos!B{4+i},"")'
    ws_dash.cell(row=r, column=2).font = Font(name='Calibri', size=9, color=GRAY_700)
    ws_dash.cell(row=r, column=2).border = thin_border
    ws_dash.cell(row=r, column=3).value = f'=IFERROR(Productos!D{4+i},0)'
    ws_dash.cell(row=r, column=3).number_format = '#,##0.00'
    ws_dash.cell(row=r, column=3).border = thin_border
    ws_dash.cell(row=r, column=4).value = f'=IFERROR(Productos!G{4+i},0)'
    ws_dash.cell(row=r, column=4).number_format = '#,##0.00'
    ws_dash.cell(row=r, column=4).border = thin_border

# Stacked bar chart: Direct + Indirect costs per product
chart1 = BarChart()
chart1.type = "col"
chart1.grouping = "stacked"
chart1.style = 10
chart1.title = "Desglose de Costos por Producto"
chart1.y_axis.title = "Bs."
chart1.legend.position = 'b'

data_d = Reference(ws_dash, min_col=3, min_row=15, max_row=30)
data_i = Reference(ws_dash, min_col=4, min_row=15, max_row=30)
cats_p = Reference(ws_dash, min_col=2, min_row=16, max_row=30)

chart1.add_data(data_d, titles_from_data=True)
chart1.add_data(data_i, titles_from_data=True)
chart1.set_categories(cats_p)

chart1.series[0].graphicalProperties.solidFill = TURQUOISE
chart1.series[0].graphicalProperties.line.noFill = True
chart1.series[1].graphicalProperties.solidFill = ORANGE
chart1.series[1].graphicalProperties.line.noFill = True

chart1.width = 22
chart1.height = 13
ws_dash.add_chart(chart1, "B32")

# Pie chart: Cost pool distribution
ws_dash['F14'].value = "DISTRIBUCION DE COSTOS INDIRECTOS"
ws_dash['F14'].font = Font(name='Calibri', bold=True, size=11, color=GRAY_700)

ws_dash['F15'].value = "Concepto"
ws_dash['G15'].value = "Monto (Bs.)"
style_header(ws_dash['F15'], bg=GRAY_700)
style_header(ws_dash['G15'], bg=GRAY_700)

for i in range(8):  # Top 8 items
    r = 16 + i
    ws_dash.cell(row=r, column=6).value = f'=IFERROR(CostosIndirectos!B{4+i},"")'
    ws_dash.cell(row=r, column=6).font = Font(name='Calibri', size=9, color=GRAY_700)
    ws_dash.cell(row=r, column=6).border = thin_border
    ws_dash.cell(row=r, column=7).value = f'=IFERROR(CostosIndirectos!D{4+i},0)'
    ws_dash.cell(row=r, column=7).number_format = '#,##0.00'
    ws_dash.cell(row=r, column=7).border = thin_border

chart2 = PieChart()
chart2.title = "Pool de Costos Indirectos"
chart2.style = 10
data2 = Reference(ws_dash, min_col=7, min_row=15, max_row=23)
cats2 = Reference(ws_dash, min_col=6, min_row=16, max_row=23)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.width = 16
chart2.height = 13

pie_colors = [RED_CORAL, TURQUOISE, ORANGE, PURPLE, "3498DB", "1ABC9C", "F39C12", LIGHT_GRAY]
for idx, color in enumerate(pie_colors):
    pt = DataPoint(idx=idx)
    pt.graphicalProperties.solidFill = color
    chart2.series[0].data_points.append(pt)

chart2.dataLabels = DataLabelList()
chart2.dataLabels.showPercent = True
ws_dash.add_chart(chart2, "F32")

# ============================================================
# SHEET 2: COSTOS INDIRECTOS
# ============================================================
ws_ci = wb.create_sheet("CostosIndirectos")
ws_ci.sheet_properties.tabColor = ORANGE
ws_ci.sheet_view.showGridLines = False

ci_widths = {1: 6, 2: 35, 3: 20, 4: 18, 5: 14}
for c, w in ci_widths.items():
    ws_ci.column_dimensions[get_column_letter(c)].width = w

ws_ci.merge_cells('A1:E1')
c = ws_ci['A1']
style_title(c, 16)
c.value = "COSTOS INDIRECTOS DE FABRICACION"
c.fill = PatternFill('solid', fgColor=WHITE)
ws_ci.row_dimensions[1].height = 35

ws_ci.merge_cells('A2:E2')
c = ws_ci['A2']
c.value = "Ingresa todos los costos indirectos del periodo. Se asignaran a productos segun la base elegida."
c.font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)
c.fill = PatternFill('solid', fgColor=WHITE)

headers_ci = [("A", "#"), ("B", "CONCEPTO"), ("C", "TIPO"), ("D", "MONTO (Bs.)"), ("E", "% DEL TOTAL")]
ws_ci.row_dimensions[3].height = 28
for col_letter, label in headers_ci:
    cell = ws_ci[f'{col_letter}3']
    style_header(cell, bg=DARK_BG, fg=GOLD, size=10)
    cell.value = label

for row in range(4, 4 + MAX_COSTS):
    # #
    c = ws_ci.cell(row=row, column=1)
    c.value = f'=IF(B{row}="","",ROW()-3)'
    c.font = Font(name='Calibri', size=10, color=GRAY_500)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border

    # Concepto (INPUT)
    c = ws_ci.cell(row=row, column=2)
    style_input(c)
    c.alignment = Alignment(horizontal='left', vertical='center')
    unlock_cell(c)

    # Tipo (INPUT dropdown)
    c = ws_ci.cell(row=row, column=3)
    style_input(c)
    unlock_cell(c)

    # Monto (INPUT)
    c = ws_ci.cell(row=row, column=4)
    style_input(c)
    c.number_format = '#,##0.00'
    unlock_cell(c)

    # % del total (CALCULATED)
    c = ws_ci.cell(row=row, column=5)
    c.value = f'=IFERROR(D{row}/SUM(D$4:D$23),0)'
    c.number_format = '0.0%'
    style_output(c)

# Total row
r_ci_tot = 4 + MAX_COSTS
ws_ci.cell(row=r_ci_tot, column=2).value = "TOTAL COSTOS INDIRECTOS"
ws_ci.cell(row=r_ci_tot, column=2).font = Font(name='Calibri', bold=True, size=11, color=BLACK)
ws_ci.cell(row=r_ci_tot, column=2).border = thin_border
ws_ci.cell(row=r_ci_tot, column=4).value = f'=SUM(D4:D{r_ci_tot-1})'
ws_ci.cell(row=r_ci_tot, column=4).number_format = '#,##0.00'
style_output(ws_ci.cell(row=r_ci_tot, column=4), bold=True)

# Type dropdown
dv_type = DataValidation(type="list",
    formula1='"Alquiler,Depreciacion,Servicios,Seguros,Mantenimiento,Supervision,Otros"',
    allow_blank=True)
dv_type.prompt = "Selecciona el tipo de costo"
dv_type.promptTitle = "Tipo de Costo"
ws_ci.add_data_validation(dv_type)
dv_type.add('C4:C23')

# Data bars
ws_ci.conditional_formatting.add('D4:D23',
    DataBarRule(start_type='min', end_type='max', color=ORANGE))

# Sample data
sample_ci = [
    ("Alquiler de planta", "Alquiler", 5000),
    ("Depreciacion maquinaria", "Depreciacion", 3000),
    ("Electricidad fabrica", "Servicios", 1200),
    ("Agua industrial", "Servicios", 400),
    ("Seguro de planta", "Seguros", 800),
    ("Mantenimiento equipos", "Mantenimiento", 1500),
    ("Supervisor de produccion", "Supervision", 4000),
    ("Material de limpieza", "Otros", 300),
]
for i, (concept, tipo, amount) in enumerate(sample_ci):
    ws_ci.cell(row=4+i, column=2).value = concept
    ws_ci.cell(row=4+i, column=3).value = tipo
    ws_ci.cell(row=4+i, column=4).value = amount

# ============================================================
# SHEET 3: BASE DE ASIGNACION
# ============================================================
ws_base = wb.create_sheet("BaseAsignacion")
ws_base.sheet_properties.tabColor = TURQUOISE
ws_base.sheet_view.showGridLines = False

base_widths = {1: 6, 2: 30, 3: 22, 4: 18, 5: 16}
for c, w in base_widths.items():
    ws_base.column_dimensions[get_column_letter(c)].width = w

ws_base.merge_cells('A1:E1')
c = ws_base['A1']
style_title(c, 16)
c.value = "BASE DE ASIGNACION"
c.fill = PatternFill('solid', fgColor=WHITE)
ws_base.row_dimensions[1].height = 35

ws_base.merge_cells('A2:E2')
c = ws_base['A2']
c.value = "Define la base para distribuir costos indirectos (horas maquina, horas MOD, unidades, etc.)."
c.font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)
c.fill = PatternFill('solid', fgColor=WHITE)

# Base type selector
ws_base['A3'].value = "Base:"
ws_base['A3'].font = Font(name='Calibri', size=10, bold=True, color=GRAY_700)
ws_base['B3'].value = "Horas Maquina"
style_input(ws_base['B3'])
ws_base['B3'].alignment = Alignment(horizontal='left', vertical='center')
unlock_cell(ws_base['B3'])

dv_base_type = DataValidation(type="list",
    formula1='"Horas Maquina,Horas MOD,Unidades Producidas,Costo MOD,Otro"',
    allow_blank=True)
dv_base_type.prompt = "Selecciona la base de asignacion"
dv_base_type.promptTitle = "Base de Asignacion"
ws_base.add_data_validation(dv_base_type)
dv_base_type.add('B3')

headers_base = [("A", "#"), ("B", "PRODUCTO"), ("C", "BASE ASIGNADA"), ("D", "UNIDADES BASE"),
                ("E", "% DEL TOTAL")]
ws_base.row_dimensions[4].height = 28
# Actually use row 4 as header
# Wait, we had row 3 used for base type selector. Let's use row 5 as headers
# Fix: use row 5

# Row 5: Headers
for col_letter, label in [("A", "#"), ("B", "PRODUCTO"), ("C", "BASE ASIGNADA"), ("D", "UNIDADES BASE"),
                           ("E", "% DEL TOTAL")]:
    cell = ws_base[f'{col_letter}5']
    style_header(cell, bg=DARK_BG, fg=GOLD, size=10)
    cell.value = label

# Tasa de asignacion
ws_base['A4'].value = ""
ws_base.merge_cells('C4:E4')
c = ws_base['C4']
c.value = '=IFERROR(SUM(CostosIndirectos!D4:D23)/SUM(D6:D20),0)'
c.number_format = '"Tasa: Bs. "#,##0.00" por unidad de base"'
c.font = Font(name='Calibri', size=11, bold=True, color=ACCENT_BLUE)
c.alignment = Alignment(horizontal='center', vertical='center')

for row in range(6, 6 + MAX_PRODUCTS):
    # #
    c = ws_base.cell(row=row, column=1)
    c.value = f'=IF(B{row}="","",ROW()-5)'
    c.font = Font(name='Calibri', size=10, color=GRAY_500)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border

    # Producto (auto from Productos sheet)
    c = ws_base.cell(row=row, column=2)
    c.value = f'=IFERROR(Productos!B{row-2},"")'
    c.font = Font(name='Calibri', size=10, color=GRAY_700)
    c.border = thin_border

    # Base (auto reference)
    c = ws_base.cell(row=row, column=3)
    c.value = f'=$B$3'
    c.font = Font(name='Calibri', size=10, color=GRAY_500)
    c.border = thin_border

    # Unidades base (INPUT)
    c = ws_base.cell(row=row, column=4)
    style_input(c)
    c.number_format = '#,##0.00'
    unlock_cell(c)

    # % del total
    c = ws_base.cell(row=row, column=5)
    c.value = f'=IFERROR(D{row}/SUM(D$6:D$20),0)'
    c.number_format = '0.0%'
    style_output(c)

# Total
r_base_tot = 6 + MAX_PRODUCTS
ws_base.cell(row=r_base_tot, column=2).value = "TOTAL"
ws_base.cell(row=r_base_tot, column=2).font = Font(name='Calibri', bold=True, size=11, color=BLACK)
ws_base.cell(row=r_base_tot, column=2).border = thin_border
ws_base.cell(row=r_base_tot, column=4).value = f'=SUM(D6:D{r_base_tot-1})'
ws_base.cell(row=r_base_tot, column=4).number_format = '#,##0.00'
style_output(ws_base.cell(row=r_base_tot, column=4), bold=True)

# Sample allocation data
sample_base = [100, 150, 80, 200, 120]
for i, units in enumerate(sample_base):
    ws_base.cell(row=6+i, column=4).value = units

# ============================================================
# SHEET 4: PRODUCTOS (Product Cost Buildup)
# ============================================================
ws_prod = wb.create_sheet("Productos")
ws_prod.sheet_properties.tabColor = PURPLE
ws_prod.sheet_view.showGridLines = False

prod_widths = {1: 6, 2: 30, 3: 14, 4: 18, 5: 18, 6: 14, 7: 18, 8: 18, 9: 16}
for c, w in prod_widths.items():
    ws_prod.column_dimensions[get_column_letter(c)].width = w

ws_prod.merge_cells('A1:I1')
c = ws_prod['A1']
style_title(c, 16)
c.value = "COSTO TOTAL POR PRODUCTO"
c.fill = PatternFill('solid', fgColor=WHITE)
ws_prod.row_dimensions[1].height = 35

ws_prod.merge_cells('A2:I2')
c = ws_prod['A2']
c.value = "El costo total = costo directo + costo indirecto asignado. Ingresa materiales y MOD por producto."
c.font = Font(name='Calibri', size=10, color=GRAY_500, italic=True)
c.fill = PatternFill('solid', fgColor=WHITE)

headers_prod = [
    ("A", "#"), ("B", "PRODUCTO"), ("C", "UNIDADES"),
    ("D", "COSTO DIRECTO (Bs.)"), ("E", "COSTO DIRECTO UNIT."),
    ("F", "BASE ASIGN."), ("G", "COSTO INDIRECTO ASIGN."),
    ("H", "COSTO TOTAL (Bs.)"), ("I", "COSTO UNIT. TOTAL")
]
ws_prod.row_dimensions[3].height = 30
for col_letter, label in headers_prod:
    cell = ws_prod[f'{col_letter}3']
    style_header(cell, bg=DARK_BG, fg=GOLD, size=9)
    cell.value = label

for row in range(4, 4 + MAX_PRODUCTS):
    i = row - 4
    # #
    c = ws_prod.cell(row=row, column=1)
    c.value = f'=IF(B{row}="","",ROW()-3)'
    c.font = Font(name='Calibri', size=10, color=GRAY_500)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border

    # Producto (INPUT)
    c = ws_prod.cell(row=row, column=2)
    style_input(c)
    c.alignment = Alignment(horizontal='left', vertical='center')
    unlock_cell(c)

    # Unidades (INPUT)
    c = ws_prod.cell(row=row, column=3)
    style_input(c)
    c.number_format = '#,##0'
    unlock_cell(c)

    # Costo Directo total (INPUT)
    c = ws_prod.cell(row=row, column=4)
    style_input(c)
    c.number_format = '#,##0.00'
    unlock_cell(c)

    # Costo Directo Unitario (CALCULATED)
    c = ws_prod.cell(row=row, column=5)
    c.value = f'=IFERROR(D{row}/C{row},0)'
    c.number_format = '#,##0.00'
    style_output(c)

    # Base asignacion (from BaseAsignacion sheet)
    c = ws_prod.cell(row=row, column=6)
    c.value = f'=IFERROR(BaseAsignacion!D{row+2},0)'
    c.number_format = '#,##0.00'
    c.font = Font(name='Calibri', size=10, color=GRAY_500)
    c.border = thin_border

    # Costo Indirecto asignado (CALCULATED)
    c = ws_prod.cell(row=row, column=7)
    c.value = f'=IFERROR(F{row}*SUM(CostosIndirectos!D4:D23)/SUM(BaseAsignacion!D6:D20),0)'
    c.number_format = '#,##0.00'
    style_output(c)

    # Costo Total
    c = ws_prod.cell(row=row, column=8)
    c.value = f'=D{row}+G{row}'
    c.number_format = '#,##0.00'
    style_output(c, bold=True)

    # Costo Unitario Total
    c = ws_prod.cell(row=row, column=9)
    c.value = f'=IFERROR(H{row}/C{row},0)'
    c.number_format = '#,##0.00'
    style_output(c, bold=True)

# Total row
r_prod_tot = 4 + MAX_PRODUCTS
ws_prod.cell(row=r_prod_tot, column=2).value = "TOTAL"
ws_prod.cell(row=r_prod_tot, column=2).font = Font(name='Calibri', bold=True, size=11, color=BLACK)
ws_prod.cell(row=r_prod_tot, column=2).border = thin_border
for col in [3, 4, 7, 8]:
    c = ws_prod.cell(row=r_prod_tot, column=col)
    c.value = f'=SUM({get_column_letter(col)}4:{get_column_letter(col)}{r_prod_tot-1})'
    c.number_format = '#,##0.00'
    style_output(c, bold=True)

# Conditional formatting: data bars on cost
ws_prod.conditional_formatting.add(f'H4:H{r_prod_tot-1}',
    DataBarRule(start_type='min', end_type='max', color=PURPLE))

# Sample product data
sample_prods = [
    ("Producto A - Basico", 500, 15000),
    ("Producto B - Intermedio", 300, 12000),
    ("Producto C - Premium", 200, 20000),
    ("Producto D - Especial", 400, 18000),
    ("Producto E - Economico", 600, 10000),
]
for i, (name, units, cost) in enumerate(sample_prods):
    ws_prod.cell(row=4+i, column=2).value = name
    ws_prod.cell(row=4+i, column=3).value = units
    ws_prod.cell(row=4+i, column=4).value = cost

# ============================================================
# SHEET 5: CONFIG
# ============================================================
ws_conf = wb.create_sheet("Config")
ws_conf.sheet_properties.tabColor = GRAY_500
ws_conf.sheet_view.showGridLines = False
conf_widths = {1: 25, 2: 45}
for c, w in conf_widths.items():
    ws_conf.column_dimensions[get_column_letter(c)].width = w
fill_white(ws_conf, 1, 25, 1, 3)

ws_conf['A1'].value = "v1.0.0"
ws_conf['A1'].font = Font(name='Calibri', size=9, color=GRAY_400)
ws_conf.merge_cells('A3:B3')
style_title(ws_conf['A3'], 14)
ws_conf['A3'].value = "CONFIGURACION"

settings = [
    ("Producto", "Calculadora de Costos Indirectos"),
    ("Version", "v1.0.0"),
    ("Marca", "No Somos Ignorantes"),
    ("Precio", "Bs. 69"),
    ("Proteccion", "nsi2024"),
    ("", ""),
    ("INSTRUCCIONES", ""),
    ("1.", "En 'CostosIndirectos', registra todos los costos indirectos del periodo."),
    ("2.", "En 'BaseAsignacion', define las unidades base por producto (horas maquina, MOD, etc.)."),
    ("3.", "En 'Productos', ingresa cada producto con su costo directo y unidades."),
    ("4.", "El sistema calcula automaticamente la tasa de asignacion y el costo total."),
    ("5.", "El Dashboard muestra KPIs y graficos de desglose de costos."),
    ("6.", "Para desbloquear: Revisar -> Desproteger hoja -> nsi2024"),
]
for i, (label, value) in enumerate(settings):
    r = 5 + i
    ws_conf.cell(row=r, column=1).value = label
    ws_conf.cell(row=r, column=1).font = Font(name='Calibri', size=10, bold=True, color=GRAY_700)
    ws_conf.cell(row=r, column=2).value = value
    ws_conf.cell(row=r, column=2).font = Font(name='Calibri', size=10, color=GRAY_600)

# ============================================================
# PROTECTION
# ============================================================
for ws in wb.worksheets:
    ws.protection.sheet = True
    ws.protection.password = "nsi2024"
    ws.protection.enable()

# Unlock inputs
for row in range(4, 4 + MAX_COSTS):
    for col in [2, 3, 4]:
        unlock_cell(ws_ci.cell(row=row, column=col))

unlock_cell(ws_base['B3'])
for row in range(6, 6 + MAX_PRODUCTS):
    unlock_cell(ws_base.cell(row=row, column=4))

for row in range(4, 4 + MAX_PRODUCTS):
    for col in [2, 3, 4]:
        unlock_cell(ws_prod.cell(row=row, column=col))

# ============================================================
# SAVE & VERIFY
# ============================================================
output_path = r"D:\Landing-Page_marketplace\excel_products\Calculadora_Costos_Indirectos_NSI.xlsx"
wb.save(output_path)
print(f"[OK] Saved: {output_path}")

from openpyxl import load_workbook
wb2 = load_workbook(output_path)
formula_count = 0
for ws in wb2.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formula_count += 1

print(f"[OK] Sheets: {wb2.sheetnames}")
print(f"[OK] Total formulas: {formula_count}")
print(f"[OK] Charts: Dashboard={len(ws_dash._charts)}")
print(f"[OK] Max cost items: {MAX_COSTS}, Max products: {MAX_PRODUCTS}")
